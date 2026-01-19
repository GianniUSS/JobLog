from __future__ import annotations

import atexit
import base64
import csv
import hashlib
import io
import json
import logging
import os
import random
import secrets
import sqlite3
import time
import re
from decimal import Decimal
from copy import deepcopy
from datetime import date, datetime, timedelta, timezone
from functools import wraps
from pathlib import Path


def format_time_value(value) -> Optional[str]:
    """Formatta un valore tempo (timedelta, datetime.time o stringa) nel formato HH:MM."""
    if value is None:
        return None
    if isinstance(value, timedelta):
        # MySQL TIME restituisce timedelta
        total_seconds = int(value.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return f"{hours:02d}:{minutes:02d}"
    if hasattr(value, 'strftime'):
        # datetime.time object
        return value.strftime("%H:%M")
    # Stringa - assicuriamo formato HH:MM
    s = str(value)
    if len(s) >= 5:
        # Potrebbe essere "6:00:00" o "06:00:00" o "06:00"
        parts = s.split(':')
        if len(parts) >= 2:
            return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
    return s[:5] if len(s) >= 5 else s


# Cache per geocoding (evita richieste ripetute a Nominatim)
_geocode_cache: Dict[str, Optional[Tuple[float, float]]] = {}
_geocode_last_request = 0.0  # Timestamp ultima richiesta (rate limiting)


def geocode_address(address: str) -> Optional[Tuple[float, float]]:
    """
    Converte un indirizzo in coordinate GPS usando Nominatim (OpenStreetMap).
    Ritorna (latitude, longitude) o None se non trovato.
    NON fa fallback a indirizzi approssimati per evitare errori di timbratura.
    """
    global _geocode_last_request
    
    if not address or not address.strip():
        return None
    
    # Normalizza l'indirizzo per la cache
    cache_key = address.strip().lower()
    if cache_key in _geocode_cache:
        return _geocode_cache[cache_key]
    
    import time as time_module
    import urllib.request
    import urllib.parse
    import urllib.error
    
    # Rate limiting: aspetta almeno 1 secondo tra le richieste
    elapsed = time_module.time() - _geocode_last_request
    if elapsed < 1.0:
        time_module.sleep(1.0 - elapsed)
    
    try:
        params = urllib.parse.urlencode({
            'q': address,
            'format': 'json',
            'limit': 1,
            'addressdetails': 0,
            'countrycodes': 'it'
        })
        url = f"https://nominatim.openstreetmap.org/search?{params}"
        
        req = urllib.request.Request(url, headers={
            'User-Agent': 'JobLogApp/1.0 (geocoding for work shifts)'
        })
        
        _geocode_last_request = time_module.time()
        
        with urllib.request.urlopen(req, timeout=10) as response:
            data = json.loads(response.read().decode('utf-8'))
            
            if data and len(data) > 0:
                lat = float(data[0]['lat'])
                lon = float(data[0]['lon'])
                result = (lat, lon)
                _geocode_cache[cache_key] = result
                logging.getLogger(__name__).info(f"Geocoding OK: '{address}' -> {result}")
                return result
    except Exception as e:
        logging.getLogger(__name__).warning(f"Geocoding fallito per '{address}': {e}")
    
    _geocode_cache[cache_key] = None
    logging.getLogger(__name__).warning(f"Geocoding: nessun risultato per '{address}'")
    return None


from threading import Event, Thread
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Set, Tuple, TypeAlias, cast

try:
    import pymysql  # type: ignore[import]
    from pymysql import err as pymysql_err  # type: ignore[import]
    from pymysql.cursors import DictCursor  # type: ignore[import]
except ImportError:  # pragma: no cover - fallback when MySQL client not installed
    pymysql = None
    pymysql_err = None
    DictCursor = None

from flask import Flask, abort, flash, g, jsonify, redirect, render_template, request, send_file, send_from_directory, session, url_for
from flask_session import Session
from flask.typing import ResponseReturnValue
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pywebpush import WebPushException, webpush
import qrcode
from rentman_client import (
    RentmanAPIError,
    RentmanAuthError,
    RentmanClient,
    RentmanError,
    RentmanNotFound,
    MAX_LIMIT,
)


SECRET_FILE = Path(__file__).with_name('.flask_secret')


def _load_or_create_secret_key() -> str:
    override = os.environ.get('FLASK_SECRET_KEY')
    if override:
        return override
    if SECRET_FILE.exists():
        try:
            value = SECRET_FILE.read_text(encoding='utf-8').strip()
            if value:
                return value
        except OSError:
            pass
    generated = secrets.token_hex(32)
    try:
        SECRET_FILE.write_text(generated, encoding='utf-8')
    except OSError:
        logging.getLogger(__name__).warning("Impossibile salvare la chiave segreta su %s", SECRET_FILE)
    return generated


app = Flask(__name__)
app.secret_key = _load_or_create_secret_key()
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB max upload

PERSISTENT_SESSION_COOKIE_NAME = os.environ.get('JOBLOG_PERSISTENT_COOKIE_NAME', 'joblog_auth')
PERSISTENT_SESSION_MAX_AGE = int(os.environ.get('JOBLOG_PERSISTENT_SESSION_MAX_AGE', str(30 * 86400)))
app.config['PERSISTENT_SESSION_COOKIE_NAME'] = PERSISTENT_SESSION_COOKIE_NAME
app.config['PERSISTENT_SESSION_MAX_AGE'] = PERSISTENT_SESSION_MAX_AGE

SESSION_STORAGE = Path(__file__).with_name('.flask_session')
SESSION_STORAGE.mkdir(parents=True, exist_ok=True)
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = str(SESSION_STORAGE)
app.config['SESSION_FILE_THRESHOLD'] = 1024
app.config['SESSION_PERMANENT'] = True
session_manager = Session(app)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
    force=True,
)
app.logger.handlers = logging.getLogger().handlers
app.logger.setLevel(logging.INFO)
app.logger.propagate = True
DATABASE = Path(__file__).with_name("joblog.db")
PROJECTS_FILE = Path(__file__).with_name("projects.json")
CONFIG_FILE = Path(__file__).with_name("config.json")
USERS_FILE = Path(__file__).with_name("users.json")
DEMO_PROJECT_CODE = "1001"

# ═══════════════════════════════════════════════════════════════════
# DATE SIMULATION FOR TESTING
# ═══════════════════════════════════════════════════════════════════
# Per simulare una data diversa, imposta SIMULATED_DATE nel formato "YYYY-MM-DD"
# Es: SIMULATED_DATE = "2026-01-10"
# Per usare la data reale, lascia None
SIMULATED_DATE: Optional[str] = None


def get_simulated_now() -> datetime:
    """
    Ritorna datetime.now() o la data simulata se impostata.
    Usa questa funzione invece di datetime.now() per testare date diverse.
    """
    if SIMULATED_DATE:
        try:
            simulated = datetime.strptime(SIMULATED_DATE, "%Y-%m-%d")
            # Mantieni l'ora corrente ma con la data simulata
            now = datetime.now()
            return simulated.replace(hour=now.hour, minute=now.minute, second=now.second, microsecond=now.microsecond)
        except ValueError:
            app.logger.warning(f"SIMULATED_DATE '{SIMULATED_DATE}' non valida, uso data reale")
    return datetime.now()


def get_simulated_today() -> date:
    """Ritorna la data di oggi o la data simulata se impostata."""
    return get_simulated_now().date()


def static_version(filename: str) -> str:
    """Return cache-busted static URL using file mtime as version."""
    safe_path = filename.lstrip("/")
    static_folder = app.static_folder or os.path.join(app.root_path, "static")
    file_path = os.path.join(static_folder, safe_path)
    try:
        version = int(os.path.getmtime(file_path))
    except (OSError, TypeError):
        version = int(time.time())
    return url_for("static", filename=safe_path, v=version)


app.jinja_env.globals["static_version"] = static_version


@app.context_processor
def inject_company_logo():
    """Inietta il logo aziendale in tutti i template."""
    logo_path = None
    try:
        db = get_db()
        cursor = db.execute("SELECT logo_path FROM company_settings WHERE id = 1")
        row = cursor.fetchone()
        if row:
            if isinstance(row, dict):
                logo_path = row.get('logo_path')
            else:
                logo_path = row[0]
    except Exception:
        pass
    return {'company_logo': logo_path}


MOCK_PROJECTS: Dict[str, Dict[str, Any]] = {
    "1001": {
        "project_code": "1001",
        "project_name": "Allestimento Stand Futuro",
        "activities": [
            {"id": "PREP", "label": "Preparazione materiali"},
            {"id": "MONT", "label": "Montaggio struttura"},
            {"id": "LGT", "label": "Illuminazione e cablaggi"},
            {"id": "FIN", "label": "Finiture e collaudo"},
        ],
        "team": [
            {"key": "anna", "name": "Anna Rossi"},
            {"key": "luca", "name": "Luca Bianchi"},
            {"key": "mario", "name": "Mario Verdi"},
            {"key": "giulia", "name": "Giulia Neri"},
        ],
    },
    "1002": {
        "project_code": "1002",
        "project_name": "Evento Corporate Milano",
        "activities": [
            {"id": "LOAD", "label": "Carico materiali"},
            {"id": "STAGE", "label": "Montaggio palco"},
            {"id": "AUDIO", "label": "Audio & video"},
            {"id": "DECOR", "label": "Decorazioni"},
        ],
        "team": [
            {"key": "federico", "name": "Federico Cattaneo"},
            {"key": "ilaria", "name": "Ilaria Riva"},
            {"key": "paolo", "name": "Paolo Gallo"},
            {"key": "roberta", "name": "Roberta Colombo"},
        ],
    },
}


@app.get("/sw.js")
def service_worker() -> ResponseReturnValue:
    response = app.send_static_file("sw.js")
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response


class RowMapping(dict):
    """Row helper that mimics sqlite3.Row access for dict-based cursors."""

    __slots__ = ("_ordered",)

    def __init__(self, data: Dict[str, Any], columns: Sequence[str]):
        super().__init__(data)
        self._ordered = [data.get(col) for col in columns]

    def __getitem__(self, key: Any) -> Any:  # type: ignore[override]
        if isinstance(key, int):
            return self._ordered[key]
        return super().__getitem__(key)


class CursorWrapper:
    """Minimal cursor wrapper to align PyMySQL cursor behaviour with sqlite3."""

    __slots__ = ("_cursor", "_columns", "_closed")

    def __init__(self, cursor):
        self._cursor = cursor
        self._columns = [col[0] for col in cursor.description] if cursor.description else []
        self._closed = False

    @property
    def description(self):
        return self._cursor.description

    @property
    def lastrowid(self):
        return self._cursor.lastrowid

    def fetchone(self):
        row = self._cursor.fetchone()
        if row is None:
            return None
        if isinstance(row, dict):
            return RowMapping(row, self._columns)
        return row

    def fetchall(self):
        rows = self._cursor.fetchall()
        try:
            if rows and isinstance(rows[0], dict):
                mapped = [RowMapping(row, self._columns) for row in rows]
            else:
                mapped = rows
        finally:
            self.close()
        return mapped

    def __iter__(self):
        for row in self._cursor:
            if isinstance(row, dict):
                yield RowMapping(row, self._columns)
            else:
                yield row
        self.close()

    def close(self):
        if not self._closed:
            try:
                self._cursor.close()
            finally:
                self._closed = True

    def __del__(self):  # pragma: no cover - best effort cleanup
        self.close()


class MySQLConnection:
    """Adapter to expose a sqlite-like interface backed by PyMySQL."""

    def __init__(self, settings: Dict[str, Any]):
        if pymysql is None or DictCursor is None:
              raise RuntimeError("PyMySQL non è installato. Esegui 'pip install PyMySQL' per usare il backend MySQL.")
        self._settings = settings
        self._conn = self._connect_with_autocreate()

    # Internal helpers -------------------------------------------------
    def _base_connect(self, include_db: bool = True):
        kwargs = {
            "host": self._settings["host"],
            "port": int(self._settings.get("port", 3306)),
            "user": self._settings["user"],
            "password": self._settings["password"],
            "charset": "utf8mb4",
            "cursorclass": DictCursor,
            "autocommit": False,
        }
        if include_db:
            kwargs["database"] = self._settings["name"]
        return pymysql.connect(**kwargs)  # type: ignore[union-attr]

    def _connect_with_autocreate(self):
        try:
                return self._base_connect(include_db=True)
        except Exception as exc:  # pragma: no cover - MySQL specific bootstrap
            if pymysql_err is not None and isinstance(exc, pymysql_err.OperationalError):
                err_code = exc.args[0] if exc.args else None
                if err_code == 1049:  # Unknown database
                    bootstrap = self._base_connect(include_db=False)
                    try:
                        with bootstrap.cursor() as cursor:
                            cursor.execute(
                                f"CREATE DATABASE IF NOT EXISTS `{self._settings['name']}` "
                                "DEFAULT CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"
                            )
                        bootstrap.commit()
                    finally:
                        bootstrap.close()
                    return self._base_connect(include_db=True)
            raise

    @staticmethod
    def _prepare_sql(sql: str) -> str:
        if "%s" in sql or "%(" in sql:
            return sql
        return sql.replace("?", "%s")

    @staticmethod
    def _prepare_params(params: Any) -> tuple:
        if params is None:
            return ()
        if isinstance(params, (list, tuple)):
            return tuple(params)
        return (params,)

    # Public API -------------------------------------------------------
    def execute(self, sql: str, params: Any = None) -> CursorWrapper:
        cursor = self._conn.cursor()
        cursor.execute(self._prepare_sql(sql), self._prepare_params(params))
        return CursorWrapper(cursor)

    def executemany(self, sql: str, seq_of_params: Iterable[Iterable[Any]]) -> CursorWrapper:
        cursor = self._conn.cursor()
        prepared_sql = self._prepare_sql(sql)
        prepared_params = [self._prepare_params(params) for params in seq_of_params]
        if prepared_params:
            cursor.executemany(prepared_sql, prepared_params)
        return CursorWrapper(cursor)

    def executescript(self, script: str) -> None:
        for statement in script.split(";"):
            stmt = statement.strip()
            if stmt:
                self.execute(stmt)

    def commit(self) -> None:
        self._conn.commit()

    def rollback(self) -> None:
        self._conn.rollback()

    def close(self) -> None:
        self._conn.close()


DatabaseLike: TypeAlias = sqlite3.Connection | MySQLConnection

_RENTMAN_CLIENT: Optional[RentmanClient] = None
_RENTMAN_CLIENT_TOKEN: Optional[str] = None
_CONFIG_CACHE: Optional[Dict[str, Any]] = None
_CONFIG_CACHE_MTIME: Optional[float] = None
_WEBPUSH_SETTINGS: Optional[Dict[str, Optional[str]]] = None
_NOTIFICATION_THREAD: Optional[Thread] = None
_NOTIFICATION_STOP: Optional[Event] = None
_CEDOLINO_RETRY_THREAD: Optional[Thread] = None
_CEDOLINO_RETRY_STOP: Optional[Event] = None

NOTIFICATION_INTERVAL_SECONDS = int(os.environ.get("JOBLOG_NOTIFICATION_INTERVAL", "60"))
CEDOLINO_RETRY_INTERVAL_SECONDS = int(os.environ.get("JOBLOG_CEDOLINO_RETRY_INTERVAL", "300"))  # 5 minuti
ACTIVITY_OVERDUE_GRACE_MS = 10 * 60 * 1000  # 10 minuti di ritardo tollerato
PUSH_NOTIFIED_STATE_KEY = "push_notified_activities"
LONG_RUNNING_STATE_KEY = "long_running_member_notifications"
LONG_RUNNING_THRESHOLD_MS = 2 * 60 * 1000  # 2 minuti
OVERDUE_PUSH_TTL_SECONDS = max(300, int(os.environ.get("JOBLOG_OVERDUE_PUSH_TTL", "3600")))

RUN_STATE_PAUSED = 0
RUN_STATE_RUNNING = 1
RUN_STATE_FINISHED = 2
VALID_RUN_STATES = {RUN_STATE_PAUSED, RUN_STATE_RUNNING, RUN_STATE_FINISHED}

ROLE_USER = "user"
ROLE_SUPERVISOR = "supervisor"
ROLE_ADMIN = "admin"
ROLE_MAGAZZINO = "magazzino"
VALID_USER_ROLES = {ROLE_USER, ROLE_SUPERVISOR, ROLE_ADMIN, ROLE_MAGAZZINO}


# Permission check helpers
def is_admin_or_supervisor() -> bool:
    """Check if current user is admin or supervisor (responsabile squadra)."""
    return bool(session.get("is_admin") or session.get("is_supervisor"))


def is_admin_only() -> bool:
    """Check if current user is admin only."""
    return bool(session.get("is_admin"))


# Authentication helpers
def hash_password(password: str) -> str:
    """Hash a password using SHA-256."""
    return hashlib.sha256(password.encode('utf-8')).hexdigest()


def verify_password(password: str, hashed: str) -> bool:
    """Verify a password against a hash."""
    return hash_password(password) == hashed


def compute_initials(value: str) -> str:
    """Return up to two initials from the provided value."""
    if not value:
        return "?"
    cleaned = value.replace("-", " ")
    tokens = [part for part in cleaned.split() if part]
    if not tokens:
        return value[:2].upper()
    initials = "".join(part[0] for part in tokens[:2]).upper()
    if initials:
        return initials
    return value[:2].upper()


def load_users_file() -> Dict[str, Dict[str, Any]]:
    """Return the legacy users.json payload (if present) for migrations."""
    if not USERS_FILE.exists():
        return {}
    try:
        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            payload = json.load(f)
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _normalize_username(value: str) -> str:
    return value.strip().lower()


def _normalize_role(value: Optional[str]) -> str:
    if not value:
        return ROLE_USER
    candidate = value.strip().lower()
    if candidate in VALID_USER_ROLES:
        return candidate
    return ROLE_USER


def _role_from_legacy_entry(entry: Mapping[str, Any]) -> str:
    roles_field = entry.get("roles")
    candidates: List[str] = []
    if isinstance(roles_field, str):
        candidates.append(roles_field)
    elif isinstance(roles_field, Iterable):
        for role in roles_field:
            if isinstance(role, str):
                candidates.append(role)

    for role in candidates:
        normalized = role.strip().lower()
        if normalized == ROLE_ADMIN:
            return ROLE_ADMIN
        if normalized == ROLE_SUPERVISOR:
            return ROLE_SUPERVISOR
        if normalized == ROLE_MAGAZZINO:
            return ROLE_MAGAZZINO

    is_admin_flag = entry.get("is_admin")
    if isinstance(is_admin_flag, bool) and is_admin_flag:
        return ROLE_ADMIN
    if isinstance(is_admin_flag, str) and is_admin_flag.strip().lower() in {"1", "true", "yes", "y"}:
        return ROLE_ADMIN
    return ROLE_USER


def fetch_user_record(db: DatabaseLike, username: str) -> Optional[Mapping[str, Any]]:
    if not username:
        return None
    return db.execute(
        """
        SELECT username, password_hash, display_name, full_name, role, is_active,
               current_project_code, current_project_name
        FROM app_users
        WHERE LOWER(username)=LOWER(?)
        """,
        (username,),
    ).fetchone()


def _persistent_cookie_name() -> str:
    return app.config.get('PERSISTENT_SESSION_COOKIE_NAME', PERSISTENT_SESSION_COOKIE_NAME)


def _persistent_session_max_age() -> int:
    value = app.config.get('PERSISTENT_SESSION_MAX_AGE', PERSISTENT_SESSION_MAX_AGE)
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return PERSISTENT_SESSION_MAX_AGE
    return max(300, parsed)


def _hash_persistent_token(token: str) -> str:
    return hashlib.sha256(token.encode('utf-8')).hexdigest()


def _request_metadata() -> Tuple[Optional[str], Optional[str]]:
    user_agent = request.headers.get('User-Agent', '').strip()
    forwarded_for = request.headers.get('X-Forwarded-For', '').strip()
    if forwarded_for:
        ip_candidate = forwarded_for.split(',')[0].strip()
    else:
        ip_candidate = request.remote_addr or ''
    user_agent = user_agent or None
    ip_candidate = ip_candidate or None
    return user_agent, ip_candidate


def _store_persistent_session(username: str) -> Tuple[str, int]:
    token = secrets.token_urlsafe(48)
    token_hash = _hash_persistent_token(token)
    now = now_ms()
    max_age_ms = _persistent_session_max_age() * 1000
    expires_ts = now + max_age_ms
    user_agent, ip_addr = _request_metadata()
    db = get_db()
    db.execute(
        """
        INSERT INTO persistent_sessions(token_hash, username, user_agent, ip_address, created_ts, last_seen_ts, expires_ts)
        VALUES(?,?,?,?,?,?,?)
        """,
        (token_hash, username, user_agent, ip_addr, now, now, expires_ts),
    )
    try:
        db.commit()
    except Exception:
        pass
    return token, expires_ts


def _delete_persistent_session(token_value: Optional[str]) -> None:
    if not token_value:
        return
    token_hash = _hash_persistent_token(token_value)
    db = get_db()
    db.execute("DELETE FROM persistent_sessions WHERE token_hash=?", (token_hash,))
    try:
        db.commit()
    except Exception:
        pass


def _load_persistent_session(token_value: str) -> Optional[Tuple[str, int]]:
    if not token_value:
        return None
    token_hash = _hash_persistent_token(token_value)
    db = get_db()
    row = db.execute(
        "SELECT username, expires_ts FROM persistent_sessions WHERE token_hash=?",
        (token_hash,),
    ).fetchone()
    if not row:
        return None
    username = row.get('username') if isinstance(row, Mapping) else row[0]
    expires_raw = row.get('expires_ts') if isinstance(row, Mapping) else row[1]
    expires_ts = _coerce_int(expires_raw) or 0
    now = now_ms()
    if expires_ts <= now:
        db.execute("DELETE FROM persistent_sessions WHERE token_hash=?", (token_hash,))
        try:
            db.commit()
        except Exception:
            pass
        return None
    new_expires = now + _persistent_session_max_age() * 1000
    db.execute(
        "UPDATE persistent_sessions SET last_seen_ts=?, expires_ts=? WHERE token_hash=?",
        (now, new_expires, token_hash),
    )
    try:
        db.commit()
    except Exception:
        pass
    return username, new_expires


def _set_persistent_cookie(response, token_value: str, expires_ts: Optional[int] = None) -> None:
    max_age = _persistent_session_max_age()
    expires = expires_ts or (now_ms() + max_age * 1000)
    expires_dt = datetime.fromtimestamp(expires / 1000, tz=timezone.utc)
    response.set_cookie(
        _persistent_cookie_name(),
        token_value,
        max_age=max_age,
        expires=expires_dt,
        httponly=True,
        secure=app.config.get('SESSION_COOKIE_SECURE', False),
        samesite=app.config.get('SESSION_COOKIE_SAMESITE', 'Lax'),
        path='/',
    )


def _apply_user_session(user_row: Mapping[str, Any]) -> None:
    username = user_row.get('username') or ''
    display = user_row.get('display_name') or username
    full_name = user_row.get('full_name') or display
    role = _normalize_role(user_row.get('role'))
    session.permanent = True
    session['user'] = username
    session['user_display'] = display
    session['user_name'] = full_name
    session['user_initials'] = compute_initials(full_name)
    session['user_role'] = role
    session['is_admin'] = role == ROLE_ADMIN
    session['is_supervisor'] = role in {ROLE_SUPERVISOR, ROLE_ADMIN}
    session['is_magazzino'] = role in {ROLE_MAGAZZINO, ROLE_ADMIN}
    
    # Ripristina il progetto salvato per i supervisor
    if role in {ROLE_SUPERVISOR, ROLE_ADMIN}:
        current_project_code = user_row.get('current_project_code')
        current_project_name = user_row.get('current_project_name')
        app.logger.info("Login %s: recupero progetto salvato = %s", username, current_project_code)
        if current_project_code:
            session['supervisor_project_code'] = current_project_code
            session['supervisor_project_name'] = current_project_name or current_project_code


def _magazzino_only() -> Optional[ResponseReturnValue]:
    """Verifica accesso al magazzino: admin, magazzinieri, o utenti se modulo abilitato."""
    role = session.get('user_role')
    
    # Admin e magazzinieri hanno sempre accesso
    if role in {ROLE_MAGAZZINO, ROLE_ADMIN}:
        return None
    
    # Per altri utenti, verifica se il modulo magazzino è abilitato
    db = get_db()
    if is_module_enabled(db, "magazzino"):
        return None  # Accesso consentito
    
    # Accesso negato
    if request.path.startswith('/api/'):
        return jsonify({"error": "forbidden"}), 403
    return ("Forbidden", 403)


@app.before_request
def _restore_persistent_session() -> None:
    if 'user' in session:
        return
    token = request.cookies.get(_persistent_cookie_name())
    if not token:
        return
    loaded = _load_persistent_session(token)
    if not loaded:
        return
    username, new_expires = loaded
    db = get_db()
    user_row = fetch_user_record(db, username)
    if not user_row or not user_row.get('is_active'):
        _delete_persistent_session(token)
        return
    _apply_user_session(user_row)
    g.persistent_cookie_refresh = (token, new_expires)


@app.after_request
def _refresh_persistent_cookie(response):
    pending = g.pop('persistent_cookie_refresh', None)
    if pending:
        token_value, expires_ts = pending
        _set_persistent_cookie(response, token_value, expires_ts)
    return response


def migrate_users_file(db: DatabaseLike) -> bool:
    legacy = load_users_file()
    if not legacy:
        return False

    rows: List[tuple] = []
    now = now_ms()
    for raw_username, entry in legacy.items():
        if not isinstance(entry, Mapping):
            continue
        username = _normalize_username(str(raw_username))
        password_hash = entry.get("password") or entry.get("password_hash")
        if not username or not isinstance(password_hash, str) or not password_hash.strip():
            continue
        display = str(entry.get("display") or raw_username).strip() or username
        full_name = str(entry.get("name") or display).strip() or display
        role = _role_from_legacy_entry(entry)
        rows.append((username, password_hash.strip(), display, full_name, role, 1, now, now))

    if not rows:
        return False

    if DB_VENDOR == "mysql":
        insert_sql = (
            "INSERT INTO app_users("
            "username, password_hash, display_name, full_name, role, is_active, created_ts, updated_ts"
            ") VALUES(?,?,?,?,?,?,?,?) "
            "ON DUPLICATE KEY UPDATE username=VALUES(username)"
        )
    else:
        insert_sql = (
            "INSERT INTO app_users("
            "username, password_hash, display_name, full_name, role, is_active, created_ts, updated_ts"
            ") VALUES(?,?,?,?,?,?,?,?) "
            "ON CONFLICT(username) DO NOTHING"
        )

    db.executemany(insert_sql, rows)
    app.logger.info("Importati %s utenti da %s", len(rows), USERS_FILE.name)
    return True


def bootstrap_user_store(db: DatabaseLike) -> None:
    try:
        ensure_app_users_table(db)
    except Exception:
        app.logger.exception("Impossibile preparare la tabella app_users")
        return

    try:
        row = db.execute("SELECT COUNT(*) AS total FROM app_users").fetchone()
        existing = int(row["total"] if row else 0)
    except Exception:
        app.logger.exception("Impossibile verificare il conteggio degli utenti")
        return

    if existing > 0:
        return

    try:
        migrated = migrate_users_file(db)
    except Exception:
        app.logger.exception("Migrazione utenti legacy fallita")
        migrated = False

    if migrated:
        return

    bootstrap_user = os.environ.get("JOBLOG_BOOTSTRAP_ADMIN_USER", "admin")
    bootstrap_password = os.environ.get("JOBLOG_BOOTSTRAP_ADMIN_PASSWORD")
    if not bootstrap_password:
        app.logger.warning(
            "Nessun utente configurato: crea un account con 'python manage_users.py create <utente> --role admin'"
        )
        return

    now = now_ms()
    username = _normalize_username(bootstrap_user)
    db.execute(
        """
        INSERT INTO app_users(username, password_hash, display_name, full_name, role, is_active, created_ts, updated_ts)
        VALUES(?,?,?,?,?,?,?,?)
        """,
        (
            username,
            hash_password(bootstrap_password),
            bootstrap_user,
            bootstrap_user,
            ROLE_ADMIN,
            1,
            now,
            now,
        ),
    )
    app.logger.warning(
        "Creato automaticamente l'utente amministratore '%s' dalle variabili JOBLOG_BOOTSTRAP_ADMIN_*",
        bootstrap_user,
    )


def login_required(f):
    """Decorator to require login for a route."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Authentication required'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def _is_truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "t", "yes", "y", "si"}
    if isinstance(value, (int, float)):
        return value != 0
    return False




def _normalize_datetime(value: Any) -> Optional[str]:
    """Normalizza un valore data/ora in formato ISO 8601 (se possibile)."""

    if value is None:
        return None

    if isinstance(value, (int, float)):
        try:
            return datetime.fromtimestamp(float(value), tz=timezone.utc).isoformat()
        except (ValueError, OSError):
            return None

    if isinstance(value, str):
        slug = value.strip()
        if not slug:
            return None
        normalized = slug.replace("Z", "+00:00")
        try:
            dt = datetime.fromisoformat(normalized)
        except ValueError:
            return slug
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.isoformat()

    return None


def _normalize_date(value: Any) -> Optional[str]:
    """Normalizza un valore data (YYYY-MM-DD)."""

    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, (int, float)):
        try:
            return datetime.fromtimestamp(float(value), tz=timezone.utc).date().isoformat()
        except (ValueError, OSError):
            return None

    if isinstance(value, str):
        slug = value.strip()
        if not slug:
            return None
        for candidate in (slug, slug.replace("/", "-")):
            try:
                dt = datetime.fromisoformat(candidate)
            except ValueError:
                try:
                    dt = datetime.strptime(candidate, "%Y-%m-%d")
                except ValueError:
                    continue
            return dt.date().isoformat()
        return None

    return None


def _extract_iso_date(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    cleaned = value.replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(cleaned)
    except ValueError:
        if len(value) >= 10:
            return value[:10]
        return None
    return dt.date().isoformat()


def _parse_date_any(value: Any) -> Optional[date]:
    """Parsa una data da vari formati (ISO, timestamp, dd/mm/yyyy, dd-mm-yyyy)."""

    if value is None:
        return None

    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, (int, float)):
        try:
            return datetime.fromtimestamp(float(value), tz=timezone.utc).date()
        except (ValueError, OSError):
            return None

    if isinstance(value, str):
        slug = value.strip()
        if not slug:
            return None
        slug = slug.replace("Z", "+00:00")
        # ISO 8601 (date o datetime)
        try:
            return datetime.fromisoformat(slug).date()
        except ValueError:
            pass
        # Se contiene un datetime ma non ISO perfetto, prova a prendere YYYY-MM-DD
        if len(slug) >= 10:
            head = slug[:10]
            try:
                return datetime.strptime(head, "%Y-%m-%d").date()
            except ValueError:
                pass
        # Formati europei
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(slug[:10], fmt).date()
            except ValueError:
                continue
        return None

    return None


def _coerce_int(value: Any) -> Optional[int]:
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float, Decimal)):
        return int(value)
    if isinstance(value, str):
        slug = value.strip()
        if not slug:
            return None
        try:
            return int(slug)
        except ValueError:
            return None
    return None


def _normalize_running(value: Any, default: int | None = None) -> int:
    coerced = _coerce_int(value)
    if coerced in VALID_RUN_STATES:
        return int(coerced)
    if default is None:
        return RUN_STATE_PAUSED
    return default


def _slugify(value: str) -> str:
    """Crea uno slug alfanumerico minuscolo."""

    normalized = value.strip().lower()
    if not normalized:
        return ""
    buffer: List[str] = []
    previous_dash = False
    for char in normalized:
        if char.isalnum():
            buffer.append(char)
            previous_dash = False
        elif not previous_dash:
            buffer.append("-")
            previous_dash = True
    slug = "".join(buffer).strip("-")
    return slug


def _normalize_activity_id(value: str) -> str:
    slug = _slugify(value)
    if not slug:
        return ""
    return slug.replace("-", "_").upper()


def _generate_activity_id(db: DatabaseLike, label: str) -> str:
    base = _normalize_activity_id(label) or "ATTIVITA"
    base = base[:24]
    if not base:
        base = "ATTIVITA"
    candidate = base
    suffix = 1
    while True:
        row = db.execute("SELECT 1 FROM activities WHERE activity_id=?", (candidate,)).fetchone()
        if row is None:
            return candidate
        suffix += 1
        padded = f"{suffix:02d}"
        truncated = base[: max(4, 24 - len(padded) - 1)]
        candidate = f"{truncated}_{padded}"


def compute_planned_duration_ms(
    plan_start: Optional[str],
    plan_end: Optional[str],
    planned_members: Optional[int],
) -> Optional[int]:
    start_ms = parse_iso_to_ms(plan_start)
    end_ms = parse_iso_to_ms(plan_end)
    if start_ms is None or end_ms is None:
        return None
    base = max(0, end_ms - start_ms)
    if base == 0:
        return 0
    normalized_members = planned_members if isinstance(planned_members, int) else None
    if normalized_members is None:
        normalized_members = _coerce_int(planned_members)
    if normalized_members is None or normalized_members <= 0:
        normalized_members = 1
    return base * normalized_members


def load_activity_meta(db: DatabaseLike) -> Dict[str, Any]:
    raw = get_app_state(db, "activity_plan_meta")
    if not raw:
        return {}
    try:
        meta = json.loads(raw)
    except json.JSONDecodeError:
        return {}
    if isinstance(meta, dict):
        return meta
    return {}


def save_activity_meta(db: DatabaseLike, meta: Mapping[str, Any]) -> None:
    set_app_state(db, "activity_plan_meta", json.dumps(meta))


def refresh_activity_meta(db: DatabaseLike) -> Dict[str, Any]:
    rows = db.execute(
        "SELECT activity_id, plan_start, plan_end, planned_members, planned_duration_ms FROM activities"
    ).fetchall()
    current = load_activity_meta(db)
    next_meta: Dict[str, Any] = {}
    for row in rows:
        activity_id = row["activity_id"]
        if not activity_id:
            continue
        previous = current.get(activity_id)
        previous_runtime = 0
        if isinstance(previous, Mapping):
            try:
                previous_runtime = int(previous.get("actual_runtime_ms") or 0)
            except (TypeError, ValueError):
                previous_runtime = 0
        next_meta[str(activity_id)] = {
            "plan_start": row["plan_start"],
            "plan_end": row["plan_end"],
            "planned_members": row["planned_members"],
            "planned_duration_ms": row["planned_duration_ms"],
            "actual_runtime_ms": previous_runtime,
        }
    save_activity_meta(db, next_meta)
    return next_meta


def increment_activity_runtime(meta: Dict[str, Any], activity_id: Optional[str], delta_ms: int) -> bool:
    if not activity_id:
        return False
    try:
        contribution = int(delta_ms)
    except (TypeError, ValueError):
        return False
    if contribution <= 0:
        return False

    entry = meta.get(activity_id)
    if not isinstance(entry, dict):
        entry = {}
    current_value = entry.get("actual_runtime_ms")
    try:
        current_int = int(current_value)
    except (TypeError, ValueError):
        current_int = 0
    entry["actual_runtime_ms"] = current_int + contribution
    meta[activity_id] = entry
    return True


def _activity_matches_date(plan_start: Optional[str], plan_end: Optional[str], expected: str) -> bool:
    """Verifica se la data selezionata cade nell'intervallo dell'attività."""
    expected_date = _parse_date_any(expected)
    start = _parse_date_any(plan_start)
    end = _parse_date_any(plan_end)

    if expected_date is None:
        return False

    # Se non abbiamo date valide, escludiamo l'attività
    if start is None and end is None:
        return False

    # Se abbiamo solo una data, verifichiamo l'uguaglianza
    if start is not None and end is None:
        return start == expected_date
    if end is not None and start is None:
        return end == expected_date

    assert start is not None and end is not None
    if end < start:
        start, end = end, start
    return start <= expected_date <= end


def load_external_projects() -> Dict[str, Dict[str, Any]]:
    """Legge un catalogo di progetti personalizzati da projects.json."""

    if not PROJECTS_FILE.exists():
        return {}

    try:
        content = PROJECTS_FILE.read_text(encoding="utf-8")
    except OSError:
        return {}

    try:
        payload = json.loads(content) if content.strip() else {}
    except json.JSONDecodeError:
        return {}

    def normalize(entry: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        code_raw = entry.get("project_code") or entry.get("code")
        code = str(code_raw or "").strip().upper()
        if not code:
            return None

        name_raw = entry.get("project_name") or entry.get("name")
        project_name = str(name_raw or code).strip()

        activities_payload = entry.get("activities") or []
        activities: List[Dict[str, Any]] = []
        for index, item in enumerate(activities_payload, start=1):
            if not isinstance(item, dict):
                continue
            raw_id = item.get("id") or item.get("code")
            activity_id = str(raw_id or "").strip().upper() or f"ACT{index:02}"
            raw_label = item.get("label") or item.get("name")
            label = str(raw_label or activity_id).strip() or activity_id
            activities.append(
                {
                    "id": activity_id,
                    "label": label,
                    "plan_start": _normalize_datetime(item.get("plan_start")),
                    "plan_end": _normalize_datetime(item.get("plan_end")),
                }
            )

        team_payload = entry.get("team") or entry.get("members") or []
        team: List[Dict[str, Any]] = []
        for member in team_payload:
            if not isinstance(member, dict):
                continue
            raw_key = member.get("key") or member.get("id")
            key = str(raw_key or "").strip()
            raw_name = member.get("name") or member.get("full_name")
            name = str(raw_name or key or "Operatore").strip()
            activity_id = member.get("activity_id")
            if isinstance(activity_id, str):
                activity_id = activity_id.strip().upper() or None
            else:
                activity_id = None
            team.append({
                "key": key,
                "name": name,
                "activity_id": activity_id,
            })

        return {
            "project_code": code,
            "project_name": project_name,
            "activities": activities,
            "team": team,
        }

    catalog: Dict[str, Dict[str, Any]] = {}

    if isinstance(payload, dict):
        for key, value in payload.items():
            if not isinstance(value, dict):
                continue
            entry = dict(value)
            entry.setdefault("project_code", key)
            normalized = normalize(entry)
            if normalized:
                catalog[normalized["project_code"]] = normalized
    elif isinstance(payload, list):
        for entry in payload:
            if not isinstance(entry, dict):
                continue
            normalized = normalize(entry)
            if normalized:
                catalog[normalized["project_code"]] = normalized

    return catalog


def load_config() -> Dict[str, Any]:
    """Carica config.json quando disponibile e mantiene una cache in memoria."""

    global _CONFIG_CACHE, _CONFIG_CACHE_MTIME, _DATABASE_SETTINGS, _WEBPUSH_SETTINGS

    if not CONFIG_FILE.exists():
        _CONFIG_CACHE = {}
        _CONFIG_CACHE_MTIME = None
        return {}

    try:
        mtime = CONFIG_FILE.stat().st_mtime
    except OSError:
        return {}

    if _CONFIG_CACHE is not None and _CONFIG_CACHE_MTIME == mtime:
        return _CONFIG_CACHE

    try:
        content = CONFIG_FILE.read_text(encoding="utf-8")
        data = json.loads(content) if content.strip() else {}
        if not isinstance(data, dict):
            raise ValueError("Config root deve essere un oggetto JSON")
    except (OSError, ValueError, json.JSONDecodeError) as exc:
        app.logger.warning("Config: impossibile leggere %s (%s)", CONFIG_FILE.name, exc)
        _CONFIG_CACHE = {}
        _CONFIG_CACHE_MTIME = mtime
        return {}

    _CONFIG_CACHE = data
    _CONFIG_CACHE_MTIME = mtime
    _DATABASE_SETTINGS = None
    _WEBPUSH_SETTINGS = None
    return data


APP_USERS_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS app_users (
    username VARCHAR(190) PRIMARY KEY,
    password_hash VARCHAR(255) NOT NULL,
    display_name VARCHAR(255) NOT NULL,
    full_name VARCHAR(255) DEFAULT NULL,
    role VARCHAR(32) NOT NULL DEFAULT 'user',
    is_active TINYINT(1) NOT NULL DEFAULT 1,
    external_id VARCHAR(255) DEFAULT NULL,
    external_group_id VARCHAR(255) DEFAULT NULL,
    created_ts BIGINT NOT NULL,
    updated_ts BIGINT NOT NULL
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

APP_USERS_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS app_users (
    username TEXT PRIMARY KEY,
    password_hash TEXT NOT NULL,
    display_name TEXT NOT NULL,
    full_name TEXT,
    role TEXT NOT NULL DEFAULT 'user',
    is_active INTEGER NOT NULL DEFAULT 1,
    external_id TEXT DEFAULT NULL,
    external_group_id TEXT DEFAULT NULL,
    created_ts INTEGER NOT NULL,
    updated_ts INTEGER NOT NULL
)
"""

# ═══════════════════════════════════════════════════════════════════
# TABELLA GRUPPI UTENTI
# ═══════════════════════════════════════════════════════════════════
USER_GROUPS_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS user_groups (
    id INT AUTO_INCREMENT PRIMARY KEY,
    name VARCHAR(255) NOT NULL,
    description TEXT DEFAULT NULL,
    cedolino_group_id VARCHAR(255) DEFAULT NULL,
    gps_location_name VARCHAR(255) DEFAULT NULL COMMENT 'Nome sede GPS associata al gruppo',
    is_active TINYINT(1) NOT NULL DEFAULT 1,
    created_ts BIGINT NOT NULL,
    updated_ts BIGINT NOT NULL,
    UNIQUE KEY uk_group_name (name)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

USER_GROUPS_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS user_groups (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    description TEXT DEFAULT NULL,
    cedolino_group_id TEXT DEFAULT NULL,
    gps_location_name TEXT DEFAULT NULL,
    is_active INTEGER NOT NULL DEFAULT 1,
    created_ts INTEGER NOT NULL,
    updated_ts INTEGER NOT NULL
)
"""

SESSION_OVERRIDES_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS activity_session_overrides (
    id INT AUTO_INCREMENT PRIMARY KEY,
    member_key VARCHAR(255) NOT NULL,
    member_name VARCHAR(255) NOT NULL,
    activity_id VARCHAR(255) NOT NULL,
    activity_label VARCHAR(255) DEFAULT NULL,
    project_code VARCHAR(64) DEFAULT NULL,
    start_ts BIGINT NOT NULL,
    end_ts BIGINT DEFAULT NULL,
    net_ms BIGINT DEFAULT NULL,
    pause_ms BIGINT DEFAULT NULL,
    pause_count INT NOT NULL DEFAULT 0,
    status VARCHAR(32) NOT NULL DEFAULT 'completed',
    source_member_key VARCHAR(255) DEFAULT NULL,
    source_activity_id VARCHAR(255) DEFAULT NULL,
    source_start_ts BIGINT DEFAULT NULL,
    manual_entry TINYINT(1) NOT NULL DEFAULT 0,
    note TEXT,
    created_by VARCHAR(190) DEFAULT NULL,
    updated_by VARCHAR(190) DEFAULT NULL,
    created_ts BIGINT NOT NULL,
    updated_ts BIGINT NOT NULL,
    INDEX idx_session_override_source (source_member_key, source_activity_id, source_start_ts),
    INDEX idx_session_override_member (member_key),
    INDEX idx_session_override_activity (activity_id),
    INDEX idx_session_override_project (project_code)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

SESSION_OVERRIDES_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS activity_session_overrides (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    member_key TEXT NOT NULL,
    member_name TEXT NOT NULL,
    activity_id TEXT NOT NULL,
    activity_label TEXT,
    project_code TEXT,
    start_ts INTEGER NOT NULL,
    end_ts INTEGER,
    net_ms INTEGER,
    pause_ms INTEGER,
    pause_count INTEGER NOT NULL DEFAULT 0,
    status TEXT NOT NULL DEFAULT 'completed',
    source_member_key TEXT,
    source_activity_id TEXT,
    source_start_ts INTEGER,
    manual_entry INTEGER NOT NULL DEFAULT 0,
    note TEXT,
    created_by TEXT,
    updated_by TEXT,
    created_ts INTEGER NOT NULL,
    updated_ts INTEGER NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_session_override_source ON activity_session_overrides(source_member_key, source_activity_id, source_start_ts);
CREATE INDEX IF NOT EXISTS idx_session_override_member ON activity_session_overrides(member_key);
CREATE INDEX IF NOT EXISTS idx_session_override_activity ON activity_session_overrides(activity_id);
CREATE INDEX IF NOT EXISTS idx_session_override_project ON activity_session_overrides(project_code);
"""

PERSISTENT_SESSIONS_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS persistent_sessions (
    token_hash CHAR(64) PRIMARY KEY,
    username VARCHAR(190) NOT NULL,
    user_agent VARCHAR(255) DEFAULT NULL,
    ip_address VARCHAR(64) DEFAULT NULL,
    created_ts BIGINT NOT NULL,
    last_seen_ts BIGINT NOT NULL,
    expires_ts BIGINT NOT NULL,
    INDEX idx_persistent_sessions_user (username)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

PERSISTENT_SESSIONS_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS persistent_sessions (
    token_hash TEXT PRIMARY KEY,
    username TEXT NOT NULL,
    user_agent TEXT,
    ip_address TEXT,
    created_ts INTEGER NOT NULL,
    last_seen_ts INTEGER NOT NULL,
    expires_ts INTEGER NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_persistent_sessions_user ON persistent_sessions(username);
"""


EQUIPMENT_CHECKS_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS equipment_checks (
    id INT AUTO_INCREMENT PRIMARY KEY,
    project_code VARCHAR(128) NOT NULL,
    item_key VARCHAR(512) NOT NULL,
    checked_ts BIGINT NOT NULL,
    username VARCHAR(190) DEFAULT NULL,
    created_ts BIGINT NOT NULL,
    updated_ts BIGINT NOT NULL,
    UNIQUE KEY uq_equipment_project_item (project_code, item_key),
    INDEX idx_equipment_project (project_code)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""


EQUIPMENT_CHECKS_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS equipment_checks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_code TEXT NOT NULL,
    item_key TEXT NOT NULL,
    checked_ts INTEGER NOT NULL,
    username TEXT,
    created_ts INTEGER NOT NULL,
    updated_ts INTEGER NOT NULL
);
CREATE UNIQUE INDEX IF NOT EXISTS uq_equipment_project_item ON equipment_checks(project_code, item_key);
CREATE INDEX IF NOT EXISTS idx_equipment_project ON equipment_checks(project_code);
"""


LOCAL_EQUIPMENT_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS local_equipment (
    id INT AUTO_INCREMENT PRIMARY KEY,
    project_code VARCHAR(128) NOT NULL,
    name VARCHAR(255) NOT NULL,
    quantity INT NOT NULL DEFAULT 1,
    notes TEXT,
    group_name VARCHAR(255) DEFAULT 'Attrezzature extra',
    created_ts BIGINT NOT NULL,
    updated_ts BIGINT NOT NULL,
    INDEX idx_local_equipment_project (project_code)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""


LOCAL_EQUIPMENT_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS local_equipment (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_code TEXT NOT NULL,
    name TEXT NOT NULL,
    quantity INTEGER NOT NULL DEFAULT 1,
    notes TEXT,
    group_name TEXT DEFAULT 'Attrezzature extra',
    created_ts INTEGER NOT NULL,
    updated_ts INTEGER NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_local_equipment_project ON local_equipment(project_code);
"""


PROJECT_PHOTOS_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS project_photos (
    id INT AUTO_INCREMENT PRIMARY KEY,
    project_code VARCHAR(128) NOT NULL,
    filename VARCHAR(255) NOT NULL,
    original_name VARCHAR(255) NOT NULL,
    mime_type VARCHAR(100) NOT NULL,
    file_size INT NOT NULL,
    caption TEXT,
    created_ts BIGINT NOT NULL,
    INDEX idx_project_photos_project (project_code)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""


PROJECT_PHOTOS_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS project_photos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_code TEXT NOT NULL,
    filename TEXT NOT NULL,
    original_name TEXT NOT NULL,
    mime_type TEXT NOT NULL,
    file_size INTEGER NOT NULL,
    caption TEXT,
    created_ts INTEGER NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_project_photos_project ON project_photos(project_code);
"""


PROJECT_MATERIALS_CACHE_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS project_materials_cache (
    project_code VARCHAR(128) PRIMARY KEY,
    project_name VARCHAR(255) NOT NULL,
    data_json LONGTEXT NOT NULL,
    created_ts BIGINT NOT NULL,
    updated_ts BIGINT NOT NULL
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""


PROJECT_MATERIALS_CACHE_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS project_materials_cache (
    project_code TEXT PRIMARY KEY,
    project_name TEXT NOT NULL,
    data_json TEXT NOT NULL,
    created_ts INTEGER NOT NULL,
    updated_ts INTEGER NOT NULL
);
"""


# ═══════════════════════════════════════════════════════════════════════════════
#  CEDOLINO TIMBRATURE - Integrazione CedolinoWeb
# ═══════════════════════════════════════════════════════════════════════════════

CEDOLINO_TIMBRATURE_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS cedolino_timbrature (
    id INT AUTO_INCREMENT PRIMARY KEY,
    member_key VARCHAR(255) DEFAULT NULL,
    member_name VARCHAR(255) NOT NULL,
    username VARCHAR(190) DEFAULT NULL,
    external_id VARCHAR(255) NOT NULL,
    timeframe_id INT NOT NULL COMMENT '1=inizio giornata, 4=inizio pausa, 5=fine pausa, 8=fine giornata',
    timestamp_ms BIGINT NOT NULL,
    data_riferimento DATE NOT NULL,
    ora_originale TIME NOT NULL,
    ora_modificata TIME NOT NULL,
    project_code VARCHAR(128) DEFAULT NULL,
    activity_id VARCHAR(255) DEFAULT NULL,
    synced_ts BIGINT DEFAULT NULL,
    sync_error TEXT DEFAULT NULL,
    sync_attempts INT NOT NULL DEFAULT 0,
    overtime_request_id INT DEFAULT NULL COMMENT 'ID richiesta straordinario collegata (blocca sync fino a revisione)',
    created_ts BIGINT NOT NULL,
    INDEX idx_cedolino_member (member_key),
    INDEX idx_cedolino_username (username),
    INDEX idx_cedolino_external (external_id),
    INDEX idx_cedolino_synced (synced_ts),
    INDEX idx_cedolino_data (data_riferimento),
    INDEX idx_cedolino_overtime (overtime_request_id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

CEDOLINO_TIMBRATURE_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS cedolino_timbrature (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    member_key TEXT,
    member_name TEXT NOT NULL,
    username TEXT,
    external_id TEXT NOT NULL,
    timeframe_id INTEGER NOT NULL,
    timestamp_ms INTEGER NOT NULL,
    data_riferimento TEXT NOT NULL,
    ora_originale TEXT NOT NULL,
    ora_modificata TEXT NOT NULL,
    project_code TEXT,
    activity_id TEXT,
    synced_ts INTEGER,
    sync_error TEXT,
    sync_attempts INTEGER NOT NULL DEFAULT 0,
    overtime_request_id INTEGER DEFAULT NULL,
    created_ts INTEGER NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_cedolino_member ON cedolino_timbrature(member_key);
CREATE INDEX IF NOT EXISTS idx_cedolino_username ON cedolino_timbrature(username);
CREATE INDEX IF NOT EXISTS idx_cedolino_external ON cedolino_timbrature(external_id);
CREATE INDEX IF NOT EXISTS idx_cedolino_synced ON cedolino_timbrature(synced_ts);
CREATE INDEX IF NOT EXISTS idx_cedolino_data ON cedolino_timbrature(data_riferimento);
CREATE INDEX IF NOT EXISTS idx_cedolino_overtime ON cedolino_timbrature(overtime_request_id);
"""

# Costanti timeframe CedolinoWeb
TIMEFRAME_INIZIO_GIORNATA = 1
TIMEFRAME_INIZIO_PAUSA = 4
TIMEFRAME_FINE_PAUSA = 5
TIMEFRAME_FINE_GIORNATA = 8

CEDOLINO_WEB_ENDPOINT = "http://80.211.18.30/WebServices/crea_timbrata_elaborata"
CEDOLINO_CODICE_TERMINALE = "musa_mobile"


# ═══════════════════════════════════════════════════════════════════════════════
#  EMPLOYEE SHIFTS - Turni settimanali per impiegati non-Rentman
# ═══════════════════════════════════════════════════════════════════════════════

EMPLOYEE_SHIFTS_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS employee_shifts (
    id INT AUTO_INCREMENT PRIMARY KEY,
    username VARCHAR(190) NOT NULL,
    day_of_week TINYINT NOT NULL COMMENT '0=Lunedì, 1=Martedì, ..., 6=Domenica',
    start_time TIME NOT NULL,
    end_time TIME NOT NULL,
    break_start TIME DEFAULT NULL,
    break_end TIME DEFAULT NULL,
    shift_name VARCHAR(100) DEFAULT NULL COMMENT 'Nome identificativo del turno (es. Turno Mattina)',
    location_name VARCHAR(255) DEFAULT NULL COMMENT 'Nome sede GPS associata al turno',
    is_active TINYINT(1) NOT NULL DEFAULT 1,
    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
    UNIQUE KEY uk_user_day (username, day_of_week),
    INDEX idx_shifts_username (username),
    INDEX idx_shifts_active (is_active)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

EMPLOYEE_SHIFTS_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS employee_shifts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT NOT NULL,
    day_of_week INTEGER NOT NULL,
    start_time TEXT NOT NULL,
    end_time TEXT NOT NULL,
    break_start TEXT,
    break_end TEXT,
    shift_name TEXT,
    location_name TEXT,
    is_active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
    updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
    UNIQUE (username, day_of_week)
);
CREATE INDEX IF NOT EXISTS idx_shifts_username ON employee_shifts(username);
CREATE INDEX IF NOT EXISTS idx_shifts_active ON employee_shifts(is_active);
"""


MYSQL_SCHEMA_STATEMENTS: List[str] = [
    """
    CREATE TABLE IF NOT EXISTS activities (
        activity_id VARCHAR(255) NOT NULL,
        project_code VARCHAR(64) NOT NULL DEFAULT '',
        label VARCHAR(255) NOT NULL,
        sort_order INT NOT NULL,
        plan_start TEXT NULL,
        plan_end TEXT NULL,
        planned_members INT NULL,
        planned_duration_ms BIGINT NULL,
        notes TEXT NULL,
        PRIMARY KEY (activity_id, project_code)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """,
    """
    CREATE TABLE IF NOT EXISTS member_state (
        member_key VARCHAR(255) NOT NULL,
        project_code VARCHAR(64) NOT NULL DEFAULT '',
        member_name VARCHAR(255) NOT NULL,
        activity_id VARCHAR(255),
        running TINYINT(1) NOT NULL DEFAULT 0,
        start_ts BIGINT,
        elapsed_cached BIGINT NOT NULL DEFAULT 0,
        pause_start BIGINT,
        entered_ts BIGINT,
        PRIMARY KEY (member_key, project_code)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """,
    """
    CREATE TABLE IF NOT EXISTS event_log (
        id INT AUTO_INCREMENT PRIMARY KEY,
        project_code VARCHAR(64) NOT NULL DEFAULT '',
        ts BIGINT NOT NULL,
        kind VARCHAR(64) NOT NULL,
        member_key VARCHAR(255),
        details LONGTEXT,
        INDEX idx_event_project (project_code)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """,
    """
    CREATE TABLE IF NOT EXISTS app_state (
        `key` VARCHAR(128) PRIMARY KEY,
        value TEXT NOT NULL
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """,
    """
    CREATE TABLE IF NOT EXISTS push_subscriptions (
        id INT AUTO_INCREMENT PRIMARY KEY,
        username VARCHAR(190) NOT NULL,
        endpoint VARCHAR(500) NOT NULL UNIQUE,
        p256dh VARCHAR(255) NOT NULL,
        auth VARCHAR(128) NOT NULL,
        content_encoding VARCHAR(32) DEFAULT NULL,
        user_agent VARCHAR(255) DEFAULT NULL,
        expiration_time BIGINT DEFAULT NULL,
        created_ts BIGINT NOT NULL,
        updated_ts BIGINT NOT NULL,
        INDEX idx_push_username (username)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """,
    """
    CREATE TABLE IF NOT EXISTS push_notification_log (
        id INT AUTO_INCREMENT PRIMARY KEY,
        kind VARCHAR(32) NOT NULL,
        activity_id VARCHAR(255) DEFAULT NULL,
        username VARCHAR(190) DEFAULT NULL,
        title VARCHAR(255) NOT NULL,
        body TEXT,
        payload LONGTEXT,
        sent_ts BIGINT NOT NULL,
        created_ts BIGINT NOT NULL,
        read_at BIGINT DEFAULT NULL,
        INDEX idx_push_log_user (username),
        INDEX idx_push_log_sent (sent_ts)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """,
    APP_USERS_TABLE_MYSQL,
    SESSION_OVERRIDES_TABLE_MYSQL,
    EQUIPMENT_CHECKS_TABLE_MYSQL,
    PROJECT_MATERIALS_CACHE_TABLE_MYSQL,
    LOCAL_EQUIPMENT_TABLE_MYSQL,
]


_DATABASE_SETTINGS: Optional[Dict[str, Any]] = None


REQUIRED_ACTIVITY_COLUMNS: Dict[str, str] = {
    "plan_start": "TEXT",
    "plan_end": "TEXT",
    "planned_members": "INTEGER",
    "planned_duration_ms": "BIGINT",
    "notes": "TEXT",
}

_ACTIVITY_SCHEMA_READY = False


def _get_existing_columns(db: DatabaseLike, table: str) -> Set[str]:
    columns: Set[str] = set()
    if DB_VENDOR == "mysql":
        query = (
            "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
            "WHERE TABLE_SCHEMA=? AND TABLE_NAME=?"
        )
        rows = db.execute(query, (DATABASE_SETTINGS["name"], table)).fetchall()
        for row in rows:
            if isinstance(row, Mapping):
                columns.add(str(row.get("COLUMN_NAME")))
            elif isinstance(row, Sequence) and row:
                columns.add(str(row[0]))
    else:
        rows = db.execute(f"PRAGMA table_info({table})").fetchall()
        for row in rows:
            if isinstance(row, Mapping):
                columns.add(str(row.get("name")))
            elif isinstance(row, Sequence) and len(row) > 1:
                columns.add(str(row[1]))
    return {str(col).lower() for col in columns if col}


def ensure_activity_schema(db: DatabaseLike) -> None:
    global _ACTIVITY_SCHEMA_READY
    if _ACTIVITY_SCHEMA_READY:
        return
    try:
        existing = _get_existing_columns(db, "activities")
    except Exception:
        return
    alter_template = (
        "ALTER TABLE activities ADD COLUMN {name} {definition}"
        if DB_VENDOR == "mysql"
        else "ALTER TABLE activities ADD COLUMN {name} {definition}"
    )
    for name, definition in REQUIRED_ACTIVITY_COLUMNS.items():
        if name.lower() in existing:
            continue
        db.execute(alter_template.format(name=name, definition=definition))
    _ACTIVITY_SCHEMA_READY = True


_PROJECT_CODE_MIGRATION_DONE = False


def ensure_project_code_columns(db: DatabaseLike) -> None:
    """Migra le tabelle esistenti per aggiungere la colonna project_code."""
    global _PROJECT_CODE_MIGRATION_DONE
    if _PROJECT_CODE_MIGRATION_DONE:
        return
    
    tables_to_migrate = {
        "activities": "VARCHAR(64) NOT NULL DEFAULT ''" if DB_VENDOR == "mysql" else "TEXT NOT NULL DEFAULT ''",
        "member_state": "VARCHAR(64) NOT NULL DEFAULT ''" if DB_VENDOR == "mysql" else "TEXT NOT NULL DEFAULT ''",
        "event_log": "VARCHAR(64) NOT NULL DEFAULT ''" if DB_VENDOR == "mysql" else "TEXT NOT NULL DEFAULT ''",
    }
    
    for table, col_def in tables_to_migrate.items():
        try:
            existing = _get_existing_columns(db, table)
            if "project_code" not in existing:
                db.execute(f"ALTER TABLE {table} ADD COLUMN project_code {col_def}")
                app.logger.info("Aggiunta colonna project_code a tabella %s", table)
        except Exception as e:
            app.logger.warning("Impossibile aggiungere project_code a %s: %s", table, e)
    
    # Aggiungi indice su event_log se non esiste
    try:
        if DB_VENDOR == "mysql":
            # MySQL: verifica se l'indice esiste
            idx_check = db.execute(
                "SELECT COUNT(*) as cnt FROM INFORMATION_SCHEMA.STATISTICS WHERE TABLE_SCHEMA=%s AND TABLE_NAME='event_log' AND INDEX_NAME='idx_event_project'",
                (DATABASE_SETTINGS["name"],)
            ).fetchone()
            cnt = idx_check["cnt"] if isinstance(idx_check, Mapping) else idx_check[0]
            if cnt == 0:
                db.execute("CREATE INDEX idx_event_project ON event_log(project_code)")
        else:
            db.execute("CREATE INDEX IF NOT EXISTS idx_event_project ON event_log(project_code)")
    except Exception as e:
        app.logger.warning("Impossibile creare indice idx_event_project: %s", e)
    
    _PROJECT_CODE_MIGRATION_DONE = True


def get_database_settings(force_refresh: bool = False) -> Dict[str, Any]:
    """Restituisce le impostazioni DB combinando env e config.json."""

    global _DATABASE_SETTINGS
    if _DATABASE_SETTINGS is not None and not force_refresh:
        return _DATABASE_SETTINGS

    config = load_config()
    db_section = config.get("database")
    raw_db: Dict[str, Any] = db_section if isinstance(db_section, dict) else {}

    def read(key: str, env_key: str, default: Any = None) -> Any:
        env_value = os.environ.get(env_key)
        if env_value not in (None, ""):
            return env_value
        if key in raw_db:
            value = raw_db.get(key)
            if value not in (None, ""):
                return value
        return default

    vendor = str(read("vendor", "JOBLOG_DB_VENDOR", "sqlite") or "sqlite").lower()
    port_value = read("port", "JOBLOG_DB_PORT", 3306)
    try:
        port = int(port_value)
    except (TypeError, ValueError):  # pragma: no cover - configurazione invalida
        port = 3306

    settings = {
        "vendor": vendor,
        "host": read("host", "JOBLOG_DB_HOST", "localhost"),
        "port": port,
        "user": read("user", "JOBLOG_DB_USER", "root"),
        "password": read("password", "JOBLOG_DB_PASSWORD", ""),
        "name": read("name", "JOBLOG_DB_NAME", "joblog"),
    }

    _DATABASE_SETTINGS = settings
    return settings


DATABASE_SETTINGS = get_database_settings()
DB_VENDOR = DATABASE_SETTINGS["vendor"]
APP_STATE_KEY_COLUMN = "`key`" if DB_VENDOR == "mysql" else "key"


def get_webpush_settings(force_refresh: bool = False) -> Optional[Dict[str, str]]:
    """Restituisce le impostazioni VAPID per il Web Push, se configurate."""

    global _WEBPUSH_SETTINGS
    if _WEBPUSH_SETTINGS is not None and not force_refresh:
        return cast(Optional[Dict[str, str]], _WEBPUSH_SETTINGS)

    config = load_config()
    section = config.get("webpush")
    raw_section: Dict[str, Any] = section if isinstance(section, dict) else {}

    def read(key: str, env_key: str) -> Optional[str]:
        value = os.environ.get(env_key)
        if value not in (None, ""):
            return value.strip()
        candidate = raw_section.get(key)
        if isinstance(candidate, str) and candidate.strip():
            return candidate.strip()
        return None

    public = read("vapid_public", "WEBPUSH_VAPID_PUBLIC")
    private = read("vapid_private", "WEBPUSH_VAPID_PRIVATE")
    subject = read("subject", "WEBPUSH_VAPID_SUBJECT")

    if not public or not private or not subject:
        _WEBPUSH_SETTINGS = None
        return None

    _WEBPUSH_SETTINGS = {
        "vapid_public": public,
        "vapid_private": private,
        "subject": subject,
    }
    return cast(Dict[str, str], _WEBPUSH_SETTINGS)


def get_rentman_client() -> Optional[RentmanClient]:
    """Istanzia il client Rentman solo quando disponibile un token valido."""

    global _RENTMAN_CLIENT, _RENTMAN_CLIENT_TOKEN

    token = (os.environ.get("RENTMAN_API_TOKEN") or "").strip()
    if not token:
        config = load_config()
        token = str(config.get("rentman_api_token") or "").strip()
        if token:
            app.logger.warning("Rentman: uso token da config.json")

    if not token:
        app.logger.warning("Rentman: token non trovato (env o config)")
        return None

    if _RENTMAN_CLIENT and _RENTMAN_CLIENT_TOKEN == token:
        return _RENTMAN_CLIENT

    try:
        client = RentmanClient(token=token)
    except RentmanAuthError as exc:
        app.logger.warning("Rentman: token non valido (%s)", exc)
        return None

    app.logger.debug("Rentman: client inizializzato con token attivo")

    _RENTMAN_CLIENT = client
    _RENTMAN_CLIENT_TOKEN = token
    return client


def parse_reference(reference: Any) -> Optional[int]:
    """Estrae l'identificativo numerico da un riferimento Rentman."""

    if reference is None:
        return None

    if isinstance(reference, int):
        return reference

    if isinstance(reference, dict):
        candidate = reference.get("id") or reference.get("href") or reference.get("resource")
        if isinstance(candidate, int):
            return candidate
        reference = candidate

    if isinstance(reference, str):
        slug = reference.strip()
        if not slug:
            return None
        slug = slug.strip("/")
        last_segment = slug.split("/")[-1]
        last_segment = last_segment.split("?")[0]
        try:
            return int(last_segment)
        except ValueError:
            return None

    return None


def fetch_rentman_plan(project_code: str, project_date: Optional[str] = None) -> Optional[Dict[str, Any]]:
    """Recupera da Rentman le funzioni equiparate alle attività e il relativo crew."""

    client = get_rentman_client()
    if not client:
        return None

    app.logger.warning("Rentman: ricerca progetto per codice '%s' (data: %s)", project_code, project_date)

    try:
        project = client.find_project(project_code)
        app.logger.info(
            "Rentman: payload progetto=\n%s",
            json.dumps(project, ensure_ascii=False, indent=2) if project else "{}",
        )
    except RentmanNotFound:
        app.logger.warning("Rentman: progetto %s non trovato", project_code)
        return None
    except RentmanAuthError as exc:
        app.logger.error("Rentman: autenticazione fallita (%s)", exc)
        return None
    except RentmanAPIError as exc:
        app.logger.error("Rentman: errore recuperando il progetto %s: %s", project_code, exc)
        return None

    if not project:
        app.logger.warning(
            "Rentman: nessun progetto associato al codice %s (number/reference)",
            project_code,
        )
        return None

    project_id = parse_reference(project.get("id")) or project.get("id")
    if not isinstance(project_id, int):
        app.logger.error("Rentman: progetto %s senza id valido", project_code)
        return None

    try:
        subprojects = client.get_project_subprojects(project_id)
        app.logger.info(
            "Rentman: payload subprojects=\n%s",
            json.dumps(subprojects, ensure_ascii=False, indent=2),
        )
    except RentmanNotFound:
        subprojects = []
    except RentmanAPIError as exc:
        app.logger.error(
            "Rentman: errore leggendo i subprojects del progetto %s: %s",
            project_code,
            exc,
        )
        subprojects = []

    allowed_subprojects: Set[int] = set()
    subproject_labels: Dict[int, str] = {}
    for sub in subprojects:
        sub_id = parse_reference(sub.get("id")) or sub.get("id")
        if not isinstance(sub_id, int):
            continue

        label = (
            sub.get("displayname")
            or sub.get("name")
            or sub.get("description")
            or str(sub_id)
        )
        subproject_labels[sub_id] = str(label)

        if _is_truthy(sub.get("is_planning")):
            allowed_subprojects.add(sub_id)

    app.logger.info("Rentman: subprojects pianificati=%s", sorted(allowed_subprojects))
    filter_subprojects = bool(allowed_subprojects)
    if subprojects and not allowed_subprojects:
        app.logger.warning(
            "Rentman: subprojects presenti ma nessuno con is_planning=true per %s",
            project_code,
        )

    try:
        functions = client.get_project_functions(project_id)
        app.logger.info(
            "Rentman: payload funzioni=\n%s",
            json.dumps(functions, ensure_ascii=False, indent=2),
        )
    except RentmanNotFound:
        functions = []
    except RentmanAPIError as exc:
        app.logger.error(
            "Rentman: errore leggendo le funzioni del progetto %s: %s",
            project_code,
            exc,
        )
        functions = []

    filtered_functions: List[Dict[str, Any]] = []
    function_ids: List[int] = []
    seen_function_ids: Set[int] = set()
    for entry in functions:
        func_id = parse_reference(entry.get("id")) or entry.get("id")
        if not isinstance(func_id, int):
            continue
        subproject_id = parse_reference(entry.get("subproject")) or entry.get("subproject")
        if filter_subprojects:
            if not isinstance(subproject_id, int):
                continue
            if subproject_id not in allowed_subprojects:
                continue
        if func_id in seen_function_ids:
            continue
        filtered_functions.append(entry)
        function_ids.append(func_id)
        seen_function_ids.add(func_id)

    crew_assignments: List[Dict[str, Any]] = []
    if function_ids:
        try:
            crew_assignments = client.get_project_crew_by_function_ids(function_ids)
            app.logger.info(
                "Rentman: payload crew assignments=\n%s",
                json.dumps(crew_assignments, ensure_ascii=False, indent=2),
            )
        except RentmanNotFound:
            crew_assignments = []
        except RentmanAPIError as exc:
            app.logger.error(
                "Rentman: errore leggendo il planning crew del progetto %s: %s",
                project_code,
                exc,
            )
            crew_assignments = []

    activity_lookup: Dict[int, str] = {}
    activities: List[Dict[str, Any]] = []
    for entry in filtered_functions:
        func_id = parse_reference(entry.get("id")) or entry.get("id")
        if not isinstance(func_id, int) or func_id in activity_lookup:
            continue

        subproject_ref = entry.get("subproject")
        subproject_id = parse_reference(subproject_ref)
        subproject_label: Optional[str] = None
        if isinstance(subproject_id, int):
            subproject_label = subproject_labels.get(subproject_id) or str(subproject_id)
        elif isinstance(subproject_ref, str):
            if subproject_ref.startswith("/subprojects/"):
                subproject_label = subproject_ref.split("/")[-1]
            else:
                subproject_label = subproject_ref

        normalized_subproject = (
            str(subproject_label).strip().lower() if subproject_label else ""
        )
        if "produzione" not in normalized_subproject:
            app.logger.debug(
                "Rentman: funzione %s esclusa per sottoprogetto '%s'",
                func_id,
                subproject_label,
            )
            continue

        plan_start = (
            _normalize_datetime(entry.get("planperiod_start"))
            or _normalize_datetime(entry.get("usageperiod_start"))
        )
        plan_end = (
            _normalize_datetime(entry.get("planperiod_end"))
            or _normalize_datetime(entry.get("usageperiod_end"))
        )

        if project_date:
            if not _activity_matches_date(plan_start, plan_end, project_date):
                app.logger.debug(
                    "Rentman: funzione %s esclusa per data '%s' (start: %s, end: %s)",
                    func_id,
                    project_date,
                    plan_start,
                    plan_end,
                )
                continue

        activity_id = f"rentman-f-{func_id}"
        label = (
            entry.get("name")
            or entry.get("displayname")
            or entry.get("description")
            or f"Funzione {func_id}"
        )

        label = f"{label} [ID {func_id}]"

        activities.append(
            {
                "id": activity_id,
                "label": str(label),
                "plan_start": plan_start,
                "plan_end": plan_end,
            }
        )
        activity_lookup[func_id] = activity_id

    activities.sort(key=lambda item: item["label"].lower())
    app.logger.info(
        "Rentman: funzioni considerate=%s",
        json.dumps(activities, ensure_ascii=False, indent=2),
    )

    valid_function_ids: Set[int] = set(activity_lookup)
    crew_ids: Set[int] = set()
    for assignment in crew_assignments:
        member_id = parse_reference(assignment.get("crewmember"))
        function_id = parse_reference(assignment.get("function"))
        if not isinstance(function_id, int) or function_id not in valid_function_ids:
            continue
        if isinstance(member_id, int):
            crew_ids.add(member_id)

    crew_details: List[Dict[str, Any]] = []
    if crew_ids:
        try:
            crew_details = client.get_crew_members_by_ids(crew_ids)
            app.logger.info(
                "Rentman: payload crew details=\n%s",
                json.dumps(crew_details, ensure_ascii=False, indent=2),
            )
        except RentmanNotFound:
            crew_details = []
        except RentmanAPIError as exc:
            app.logger.error(
                "Rentman: errore leggendo i membri crew del progetto %s: %s",
                project_code,
                exc,
            )
            crew_details = []

    crew_map: Dict[int, Dict[str, Any]] = {}
    for member in crew_details:
        member_id = parse_reference(member.get("id")) or member.get("id")
        if isinstance(member_id, int):
            crew_map[member_id] = member

    team: List[Dict[str, Any]] = []
    seen_members: Set[str] = set()
    for assignment in crew_assignments:
        assignment_id = assignment.get("id")
        member_id = parse_reference(assignment.get("crewmember"))
        function_id = parse_reference(assignment.get("function"))

        if (
            not isinstance(assignment_id, int)
            or member_id is None
            or function_id is None
            or function_id not in valid_function_ids
        ):
            continue

        activity_id = activity_lookup.get(function_id)
        if not activity_id:
            continue

        crew_info = crew_map.get(member_id, {})
        display_name = (
            crew_info.get("displayname")
            or crew_info.get("name")
            or assignment.get("displayname")
            or assignment.get("name")
            or "Operatore"
        )

        member_key = f"rentman-crew-{assignment_id}"
        if member_key in seen_members:
            continue
        seen_members.add(member_key)

        team.append(
            {
                "key": member_key,
                "name": str(display_name),
                "activity_id": activity_id,
            }
        )

    project_name = (
        project.get("name")
        or project.get("displayname")
        or project.get("description")
        or project_code
    )

    plan = {
        "project_code": str(project.get("number") or project_code),
        "project_name": str(project_name),
        "activities": activities,
        "team": team,
    }

    return plan


def _normalize_attachment_name(entry: Mapping[str, Any]) -> str:
    for key in (
        "readable_name",
        "displayname",
        "friendly_name_without_extension",
        "name_without_extension",
    ):
        value = entry.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    identifier = entry.get("id")
    return f"Allegato {identifier}" if identifier is not None else "Allegato"


def _normalize_attachment_extension(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    slug = value.strip().lstrip(".")
    return slug.upper()

ATTACHMENT_IMAGE_EXTENSIONS: Set[str] = {
    "JPG",
    "JPEG",
    "PNG",
    "GIF",
    "WEBP",
    "BMP",
    "TIFF",
    "HEIC",
}


def _attachment_is_image(entry: Mapping[str, Any]) -> bool:
    if _is_truthy(entry.get("image")):
        return True
    extension = _normalize_attachment_extension(entry.get("extension") or entry.get("type"))
    if extension and extension in ATTACHMENT_IMAGE_EXTENSIONS:
        return True
    description = entry.get("description")
    if isinstance(description, str):
        slug = description.lower()
        if any(token in slug for token in ("photo", "immagine", "preview")):
            return True
    return False


def _folder_display_name(entry: Mapping[str, Any]) -> str:
    for key in ("displayname", "name", "readable_name"):
        value = entry.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    identifier = entry.get("id")
    return f"Cartella {identifier}" if identifier is not None else "Cartella"


def _build_folder_path(folder_id: int, lookup: Mapping[int, Mapping[str, Any]], max_depth: int = 20) -> str:
    parts: List[str] = []
    current = folder_id
    visited: Set[int] = set()
    depth = 0
    while isinstance(current, int) and current not in visited and depth < max_depth:
        visited.add(current)
        entry = lookup.get(current)
        if not entry:
            break
        parts.append(_folder_display_name(entry))
        current = parse_reference(entry.get("parent"))
        depth += 1
    if not parts:
        return ""
    return " / ".join(reversed(parts))


def _collect_project_folders(client: RentmanClient, project_id: int) -> List[Dict[str, Any]]:
    try:
        raw_folders = client.get_project_file_folders(project_id)
    except RentmanError as exc:
        app.logger.error("Rentman: errore recuperando le cartelle file del progetto %s: %s", project_id, exc)
        return []

    folder_lookup: Dict[int, Mapping[str, Any]] = {}
    for entry in raw_folders:
        folder_id = parse_reference(entry.get("id")) or entry.get("id")
        if isinstance(folder_id, int):
            folder_lookup[folder_id] = entry

    try:
        raw_files = client.get_project_files(project_id, exhaustive=False)
    except RentmanError as exc:
        app.logger.error("Rentman: errore recuperando i file per il progetto %s: %s", project_id, exc)
        raw_files = []

    folder_files: Dict[int, List[Dict[str, Any]]] = {}
    for entry in raw_files:
        folder_id = parse_reference(entry.get("folder"))
        if not isinstance(folder_id, int):
            continue
        normalized = {
            "id": entry.get("id"),
            "name": _normalize_attachment_name(entry),
            "extension": _normalize_attachment_extension(entry.get("extension") or entry.get("type")),
            "url": entry.get("url"),
            "preview_url": entry.get("proxy_url") or entry.get("url"),
            "image": _attachment_is_image(entry),
        }
        folder_files.setdefault(folder_id, []).append(normalized)

    folders: List[Dict[str, Any]] = []
    for folder_id, entry in folder_lookup.items():
        parent_id = parse_reference(entry.get("parent"))
        path_value = entry.get("path") or _build_folder_path(folder_id, folder_lookup)
        files = folder_files.get(folder_id, [])
        image_file = next((item for item in files if item.get("image")), None)
        if not image_file and files:
            image_file = files[0]
        folders.append(
            {
                "id": folder_id,
                "name": _folder_display_name(entry),
                "parent_id": parent_id,
                "path": path_value or _folder_display_name(entry),
                "file_count": len(files),
                "photo": {
                    "name": image_file.get("name"),
                    "url": image_file.get("url"),
                    "preview_url": image_file.get("preview_url"),
                    "extension": image_file.get("extension"),
                }
                if image_file
                else None,
            }
        )

    folders.sort(key=lambda item: str(item.get("path") or item.get("name") or "").lower())
    return folders


def _equipment_group_display_name(entry: Mapping[str, Any]) -> str:
    for key in ("path", "displayname", "name", "description"):
        value = entry.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    identifier = entry.get("id")
    return f"Gruppo {identifier}" if identifier is not None else "Gruppo"


def _build_equipment_group_path(
    group_id: int,
    lookup: Mapping[int, Mapping[str, Any]],
    *,
    max_depth: int = 20,
) -> str:
    parts: List[str] = []
    current = group_id
    visited: Set[int] = set()
    depth = 0
    while isinstance(current, int) and current not in visited and depth < max_depth:
        visited.add(current)
        entry = lookup.get(current)
        if not entry:
            break
        parts.append(_equipment_group_display_name(entry))
        current = parse_reference(entry.get("parent"))
        depth += 1
    if not parts:
        return ""
    return " / ".join(reversed(parts))


def _collect_material_groups(client: RentmanClient, project_id: int) -> Dict[int, Dict[str, Any]]:
    try:
        raw_groups = client.get_project_equipment_groups(project_id)
    except RentmanError as exc:
        app.logger.error("Rentman: errore recuperando i gruppi materiali del progetto %s: %s", project_id, exc)
        return {}

    group_lookup: Dict[int, Mapping[str, Any]] = {}
    for entry in raw_groups:
        group_id = parse_reference(entry.get("id")) or entry.get("id")
        if isinstance(group_id, int):
            group_lookup[group_id] = entry

    result: Dict[int, Dict[str, Any]] = {}
    for group_id, entry in group_lookup.items():
        parent_id = parse_reference(entry.get("parent"))
        path_value = entry.get("path")
        if not isinstance(path_value, str) or not path_value.strip():
            path_value = _build_equipment_group_path(group_id, group_lookup)
        result[group_id] = {
            "id": group_id,
            "name": _equipment_group_display_name(entry),
            "parent_id": parent_id,
            "path": path_value or _equipment_group_display_name(entry),
        }

    return result


def fetch_project_attachments(project_code: Optional[str], *, exhaustive: bool = False) -> List[Dict[str, Any]]:
    attachments: List[Dict[str, Any]] = []
    code = (project_code or "").strip()
    if not code:
        return attachments

    client = get_rentman_client()
    if not client:
        return attachments

    try:
        project = client.find_project(code)
    except RentmanNotFound:
        app.logger.warning("Rentman: nessun progetto per allegati %s", code)
        return attachments
    except (RentmanAuthError, RentmanAPIError) as exc:
        app.logger.error("Rentman: errore durante la ricerca degli allegati per %s: %s", code, exc)
        return attachments

    if not project:
        app.logger.info("Rentman: progetto %s non trovato, nessun allegato", code)
        return attachments

    project_id = parse_reference(project.get("id")) or project.get("id")
    if not isinstance(project_id, int):
        app.logger.warning("Rentman: allegati non disponibili, id progetto non valido per %s (%s)", code, project.get("id"))
        return attachments

    try:
        app.logger.info(
            "Rentman: recupero allegati progetto %s (id=%s, exhaustive=%s)",
            code,
            project_id,
            exhaustive,
        )
        files = client.get_project_files(project_id, exhaustive=exhaustive)
        app.logger.info(
            "Rentman: payload files raw (primi 3)=\n%s",
            json.dumps(files[:3], ensure_ascii=False, indent=2) if files else "[]",
        )
    except RentmanNotFound:
        app.logger.warning("Rentman: endpoint files non trovato (404) per progetto %s", code)
        files = []
    except RentmanAuthError as exc:
        app.logger.error("Rentman: autenticazione fallita leggendo allegati per %s: %s", code, exc)
        files = []
    except RentmanAPIError as exc:
        app.logger.error(
            "Rentman: errore leggendo gli allegati del progetto %s: %s",
            code,
            exc,
        )
        files = []

    app.logger.info("Rentman: ricevuti %s allegati per progetto %s", len(files), code)

    for entry in files:
        if not isinstance(entry, Mapping):
            continue
        attachments.append(
            {
                "id": entry.get("id"),
                "name": _normalize_attachment_name(entry),
                "created": entry.get("created"),
                "type": _normalize_attachment_extension(entry.get("extension") or entry.get("type")),
                "size": entry.get("size"),
                "url": entry.get("url"),
                "preview_url": entry.get("proxy_url") or entry.get("url"),
            }
        )

    attachments.sort(key=lambda item: str(item.get("name") or "").lower())
    return attachments


def _normalize_material_name(entry: Mapping[str, Any]) -> str:
    for key in ("displayname", "name", "description"):
        value = entry.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    identifier = entry.get("id")
    return f"Materiale {identifier}" if identifier is not None else "Materiale"


def _extract_material_quantity(entry: Mapping[str, Any]) -> Tuple[Optional[float], str]:
    numeric_candidate = entry.get("quantity_total")
    quantity_label = ""
    quantity_value: Optional[float] = None
    if isinstance(numeric_candidate, (int, float)):
        quantity_value = float(numeric_candidate)

    raw_quantity = entry.get("quantity")
    if quantity_value is None:
        if isinstance(raw_quantity, (int, float)):
            quantity_value = float(raw_quantity)
        elif isinstance(raw_quantity, str):
            slug = raw_quantity.strip().replace(",", ".")
            try:
                quantity_value = float(slug)
            except ValueError:
                quantity_value = None

    if quantity_value is not None:
        if quantity_value.is_integer():
            quantity_label = str(int(quantity_value))
        else:
            quantity_label = f"{quantity_value:.2f}".rstrip("0").rstrip(".")

    if not quantity_label and isinstance(raw_quantity, str) and raw_quantity.strip():
        quantity_label = raw_quantity.strip()
    elif not quantity_label and isinstance(numeric_candidate, (int, float)):
        quantity_label = str(numeric_candidate)

    return quantity_value, quantity_label or "0"


def _material_status(entry: Mapping[str, Any]) -> Tuple[str, str]:
    if _is_truthy(entry.get("has_missings")):
        return "missing", "Mancanze"
    if _is_truthy(entry.get("delay_notified")):
        return "delayed", "In ritardo"
    subrent = _coerce_int(entry.get("subrent_reservations")) or 0
    if subrent > 0:
        return "subrent", "Subnoleggio"
    reserved = _coerce_int(entry.get("warehouse_reservations")) or 0
    if reserved > 0:
        return "reserved", "Riservato"
    if _is_truthy(entry.get("is_option")):
        return "option", "Opzione"
    return "planned", "Pianificato"


def _coerce_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        slug = value.strip().replace(",", ".")
        if not slug:
            return None
        try:
            return float(slug)
        except ValueError:
            return None
    return None


def _format_dimension_value(value: Optional[float]) -> Optional[str]:
    if value is None:
        return None
    if value.is_integer():
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def _format_dimensions_label(length: Optional[float], width: Optional[float], height: Optional[float]) -> str:
    components: List[str] = []
    for component in (length, width, height):
        label = _format_dimension_value(component)
        components.append(label or "0")
    if not any(component for component in components if component != "0"):
        return "---"
    return "x".join(components)


def _format_weight_label(value: Optional[float]) -> str:
    if value is None:
        return "---"
    label = _format_dimension_value(value)
    if not label:
        return "---"
    return f"{label} kg"


def _resolve_equipment_meta(
    equipment_ref: Any,
    client: RentmanClient,
    cache: Dict[int, Optional[Mapping[str, Any]]],
) -> Optional[Mapping[str, Any]]:
    equipment_id = parse_reference(equipment_ref)
    if not isinstance(equipment_id, int):
        return None
    if equipment_id in cache:
        return cache[equipment_id]
    try:
        meta = client.get_equipment(equipment_id)
    except RentmanError as exc:
        app.logger.error("Rentman: errore recuperando equipment %s: %s", equipment_id, exc)
        cache[equipment_id] = None
        return None
    cache[equipment_id] = meta
    return meta


def _resolve_photo_payload(
    reference: Any,
    client: RentmanClient,
    cache: Dict[int, Optional[Mapping[str, Any]]],
) -> Optional[Dict[str, Any]]:
    if not reference:
        return None

    if isinstance(reference, str) and reference.startswith("http"):
        return {"name": "Foto materiale", "url": reference, "preview_url": reference}

    file_id = parse_reference(reference)
    if not isinstance(file_id, int):
        return None

    if file_id in cache:
        file_entry = cache[file_id]
    else:
        try:
            file_entry = client.get_file(file_id)
        except RentmanError as exc:
            app.logger.error("Rentman: errore recuperando file %s: %s", file_id, exc)
            file_entry = None
        cache[file_id] = file_entry

    if not file_entry:
        return None

    url = file_entry.get("url")
    preview = file_entry.get("proxy_url") or url
    if not url and not preview:
        return None

    return {
        "name": _normalize_attachment_name(file_entry),
        "url": url or preview,
        "preview_url": preview or url,
        "extension": _normalize_attachment_extension(file_entry.get("extension") or file_entry.get("type")),
    }


def fetch_project_materials(project_code: Optional[str]) -> Dict[str, List[Dict[str, Any]]]:
    result: Dict[str, List[Dict[str, Any]]] = {"items": [], "folders": []}
    code = (project_code or "").strip()
    if not code:
        return result

    client = get_rentman_client()
    if not client:
        return result

    try:
        project = client.find_project(code)
    except RentmanNotFound:
        app.logger.warning("Rentman: nessun progetto per materiali %s", code)
        return result
    except (RentmanAuthError, RentmanAPIError) as exc:
        app.logger.error("Rentman: errore durante la ricerca dei materiali per %s: %s", code, exc)
        return result

    if not project:
        app.logger.info("Rentman: progetto %s non trovato, nessun materiale", code)
        return result

    project_id = parse_reference(project.get("id")) or project.get("id")
    if not isinstance(project_id, int):
        app.logger.warning("Rentman: materiali non disponibili, id progetto non valido per %s (%s)", code, project.get("id"))
        return result

    try:
        records = client.get_project_planned_equipment(project_id)
        app.logger.info(
            "Rentman: materiali pianificati raw (primi 3)=\n%s",
            json.dumps(records[:3], ensure_ascii=False, indent=2) if records else "[]",
        )
    except RentmanError as exc:
        app.logger.error("Rentman: errore leggendo i materiali del progetto %s: %s", code, exc)
        return result

    equipment_cache: Dict[int, Optional[Mapping[str, Any]]] = {}
    file_cache: Dict[int, Optional[Mapping[str, Any]]] = {}
    group_lookup = _collect_material_groups(client, project_id)
    default_group_label = "Altri materiali"
    materials: List[Dict[str, Any]] = []
    for entry in records:
        if not isinstance(entry, Mapping):
            continue
        quantity_value, quantity_label = _extract_material_quantity(entry)
        status_code, status_label = _material_status(entry)
        equipment_meta = _resolve_equipment_meta(entry.get("equipment"), client, equipment_cache)
        length = _coerce_float(entry.get("length")) or ( _coerce_float(equipment_meta.get("length")) if equipment_meta else None)
        width = _coerce_float(entry.get("width")) or ( _coerce_float(equipment_meta.get("width")) if equipment_meta else None)
        height = _coerce_float(entry.get("height")) or ( _coerce_float(equipment_meta.get("height")) if equipment_meta else None)
        weight_value = _coerce_float(entry.get("weight"))
        if weight_value is None and equipment_meta:
            weight_value = _coerce_float(equipment_meta.get("weight"))
        dimensions_label = _format_dimensions_label(length, width, height)
        weight_label = _format_weight_label(weight_value)
        image_reference = entry.get("image") or (equipment_meta.get("image") if equipment_meta else None)
        photo_payload = _resolve_photo_payload(image_reference, client, file_cache)
        group_id = parse_reference(entry.get("equipment_group"))
        group_entry = group_lookup.get(group_id) if isinstance(group_id, int) else None
        group_name = group_entry.get("name") if group_entry else default_group_label
        group_path = group_entry.get("path") if group_entry else group_name
        notes: List[str] = []
        for key in ("internal_remark", "external_remark"):
            value = entry.get(key)
            if isinstance(value, str):
                stripped = value.strip()
                if stripped:
                    notes.append(stripped)
        note_text = " · ".join(dict.fromkeys(notes)) if notes else ""
        materials.append(
            {
                "id": entry.get("id"),
                "name": _normalize_material_name(entry),
                "quantity": quantity_value,
                "quantity_label": quantity_label,
                "period_start": entry.get("planperiod_start"),
                "period_end": entry.get("planperiod_end"),
                "note": note_text,
                "status": status_label,
                "status_code": status_code,
                "has_missings": bool(_is_truthy(entry.get("has_missings"))),
                "is_option": bool(_is_truthy(entry.get("is_option"))),
                "dimensions_label": dimensions_label,
                "weight_label": weight_label,
                "photo": photo_payload,
                "group_id": group_id,
                "group_name": group_name,
                "group_path": group_path,
            }
        )

    materials.sort(
        key=lambda item: (
            str(item.get("group_path") or item.get("group_name") or "").lower(),
            item.get("status_code"),
            str(item.get("name") or "").lower(),
        )
    )

    folders = _collect_project_folders(client, project_id)
    result["items"] = materials
    result["folders"] = folders
    return result


def now_ms() -> int:
    return int(time.time() * 1000)


def ensure_app_users_table(db: DatabaseLike) -> None:
    statement = APP_USERS_TABLE_MYSQL if DB_VENDOR == "mysql" else APP_USERS_TABLE_SQLITE
    cursor = db.execute(statement)
    try:
        cursor.close()
    except AttributeError:
        pass
    
    # Migrazione: aggiungi colonna full_name se non esiste
    if DB_VENDOR == "mysql":
        try:
            db.execute("ALTER TABLE app_users ADD COLUMN full_name VARCHAR(255) DEFAULT NULL")
            db.commit()
        except Exception:
            pass  # Colonna già esistente
        # Migrazione: aggiungi colonna rentman_crew_id se non esiste
        try:
            db.execute("ALTER TABLE app_users ADD COLUMN rentman_crew_id INT DEFAULT NULL")
            db.commit()
        except Exception:
            pass  # Colonna già esistente
        # Migrazione: aggiungi colonne per progetto corrente supervisor
        try:
            db.execute("ALTER TABLE app_users ADD COLUMN current_project_code VARCHAR(64) DEFAULT NULL")
            db.commit()
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE app_users ADD COLUMN current_project_name VARCHAR(255) DEFAULT NULL")
            db.commit()
        except Exception:
            pass
        # Migrazione: aggiungi colonne external_id e external_group_id per CedolinoWeb
        try:
            db.execute("ALTER TABLE app_users ADD COLUMN external_id VARCHAR(255) DEFAULT NULL")
            db.commit()
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE app_users ADD COLUMN external_group_id VARCHAR(255) DEFAULT NULL")
            db.commit()
        except Exception:
            pass
    else:
        # SQLite
        try:
            db.execute("ALTER TABLE app_users ADD COLUMN full_name TEXT")
            db.commit()
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE app_users ADD COLUMN rentman_crew_id INTEGER")
            db.commit()
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE app_users ADD COLUMN current_project_code TEXT")
            db.commit()
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE app_users ADD COLUMN current_project_name TEXT")
            db.commit()
        except Exception:
            pass
        # Migrazione: aggiungi colonne external_id e external_group_id per CedolinoWeb
        try:
            db.execute("ALTER TABLE app_users ADD COLUMN external_id TEXT")
            db.commit()
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE app_users ADD COLUMN external_group_id TEXT")
            db.commit()
        except Exception:
            pass
    
    # Migrazione: aggiungi colonna group_id per collegamento a user_groups
    if DB_VENDOR == "mysql":
        try:
            db.execute("ALTER TABLE app_users ADD COLUMN group_id INT DEFAULT NULL")
            db.commit()
        except Exception:
            pass
    else:
        try:
            db.execute("ALTER TABLE app_users ADD COLUMN group_id INTEGER DEFAULT NULL")
            db.commit()
        except Exception:
            pass


def ensure_user_groups_table(db: DatabaseLike) -> None:
    """Crea la tabella user_groups se non esiste."""
    statement = USER_GROUPS_TABLE_MYSQL if DB_VENDOR == "mysql" else USER_GROUPS_TABLE_SQLITE
    cursor = db.execute(statement)
    try:
        cursor.close()
    except AttributeError:
        pass
    
    # Migrazione: aggiunge colonna gps_location_name se non esiste
    try:
        if DB_VENDOR == "mysql":
            db.execute("ALTER TABLE user_groups ADD COLUMN gps_location_name VARCHAR(255) DEFAULT NULL COMMENT 'Nome sede GPS associata al gruppo'")
        else:
            db.execute("ALTER TABLE user_groups ADD COLUMN gps_location_name TEXT DEFAULT NULL")
        db.commit()
    except Exception:
        pass  # Colonna già esistente


def ensure_session_override_table(db: DatabaseLike) -> None:
    if DB_VENDOR == "mysql":
        cursor = db.execute(SESSION_OVERRIDES_TABLE_MYSQL)
        try:
            cursor.close()
        except AttributeError:
            pass
        # Migrazione: aggiungi colonna project_code se non esiste
        try:
            db.execute("ALTER TABLE activity_session_overrides ADD COLUMN project_code VARCHAR(64) DEFAULT NULL")
            db.commit()
        except Exception:
            pass  # Colonna già esistente
        try:
            db.execute("CREATE INDEX idx_session_override_project ON activity_session_overrides(project_code)")
            db.commit()
        except Exception:
            pass  # Indice già esistente


def ensure_persistent_session_table(db: DatabaseLike) -> None:
    statement = (
        PERSISTENT_SESSIONS_TABLE_MYSQL if DB_VENDOR == "mysql" else PERSISTENT_SESSIONS_TABLE_SQLITE
    )
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass


def ensure_equipment_checks_table(db: DatabaseLike) -> None:
    statement = EQUIPMENT_CHECKS_TABLE_MYSQL if DB_VENDOR == "mysql" else EQUIPMENT_CHECKS_TABLE_SQLITE
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass


def ensure_project_materials_cache_table(db: DatabaseLike) -> None:
    statement = (
        PROJECT_MATERIALS_CACHE_TABLE_MYSQL if DB_VENDOR == "mysql" else PROJECT_MATERIALS_CACHE_TABLE_SQLITE
    )
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass


def ensure_push_notification_read_column(db: DatabaseLike) -> None:
    """Assicura che la colonna read_at esista in push_notification_log."""
    if DB_VENDOR == "mysql":
        try:
            db.execute("ALTER TABLE push_notification_log ADD COLUMN read_at BIGINT DEFAULT NULL")
            db.commit()
        except Exception:
            pass  # Colonna già esistente
    else:
        try:
            db.execute("ALTER TABLE push_notification_log ADD COLUMN read_at INTEGER DEFAULT NULL")
            db.commit()
        except Exception:
            pass  # Colonna già esistente


def ensure_local_equipment_table(db: DatabaseLike) -> None:
    statement = LOCAL_EQUIPMENT_TABLE_MYSQL if DB_VENDOR == "mysql" else LOCAL_EQUIPMENT_TABLE_SQLITE
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass


def ensure_project_photos_table(db: DatabaseLike) -> None:
    statement = PROJECT_PHOTOS_TABLE_MYSQL if DB_VENDOR == "mysql" else PROJECT_PHOTOS_TABLE_SQLITE
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass


def ensure_employee_shifts_table(db: DatabaseLike) -> None:
    """Assicura l'esistenza della tabella employee_shifts per turni impiegati non-Rentman."""
    statement = EMPLOYEE_SHIFTS_TABLE_MYSQL if DB_VENDOR == "mysql" else EMPLOYEE_SHIFTS_TABLE_SQLITE
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass
    
    # Migrazione: aggiunge colonna location_name se non esiste
    try:
        if DB_VENDOR == "mysql":
            db.execute("ALTER TABLE employee_shifts ADD COLUMN location_name VARCHAR(255) DEFAULT NULL COMMENT 'Nome sede GPS associata al turno'")
        else:
            db.execute("ALTER TABLE employee_shifts ADD COLUMN location_name TEXT")
        db.commit()
    except Exception:
        pass  # Colonna già esistente
    
    # Migrazione: aggiunge colonna shift_name se non esiste
    try:
        if DB_VENDOR == "mysql":
            db.execute("ALTER TABLE employee_shifts ADD COLUMN shift_name VARCHAR(100) DEFAULT NULL COMMENT 'Nome identificativo del turno' AFTER break_end")
        else:
            db.execute("ALTER TABLE employee_shifts ADD COLUMN shift_name TEXT")
        db.commit()
    except Exception:
        pass  # Colonna già esistente


# Cartella per salvare le foto del progetto
PHOTOS_UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads", "photos")
os.makedirs(PHOTOS_UPLOAD_FOLDER, exist_ok=True)

ALLOWED_PHOTO_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp", "heic", "heif"}
MAX_PHOTO_SIZE = 10 * 1024 * 1024  # 10 MB


def allowed_photo_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_PHOTO_EXTENSIONS


def _last_insert_id(db: DatabaseLike) -> Optional[int]:
    query = "SELECT LAST_INSERT_ID() AS lid" if DB_VENDOR == "mysql" else "SELECT last_insert_rowid() AS lid"
    row = db.execute(query).fetchone()
    if not row:
        return None
    value = row.get("lid") if isinstance(row, Mapping) else row[0]
    try:
        return int(value)
    except (TypeError, ValueError):
        return None
        return

    for stmt in SESSION_OVERRIDES_TABLE_SQLITE.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass


def init_db() -> None:
    if DB_VENDOR == "mysql":
        db = MySQLConnection(DATABASE_SETTINGS)
        try:
            for statement in MYSQL_SCHEMA_STATEMENTS:
                cursor = db.execute(statement)
                cursor.close()
            _ensure_entered_ts_column(db, "BIGINT")
            purge_legacy_seed(db)
            ensure_app_users_table(db)
            ensure_session_override_table(db)
            ensure_persistent_session_table(db)
            ensure_equipment_checks_table(db)
            ensure_project_materials_cache_table(db)
            ensure_employee_shifts_table(db)
            ensure_user_groups_table(db)
            bootstrap_user_store(db)
            db.commit()
        finally:
            db.close()
        return

    db = sqlite3.connect(DATABASE)
    try:
        db.row_factory = sqlite3.Row
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS activities (
                activity_id TEXT NOT NULL,
                project_code TEXT NOT NULL DEFAULT '',
                label TEXT NOT NULL,
                sort_order INTEGER NOT NULL,
                plan_start TEXT,
                plan_end TEXT,
                planned_members INTEGER,
                planned_duration_ms INTEGER,
                notes TEXT,
                PRIMARY KEY (activity_id, project_code)
            );

            CREATE TABLE IF NOT EXISTS member_state (
                member_key TEXT NOT NULL,
                project_code TEXT NOT NULL DEFAULT '',
                member_name TEXT NOT NULL,
                activity_id TEXT,
                running INTEGER NOT NULL DEFAULT 0,
                start_ts INTEGER,
                elapsed_cached INTEGER NOT NULL DEFAULT 0,
                pause_start INTEGER,
                entered_ts INTEGER,
                PRIMARY KEY (member_key, project_code)
            );

            CREATE TABLE IF NOT EXISTS event_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_code TEXT NOT NULL DEFAULT '',
                ts INTEGER NOT NULL,
                kind TEXT NOT NULL,
                member_key TEXT,
                details TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_event_project ON event_log(project_code);

            CREATE TABLE IF NOT EXISTS app_state (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS push_subscriptions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL,
                endpoint TEXT NOT NULL UNIQUE,
                p256dh TEXT NOT NULL,
                auth TEXT NOT NULL,
                content_encoding TEXT,
                user_agent TEXT,
                expiration_time INTEGER,
                created_ts INTEGER NOT NULL,
                updated_ts INTEGER NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_push_username ON push_subscriptions(username);

            CREATE TABLE IF NOT EXISTS push_notification_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                kind TEXT NOT NULL,
                activity_id TEXT,
                username TEXT,
                title TEXT NOT NULL,
                body TEXT,
                payload TEXT,
                sent_ts INTEGER NOT NULL,
                created_ts INTEGER NOT NULL,
                read_at INTEGER DEFAULT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_push_log_user ON push_notification_log(username);
            CREATE INDEX IF NOT EXISTS idx_push_log_sent ON push_notification_log(sent_ts);

            CREATE TABLE IF NOT EXISTS app_users (
                username TEXT PRIMARY KEY,
                password_hash TEXT NOT NULL,
        ensure_app_users_table(db)
                display_name TEXT NOT NULL,
                full_name TEXT,
                role TEXT NOT NULL DEFAULT 'user',
                is_active INTEGER NOT NULL DEFAULT 1,
                created_ts INTEGER NOT NULL,
                updated_ts INTEGER NOT NULL
            );
            CREATE TABLE IF NOT EXISTS equipment_checks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_code TEXT NOT NULL,
                item_key TEXT NOT NULL,
                checked_ts INTEGER NOT NULL,
                username TEXT,
                created_ts INTEGER NOT NULL,
                updated_ts INTEGER NOT NULL
            );
            CREATE UNIQUE INDEX IF NOT EXISTS uq_equipment_project_item ON equipment_checks(project_code, item_key);
            CREATE INDEX IF NOT EXISTS idx_equipment_project ON equipment_checks(project_code);
            CREATE TABLE IF NOT EXISTS project_materials_cache (
                project_code TEXT PRIMARY KEY,
                project_name TEXT NOT NULL,
                data_json TEXT NOT NULL,
                created_ts INTEGER NOT NULL,
                updated_ts INTEGER NOT NULL
            );
            """
        )
        _ensure_entered_ts_column(db, "INTEGER")
        purge_legacy_seed(db)
        ensure_persistent_session_table(db)
        ensure_equipment_checks_table(db)
        ensure_project_materials_cache_table(db)
        ensure_employee_shifts_table(db)
        ensure_user_groups_table(db)
        bootstrap_user_store(db)
        db.commit()
    finally:
        db.close()


def _ensure_entered_ts_column(db: DatabaseLike, column_type: str) -> None:
    try:
        db.execute(f"ALTER TABLE member_state ADD COLUMN entered_ts {column_type}")
    except Exception:
        return


def purge_legacy_seed(db: DatabaseLike) -> None:
    try:
        project_code = get_app_state(db, "project_code")
    except Exception:  # pragma: no cover - ignora errori iniziali
        return

    if project_code != DEMO_PROJECT_CODE:
        return

    db.execute("DELETE FROM activities")
    db.execute("DELETE FROM member_state")
    db.execute("DELETE FROM event_log")
    db.execute(
        f"DELETE FROM app_state WHERE {APP_STATE_KEY_COLUMN} IN ('project_code','project_name','activity_plan_meta','push_notified_activities','long_running_member_notifications')"
    )


def mock_fetch_project(project_code: str, project_date: Optional[str] = None) -> Optional[Dict[str, Any]]:
    code = (project_code or "").strip().upper()
    if not code:
        return None
    plan = fetch_rentman_plan(code, project_date)
    if plan:
        app.logger.warning("Rentman: piano caricato da API per %s (data: %s)", code, project_date)
        return plan
    external = load_external_projects().get(code)
    plan = external or MOCK_PROJECTS.get(code)
    if plan is None:
        app.logger.warning("Rentman: nessun piano disponibile per %s", code)
        return None
    app.logger.warning("Rentman: uso piano locale per %s", code)
    result = deepcopy(plan)
    result["project_code"] = code
    return result


def set_app_state(db: DatabaseLike, key: str, value: str) -> None:
    if DB_VENDOR == "mysql":
        db.execute(
            """
            INSERT INTO app_state(`key`, value) VALUES(?, ?)
            ON DUPLICATE KEY UPDATE value=VALUES(value)
            """,
            (key, value),
        )
        return

    db.execute(
        """
        INSERT INTO app_state(key, value) VALUES(?, ?)
        ON CONFLICT(key) DO UPDATE SET value=excluded.value
        """,
        (key, value),
    )


def get_app_state(db: DatabaseLike, key: str) -> Optional[str]:
    try:
        query = f"SELECT value FROM app_state WHERE {APP_STATE_KEY_COLUMN}=?"
        row = db.execute(query, (key,)).fetchone()
    except sqlite3.OperationalError:
        return None
    except Exception as exc:  # pragma: no cover - gestione MySQL
        if pymysql_err is not None and isinstance(exc, pymysql_err.ProgrammingError):
            return None
        raise
    return row["value"] if row else None


def get_push_notified_map(db: DatabaseLike) -> Dict[str, Any]:
    raw = get_app_state(db, PUSH_NOTIFIED_STATE_KEY)
    if not raw:
        return {}
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return {}
    if isinstance(data, dict):
        return data
    return {}


def save_push_notified_map(db: DatabaseLike, payload: Mapping[str, Any]) -> None:
    set_app_state(db, PUSH_NOTIFIED_STATE_KEY, json.dumps(payload))


def get_long_running_notified_map(db: DatabaseLike) -> Dict[str, Any]:
    raw = get_app_state(db, LONG_RUNNING_STATE_KEY)
    if not raw:
        return {}
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return {}
    if isinstance(data, dict):
        return data
    return {}


def save_long_running_notified_map(db: DatabaseLike, payload: Mapping[str, Any]) -> None:
    set_app_state(db, LONG_RUNNING_STATE_KEY, json.dumps(payload))


def parse_iso_to_ms(value: Optional[str]) -> Optional[int]:
    if not value:
        return None
    slug = value.replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(slug)
    except ValueError:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return int(dt.timestamp() * 1000)


def clear_project_state(db: DatabaseLike, project_code: Optional[str] = None) -> None:
    """Rimuove il progetto e i relativi dati dal database."""
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    if project_code:
        # Cancella solo i dati del progetto specifico
        db.execute(f"DELETE FROM activities WHERE project_code = {placeholder}", (project_code,))
        db.execute(f"DELETE FROM member_state WHERE project_code = {placeholder}", (project_code,))
        db.execute(f"DELETE FROM event_log WHERE project_code = {placeholder}", (project_code,))
        delete_project_materials_cache(db, project_code)
    else:
        # Fallback: cancella tutto (per retrocompatibilità)
        db.execute("DELETE FROM activities")
        db.execute("DELETE FROM member_state")
        db.execute("DELETE FROM event_log")
        db.execute(
            f"DELETE FROM app_state WHERE {APP_STATE_KEY_COLUMN} IN ('project_code','project_name','activity_plan_meta','push_notified_activities','long_running_member_notifications')"
        )


def has_active_member_sessions(db: DatabaseLike) -> bool:
    """Restituisce True se esistono timer in corso o posti in pausa."""

    row = db.execute(
        """
        SELECT 1 FROM member_state
        WHERE running=? OR pause_start IS NOT NULL OR COALESCE(elapsed_cached, 0) > 0
        LIMIT 1
        """,
        (RUN_STATE_RUNNING,),
    ).fetchone()
    return row is not None


def apply_project_plan(db: DatabaseLike, plan: Dict[str, Any]) -> None:
    activities = list(plan.get("activities") or [])
    team = list(plan.get("team") or [])
    project_code = str(plan.get("project_code") or "UNKNOWN")
    project_name = str(plan.get("project_name") or project_code)
    
    # Cancella solo i dati del progetto specifico (non tocca altri progetti)
    clear_project_state(db, project_code)

    planned_counts: Dict[str, int] = {}
    for member in team:
        activity_id = (member.get("activity_id") or "").strip()
        if activity_id:
            planned_counts[activity_id] = planned_counts.get(activity_id, 0) + 1

    activity_rows: List[tuple] = []
    for index, activity in enumerate(activities, start=1):
        activity_id = activity.get("id")
        activity_key = (str(activity_id).strip() if activity_id is not None else "")
        plan_start = activity.get("plan_start")
        plan_end = activity.get("plan_end")
        planned_members = planned_counts.get(activity_key, 0)
        planned_duration_ms = compute_planned_duration_ms(
            plan_start,
            plan_end,
            planned_members,
        )
        activity_rows.append(
            (
                activity_id,
                project_code,
                activity.get("label"),
                index,
                plan_start,
                plan_end,
                planned_members,
                planned_duration_ms,
                activity.get("notes"),
            )
        )

    db.executemany(
        """
        INSERT INTO activities(
            activity_id, project_code, label, sort_order, plan_start, plan_end,
            planned_members, planned_duration_ms, notes
        ) VALUES(?,?,?,?,?,?,?,?,?)
        """,
        activity_rows,
    )

    now = now_ms()
    member_rows: List[tuple] = []
    seen_keys = set()
    for member in team:
        raw_key = (member.get("key") or "").strip()
        name = (member.get("name") or raw_key or "Operatore").strip()
        key = raw_key or name.lower().replace(" ", "-")
        while key in seen_keys:
            key = f"{key}-dup"
        seen_keys.add(key)
        activity_id = (member.get("activity_id") or "").strip() or None
        # Non avviare automaticamente il timer, il capo squadra lo avvierà manualmente
        running = RUN_STATE_PAUSED
        start_ts = None
        member_rows.append(
            (key, project_code, name, activity_id, running, start_ts, 0, None, None)
        )

    if member_rows:
        db.executemany(
            """
            INSERT INTO member_state(
                member_key, project_code, member_name, activity_id, running, start_ts, elapsed_cached, pause_start, entered_ts
            ) VALUES(?,?,?,?,?,?,?,?,?)
            """,
            member_rows,
        )

    activity_meta = {}
    for activity in activities:
        activity_id = activity.get("id")
        if not activity_id:
            continue
        key = str(activity_id).strip()
        if not key:
            continue
        plan_start = activity.get("plan_start")
        plan_end = activity.get("plan_end")
        planned_members = planned_counts.get(key, 0)
        activity_meta[key] = {
            "plan_start": plan_start,
            "plan_end": plan_end,
            "planned_members": planned_members,
            "planned_duration_ms": compute_planned_duration_ms(
                plan_start,
                plan_end,
                planned_members,
            ),
            "actual_runtime_ms": 0,
        }

    set_app_state(db, "project_code", project_code)
    set_app_state(db, "project_name", project_name)
    set_app_state(db, "activity_plan_meta", json.dumps(activity_meta))
    set_app_state(db, PUSH_NOTIFIED_STATE_KEY, json.dumps({}))
    set_app_state(db, LONG_RUNNING_STATE_KEY, json.dumps({}))

    db.execute(
        "INSERT INTO event_log(ts, kind, details) VALUES(?,?,?)",
        (
            now,
            "project_load",
            json.dumps({"project_code": project_code, "project_name": project_name}),
        ),
    )


def seed_demo_data(db: DatabaseLike) -> None:
    plan = mock_fetch_project(DEMO_PROJECT_CODE, None)
    if plan is None:
        raise RuntimeError("Demo project configuration missing")
    apply_project_plan(db, plan)


def get_db() -> DatabaseLike:
    if "db" not in g:
        if DB_VENDOR == "mysql":
            g.db = MySQLConnection(DATABASE_SETTINGS)
        else:
            conn = sqlite3.connect(DATABASE)
            conn.row_factory = sqlite3.Row
            g.db = conn
        try:
            ensure_activity_schema(g.db)
            ensure_project_code_columns(g.db)
            ensure_app_users_table(g.db)
            ensure_session_override_table(g.db)
            ensure_persistent_session_table(g.db)
            ensure_equipment_checks_table(g.db)
            ensure_project_materials_cache_table(g.db)
            ensure_warehouse_manual_projects_table(g.db)
        except Exception:
            app.logger.exception("Impossibile aggiornare lo schema attività")
    return g.db


@app.teardown_appcontext
def close_db(_: BaseException | None) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


def compute_elapsed(row: Mapping[str, Any], reference: int) -> int:
    elapsed = row["elapsed_cached"] or 0
    if row["running"] == RUN_STATE_RUNNING:
        start_ts = row["start_ts"] or reference
        elapsed += max(0, reference - start_ts)
    return elapsed


def row_value(row: Mapping[str, Any], key: str) -> Optional[Any]:
    if hasattr(row, "get"):
        try:
            return row.get(key)  # type: ignore[call-arg]
        except Exception:
            pass
    try:
        return row[key]
    except Exception:
        return None


def fetch_equipment_checks(db: DatabaseLike, project_code: Optional[str]) -> Dict[str, int]:
    if not project_code:
        return {}
    rows = db.execute(
        "SELECT item_key, checked_ts FROM equipment_checks WHERE project_code=?",
        (project_code,),
    ).fetchall()
    result: Dict[str, int] = {}
    for row in rows:
        item_key = row_value(row, "item_key")
        if not item_key:
            continue
        timestamp_raw = row_value(row, "checked_ts")
        try:
            timestamp_int = int(timestamp_raw)
        except (TypeError, ValueError):
            continue
        result[str(item_key)] = timestamp_int
    return result


def persist_equipment_check(
    db: DatabaseLike,
    *,
    project_code: str,
    item_key: str,
    checked: bool,
    username: Optional[str] = None,
) -> Optional[int]:
    normalized_project = (project_code or "").strip()
    normalized_item = (item_key or "").strip()
    if not normalized_project or not normalized_item:
        return None
    if checked:
        now = now_ms()
        if DB_VENDOR == "mysql":
            db.execute(
                """
                INSERT INTO equipment_checks(project_code, item_key, checked_ts, username, created_ts, updated_ts)
                VALUES(?,?,?,?,?,?)
                ON DUPLICATE KEY UPDATE checked_ts=VALUES(checked_ts), username=VALUES(username), updated_ts=VALUES(updated_ts)
                """,
                (normalized_project, normalized_item, now, username, now, now),
            )
        else:
            db.execute(
                """
                INSERT INTO equipment_checks(project_code, item_key, checked_ts, username, created_ts, updated_ts)
                VALUES(?,?,?,?,?,?)
                ON CONFLICT(project_code, item_key) DO UPDATE SET
                    checked_ts=excluded.checked_ts,
                    username=excluded.username,
                    updated_ts=excluded.updated_ts
                """,
                (normalized_project, normalized_item, now, username, now, now),
            )
        return now

    db.execute(
        "DELETE FROM equipment_checks WHERE project_code=? AND item_key=?",
        (normalized_project, normalized_item),
    )
    return None


def load_project_materials_cache(db: DatabaseLike, project_code: Optional[str]) -> Optional[Dict[str, Any]]:
    if not project_code:
        return None
    row = db.execute(
        "SELECT project_name, data_json, updated_ts FROM project_materials_cache WHERE project_code=?",
        (project_code,),
    ).fetchone()
    if not row:
        return None
    raw_payload = row_value(row, "data_json")
    try:
        payload = json.loads(raw_payload) if raw_payload else {}
    except json.JSONDecodeError:
        payload = {}
    items = payload.get("items")
    if not isinstance(items, list):
        items = []
    folders = payload.get("folders")
    if not isinstance(folders, list):
        folders = []
    return {
        "project": {
            "code": project_code,
            "name": row_value(row, "project_name") or project_code,
        },
        "items": items,
        "folders": folders,
        "updated_ts": row_value(row, "updated_ts"),
    }


def save_project_materials_cache(
    db: DatabaseLike,
    project_code: Optional[str],
    project_name: Optional[str],
    *,
    items: Sequence[Mapping[str, Any]] | Sequence[Any],
    folders: Sequence[Mapping[str, Any]] | Sequence[Any],
) -> int:
    if not project_code:
        return 0
    normalized_name = project_name or project_code
    sanitized_items = list(items or [])
    sanitized_folders = list(folders or [])
    payload = json.dumps({"items": sanitized_items, "folders": sanitized_folders}, ensure_ascii=False)
    now = now_ms()
    if DB_VENDOR == "mysql":
        db.execute(
            """
            INSERT INTO project_materials_cache(project_code, project_name, data_json, created_ts, updated_ts)
            VALUES(?,?,?,?,?)
            ON DUPLICATE KEY UPDATE project_name=VALUES(project_name), data_json=VALUES(data_json), updated_ts=VALUES(updated_ts)
            """,
            (project_code, normalized_name, payload, now, now),
        )
    else:
        db.execute(
            """
            INSERT INTO project_materials_cache(project_code, project_name, data_json, created_ts, updated_ts)
            VALUES(?,?,?,?,?)
            ON CONFLICT(project_code) DO UPDATE SET
                project_name=excluded.project_name,
                data_json=excluded.data_json,
                updated_ts=excluded.updated_ts
            """,
            (project_code, normalized_name, payload, now, now),
        )
    return now


def delete_project_materials_cache(db: DatabaseLike, project_code: Optional[str]) -> None:
    if not project_code:
        return
    db.execute("DELETE FROM project_materials_cache WHERE project_code=?", (project_code,))


def find_last_move_ts(db: DatabaseLike, member_key: str, activity_id: str) -> Optional[int]:
    if not member_key or not activity_id:
        return None
    rows = db.execute(
        "SELECT ts, details FROM event_log WHERE member_key=? AND kind='move' ORDER BY ts DESC LIMIT 200",
        (member_key,),
    ).fetchall()
    for row in rows:
        try:
            details = json.loads(row["details"] or "{}")
        except Exception:
            continue
        if str(details.get("to")) == activity_id:
            return row["ts"]
    return None


def fetch_member(db: DatabaseLike, member_key: str, project_code: Optional[str] = None) -> Optional[Mapping[str, Any]]:
    if not member_key:
        return None
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    if project_code is not None:
        return db.execute(
            f"SELECT * FROM member_state WHERE member_key={placeholder} AND project_code={placeholder}",
            (member_key, project_code),
        ).fetchone()
    return db.execute(
        f"SELECT * FROM member_state WHERE member_key={placeholder}",
        (member_key,),
    ).fetchone()


def fetch_push_subscriptions(db: DatabaseLike) -> List[Mapping[str, Any]]:
    rows = db.execute(
        "SELECT username, endpoint, p256dh, auth, content_encoding, user_agent FROM push_subscriptions"
    ).fetchall()
    return [dict(row) for row in rows]  # type: ignore[list-item]


def remove_push_subscription(db: DatabaseLike, endpoint: str) -> None:
    if not endpoint:
        return
    db.execute("DELETE FROM push_subscriptions WHERE endpoint=?", (endpoint,))


def record_push_notification(
    db: DatabaseLike,
    *,
    kind: str,
    title: str,
    body: Optional[str],
    payload: Mapping[str, Any],
    activity_id: Optional[str] = None,
    username: Optional[str] = None,
) -> None:
    sent_ts = now_ms()
    try:
        serialized = json.dumps(payload, ensure_ascii=False)
    except TypeError:
        serialized = json.dumps({"payload_repr": repr(payload)}, ensure_ascii=False)
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    db.execute(
        f"""
        INSERT INTO push_notification_log(
            kind, activity_id, username, title, body, payload, sent_ts, created_ts
        ) VALUES({placeholder},{placeholder},{placeholder},{placeholder},{placeholder},{placeholder},{placeholder},{placeholder})
        """,
        (
            kind,
            activity_id,
            username,
            title,
            body,
            serialized,
            sent_ts,
            sent_ts,
        ),
    )
    db.commit()


def fetch_recent_push_notifications(
    db: DatabaseLike,
    *,
    username: str,
    limit: Optional[int] = None,
) -> List[Dict[str, Any]]:
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    sql = f"""
        SELECT id, kind, activity_id, username, title, body, payload, sent_ts, created_ts, read_at
        FROM push_notification_log
        WHERE username = {placeholder}
        ORDER BY sent_ts DESC, id DESC
    """
    params: List[Any] = [username]
    if limit is not None and limit > 0:
        safe_limit = max(1, min(limit, 1000))
        sql += f" LIMIT {placeholder}"
        params.append(safe_limit)
    rows = db.execute(sql, tuple(params)).fetchall()

    items: List[Dict[str, Any]] = []
    for row in rows:
        raw_payload = row["payload"]
        try:
            payload = json.loads(raw_payload) if raw_payload else None
        except json.JSONDecodeError:
            payload = None
        items.append(
            {
                "id": row["id"],
                "kind": row["kind"],
                "activity_id": row["activity_id"],
                "username": row["username"],
                "title": row["title"],
                "body": row["body"],
                "payload": payload,
                "sent_ts": row["sent_ts"],
                "created_ts": row["created_ts"],
                "read_at": row["read_at"],
                "read": row["read_at"] is not None,
            }
        )
    return items


def format_duration_ms(ms: Any) -> Optional[str]:
    if not isinstance(ms, (int, float)):
        return None
    total_seconds = max(0, int(ms // 1000))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours:02}:{minutes:02}:{seconds:02}"
    return f"{minutes:02}:{seconds:02}"


def parse_iso_date(value: Optional[str]) -> Optional[date]:
    if not value:
        return None
    candidate = value.strip()
    if not candidate:
        return None
    try:
        return date.fromisoformat(candidate)
    except ValueError:
        return None


def _normalize_epoch_ms(value: Any) -> Optional[int]:
    ms = _coerce_int(value)
    if ms is None:
        return None
    return max(0, ms)


def _fetch_session_override_rows(
    db: DatabaseLike,
    *,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    member_filter: Optional[str] = None,
    activity_filter: Optional[str] = None,
) -> List[Dict[str, Any]]:
    clauses: List[str] = []
    params: List[Any] = []

    if member_filter:
        clauses.append("LOWER(member_key)=LOWER(?)")
        params.append(member_filter.strip())
    if activity_filter:
        clauses.append("activity_id=?")
        params.append(activity_filter.strip())

    query = "SELECT * FROM activity_session_overrides"
    if clauses:
        query += " WHERE " + " AND ".join(clauses)
    query += " ORDER BY start_ts DESC"

    rows = db.execute(query, tuple(params) if params else None).fetchall()
    results: List[Dict[str, Any]] = []
    for row in rows:
        record = dict(row)
        start_dt = datetime.fromtimestamp(record["start_ts"] / 1000, tz=timezone.utc).date()
        if start_date and start_dt < start_date:
            continue
        if end_date and start_dt > end_date:
            continue
        results.append(record)
    return results


def _override_row_to_session(row: Mapping[str, Any]) -> Dict[str, Any]:
    start_ts = int(row.get("start_ts") or 0)
    end_ts_value = row.get("end_ts")
    end_ts = int(end_ts_value) if end_ts_value is not None else start_ts
    net_ms = max(0, int(row.get("net_ms") or 0))
    pause_ms = max(0, int(row.get("pause_ms") or 0))
    pause_count = max(0, int(row.get("pause_count") or 0))
    manual_entry = bool(row.get("manual_entry"))
    note = row.get("note") or ""
    status = row.get("status") or "completed"

    payload = {
        "member_key": row.get("member_key"),
        "member_name": row.get("member_name"),
        "activity_id": row.get("activity_id"),
        "activity_label": row.get("activity_label") or row.get("activity_id"),
        "project_code": row.get("project_code") or "",
        "start_ts": start_ts,
        "end_ts": end_ts,
        "status": status if status in {"completed", "running"} else "completed",
        "net_ms": net_ms,
        "pause_ms": pause_ms,
        "pause_count": pause_count,
        "auto_closed": False,
        "override_id": row.get("id"),
        "manual_entry": manual_entry,
        "note": note,
        "source_member_key": row.get("source_member_key"),
        "source_activity_id": row.get("source_activity_id"),
        "source_start_ts": row.get("source_start_ts"),
    }
    return payload


def build_session_rows(
    db: DatabaseLike,
    *,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    member_filter: Optional[str] = None,
    activity_filter: Optional[str] = None,
    project_filter: Optional[str] = None,
) -> List[Dict[str, Any]]:
    activity_rows = db.execute(
        "SELECT activity_id, project_code, label, planned_duration_ms, notes FROM activities ORDER BY sort_order, label"
    ).fetchall()
    # Mappa con chiave (activity_id, project_code) per supportare attività multi-progetto
    activity_map = {(row["activity_id"], row["project_code"]): row["label"] for row in activity_rows}
    activity_planned_map = {(row["activity_id"], row["project_code"]): row["planned_duration_ms"] for row in activity_rows}
    activity_notes_map = {(row["activity_id"], row["project_code"]): row["notes"] or "" for row in activity_rows}
    # Fallback senza project_code per retrocompatibilità
    for row in activity_rows:
        if row["activity_id"] not in activity_map:
            activity_map[row["activity_id"]] = row["label"]
            activity_planned_map[row["activity_id"]] = row["planned_duration_ms"]
            activity_notes_map[row["activity_id"]] = row["notes"] or ""

    query = (
        "SELECT el.ts, el.kind, el.member_key, el.details, ms.member_name "
        "FROM event_log el "
        "LEFT JOIN member_state ms ON el.member_key = ms.member_key "
        "WHERE el.kind IN ('project_load', 'move', 'finish_activity', 'pause_member', 'resume_member') "
        "ORDER BY el.ts ASC"
    )
    event_rows = db.execute(query).fetchall()

    member_filter_norm = member_filter.strip().lower() if member_filter else None
    activity_filter_norm = activity_filter.strip() if activity_filter else None
    project_filter_norm = project_filter.strip() if project_filter else None

    sessions: Dict[Tuple[str, str], Dict[str, Any]] = {}
    last_project_code: Optional[str] = None

    for row in event_rows:
        try:
            details = json.loads(row["details"]) if row["details"] else {}
        except json.JSONDecodeError:
            details = {}

        if row["kind"] == "project_load":
            candidate = details.get("project_code")
            if candidate:
                last_project_code = str(candidate).strip() or last_project_code
            continue

        if last_project_code and not details.get("project_code"):
            details["project_code"] = last_project_code

        # Applica filtro progetto
        event_project = details.get("project_code") or ""
        if project_filter_norm and str(event_project) != project_filter_norm:
            continue

        member_key = row["member_key"]
        if not member_key:
            continue

        if member_filter_norm and member_key.lower() != member_filter_norm:
            continue

        if row["kind"] == "move":
            activity_id = details.get("to")
        else:
            activity_id = details.get("activity_id")

        if not activity_id:
            continue

        if activity_filter_norm and str(activity_id) != activity_filter_norm:
            continue

        ts = row["ts"]
        if ts is None:
            continue
        dt = datetime.fromtimestamp(ts / 1000, tz=timezone.utc)
        event_date = dt.date()
        if start_date and event_date < start_date:
            continue
        if end_date and event_date > end_date:
            continue

        member_name = row["member_name"] or details.get("member_name") or "Operatore"
        session_key = (member_key, str(activity_id))
        if session_key not in sessions:
            sessions[session_key] = {
                "member_key": member_key,
                "member_name": member_name,
                "activity_id": str(activity_id),
                "events": [],
            }

        sessions[session_key]["events"].append(
            {
                "ts": ts,
                "dt": dt,
                "kind": row["kind"],
                "details": details,
            }
        )

    if not sessions:
        return []

    # Ottieni lo stato corrente degli operatori per determinare quali sessioni sono ancora in corso
    current_state_query = "SELECT member_key, activity_id FROM member_state WHERE activity_id IS NOT NULL"
    current_state_rows = db.execute(current_state_query).fetchall()
    currently_active = {(row["member_key"], str(row["activity_id"])) for row in current_state_rows}

    now_utc = datetime.now(tz=timezone.utc)
    results: List[Dict[str, Any]] = []

    for session_key, session in sessions.items():
        events = sorted(session["events"], key=lambda e: e["ts"])
        if not events:
            continue

        member_key, activity_id = session_key
        member_name = session["member_name"]
        activity_label = activity_map.get(activity_id, activity_id)

        start_event = None
        end_event = None
        pause_events: List[Dict[str, Any]] = []
        project_code = None

        for event in events:
            if event["kind"] == "move" and str(event["details"].get("to") or "") == activity_id:
                if not start_event:
                    start_event = event
                project_code = project_code or event["details"].get("project_code")
            elif event["kind"] == "pause_member":
                pause_events.append(event)
            elif event["kind"] == "finish_activity":
                end_event = event
                project_code = project_code or event["details"].get("project_code")

        if not start_event:
            start_event = events[0]
        if not start_event:
            continue

        start_dt = start_event["dt"]
        end_dt = end_event["dt"] if end_event else now_utc
        total_duration_ms = max(0, int((end_dt - start_dt).total_seconds() * 1000))

        net_duration_ms = total_duration_ms
        pause_duration_ms = 0
        if end_event:
            net_duration_ms = int(end_event["details"].get("duration_ms", total_duration_ms))
            pause_duration_ms = int(end_event["details"].get("pause_ms", max(0, total_duration_ms - net_duration_ms)))
        else:
            pause_duration_ms = max(0, total_duration_ms - net_duration_ms)

        pause_count = len(pause_events)
        
        # Una sessione è "running" se l'operatore è attualmente su questa attività
        # altrimenti è "completed" (anche se non c'è un finish_activity esplicito)
        is_currently_active = session_key in currently_active
        status = "running" if is_currently_active else "completed"

        end_ts_value = int(end_dt.timestamp() * 1000)
        
        # Ore preventivate per questa attività (usa chiave composita o fallback)
        activity_key = (activity_id, project_code) if project_code else activity_id
        planned_ms = activity_planned_map.get(activity_key) or activity_planned_map.get(activity_id) or 0
        activity_note = activity_notes_map.get(activity_key) or activity_notes_map.get(activity_id) or ""

        results.append(
            {
                "member_key": member_key,
                "member_name": member_name,
                "activity_id": activity_id,
                "activity_label": activity_label,
                "start_ts": int(start_dt.timestamp() * 1000),
                "end_ts": end_ts_value,
                "status": status,
                "net_ms": max(0, net_duration_ms),
                "pause_ms": max(0, pause_duration_ms),
                "pause_count": pause_count,
                "auto_closed": bool(end_event and end_event["details"].get("auto_close")),
                "project_code": project_code,
                "planned_ms": planned_ms,
                "override_id": None,
                "manual_entry": False,
                "note": activity_note,
                "source_member_key": member_key,
                "source_activity_id": activity_id,
                "source_start_ts": int(start_dt.timestamp() * 1000),
            }
        )

    session_map: Dict[Tuple[str, str, int], Dict[str, Any]] = {
        (item["member_key"], item["activity_id"], item["start_ts"]): item
        for item in results
    }

    overrides = _fetch_session_override_rows(
        db,
        start_date=start_date,
        end_date=end_date,
        member_filter=member_filter,
        activity_filter=activity_filter,
    )

    merged: List[Dict[str, Any]] = []
    replaced_keys: Set[Tuple[str, str, int]] = set()

    for override_row in overrides:
        payload = _override_row_to_session(override_row)
        key = None
        if (
            payload.get("source_member_key")
            and payload.get("source_activity_id")
            and payload.get("source_start_ts")
        ):
            key = (
                str(payload["source_member_key"]),
                str(payload["source_activity_id"]),
                int(payload["source_start_ts"]),
            )
        if key and key in session_map:
            replaced_keys.add(key)
            session_map.pop(key, None)
        merged.append(payload)

    for key, item in session_map.items():
        if key in replaced_keys:
            continue
        merged.append(item)

    merged.sort(key=lambda item: item["start_ts"], reverse=True)
    return merged


def evaluate_overdue_activities(db: DatabaseLike) -> List[Dict[str, Any]]:
    meta = load_activity_meta(db)
    if not meta:
        return []

    now = now_ms()
    activity_labels = {
        row["activity_id"]: row["label"]
        for row in db.execute("SELECT activity_id, label FROM activities")
    }

    notified = get_push_notified_map(db)
    overdue: List[Dict[str, Any]] = []

    for activity_id, entry in meta.items():
        if not isinstance(entry, Mapping):
            app.logger.info(
                "Push worker: meta inatteso per attività %s (%s)", activity_id, type(entry)
            )
            continue

        member_rows = db.execute(
            "SELECT running, start_ts, elapsed_cached, pause_start FROM member_state WHERE activity_id=?",
            (activity_id,),
        ).fetchall()

        assigned_count = len(member_rows)
        running_rows = [row for row in member_rows if row["running"] == RUN_STATE_RUNNING]
        paused_count = sum(1 for row in member_rows if row["pause_start"] is not None)
        running_count = len(running_rows)

        if assigned_count == 0:
            continue

        plan_start_ms = parse_iso_to_ms(cast(Optional[str], entry.get("plan_start")))
        plan_end_ms = parse_iso_to_ms(cast(Optional[str], entry.get("plan_end")))
        planned_members = _coerce_int(entry.get("planned_members"))
        planned_duration_ms = _coerce_int(entry.get("planned_duration_ms"))

        if planned_duration_ms is None:
            if plan_start_ms is None or plan_end_ms is None:
                app.logger.info(
                    "Push worker: pianificazione assente/illeggibile per attività %s",
                    activity_id,
                )
                continue
            base_duration_ms = max(0, plan_end_ms - plan_start_ms)
            if base_duration_ms == 0:
                continue
            normalized_members = planned_members if planned_members and planned_members > 0 else assigned_count
            if normalized_members <= 0:
                normalized_members = 1
            planned_duration_ms = base_duration_ms * normalized_members

        if planned_duration_ms <= 0:
            continue

        running_total_ms = 0
        for row in running_rows:
            elapsed = int(row["elapsed_cached"] or 0)
            start_ts = row["start_ts"]
            start_value: Optional[int]
            try:
                start_value = int(start_ts) if start_ts is not None else None
            except (TypeError, ValueError):
                start_value = None
            if start_value is not None:
                elapsed += max(0, now - start_value)
            running_total_ms += elapsed

        threshold_ms = planned_duration_ms + ACTIVITY_OVERDUE_GRACE_MS
        if running_total_ms <= threshold_ms:
            app.logger.info(
                "Push worker: attività %s ancora entro il margine (tempo %sms, soglia %sms)",
                activity_id,
                running_total_ms,
                threshold_ms,
            )
            continue

        previous = notified.get(activity_id)
        if isinstance(previous, Mapping):
            previous_signature = previous.get("planned_duration_ms")
            if previous_signature is None:
                previous_signature = previous.get("plan_end_ms")
            try:
                previous_signature_int = (
                    int(previous_signature) if previous_signature is not None else None
                )
            except (TypeError, ValueError):
                previous_signature_int = None
            if previous_signature_int == planned_duration_ms:
                app.logger.info(
                    "Push worker: attività %s già notificata per questa durata prevista",
                    activity_id,
                )
                continue

        overdue_minutes = max(1, int((running_total_ms - planned_duration_ms) // 60000))
        app.logger.info(
            "Push worker: attività %s supera la durata prevista di %s minuti",
            activity_id,
            overdue_minutes,
        )
        overdue.append(
            {
                "activity_id": activity_id,
                "activity_label": activity_labels.get(activity_id, activity_id),
                "planned_duration_ms": planned_duration_ms,
                "overdue_minutes": overdue_minutes,
                "assigned_members": assigned_count,
                "active_members": running_count,
                "paused_members": paused_count,
                "actual_running_ms": running_total_ms,
            }
        )

    return overdue


def deliver_overdue_notifications(
    db: DatabaseLike,
    overdue_items: Sequence[Mapping[str, Any]],
    settings: Mapping[str, str],
) -> Set[str]:
    if not overdue_items:
        return set()

    subscriptions = fetch_push_subscriptions(db)
    if not subscriptions:
        app.logger.info(
            "Push worker: nessuna subscription attiva, skip notifica %s",
            [item.get("activity_id") for item in overdue_items],
        )
        return set()

    invalid_endpoints: Set[str] = set()
    delivered: Set[str] = set()

    for item in overdue_items:
        activity_id = cast(str, item.get("activity_id"))
        label = cast(str, item.get("activity_label", activity_id))
        overdue_minutes = cast(int, item.get("overdue_minutes", 0))
        planned_duration_ms = int(cast(Optional[int], item.get("planned_duration_ms")) or 0)
        actual_running_ms = int(cast(Optional[int], item.get("actual_running_ms")) or 0)
        planned_label = format_duration_ms(planned_duration_ms) or "Durata prevista"
        actual_label = format_duration_ms(actual_running_ms) or "Tempo in corso"
        assigned_members = int(cast(Optional[int], item.get("assigned_members")) or 0)
        active_members = int(cast(Optional[int], item.get("active_members")) or 0)
        paused_members = int(cast(Optional[int], item.get("paused_members")) or 0)

        if active_members:
            status_suffix = f" · {active_members} operatori ancora attivi"
        elif paused_members:
            status_suffix = f" · {paused_members} operatori in pausa"
        else:
            status_suffix = " · Nessun timer attivo"

        payload = {
            "title": "Attività oltre il termine",
            "body": (
                f"{label}: il tempo in corso ({actual_label}) supera la durata prevista ({planned_label})"
                f" di {overdue_minutes} minuti{status_suffix}"
            ),
            "data": {
                "activity_id": activity_id,
                "notification_type": "overdue_activity",
                "overdue_minutes": overdue_minutes,
                "planned_duration_ms": planned_duration_ms,
                "actual_running_ms": actual_running_ms,
                "assigned_members": assigned_members,
                "active_members": active_members,
                "paused_members": paused_members,
            },
        }

        delivered_this_round = False

        for sub in subscriptions:
            endpoint = sub.get("endpoint")
            if not endpoint or endpoint in invalid_endpoints:
                continue
            key_p256dh = sub.get("p256dh")
            key_auth = sub.get("auth")
            if not key_p256dh or not key_auth:
                invalid_endpoints.add(str(endpoint))
                continue
            subscription_info = {
                "endpoint": endpoint,
                "keys": {
                    "p256dh": key_p256dh,
                    "auth": key_auth,
                },
            }
            encoding = sub.get("content_encoding") or "aes128gcm"
            try:
                webpush(
                    subscription_info=subscription_info,
                    data=json.dumps(payload),
                    vapid_private_key=settings["vapid_private"],
                    vapid_claims={"sub": settings["subject"]},
                    ttl=OVERDUE_PUSH_TTL_SECONDS,
                    content_encoding=encoding,
                )
                delivered_this_round = True
                record_push_notification(
                    db,
                    kind="overdue_activity",
                    title=payload.get("title", "Notifica"),
                    body=payload.get("body"),
                    payload=payload,
                    activity_id=activity_id,
                    username=sub.get("username"),
                )
            except WebPushException as exc:
                status = getattr(exc.response, "status_code", None)
                app.logger.warning("WebPush fallita (%s): %s", status, exc)
                if status in (404, 410):
                    invalid_endpoints.add(endpoint)
            except Exception as exc:  # pragma: no cover - logging best effort
                app.logger.exception("Errore imprevisto nell'invio push", exc_info=exc)

        if delivered_this_round:
            delivered.add(activity_id)

    if invalid_endpoints:
        for endpoint in invalid_endpoints:
            remove_push_subscription(db, endpoint)
        db.commit()
        app.logger.info(
            "Push worker: rimossa %s subscription invalida", len(invalid_endpoints)
        )

    return delivered


def evaluate_long_running_members(db: DatabaseLike) -> List[Dict[str, Any]]:
    rows = db.execute(
        """
        SELECT
            ms.member_key,
            ms.member_name,
            ms.activity_id,
            ms.start_ts,
            a.label AS activity_label
        FROM member_state ms
        LEFT JOIN activities a ON a.activity_id = ms.activity_id
        WHERE ms.running=? AND ms.start_ts IS NOT NULL
        """,
        (RUN_STATE_RUNNING,),
    ).fetchall()

    now = now_ms()
    notified = get_long_running_notified_map(db)
    active_starts = {
        row["member_key"]: int(row["start_ts"])
        for row in rows
        if row["start_ts"] is not None
    }

    # Pulisce lo stato dei long running non più validi
    cleaned_state: Dict[str, Any] = {}
    changed = False
    for key, value in notified.items():
        if not isinstance(value, Mapping):
            changed = True
            continue
        try:
            recorded_start = int(value.get("start_ts"))
        except (TypeError, ValueError):
            changed = True
            continue
        if active_starts.get(key) == recorded_start:
            cleaned_state[key] = value
        else:
            changed = True

    if changed:
        save_long_running_notified_map(db, cleaned_state)
        notified = cleaned_state

    long_running: List[Dict[str, Any]] = []

    for row in rows:
        member_key = row["member_key"]
        start_ts = row["start_ts"]
        if start_ts is None:
            continue
        duration = max(0, now - int(start_ts))
        if duration < LONG_RUNNING_THRESHOLD_MS:
            continue
        previous = notified.get(member_key)
        previous_start = None
        if isinstance(previous, Mapping):
            try:
                previous_start = int(previous.get("start_ts"))
            except (TypeError, ValueError):
                previous_start = None
        if previous_start == int(start_ts):
            continue
        long_running.append(
            {
                "member_key": member_key,
                "member_name": row["member_name"],
                "activity_id": row["activity_id"],
                "activity_label": row["activity_label"] or row["activity_id"],
                "start_ts": int(start_ts),
                "duration_ms": duration,
            }
        )

    return long_running


def deliver_long_running_notifications(
    db: DatabaseLike,
    items: Sequence[Mapping[str, Any]],
    settings: Mapping[str, str],
) -> Set[str]:
    if not items:
        return set()

    subscriptions = fetch_push_subscriptions(db)
    if not subscriptions:
        app.logger.info(
            "Push worker: nessuna subscription attiva, skip avvisi operatori oltre soglia"
        )
        return set()

    invalid_endpoints: Set[str] = set()
    delivered_members: Set[str] = set()

    for item in items:
        member_key = cast(str, item.get("member_key"))
        member_name = cast(str, item.get("member_name", member_key))
        activity_label = cast(str, item.get("activity_label"))
        duration_ms = cast(int, item.get("duration_ms", 0))
        duration_label = format_duration_ms(duration_ms) or "02:00"

        payload = {
            "title": "Operatore in attività",
            "body": f"{member_name} è su {activity_label} da {duration_label}",
            "data": {
                "notification_type": "long_running_member",
                "member_key": member_key,
                "member_name": member_name,
                "activity_id": item.get("activity_id"),
                "activity_label": activity_label,
                "start_ts": item.get("start_ts"),
                "duration_ms": duration_ms,
            },
        }

        delivered_this_round = False

        for sub in subscriptions:
            endpoint = sub.get("endpoint")
            if not endpoint or endpoint in invalid_endpoints:
                continue
            key_p256dh = sub.get("p256dh")
            key_auth = sub.get("auth")
            if not key_p256dh or not key_auth:
                invalid_endpoints.add(str(endpoint))
                continue
            subscription_info = {
                "endpoint": endpoint,
                "keys": {
                    "p256dh": key_p256dh,
                    "auth": key_auth,
                },
            }
            encoding = sub.get("content_encoding") or "aes128gcm"
            try:
                webpush(
                    subscription_info=subscription_info,
                    data=json.dumps(payload),
                    vapid_private_key=settings["vapid_private"],
                    vapid_claims={"sub": settings["subject"]},
                    ttl=120,
                    content_encoding=encoding,
                )
                delivered_this_round = True
                record_push_notification(
                    db,
                    kind="long_running_member",
                    title=payload.get("title", "Notifica"),
                    body=payload.get("body"),
                    payload=payload,
                    activity_id=cast(Optional[str], item.get("activity_id")),
                    username=sub.get("username"),
                )
            except WebPushException as exc:
                status = getattr(exc.response, "status_code", None)
                app.logger.warning("WebPush fallita (%s): %s", status, exc)
                if status in (404, 410):
                    invalid_endpoints.add(endpoint)
            except Exception as exc:  # pragma: no cover - logging best effort
                app.logger.exception("Errore imprevisto nell'invio push", exc_info=exc)

        if delivered_this_round:
            delivered_members.add(member_key)

    if invalid_endpoints:
        for endpoint in invalid_endpoints:
            remove_push_subscription(db, endpoint)
        db.commit()
        app.logger.info(
            "Push worker: rimossa %s subscription invalida (avvisi long running)",
            len(invalid_endpoints),
        )

    return delivered_members


def _notification_worker() -> None:
    stop_event = _NOTIFICATION_STOP
    if stop_event is None:
        return

    app.logger.info(
        "Push worker: avviato (intervallo %ss)", NOTIFICATION_INTERVAL_SECONDS
    )

    while not stop_event.is_set():
        try:
            with app.app_context():
                settings = get_webpush_settings()
                if not settings:
                    app.logger.info("Push worker: impostazioni VAPID mancanti")
                    continue
                db = get_db()
                overdue_items = evaluate_overdue_activities(db)
                if overdue_items:
                    overdue_ids = [str(item.get("activity_id")) for item in overdue_items]
                    app.logger.info(
                        "Push worker: trovate %s attività in ritardo %s",
                        len(overdue_items),
                        overdue_ids,
                    )
                    delivered = deliver_overdue_notifications(db, overdue_items, settings)
                    if delivered:
                        app.logger.info(
                            "Push worker: notifiche inviate a %s",
                            sorted(delivered),
                        )
                        state = get_push_notified_map(db)
                        now_sent = now_ms()
                        for item in overdue_items:
                            activity_id = item.get("activity_id")
                            if activity_id in delivered:
                                state[str(activity_id)] = {
                                    "planned_duration_ms": item.get("planned_duration_ms"),
                                    "sent_ts": now_sent,
                                }
                        save_push_notified_map(db, state)
                        db.commit()
                    else:
                        app.logger.info(
                            "Push worker: nessuna notifica consegnata per %s",
                            overdue_ids,
                        )
                else:
                    app.logger.info("Push worker: nessuna attività oltre il termine")

                # Notifiche "operatore long running" disattivate su richiesta
        except Exception as exc:  # pragma: no cover - worker should never crash
            app.logger.exception("Worker notifiche push in errore", exc_info=exc)
        finally:
            stop_event.wait(NOTIFICATION_INTERVAL_SECONDS)


def start_notification_worker() -> None:
    global _NOTIFICATION_THREAD, _NOTIFICATION_STOP

    if _NOTIFICATION_THREAD and _NOTIFICATION_THREAD.is_alive():
        return

    _NOTIFICATION_STOP = Event()
    _NOTIFICATION_THREAD = Thread(target=_notification_worker, name="joblog-push-worker", daemon=True)
    _NOTIFICATION_THREAD.start()
    app.logger.info("Push worker: thread avviato")


def stop_notification_worker() -> None:
    global _NOTIFICATION_THREAD, _NOTIFICATION_STOP

    stop_event = _NOTIFICATION_STOP
    thread = _NOTIFICATION_THREAD

    if stop_event is not None:
        stop_event.set()

    if thread and thread.is_alive():
        thread.join(timeout=5)

    _NOTIFICATION_THREAD = None
    _NOTIFICATION_STOP = None


def describe_event(kind: str, details: Dict[str, Any], activity_labels: Dict[str, str]) -> str:
    def label_for(activity_id: Optional[str]) -> str:
        if not activity_id:
            return "Disponibili"
        return activity_labels.get(activity_id, activity_id)

    if kind == "move":
        member_name = details.get("member_name") or "Operatore"
        origin = label_for(details.get("from"))
        destination = label_for(details.get("to"))
        summary = f"{member_name}: {origin} → {destination}"
        duration = format_duration_ms(details.get("duration_ms"))
        if duration:
            summary += f" · {duration}"
        return summary

    if kind == "project_load":
        code = details.get("project_code") or "--"
        name = details.get("project_name") or ""
        if name:
            return f"Progetto {code} · {name}"
        return f"Progetto {code} attivato"

    if kind == "create_activity":
        label = details.get("label") or label_for(details.get("activity_id"))
        return f"Nuova attività: {label or 'Attività'}"

    if kind == "pause_all":
        affected = details.get("affected") or 0
        return f"Pausa collettiva ({affected} operatori)"

    if kind == "resume_all":
        affected = details.get("affected") or 0
        return f"Ripresa collettiva ({affected} operatori)"

    if kind == "pause_member":
        member_name = details.get("member_name") or "Operatore"
        activity_label = label_for(details.get("activity_id"))
        duration = format_duration_ms(details.get("duration_ms"))
        summary = f"{member_name}: Pausa {activity_label}"
        if duration:
            summary += f" · {duration}"
        return summary

    if kind == "resume_member":
        member_name = details.get("member_name") or "Operatore"
        activity_label = label_for(details.get("activity_id"))
        return f"{member_name}: Ripresa {activity_label}"

    if kind == "finish_activity":
        member_name = details.get("member_name") or "Operatore"
        activity_label = label_for(details.get("activity_id"))
        summary = f"{member_name}: Fine {activity_label}"
        duration = format_duration_ms(details.get("duration_ms"))
        if duration:
            summary += f" · {duration}"
        return summary

    if kind == "finish_all":
        affected = details.get("affected") or 0
        return f"Fine attività collettiva ({affected} operatori)"

    return kind.replace("_", " ").title()


# Authentication routes
@app.route("/login")
def login():
    """Render login page. Redirect to home if already logged in."""
    if 'user' in session:
        return redirect(url_for('home'))
    return render_template("login.html")


@app.get("/api/ping")
def api_ping():
    """Simple endpoint to check server connectivity."""
    return jsonify({"status": "ok", "timestamp": int(datetime.now().timestamp() * 1000)})


@app.post("/api/login")
def api_login():
    """Handle login POST request."""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Dati non validi'}), 400
    
    username_input = data.get('username', '').strip()
    password = data.get('password', '')
    
    if not username_input or not password:
        return jsonify({'success': False, 'error': 'Username e password richiesti'}), 400
    
    db = get_db()
    user_row = fetch_user_record(db, username_input)
    if not user_row:
        return jsonify({'success': False, 'error': 'Credenziali non valide'}), 401

    if not user_row.get('is_active'):
        return jsonify({'success': False, 'error': 'Account disabilitato'}), 403

    password_hash = user_row.get('password_hash') or ''
    if not verify_password(password, password_hash):
        return jsonify({'success': False, 'error': 'Credenziali non valide'}), 401

    username = user_row.get('username') or _normalize_username(username_input)
    _apply_user_session({**dict(user_row), 'username': username})
    response = jsonify({'success': True})
    try:
        token_value, expires_ts = _store_persistent_session(username)
    except Exception:
        app.logger.exception("Impossibile creare una sessione persistente per %s", username)
    else:
        _set_persistent_cookie(response, token_value, expires_ts)
    return response


@app.route("/logout")
def logout():
    """Clear session and redirect to login."""
    token = request.cookies.get(_persistent_cookie_name())
    if token:
        _delete_persistent_session(token)
    session.clear()
    response = redirect(url_for('login'))
    response.delete_cookie(_persistent_cookie_name(), path='/')
    return response


@app.route("/")
def home() -> ResponseReturnValue:
    if 'user' not in session:
        return redirect(url_for('login'))
    if session.get('is_admin'):
        return redirect(url_for('admin_dashboard_page'))
    if session.get('user_role') == ROLE_MAGAZZINO:
        return redirect(url_for('magazzino_home'))
    # Utenti "user" senza ruolo specifico - mostrano pagina minima con solo logout
    if session.get('user_role') == ROLE_USER:
        db = get_db()
        display_name = session.get('user_display') or session.get('user_name') or session.get('user')
        primary_name = session.get('user_name') or display_name or session.get('user')
        initials = session.get('user_initials') or compute_initials(primary_name or "")
        user_role = session.get('user_role', 'user')
        username = session.get('user')
        # Verifica se il modulo magazzino è abilitato
        magazzino_enabled = is_module_enabled(db, "magazzino")
        return render_template(
            "user_home.html",
            user_name=primary_name,
            user_display=display_name,
            user_initials=initials,
            user_role=user_role,
            username=username,
            magazzino_enabled=magazzino_enabled,
        )
    # Supervisor e altri ruoli - mostrano dashboard operativa completa
    db = get_db()
    display_name = session.get('user_display') or session.get('user_name') or session.get('user')
    primary_name = session.get('user_name') or display_name or session.get('user')
    initials = session.get('user_initials') or compute_initials(primary_name or "")
    is_admin = bool(session.get('is_admin'))
    project_code = get_app_state(db, "project_code")
    project_name = get_app_state(db, "project_name")
    initial_attachments: Dict[str, Any] = {"project": None, "items": []}
    initial_materials: Dict[str, Any] = {
        "project": None,
        "items": [],
        "folders": [],
        "equipment_checks": {},
        "updated_ts": None,
        "from_cache": False,
    }
    initial_project = None
    if project_code:
        initial_project = {
            "code": project_code,
            "name": project_name or project_code,
        }
        initial_attachments["project"] = initial_project
        initial_materials["project"] = initial_project
        initial_materials["equipment_checks"] = fetch_equipment_checks(db, project_code)
        cached_materials = load_project_materials_cache(db, project_code)
        if cached_materials:
            initial_materials["items"] = cached_materials.get("items", [])
            initial_materials["folders"] = cached_materials.get("folders", [])
            cached_project = cached_materials.get("project")
            if cached_project:
                initial_materials["project"] = cached_project
            initial_materials["updated_ts"] = cached_materials.get("updated_ts")
            initial_materials["from_cache"] = True
    header_clock = datetime.now().strftime("%d/%m/%Y | %H:%M")

    # Progetto salvato del supervisor (per auto-load se necessario)
    saved_supervisor_project = None
    supervisor_project_code = session.get('supervisor_project_code')
    if supervisor_project_code:
        saved_supervisor_project = {
            "code": supervisor_project_code,
            "name": session.get('supervisor_project_name') or supervisor_project_code,
        }

    return render_template(
        "index.html",
        user_name=primary_name,
        user_display=display_name,
        user_initials=initials,
        is_admin=is_admin,
        user_role=session.get('user_role', 'user'),
        initial_attachments=initial_attachments,
        initial_materials=initial_materials,
        initial_project=initial_project,
        saved_supervisor_project=saved_supervisor_project,
        header_clock=header_clock,
    )


# ============================================================
# API TIMBRATURA
# ============================================================

@app.get("/api/timbratura/oggi")
@login_required
def api_timbratura_oggi():
    """Restituisce le timbrature dell'utente per oggi."""
    username = session.get('user')
    if not username:
        return jsonify({"error": "Non autenticato"}), 401
    
    db = get_db()
    ensure_timbrature_table(db)
    ensure_cedolino_timbrature_table(db)
    
    # Recupera le regole dell'utente per sapere se usa rounding daily
    user_rules = get_user_timbratura_rules(db, username)
    rounding_mode = user_rules.get('rounding_mode', 'single')
    
    today = datetime.now().strftime("%Y-%m-%d")
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Prima recupera le timbrature di oggi
    rows = db.execute(
        f"""
        SELECT id, tipo, ora, ora_mod, method, gps_lat, gps_lon, location_name, created_ts, created_by
        FROM timbrature
        WHERE username = {placeholder} AND data = {placeholder}
        ORDER BY created_ts ASC
        """,
        (username, today)
    ).fetchall()
    
    timbrature = []
    for row in rows:
        timbratura_id = row['id'] if isinstance(row, dict) else row[0]
        tipo = row['tipo'] if isinstance(row, dict) else row[1]
        ora_val = row['ora'] if isinstance(row, dict) else row[2]
        ora_mod_val = row['ora_mod'] if isinstance(row, dict) else row[3]
        method = row['method'] if isinstance(row, dict) else row[4]
        gps_lat = row['gps_lat'] if isinstance(row, dict) else row[5]
        gps_lon = row['gps_lon'] if isinstance(row, dict) else row[6]
        location_name = row['location_name'] if isinstance(row, dict) else row[7]
        created_ts = row['created_ts'] if isinstance(row, dict) else row[8]
        created_by = row['created_by'] if isinstance(row, dict) else row[9]
        
        # Mappa tipo timbratura -> timeframe_id
        timeframe_map = {
            'inizio_giornata': 1,
            'inizio_pausa': 4,
            'fine_pausa': 5,
            'fine_giornata': 8
        }
        timeframe_id = timeframe_map.get(tipo, 0)
        
        # Cerca il cedolino corrispondente più recente (basato su created_ts vicino a quello della timbratura)
        cedolino_row = db.execute(
            f"""
            SELECT synced_ts, sync_error, overtime_request_id
            FROM cedolino_timbrature
            WHERE username = {placeholder} 
              AND data_riferimento = {placeholder}
              AND timeframe_id = {placeholder}
              AND ABS(created_ts - {placeholder}) < 60000
            ORDER BY ABS(created_ts - {placeholder}) ASC
            LIMIT 1
            """,
            (username, today, timeframe_id, created_ts, created_ts)
        ).fetchone()
        
        synced_ts = None
        sync_error = None
        overtime_request_id = None
        if cedolino_row:
            synced_ts = cedolino_row['synced_ts'] if isinstance(cedolino_row, dict) else cedolino_row[0]
            sync_error = cedolino_row['sync_error'] if isinstance(cedolino_row, dict) else cedolino_row[1]
            overtime_request_id = cedolino_row['overtime_request_id'] if isinstance(cedolino_row, dict) else cedolino_row[2]
        
        # Gestisce sia TIME che stringa
        if hasattr(ora_val, 'strftime'):
            ora_str = ora_val.strftime("%H:%M")
        else:
            ora_str = str(ora_val)[:5] if ora_val else ""
        
        # Formatta ora_mod
        ora_mod_str = None
        if ora_mod_val:
            if hasattr(ora_mod_val, 'strftime'):
                ora_mod_str = ora_mod_val.strftime("%H:%M")
            else:
                ora_mod_str = str(ora_mod_val)[:5]
        
        # Converti Decimal a float per JSON
        if gps_lat is not None:
            gps_lat = float(gps_lat)
        if gps_lon is not None:
            gps_lon = float(gps_lon)
        
        # Determina se è in attesa di conferma Extra Turno
        pending_extra_turno = False
        if overtime_request_id and synced_ts is None:
            pending_extra_turno = True
        
        timbrature.append({
            "id": timbratura_id,
            "tipo": row['tipo'] if isinstance(row, dict) else row[1],
            "ora": ora_str,
            "ora_mod": ora_mod_str,
            "rounding_mode": rounding_mode,  # 'single' o 'daily'
            "method": method,
            "created_by": created_by,
            "gps_lat": gps_lat,
            "gps_lon": gps_lon,
            "location_name": location_name,
            "pending_extra_turno": pending_extra_turno,
            "sync_error": sync_error if pending_extra_turno else None
        })
    
    # Verifica se c'è uno straordinario pending per oggi
    pending_overtime = None
    overtime_req_id = _get_pending_overtime_request_id(db, username, today)
    if overtime_req_id:
        # Recupera i dettagli della richiesta straordinario
        overtime_type_id = get_overtime_request_type_id(db)
        ot_row = db.execute(f"""
            SELECT value_amount, extra_data FROM user_requests
            WHERE id = {placeholder} AND request_type_id = {placeholder}
        """, (overtime_req_id, overtime_type_id)).fetchone()
        if ot_row:
            minutes = int(ot_row['value_amount'] if isinstance(ot_row, dict) else ot_row[0]) if ot_row else 0
            pending_overtime = {
                "id": overtime_req_id,
                "minutes": minutes,
                "status": "pending"
            }
    
    return jsonify({"timbrature": timbrature, "pending_overtime": pending_overtime})


@app.post("/api/timbratura")
@login_required
def api_timbratura_registra():
    """Registra una nuova timbratura."""
    username = session.get('user')
    if not username:
        return jsonify({"error": "Non autenticato"}), 401
    
    data = request.get_json()
    if not data:
        return jsonify({"error": "Dati mancanti"}), 400
    
    tipo = data.get('tipo')
    if tipo not in ('inizio_giornata', 'fine_giornata', 'inizio_pausa', 'fine_pausa'):
        return jsonify({"error": "Tipo timbratura non valido"}), 400
    
    # Verifica bypass QR (solo per sviluppo)
    bypass_qr = data.get('bypass_qr', False)
    
    # Gestione dati offline (GPS o QR salvati quando offline)
    offline_timestamp = data.get('offline_timestamp')  # ISO timestamp da quando era offline
    offline_gps = data.get('offline_gps')  # {latitude, longitude, accuracy}
    offline_qr = data.get('offline_qr')  # QR code scansionato offline
    
    db = get_db()
    ensure_timbrature_table(db)
    
    # Recupera configurazione timbratura GLOBALE (non per utente)
    global_timb_config = get_timbratura_config()
    gps_enabled_globally = global_timb_config.get("gps_enabled", False)
    qr_enabled_globally = global_timb_config.get("qr_enabled", True)
    
    # Se GPS è abilitato globalmente e c'è un tentativo di bypass senza coordinate GPS,
    # BLOCCA la timbratura - richiede validazione GPS
    if gps_enabled_globally and bypass_qr and not offline_gps:
        # Verifica se ha un token GPS valido dalla sessione
        timbratura_method = session.get("timbratura_method")
        token = data.get('token')
        session_token = session.get('timbratura_token')
        token_expires = session.get('timbratura_token_expires', 0)
        
        # Se non ha un token GPS valido, blocca
        if timbratura_method != "gps" or not token or token != session_token or now_ms() > token_expires:
            app.logger.warning(f"Tentativo timbratura bypass senza GPS valido per {username} - GPS richiesto")
            return jsonify({
                "error": "Timbratura GPS richiesta. Usa la funzione GPS per timbrare.",
                "need_gps": True
            }), 403
    
    # Se presente offline_gps, valida SEMPRE le coordinate (anche con bypass_qr)
    if offline_gps:
        try:
            latitude = offline_gps.get('latitude')
            longitude = offline_gps.get('longitude')
            accuracy = offline_gps.get('accuracy', 9999)
            
            # Recupera configurazione GPS
            settings = get_company_settings(db)
            custom = settings.get('custom_settings', {})
            timbratura_config = custom.get('timbratura', {})
            gps_enabled = timbratura_config.get('gps_enabled', False)
            gps_locations = timbratura_config.get('gps_locations', [])
            gps_max_accuracy = timbratura_config.get('gps_max_accuracy_meters', 50)
            
            # NUOVA LOGICA: Verifica se l'utente ha un turno Rentman per oggi con GPS configurato
            rentman_gps_location = None
            try:
                ensure_rentman_plannings_table(db)
                today = datetime.now().date().isoformat()
                placeholder = "%s" if DB_VENDOR == "mysql" else "?"
                
                # Trova il rentman_crew_id dell'utente
                user_row = db.execute(
                    f"SELECT rentman_crew_id, group_id FROM app_users WHERE username = {placeholder}",
                    (username,)
                ).fetchone()
                
                if user_row:
                    crew_id = user_row['rentman_crew_id'] if isinstance(user_row, dict) else user_row[0]
                    user_group_id = user_row['group_id'] if isinstance(user_row, dict) else user_row[1]
                    
                    if crew_id:
                        # Cerca turno Rentman di oggi per l'utente (solo se inviato)
                        rentman_row = db.execute(f"""
                            SELECT location_name, location_id,
                                   timbratura_gps_mode, gps_timbratura_location
                            FROM rentman_plannings 
                            WHERE crew_id = {placeholder} 
                              AND planning_date = {placeholder} 
                              AND sent_to_webservice = 1
                            ORDER BY plan_start ASC LIMIT 1
                        """, (crew_id, today)).fetchone()
                        
                        if rentman_row:
                            if isinstance(rentman_row, dict):
                                loc_name = rentman_row.get('location_name')
                                loc_id = rentman_row.get('location_id')
                                gps_mode = rentman_row.get('timbratura_gps_mode') or 'group'
                                group_gps_name = rentman_row.get('gps_timbratura_location')
                            else:
                                loc_name = rentman_row[0]
                                loc_id = rentman_row[1]
                                gps_mode = rentman_row[2] or 'group'
                                group_gps_name = rentman_row[3]
                            
                            # Cerca coordinate nella cache
                            loc_lat, loc_lon, loc_radius = None, None, 300
                            ensure_location_cache_table(db)
                            cached_coords = get_location_cache(db, loc_name, loc_id)
                            if cached_coords:
                                loc_lat, loc_lon, loc_radius = cached_coords
                            
                            app.logger.info(f"Timbratura GPS: mode={gps_mode}, loc_name={loc_name}, loc_lat={loc_lat}, loc_lon={loc_lon}, loc_radius={loc_radius}, group_gps_name={group_gps_name}")
                            app.logger.info(f"GPS locations configurate: {[l.get('name') for l in gps_locations]}")
                            
                            # Decidi quale location usare in base al mode
                            if gps_mode == 'location' and loc_lat and loc_lon and (float(loc_lat) != 0 or float(loc_lon) != 0):
                                # Usa le coordinate della location del progetto
                                rentman_gps_location = {
                                    'name': loc_name or 'Location Progetto',
                                    'latitude': float(loc_lat),
                                    'longitude': float(loc_lon),
                                    'radius_meters': loc_radius  # Raggio dalla cache
                                }
                                app.logger.info(f"Timbratura GPS Rentman: usando location progetto '{loc_name}' per {username} (raggio: {loc_radius}m)")
                            elif gps_mode == 'group' and group_gps_name:
                                # Usa la sede GPS del gruppo - cerca nelle locations configurate
                                group_location = None
                                for loc in gps_locations:
                                    if loc.get('name') == group_gps_name:
                                        group_location = loc
                                        break
                                
                                if group_location:
                                    rentman_gps_location = group_location
                                    app.logger.info(f"Timbratura GPS Rentman: usando sede gruppo '{group_gps_name}' per {username}")
                                else:
                                    # Se non trova la sede nelle locations, cerca nelle coordinate del gruppo
                                    if user_group_id:
                                        group_row = db.execute(f"""
                                            SELECT gps_location_name, gps_location_lat, gps_location_lon, gps_location_radius
                                            FROM user_groups WHERE id = {placeholder}
                                        """, (user_group_id,)).fetchone()
                                        if group_row:
                                            g_lat = group_row['gps_location_lat'] if isinstance(group_row, dict) else group_row[1]
                                            g_lon = group_row['gps_location_lon'] if isinstance(group_row, dict) else group_row[2]
                                            g_radius = group_row['gps_location_radius'] if isinstance(group_row, dict) else group_row[3]
                                            if g_lat and g_lon:
                                                rentman_gps_location = {
                                                    'name': group_gps_name,
                                                    'latitude': float(g_lat),
                                                    'longitude': float(g_lon),
                                                    'radius_meters': g_radius or 300
                                                }
                                                app.logger.info(f"Timbratura GPS Rentman: usando coordinate gruppo '{group_gps_name}' per {username}")
            except Exception as e:
                app.logger.warning(f"Errore lettura turno Rentman per GPS: {e}")
            
            # Se abbiamo una location Rentman specifica, usala come unica location valida
            if rentman_gps_location:
                gps_locations = [rentman_gps_location]
                app.logger.info(f"Validazione GPS: override con location Rentman per {username}")
            
            # Verifica se il turno dell'utente ha una sede specifica (fallback employee_shifts)
            shift_location_name = data.get('shift_location_name')  # Sede specifica del turno
            if not shift_location_name and not rentman_gps_location:
                # Prova a recuperare dal turno odierno
                ensure_employee_shifts_table(db)
                day_of_week = datetime.now().weekday()
                placeholder = "%s" if DB_VENDOR == "mysql" else "?"
                shift_row = db.execute(f"""
                    SELECT location_name FROM employee_shifts
                    WHERE username = {placeholder} AND day_of_week = {placeholder} AND is_active = 1
                """, (username, day_of_week)).fetchone()
                if shift_row:
                    shift_location_name = shift_row['location_name'] if isinstance(shift_row, dict) else shift_row[0]
            
            # Se il turno ha una sede specifica, filtra le locations per usare solo quella
            if shift_location_name and gps_locations:
                filtered_locations = [loc for loc in gps_locations if loc.get('name') == shift_location_name]
                if filtered_locations:
                    gps_locations = filtered_locations
                    app.logger.info(f"Validazione GPS: usando sede specifica del turno '{shift_location_name}' per {username}")
            
            if gps_enabled and gps_locations:
                from math import radians, sin, cos, sqrt, atan2
                
                # Verifica prima l'accuratezza del GPS
                if accuracy > gps_max_accuracy:
                    app.logger.warning(f"Timbratura offline GPS rifiutata: {username} - precisione insufficiente {accuracy:.0f}m > {gps_max_accuracy}m")
                    return jsonify({"error": f"Precisione GPS insufficiente ({int(accuracy)}m). Richiesta: max {gps_max_accuracy}m"}), 400
                
                # Verifica se la posizione è entro il raggio di una delle sedi
                R = 6371000  # Raggio Terra in metri
                matched_location = None
                min_distance = float('inf')
                
                for loc in gps_locations:
                    loc_lat = loc.get('latitude')
                    loc_lon = loc.get('longitude')
                    loc_radius = loc.get('radius_meters', 300)
                    
                    if loc_lat is None or loc_lon is None:
                        continue
                    
                    lat1, lon1 = radians(float(loc_lat)), radians(float(loc_lon))
                    lat2, lon2 = radians(float(latitude)), radians(float(longitude))
                    dlat = lat2 - lat1
                    dlon = lon2 - lon1
                    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
                    c = 2 * atan2(sqrt(a), sqrt(1-a))
                    distance = R * c
                    
                    if distance < min_distance:
                        min_distance = distance
                    
                    # Considera l'accuratezza GPS: la distanza effettiva minima
                    effective_distance = max(0, distance - accuracy)
                    
                    if effective_distance <= loc_radius:
                        matched_location = loc
                        break
                
                if not matched_location:
                    app.logger.warning(f"Timbratura offline GPS rifiutata: {username} - distanza minima {min_distance:.0f}m da sedi configurate")
                    return jsonify({"error": f"Posizione GPS non valida. Distanza dalla sede più vicina: {int(min_distance)}m"}), 400
                
                app.logger.info(f"Timbratura offline GPS validata: {username} - {matched_location.get('name', 'Sede')} (distanza {min_distance:.0f}m)")
                bypass_qr = True  # GPS valido, bypassa verifica QR
        except Exception as e:
            app.logger.error(f"Errore validazione GPS offline: {e}")
    
    # Se presente offline_qr, valida il QR code
    if offline_qr and not bypass_qr:
        try:
            settings = get_company_settings(db)
            custom = settings.get('custom_settings', {})
            timbratura_config = custom.get('timbratura', {})
            qr_code = timbratura_config.get('qr_code', '')
            if qr_code and offline_qr == qr_code:
                app.logger.info(f"Timbratura offline QR validata: {username}")
                bypass_qr = True  # QR valido
            else:
                app.logger.warning(f"Timbratura offline QR rifiutata: {username} - QR non corrispondente")
                return jsonify({"error": "QR Code non valido"}), 400
        except Exception as e:
            app.logger.error(f"Errore validazione QR offline: {e}")
    
    if not bypass_qr:
        # Verifica token di validazione QR
        token = data.get('token')
        session_token = session.get('timbratura_token')
        token_expires = session.get('timbratura_token_expires', 0)
        
        if not token or token != session_token:
            return jsonify({"error": "Devi prima scansionare il QR code", "need_qr": True}), 403
        
        if now_ms() > token_expires:
            # Pulisce token scaduto
            session.pop('timbratura_token', None)
            session.pop('timbratura_token_expires', None)
            return jsonify({"error": "Token scaduto, scansiona nuovamente il QR", "need_qr": True}), 403
    
    # Usa timestamp offline se presente, altrimenti ora attuale
    if offline_timestamp:
        try:
            # Parse ISO timestamp
            from dateutil import parser as dateparser
            offline_dt = dateparser.parse(offline_timestamp)
            now = offline_dt
            app.logger.info(f"Usando timestamp offline: {offline_timestamp}")
        except Exception as e:
            app.logger.warning(f"Errore parsing offline_timestamp: {e}, uso ora attuale")
            now = datetime.now()
    else:
        now = datetime.now()
    
    today = now.strftime("%Y-%m-%d")
    ora = now.strftime("%H:%M:%S")
    created_ts = now_ms()
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Verifica regole business
    existing = db.execute(
        f"SELECT tipo FROM timbrature WHERE username = {placeholder} AND data = {placeholder} ORDER BY created_ts ASC",
        (username, today)
    ).fetchall()
    existing_types = [r['tipo'] if isinstance(r, dict) else r[0] for r in existing]
    
    if tipo == 'inizio_giornata' and 'inizio_giornata' in existing_types:
        return jsonify({"error": "Hai già registrato l'inizio giornata oggi"}), 400
    
    if tipo == 'fine_giornata':
        if 'inizio_giornata' not in existing_types:
            return jsonify({"error": "Devi prima registrare l'inizio giornata"}), 400
        if 'fine_giornata' in existing_types:
            return jsonify({"error": "Hai già registrato la fine giornata oggi"}), 400
    
    if tipo == 'inizio_pausa':
        if 'inizio_giornata' not in existing_types:
            return jsonify({"error": "Devi prima registrare l'inizio giornata"}), 400
        if 'fine_giornata' in existing_types:
            return jsonify({"error": "Non puoi iniziare una pausa dopo la fine giornata"}), 400
        # Controlla se c'è già una pausa aperta
        pause_count = existing_types.count('inizio_pausa') - existing_types.count('fine_pausa')
        if pause_count > 0:
            return jsonify({"error": "Hai già una pausa in corso"}), 400
    
    if tipo == 'fine_pausa':
        pause_count = existing_types.count('inizio_pausa') - existing_types.count('fine_pausa')
        if pause_count <= 0:
            return jsonify({"error": "Non hai nessuna pausa in corso"}), 400
    
    # Calcola ora_mod in base alle regole (usa regole specifiche per gruppo se esistono)
    ora_mod = None
    turno_start = None
    turno_end = None
    flex_warning = None  # Avviso se timbrata fuori flessibilità (per daily mode)
    flex_request_id = None  # ID richiesta Fuori Flessibilità (se creata)
    try:
        # Usa le regole specifiche dell'utente (gruppo o globali)
        rules = get_user_timbratura_rules(db, username)
        rounding_mode = rules.get('rounding_mode', 'single')
        
        app.logger.info(f"Timbratura {username}: usando regole {rules.get('source')} (mode={rounding_mode})")
        
        # Ottieni turno per normalizzazione e rilevamento Extra Turno
        # Recupera sia inizio che fine turno
        day_of_week = datetime.now().weekday()
        
        if tipo in ('inizio_giornata', 'fine_giornata'):
            # Prima cerca nei turni configurati manualmente (employee_shifts)
            try:
                ensure_employee_shifts_table(db)
                shift_row = db.execute(
                    f"""SELECT start_time, end_time FROM employee_shifts 
                       WHERE username = {placeholder} AND day_of_week = {placeholder} AND is_active = 1
                       ORDER BY start_time ASC LIMIT 1""",
                    (username, day_of_week)
                ).fetchone()
                if shift_row:
                    start_time = shift_row['start_time'] if isinstance(shift_row, dict) else shift_row[0]
                    end_time = shift_row['end_time'] if isinstance(shift_row, dict) else shift_row[1]
                    if start_time:
                        if hasattr(start_time, 'strftime'):
                            turno_start = start_time.strftime("%H:%M")
                        elif hasattr(start_time, 'total_seconds'):
                            # È un timedelta (MySQL TIME restituisce timedelta)
                            total_sec = int(start_time.total_seconds())
                            turno_start = f"{total_sec // 3600:02d}:{(total_sec % 3600) // 60:02d}"
                        else:
                            turno_start = str(start_time)[:5]
                    if end_time:
                        if hasattr(end_time, 'strftime'):
                            turno_end = end_time.strftime("%H:%M")
                        elif hasattr(end_time, 'total_seconds'):
                            # È un timedelta (MySQL TIME restituisce timedelta)
                            total_sec = int(end_time.total_seconds())
                            turno_end = f"{total_sec // 3600:02d}:{(total_sec % 3600) // 60:02d}"
                        else:
                            turno_end = str(end_time)[:5]
                    if turno_start or turno_end:
                        app.logger.info(f"Turno trovato in employee_shifts per {username}: {turno_start} - {turno_end}")
            except Exception as e:
                app.logger.warning(f"Errore lettura employee_shifts: {e}")
            
            # Se non trovato, cerca in rentman_plannings (turni da Rentman)
            if not turno_start or not turno_end:
                # Trova il rentman_crew_id dell'utente
                user_row = db.execute(
                    f"SELECT rentman_crew_id FROM app_users WHERE username = {placeholder}",
                    (username,)
                ).fetchone()
                
                if user_row:
                    crew_id = user_row['rentman_crew_id'] if isinstance(user_row, dict) else user_row[0]
                    if crew_id:
                        # Cerca turno di oggi per l'utente dalla tabella rentman_plannings
                        # Recupera PRIMO inizio (plan_start ASC) e ULTIMO fine (plan_end DESC)
                        turno_row = db.execute(
                            f"""SELECT 
                                   (SELECT plan_start FROM rentman_plannings 
                                    WHERE crew_id = {placeholder} AND planning_date = {placeholder} 
                                    ORDER BY plan_start ASC LIMIT 1) as first_start,
                                   (SELECT plan_end FROM rentman_plannings 
                                    WHERE crew_id = {placeholder} AND planning_date = {placeholder} 
                                    ORDER BY plan_end DESC LIMIT 1) as last_end
                            """, (crew_id, today, crew_id, today)
                        ).fetchone()
                        if turno_row:
                            plan_start = turno_row['first_start'] if isinstance(turno_row, dict) else turno_row[0]
                            plan_end = turno_row['last_end'] if isinstance(turno_row, dict) else turno_row[1]
                            if plan_start and not turno_start:
                                if hasattr(plan_start, 'strftime'):
                                    turno_start = plan_start.strftime("%H:%M")
                                else:
                                    plan_str = str(plan_start)
                                    if len(plan_str) > 11:
                                        turno_start = plan_str[11:16]
                                    else:
                                        turno_start = plan_str[:5]
                            if plan_end and not turno_end:
                                if hasattr(plan_end, 'strftime'):
                                    turno_end = plan_end.strftime("%H:%M")
                                else:
                                    plan_str = str(plan_end)
                                    if len(plan_str) > 11:
                                        turno_end = plan_str[11:16]
                                    else:
                                        turno_end = plan_str[:5]
                            if turno_start or turno_end:
                                app.logger.info(f"Turno trovato in rentman_plannings per {username}: {turno_start} - {turno_end}")
        
        # Per fine_pausa, applica le regole sulla durata della pausa
        if tipo == 'fine_pausa':
            # Recupera l'ora di inizio pausa (l'ultima non chiusa)
            inizio_pausa_rows = db.execute(
                f"""SELECT ora, ora_mod FROM timbrature 
                   WHERE username = {placeholder} AND data = {placeholder} AND tipo = 'inizio_pausa'
                   ORDER BY created_ts DESC LIMIT 1""",
                (username, today)
            ).fetchone()
            
            if inizio_pausa_rows:
                inizio_ora_mod = inizio_pausa_rows['ora_mod'] if isinstance(inizio_pausa_rows, dict) else inizio_pausa_rows[1]
                if not inizio_ora_mod:
                    inizio_ora_mod = inizio_pausa_rows['ora'] if isinstance(inizio_pausa_rows, dict) else inizio_pausa_rows[0]
                
                # Formatta l'ora di inizio pausa
                if hasattr(inizio_ora_mod, 'strftime'):
                    inizio_str = inizio_ora_mod.strftime("%H:%M")
                else:
                    inizio_str = str(inizio_ora_mod)[:5]
                
                # Calcola durata modificata usando le regole pause
                durata_mod = calcola_pausa_mod(inizio_str, ora[:5], rules)
                
                # Calcola ora_mod di fine_pausa = inizio_pausa_mod + durata_mod
                inizio_parts = inizio_str.split(':')
                inizio_min = int(inizio_parts[0]) * 60 + int(inizio_parts[1])
                fine_mod_min = inizio_min + durata_mod
                
                h = fine_mod_min // 60
                m = fine_mod_min % 60
                ora_mod = f"{h:02d}:{m:02d}:00"
                
                app.logger.info(f"Pausa: {inizio_str} -> {ora[:5]} (durata effettiva {int(ora[:2])*60+int(ora[3:5]) - inizio_min} min, durata mod {durata_mod} min, fine mod {ora_mod})")
            else:
                ora_mod = calcola_ora_mod(ora, tipo, turno_start, rules)
        elif rounding_mode == 'daily' and tipo == 'fine_giornata':
            # Per daily mode: calcola l'ora di uscita che porta esattamente alle ore del turno
            ora_mod = _calcola_ora_fine_daily(db, username, today, ora, turno_start, turno_end, rules, placeholder)
        else:
            ora_mod = calcola_ora_mod(ora, tipo, turno_start, rules)
        
        # Per daily mode: verifica flessibilità e gestisci azione
        flex_request_id = None  # ID richiesta fuori flessibilità (se creata)
        if rounding_mode == 'daily' and tipo in ('inizio_giornata', 'fine_giornata'):
            flex_check = verifica_flessibilita_timbrata(ora, tipo, turno_start, turno_end, rules)
            app.logger.info(f"Flex check per {username}: {flex_check}")
            
            if not flex_check['within_flex']:
                action = flex_check['action']
                if action == 'block':
                    app.logger.warning(f"BLOCCO timbratura {username}: fuori flessibilità - {flex_check['message']}")
                    return jsonify({
                        "error": f"Timbratura fuori flessibilità: {flex_check['message']}. Contatta l'amministratore."
                    }), 400
                elif action == 'warn':
                    # Crea una richiesta di autorizzazione per fuori flessibilità
                    flex_warning = flex_check['message']
                    app.logger.warning(f"Timbratura {username} fuori flessibilità: {flex_warning}")
                    
                    # Crea richiesta "Fuori Flessibilità" (type_id=17)
                    try:
                        # Prendi arrotondamento dalle regole del gruppo
                        arrot_minuti = rules.get('arrotondamento_giornaliero_minuti', 30) if rules else 30
                        flex_request_id = _create_flex_request(
                            db, username, today, tipo, ora, ora_mod,
                            flex_check['diff_minutes'], flex_warning, placeholder,
                            turno_start, turno_end, arrot_minuti
                        )
                        app.logger.info(f"Creata richiesta Fuori Flessibilità ID={flex_request_id} per {username}")
                    except Exception as e:
                        app.logger.error(f"Errore creazione richiesta fuori flessibilità: {e}")
        
        app.logger.info(f"Ora mod calcolata: {ora} -> {ora_mod} (turno: {turno_start} - {turno_end}, tipo: {tipo}, mode: {rounding_mode})")
    except Exception as e:
        app.logger.error(f"Errore calcolo ora_mod: {e}")
        import traceback
        app.logger.error(traceback.format_exc())
        ora_mod = ora  # Fallback: usa ora originale
    
    # ═══════════════════════════════════════════════════════════════════════════
    # RILEVAMENTO EXTRA TURNO
    # ═══════════════════════════════════════════════════════════════════════════
    extra_turno_request_id = None
    extra_turno_data = None
    
    # Verifica se il modulo straordinari è attivo prima di rilevare Extra Turno
    if is_module_enabled(db, "straordinari") and tipo in ('inizio_giornata', 'fine_giornata'):
        try:
            # Usa le regole specifiche dell'utente (gruppo o globali)
            rules_for_extra = get_user_timbratura_rules(db, username)
            app.logger.info(
                "Checking Extra Turno: user=%s, tipo=%s, ora=%s, ora_mod=%s, turno_start=%s, turno_end=%s, rules=%s",
                username, tipo, ora, ora_mod, turno_start, turno_end, rules_for_extra
            )
            extra_turno_data = _detect_extra_turno(
                ora_timbrata=ora,
                ora_mod=ora_mod,
                tipo=tipo,
                turno_start=turno_start,
                turno_end=turno_end,
                rules=rules_for_extra
            )
            
            if extra_turno_data:
                app.logger.info(
                    "Extra Turno rilevato per %s: type=%s, minutes=%s, turno=%s",
                    username, extra_turno_data.get("extra_type"), 
                    extra_turno_data.get("extra_minutes"),
                    extra_turno_data.get("turno_time")
                )
                
                # Crea automaticamente la richiesta di Extra Turno
                extra_turno_data["planned_start"] = turno_start
                extra_turno_data["planned_end"] = turno_end
                
                extra_turno_request_id = _create_auto_extra_turno_request(
                    db=db,
                    username=username,
                    date_str=today,
                    extra_data=extra_turno_data,
                    notes=f"Extra Turno rilevato automaticamente - {extra_turno_data.get('extra_type')}"
                )
            else:
                app.logger.info(
                    "Extra Turno NON rilevato per %s: tipo=%s, ora=%s, ora_mod=%s, turno_start=%s, turno_end=%s",
                    username, tipo, ora, ora_mod, turno_start, turno_end
                )
        except Exception as e:
            app.logger.error(f"Errore rilevamento Extra Turno: {e}")
            import traceback
            app.logger.error(f"Traceback: {traceback.format_exc()}")
    
    # Recupera dati GPS dalla sessione o da offline_gps
    method = session.get("timbratura_method", "manual")
    gps_lat = session.get("timbratura_gps_lat")
    gps_lon = session.get("timbratura_gps_lon")
    location_name = session.get("timbratura_location")
    
    # Se è una timbratura offline GPS, usa quei dati
    if offline_gps and bypass_qr:
        method = "gps_offline"
        gps_lat = offline_gps.get('latitude')
        gps_lon = offline_gps.get('longitude')
        # Cerca location name dalle coordinate
        settings = get_company_settings(db)
        custom = settings.get('custom_settings', {})
        timbratura_config = custom.get('timbratura', {})
        locations = timbratura_config.get('gps_locations', [])
        for loc in locations:
            loc_lat = loc.get("latitude")
            loc_lon = loc.get("longitude")
            loc_radius = loc.get("radius_meters", 300)
            if loc_lat and loc_lon:
                from math import radians, sin, cos, sqrt, atan2
                R = 6371000
                lat1, lon1 = radians(float(loc_lat)), radians(float(loc_lon))
                lat2, lon2 = radians(float(gps_lat)), radians(float(gps_lon))
                dlat = lat2 - lat1
                dlon = lon2 - lon1
                a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
                c = 2 * atan2(sqrt(a), sqrt(1-a))
                distance = R * c
                if distance <= loc_radius:
                    location_name = loc.get("name", "Sede")
                    break
    elif offline_qr and bypass_qr:
        method = "qr_offline"
    
    # Inserisce la timbratura con dati GPS
    cursor = db.execute(
        f"""
        INSERT INTO timbrature (username, tipo, data, ora, ora_mod, created_ts, method, gps_lat, gps_lon, location_name)
        VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
        """,
        (username, tipo, today, ora, ora_mod, created_ts, method, gps_lat, gps_lon, location_name)
    )
    new_timbratura_id = cursor.lastrowid
    
    # Se fine_giornata, verifica che ora_mod non sia prima di fine_pausa.ora_mod
    # (la giornata non può finire prima che finisca la pausa normalizzata)
    if tipo == 'fine_giornata':
        fine_pausa_row = db.execute(
            f"""
            SELECT ora_mod FROM timbrature 
            WHERE username = {placeholder} AND data = {placeholder} AND tipo = 'fine_pausa'
            ORDER BY ora DESC LIMIT 1
            """,
            (username, today)
        ).fetchone()
        
        if fine_pausa_row:
            fp_ora_mod = fine_pausa_row['ora_mod'] if isinstance(fine_pausa_row, dict) else fine_pausa_row[0]
            
            if fp_ora_mod and ora_mod:
                # Converti in minuti per confronto
                fp_mod_parts = str(fp_ora_mod).split(':')
                fg_parts = str(ora_mod).split(':')
                fp_mod_min = int(fp_mod_parts[0]) * 60 + int(fp_mod_parts[1])
                fg_min = int(fg_parts[0]) * 60 + int(fg_parts[1])
                
                if fp_mod_min > fg_min:
                    # fine_giornata.ora_mod deve essere almeno = fine_pausa.ora_mod
                    app.logger.info(f"Aggiusto fine_giornata: ora_mod {ora_mod} < fine_pausa.ora_mod {fp_ora_mod}, aggiorno a {fp_ora_mod}")
                    db.execute(
                        f"UPDATE timbrature SET ora_mod = {placeholder} WHERE id = {placeholder}",
                        (fp_ora_mod, new_timbratura_id)
                    )
                    ora_mod = fp_ora_mod  # Aggiorna per CedolinoWeb
    
    # Pulisce dati timbratura dalla sessione
    session.pop("timbratura_method", None)
    session.pop("timbratura_gps_lat", None)
    session.pop("timbratura_gps_lon", None)
    session.pop("timbratura_location", None)
    
    # CedolinoWeb: invia timbrata con ora originale e modificata
    # Mappa tipo timbratura -> timeframe CedolinoWeb
    TIPO_TO_TIMEFRAME = {
        'inizio_giornata': TIMEFRAME_INIZIO_GIORNATA,  # 1
        'inizio_pausa': TIMEFRAME_INIZIO_PAUSA,        # 4
        'fine_pausa': TIMEFRAME_FINE_PAUSA,            # 5
        'fine_giornata': TIMEFRAME_FINE_GIORNATA,      # 8
    }
    
    timeframe_id = TIPO_TO_TIMEFRAME.get(tipo)
    cedolino_url = None
    
    if timeframe_id:
        # Recupera il nome dell'utente per il log
        user_row = db.execute(
            f"SELECT display_name FROM app_users WHERE username = {placeholder}",
            (username,)
        ).fetchone()
        display_name = username
        if user_row:
            display_name = (user_row['display_name'] if isinstance(user_row, dict) else user_row[0]) or username
        
        # Genera sempre l'URL di debug (anche se CedolinoWeb è disabilitato)
        from urllib.parse import urlencode
        external_id_debug = get_external_id_for_username(db, username) or "N/A"
        external_group_id_debug = get_external_group_id_for_username(db, username) or "NULL"
        debug_params = {
            "data_riferimento": today,
            "data_originale": f"{today} {ora}",
            "data_modificata": f"{today} {ora_mod or ora}",
            "codice_utente": external_id_debug,
            "codice_terminale": CEDOLINO_CODICE_TERMINALE,
            "timeframe_id": str(timeframe_id),
            "assunzione_id": external_id_debug,
            "terminale_id": "NULL",
            "gruppo_id": external_group_id_debug,
            "turno_id": "NULL",
            "note": "",
            "validata": "true",
        }
        cedolino_url = f"{CEDOLINO_WEB_ENDPOINT}?{urlencode(debug_params)}"
        
        # Determina se bloccare la sync per Extra Turno o Fuori Flessibilità:
        # 1. Se abbiamo appena creato una richiesta Extra Turno automatica, usa quell'ID
        # 2. Se abbiamo appena creato una richiesta Fuori Flessibilità automatica, usa quell'ID
        # 3. Altrimenti verifica se c'è una richiesta Extra Turno pending esistente
        overtime_request_id = extra_turno_request_id  # Usa l'ID Extra Turno appena creato se presente
        
        # Se c'è una richiesta Fuori Flessibilità, quella ha la priorità (blocca prima dell'Extra Turno)
        if flex_request_id:
            overtime_request_id = flex_request_id
        elif not overtime_request_id:
            # Cerca richieste Extra Turno pending esistenti per oggi
            overtime_request_id = _get_pending_overtime_request_id(db, username, today)
        
        if overtime_request_id:
            request_type = "Fuori Flessibilità" if flex_request_id else "Extra Turno"
            app.logger.info(
                "Timbratura %s per %s bloccata in attesa revisione %s (request_id=%s)",
                tipo, username, request_type, overtime_request_id
            )
        
        timbrata_ok, external_id, timbrata_error, _ = send_timbrata_utente(
            db,
            username=username,
            member_name=display_name,
            timeframe_id=timeframe_id,
            data_riferimento=today,
            ora_originale=ora,
            ora_modificata=ora_mod or ora,
            overtime_request_id=overtime_request_id,
        )
        
        if not timbrata_ok and external_id is None:
            # Utente senza ID esterno - blocca l'operazione (rollback implicito)
            return jsonify({
                "error": timbrata_error or "Utente senza ID esterno CedolinoWeb. Contattare l'amministratore.",
                "missing_external_id": True,
                "cedolino_url": cedolino_url  # Mostra comunque l'URL per debug
            }), 400
    
    db.commit()
    
    # Invalida il token dopo l'uso (ogni timbratura richiede nuova scansione)
    session.pop('timbratura_token', None)
    session.pop('timbratura_token_expires', None)
    
    app.logger.info(f"Timbratura registrata: {username} - {tipo} alle {ora}")
    
    # Restituisci anche l'URL CedolinoWeb per debug
    response_data = {"success": True, "tipo": tipo, "ora": ora[:5]}
    if cedolino_url:
        response_data["cedolino_url"] = cedolino_url
    
    # Se c'è un Extra Turno rilevato, informa il frontend
    if extra_turno_data:
        response_data["extra_turno"] = {
            "detected": True,
            "type": extra_turno_data.get("extra_type"),
            "minutes": extra_turno_data.get("extra_minutes"),
            "request_id": extra_turno_request_id
        }
    
    # Se c'è un warning Fuori Flessibilità, informa il frontend
    if flex_warning and flex_request_id:
        response_data["flex_warning"] = {
            "detected": True,
            "message": flex_warning,
            "request_id": flex_request_id,
            "tipo": tipo
        }
    
    app.logger.info(f"Risposta timbratura per {username}: extra_turno_data={extra_turno_data}, flex_warning={flex_warning}, response_data={response_data}")
    
    return jsonify(response_data)


@app.post("/api/user/change-password")
@login_required
def api_user_change_password():
    """Permette all'utente di cambiare la propria password."""
    username = session.get('user')
    if not username:
        return jsonify({"error": "Non autenticato"}), 401
    
    data = request.get_json()
    if not data:
        return jsonify({"error": "Dati mancanti"}), 400
    
    current_password = data.get('current_password', '')
    new_password = data.get('new_password', '')
    
    if not current_password or not new_password:
        return jsonify({"error": "Password attuale e nuova password richieste"}), 400
    
    if len(new_password) < 4:
        return jsonify({"error": "La nuova password deve essere di almeno 4 caratteri"}), 400
    
    db = get_db()
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Verifica password attuale
    user = db.execute(
        f"SELECT password_hash FROM app_users WHERE username = {placeholder}",
        (username,)
    ).fetchone()
    
    if not user:
        return jsonify({"error": "Utente non trovato"}), 404
    
    stored_hash = user['password_hash'] if isinstance(user, dict) else user[0]
    
    # Verifica la password attuale
    if not check_password_hash(stored_hash, current_password):
        return jsonify({"error": "Password attuale non corretta"}), 400
    
    # Genera hash per la nuova password
    new_hash = generate_password_hash(new_password)
    
    # Aggiorna nel database
    db.execute(
        f"UPDATE app_users SET password_hash = {placeholder}, updated_ts = {placeholder} WHERE username = {placeholder}",
        (new_hash, now_ms(), username)
    )
    db.commit()
    
    app.logger.info(f"Utente {username} ha cambiato la propria password")
    
    return jsonify({"success": True, "message": "Password aggiornata con successo"})


@app.get("/api/user/turno-oggi")
@login_required
def api_user_turno_oggi():
    """Restituisce il turno dell'utente per oggi (da Rentman o da employee_shifts)."""
    try:
        username = session.get('user')
        if not username:
            return jsonify({"error": "Non autenticato"}), 401
        
        db = get_db()
        placeholder = "%s" if DB_VENDOR == "mysql" else "?"
        
        # Trova il rentman_crew_id dell'utente
        user_row = db.execute(
            f"SELECT rentman_crew_id FROM app_users WHERE username = {placeholder}",
            (username,)
        ).fetchone()
        
        if not user_row:
            return jsonify({"turno": None, "turni": [], "message": "Utente non trovato"})
        
        crew_id = user_row['rentman_crew_id'] if isinstance(user_row, dict) else user_row[0]
        
        # Se non ha crew_id Rentman, cerca nella tabella employee_shifts
        if not crew_id:
            ensure_employee_shifts_table(db)
            
            # Carica le gps_locations dalla configurazione per associare coordinate ai turni
            settings = get_company_settings(db)
            custom = settings.get('custom_settings', {})
            timbratura_config = custom.get('timbratura', {})
            gps_locations = timbratura_config.get('gps_locations', [])
            
            # Trova il giorno della settimana (0=Lunedì, 6=Domenica)
            today = get_simulated_now()
            day_of_week = today.weekday()
            
            shift_row = db.execute(f"""
                SELECT start_time, end_time, break_start, break_end, location_name
                FROM employee_shifts
                WHERE username = {placeholder} AND day_of_week = {placeholder} AND is_active = 1
            """, (username, day_of_week)).fetchone()
            
            if shift_row:
                if isinstance(shift_row, dict):
                    start_time = str(shift_row['start_time'])[:5] if shift_row['start_time'] else None
                    end_time = str(shift_row['end_time'])[:5] if shift_row['end_time'] else None
                    break_start = str(shift_row['break_start'])[:5] if shift_row['break_start'] else None
                    break_end = str(shift_row['break_end'])[:5] if shift_row['break_end'] else None
                    location_name = shift_row.get('location_name')
                else:
                    start_time = str(shift_row[0])[:5] if shift_row[0] else None
                    end_time = str(shift_row[1])[:5] if shift_row[1] else None
                    break_start = str(shift_row[2])[:5] if shift_row[2] else None
                    break_end = str(shift_row[3])[:5] if shift_row[3] else None
                    location_name = shift_row[4] if len(shift_row) > 4 else None
                
                # Cerca coordinate della location nelle gps_locations configurate
                timbratura_lat = None
                timbratura_lon = None
                timbratura_radius = 300
                if location_name:
                    for loc in gps_locations:
                        if loc.get('name') == location_name:
                            timbratura_lat = loc.get('latitude')
                            timbratura_lon = loc.get('longitude')
                            timbratura_radius = loc.get('radius_meters', 300)
                            app.logger.info(f"✅ employee_shifts: Location '{location_name}' trovata con coordinate: {timbratura_lat}, {timbratura_lon}, raggio={timbratura_radius}")
                            break
                    if not timbratura_lat:
                        app.logger.warning(f"⚠️ employee_shifts: Location '{location_name}' NON trovata nelle gps_locations: {[l.get('name') for l in gps_locations]}")
                
                turno = {
                    "project_code": "UFFICIO",
                    "project_name": "Lavoro in ufficio",
                    "function": "Impiegato",
                    "start": start_time,
                    "end": end_time,
                    "break_start": break_start,
                    "break_end": break_end,
                    "hours": None,
                    "note": None,
                    "is_leader": False,
                    "transport": None,
                    "source": "employee_shifts",
                    "location_name": location_name,
                    "timbratura_location": location_name,
                    "timbratura_lat": timbratura_lat,
                    "timbratura_lon": timbratura_lon,
                    "timbratura_radius": timbratura_radius,
                    "gps_mode": "group",
                }
                return jsonify({"turno": turno, "turni": [turno]})
            
            return jsonify({"turno": None, "turni": [], "message": "Nessun turno configurato"})
        
        # Se ha crew_id, cerca in rentman_plannings
        today = get_simulated_now().strftime("%Y-%m-%d")
        ensure_rentman_plannings_table(db)
        
        # DEBUG: Mostra cosa c'è nel database per questo utente oggi
        debug_rows = db.execute(
            f"""SELECT crew_id, planning_date, project_name, location_name, 
                      timbratura_gps_mode, gps_timbratura_location, location_id
               FROM rentman_plannings WHERE crew_id = {placeholder} AND planning_date = {placeholder}
               AND (is_obsolete = 0 OR is_obsolete IS NULL)""",
            (crew_id, today)
        ).fetchall()
        
        if debug_rows:
            for debug_row in debug_rows:
                if isinstance(debug_row, dict):
                    app.logger.info(f"DEBUG DB: project={debug_row.get('project_name')}, "
                                  f"location_name={debug_row.get('location_name')}, "
                                  f"gps_mode={debug_row.get('timbratura_gps_mode')}, "
                                  f"gps_timbratura_location={debug_row.get('gps_timbratura_location')}, "
                                  f"location_id={debug_row.get('location_id')}")
                else:
                    app.logger.info(f"DEBUG DB (tuple): {debug_row}")
        else:
            app.logger.warning(f"DEBUG: Nessun turno trovato per crew_id={crew_id}, date={today}")
        
        planning = db.execute(
            f"""
            SELECT project_code, project_name, function_name, plan_start, plan_end,
                   hours_planned, remark, is_leader, transport, break_start, break_end, break_minutes,
                   location_name, timbratura_gps_mode, gps_timbratura_location, location_id, remark_planner
            FROM rentman_plannings
            WHERE crew_id = {placeholder} AND planning_date = {placeholder} AND sent_to_webservice = 1
              AND (is_obsolete = 0 OR is_obsolete IS NULL)
            ORDER BY plan_start ASC
            """,
            (crew_id, today)
        ).fetchall()
        
        if not planning:
            return jsonify({"turno": None, "turni": [], "message": "Nessun turno per oggi"})
        
        turni = []
        for row in planning:
            if isinstance(row, dict):
                plan_start = row['plan_start']
                plan_end = row['plan_end']
            else:
                plan_start = row[3]
                plan_end = row[4]
            
            start_str = ""
            end_str = ""
            if plan_start:
                if hasattr(plan_start, 'strftime'):
                    start_str = plan_start.strftime("%H:%M")
                else:
                    start_str = str(plan_start)[11:16] if len(str(plan_start)) > 11 else str(plan_start)[:5]
            if plan_end:
                if hasattr(plan_end, 'strftime'):
                    end_str = plan_end.strftime("%H:%M")
                else:
                    end_str = str(plan_end)[11:16] if len(str(plan_end)) > 11 else str(plan_end)[:5]
            
            if isinstance(row, dict):
                break_start = format_time_value(row.get('break_start'))
                break_end = format_time_value(row.get('break_end'))
                break_minutes = row.get('break_minutes')
                location_name = row.get('location_name')
                gps_mode = row.get('timbratura_gps_mode') or 'group'
                gps_timbratura_location = row.get('gps_timbratura_location')
                location_id = row.get('location_id')
                remark_planner = row.get('remark_planner')
            else:
                break_start = format_time_value(row[9] if len(row) > 9 else None)
                break_end = format_time_value(row[10] if len(row) > 10 else None)
                break_minutes = row[11] if len(row) > 11 else None
                location_name = row[12] if len(row) > 12 else None
                gps_mode = (row[13] if len(row) > 13 else None) or 'group'
                gps_timbratura_location = row[14] if len(row) > 14 else None
                location_id = row[15] if len(row) > 15 else None
                remark_planner = row[16] if len(row) > 16 else None
            
            # Coordinate dalla cache globale
            location_lat, location_lon, location_radius = None, None, 300
            if location_name:
                ensure_location_cache_table(db)
                cached_coords = get_location_cache(db, location_name, location_id)
                if cached_coords:
                    location_lat, location_lon, location_radius = cached_coords
                    app.logger.info(f"✅ Location '{location_name}' (id={location_id}): usando coordinate dalla cache globale (turno-oggi): {location_lat}, {location_lon}, raggio={location_radius}m")
                else:
                    app.logger.info(f"⚠️ Location '{location_name}' (id={location_id}): nessuna cache (turno-oggi)")
            
            # Determina dove timbrare e le coordinate GPS
            timbratura_location = None
            timbratura_lat = None
            timbratura_lon = None
            timbratura_radius = 300
            
            # DEBUG
            app.logger.info(f"DEBUG TURNO OGGI: gps_mode={gps_mode}, location_name={location_name}, location_lat={location_lat}, location_lon={location_lon}, gps_timbratura_location={gps_timbratura_location}")
            
            # LOGICA TIMBRATURA:
            # - timbratura_location: dove registrare la timbratura (il nome della sede)
            # - timbratura_lat/lon: le coordinate di quella sede
            
            if gps_mode == 'location':
                # Modalità LOCATION: timbrare presso la location del progetto (da Rentman)
                timbratura_location = location_name
                if location_lat and location_lon:
                    timbratura_lat = location_lat
                    timbratura_lon = location_lon
                    timbratura_radius = location_radius  # Usa il raggio dalla cache
                    app.logger.info(f"DEBUG: Modo LOCATION - Timbratura presso '{timbratura_location}' con coordinate da Rentman: {timbratura_lat}, {timbratura_lon}, raggio={timbratura_radius}m")
                else:
                    app.logger.warning(f"DEBUG: Modo LOCATION - Location '{location_name}' senza coordinate in Rentman")
            
            elif gps_mode == 'group':
                # Modalità GROUP: timbrare presso la sede del gruppo (da config aziendale)
                # timbratura_location è il nome della sede dove timbrare (gps_timbratura_location)
                timbratura_location = gps_timbratura_location or location_name
                
                app.logger.info(f"DEBUG: Modo GROUP - timbratura_location='{timbratura_location}' (gps_timbratura_location='{gps_timbratura_location}'), location_name='{location_name}'")
                
                if gps_timbratura_location:
                    # Usa la sede configurata per il gruppo
                    try:
                        settings = get_company_settings(db)
                        app.logger.info(f"DEBUG: Company settings retrieved: {type(settings)}")
                        
                        custom = settings.get('custom_settings', {})
                        app.logger.info(f"DEBUG: custom_settings keys: {custom.keys() if custom else 'empty'}")
                        
                        timbratura_config = custom.get('timbratura', {})
                        app.logger.info(f"DEBUG: timbratura_config keys: {timbratura_config.keys() if timbratura_config else 'empty'}")
                        
                        gps_locations = timbratura_config.get('gps_locations', [])
                        app.logger.info(f"DEBUG: gps_locations count: {len(gps_locations)}, names: {[l.get('name') for l in gps_locations]}")
                        
                        for loc in gps_locations:
                            loc_name = loc.get('name')
                            app.logger.info(f"DEBUG: Comparing '{gps_timbratura_location}' == '{loc_name}' ? {gps_timbratura_location == loc_name}")
                            if loc_name == gps_timbratura_location:
                                timbratura_lat = loc.get('latitude')
                                timbratura_lon = loc.get('longitude')
                                timbratura_radius = loc.get('radius_meters', 300)
                                app.logger.info(f"DEBUG: ✓ Modo GROUP - TROVATO! Timbratura presso '{gps_timbratura_location}' con coordinate: {timbratura_lat}, {timbratura_lon}, raggio={timbratura_radius}m")
                                break
                        if not timbratura_lat or not timbratura_lon:
                            app.logger.warning(f"DEBUG: ✗ Modo GROUP - Sede gruppo '{gps_timbratura_location}' NON TROVATA in config. Sedi disponibili: {[l.get('name') for l in gps_locations]}")
                    except Exception as e:
                        app.logger.error(f"DEBUG: Errore lettura config aziendale (gps_mode=group): {e}", exc_info=True)
                else:
                    app.logger.warning(f"DEBUG: Modo GROUP - gps_timbratura_location È VUOTO! Non posso cercare le coordinate.")
            
            else:
                app.logger.warning(f"DEBUG: gps_mode sconosciuto: '{gps_mode}'")
            
            turni.append({
                "project_code": row['project_code'] if isinstance(row, dict) else row[0],
                "project_name": row['project_name'] if isinstance(row, dict) else row[1],
                "function": row['function_name'] if isinstance(row, dict) else row[2],
                "start": start_str,
                "end": end_str,
                "hours": float(row['hours_planned'] if isinstance(row, dict) else row[5] or 0),
                "note": row['remark'] if isinstance(row, dict) else row[6],
                "note_planner": remark_planner,
                "is_leader": bool(row['is_leader'] if isinstance(row, dict) else row[7]),
                "transport": row['transport'] if isinstance(row, dict) else row[8],
                "break_start": break_start,
                "break_end": break_end,
                "break_minutes": break_minutes,
                "location_name": location_name,
                "gps_mode": gps_mode,
                "gps_timbratura_location": gps_timbratura_location,
                "timbratura_location": timbratura_location,
                "timbratura_lat": timbratura_lat,
                "timbratura_lon": timbratura_lon,
                "timbratura_radius": timbratura_radius,
            })
        
        return jsonify({"turno": turni[0] if turni else None, "turni": turni})
    
    except Exception as e:
        app.logger.exception(f"Errore in api_user_turno_oggi: {str(e)}")
        return jsonify({"error": f"Errore: {str(e)}", "turno": None, "turni": []}), 500


@app.get("/api/user/turni")
@login_required
def api_user_turni():
    """Restituisce tutti i turni dell'utente (7 giorni indietro + 60 giorni avanti da Rentman o employee_shifts)."""
    try:
        username = session.get('user')
        if not username:
            return jsonify({"error": "Non autenticato", "turni": []}), 401
        
        db = get_db()
        placeholder = "%s" if DB_VENDOR == "mysql" else "?"
        
        # Trova il rentman_crew_id dell'utente
        user_row = db.execute(
            f"SELECT rentman_crew_id FROM app_users WHERE username = {placeholder}",
            (username,)
        ).fetchone()
        
        if not user_row:
            return jsonify({"turni": [], "message": "Utente non trovato"})
        
        crew_id = user_row['rentman_crew_id'] if isinstance(user_row, dict) else user_row[0]
        
        # Se non ha crew_id Rentman, usa employee_shifts ricorrenti
        if not crew_id:
            ensure_employee_shifts_table(db)
            
            # Carica le gps_locations dalla configurazione per associare coordinate ai turni
            settings = get_company_settings(db)
            custom = settings.get('custom_settings', {})
            timbratura_config = custom.get('timbratura', {})
            gps_locations = timbratura_config.get('gps_locations', [])
            
            turni = []
            # Genera turni per i 7 giorni precedenti + 60 giorni successivi basandosi su employee_shifts
            today = get_simulated_now()
            for days_offset in range(-7, 60):  # 7 giorni indietro + 60 giorni avanti
                check_date = today + timedelta(days=days_offset)
                day_of_week = check_date.weekday()
                
                shift_row = db.execute(f"""
                    SELECT start_time, end_time, break_start, break_end, location_name
                    FROM employee_shifts
                    WHERE username = {placeholder} AND day_of_week = {placeholder} AND is_active = 1
                """, (username, day_of_week)).fetchone()
                
                if shift_row:
                    if isinstance(shift_row, dict):
                        start_time = str(shift_row['start_time'])[:5] if shift_row['start_time'] else None
                        end_time = str(shift_row['end_time'])[:5] if shift_row['end_time'] else None
                        break_start = str(shift_row['break_start'])[:5] if shift_row['break_start'] else None
                        break_end = str(shift_row['break_end'])[:5] if shift_row['break_end'] else None
                        location_name = shift_row.get('location_name')
                    else:
                        start_time = str(shift_row[0])[:5] if shift_row[0] else None
                        end_time = str(shift_row[1])[:5] if shift_row[1] else None
                        break_start = str(shift_row[2])[:5] if shift_row[2] else None
                        break_end = str(shift_row[3])[:5] if shift_row[3] else None
                        location_name = shift_row[4] if len(shift_row) > 4 else None
                    
                    # Cerca coordinate della location nelle gps_locations configurate
                    timbratura_lat = None
                    timbratura_lon = None
                    timbratura_radius = 300
                    if location_name:
                        for loc in gps_locations:
                            if loc.get('name') == location_name:
                                timbratura_lat = loc.get('latitude')
                                timbratura_lon = loc.get('longitude')
                                timbratura_radius = loc.get('radius_meters', 300)
                                break
                    
                    # Calcola ore e minuti pausa
                    hours_val = None
                    break_minutes_val = 0
                    
                    if start_time and end_time:
                        try:
                            start_parts = start_time.split(':')
                            end_parts = end_time.split(':')
                            start_mins = int(start_parts[0]) * 60 + int(start_parts[1])
                            end_mins = int(end_parts[0]) * 60 + int(end_parts[1])
                            total_mins = end_mins - start_mins
                            if total_mins > 0:
                                hours_val = total_mins / 60.0
                        except:
                            pass
                    
                    if break_start and break_end:
                        try:
                            bs_parts = break_start.split(':')
                            be_parts = break_end.split(':')
                            bs_mins = int(bs_parts[0]) * 60 + int(bs_parts[1])
                            be_mins = int(be_parts[0]) * 60 + int(be_parts[1])
                            break_minutes_val = be_mins - bs_mins
                            if break_minutes_val < 0:
                                break_minutes_val = 0
                        except:
                            pass
                    
                    turni.append({
                        "date": check_date.strftime("%Y-%m-%d"),
                        "project_code": "UFFICIO",
                        "project_name": "Lavoro in ufficio",
                        "function": "Impiegato",
                        "start": start_time,
                        "end": end_time,
                        "break_start": break_start,
                        "break_end": break_end,
                        "hours": hours_val,
                        "break_minutes": break_minutes_val,
                        "note": None,
                        "is_leader": False,
                        "transport": None,
                        "source": "employee_shifts",
                        "location_name": location_name,
                        "timbratura_location": location_name,
                        "timbratura_lat": timbratura_lat,
                        "timbratura_lon": timbratura_lon,
                        "timbratura_radius": timbratura_radius,
                    })
            
            return jsonify({"turni": turni})
        
        # Se ha crew_id, cerca in rentman_plannings (ultimi 30 giorni + prossimi 60 giorni)
        ensure_rentman_plannings_table(db)
        today = get_simulated_now()
        sixty_days_future = (today + timedelta(days=60)).strftime("%Y-%m-%d")
        thirty_days_past = (today - timedelta(days=30)).strftime("%Y-%m-%d")
        
        # Carica le gps_locations dalla configurazione per associare coordinate ai turni
        settings = get_company_settings(db)
        custom = settings.get('custom_settings', {})
        timbratura_config = custom.get('timbratura', {})
        gps_locations = timbratura_config.get('gps_locations', [])
        
        planning = db.execute(
            f"""
            SELECT planning_date, project_code, project_name, function_name, plan_start, plan_end,
                   hours_planned, remark, is_leader, transport, break_start, break_end, break_minutes,
                   location_name, timbratura_gps_mode, gps_timbratura_location
            FROM rentman_plannings
            WHERE crew_id = {placeholder} AND planning_date >= {placeholder} AND planning_date <= {placeholder}
              AND sent_to_webservice = 1
              AND (is_obsolete = 0 OR is_obsolete IS NULL)
            ORDER BY planning_date ASC, plan_start ASC
            """,
            (crew_id, thirty_days_past, sixty_days_future)
        ).fetchall()
        
        turni = []
        for row in planning:
            if isinstance(row, dict):
                planning_date = row['planning_date']
                plan_start = row['plan_start']
                plan_end = row['plan_end']
            else:
                planning_date = row[0]
                plan_start = row[4]
                plan_end = row[5]
            
            # Normalizza la data
            if hasattr(planning_date, 'isoformat'):
                date_str = planning_date.isoformat()
            else:
                date_str = str(planning_date)[:10]
            
            # Formatta orari
            start_str = ""
            end_str = ""
            if plan_start:
                if hasattr(plan_start, 'strftime'):
                    start_str = plan_start.strftime("%H:%M")
                else:
                    start_str = str(plan_start)[11:16] if len(str(plan_start)) > 11 else str(plan_start)[:5]
            if plan_end:
                if hasattr(plan_end, 'strftime'):
                    end_str = plan_end.strftime("%H:%M")
                else:
                    end_str = str(plan_end)[11:16] if len(str(plan_end)) > 11 else str(plan_end)[:5]
            
            # Estrai info break
            if isinstance(row, dict):
                break_start = format_time_value(row.get('break_start'))
                break_end = format_time_value(row.get('break_end'))
                break_minutes = row.get('break_minutes')
                location_name = row.get('location_name')
                gps_mode = row.get('timbratura_gps_mode') or 'group'
                gps_timbratura_location = row.get('gps_timbratura_location')
            else:
                break_start = format_time_value(row[10] if len(row) > 10 else None)
                break_end = format_time_value(row[11] if len(row) > 11 else None)
                break_minutes = row[12] if len(row) > 12 else None
                location_name = row[13] if len(row) > 13 else None
                gps_mode = row[14] if len(row) > 14 else 'group'
                gps_timbratura_location = row[15] if len(row) > 15 else None
            
            # Determina dove timbrare
            timbratura_location = None
            timbratura_lat = None
            timbratura_lon = None
            timbratura_radius = 300
            
            if gps_mode == 'location' and location_name:
                timbratura_location = location_name
                # Cerca coordinate nella cache location di Rentman
                try:
                    ensure_location_cache_table(db)
                    loc_id = row.get('location_id') if isinstance(row, dict) else None
                    cached = get_location_cache(db, location_name, loc_id)
                    if cached:
                        timbratura_lat, timbratura_lon, timbratura_radius = cached
                except:
                    pass
            elif gps_timbratura_location:
                timbratura_location = gps_timbratura_location
                # Cerca coordinate nelle gps_locations configurate
                for loc in gps_locations:
                    if loc.get('name') == gps_timbratura_location:
                        timbratura_lat = loc.get('latitude')
                        timbratura_lon = loc.get('longitude')
                        timbratura_radius = loc.get('radius_meters', 300)
                        break
            
            turni.append({
                "date": date_str,
                "project_code": row['project_code'] if isinstance(row, dict) else row[1],
                "project_name": row['project_name'] if isinstance(row, dict) else row[2],
                "function": row['function_name'] if isinstance(row, dict) else row[3],
                "start": start_str,
                "end": end_str,
                "hours": float(row['hours_planned'] if isinstance(row, dict) else row[6] or 0),
                "note": row['remark'] if isinstance(row, dict) else row[7],
                "is_leader": bool(row['is_leader'] if isinstance(row, dict) else row[8]),
                "transport": row['transport'] if isinstance(row, dict) else row[9],
                "break_start": break_start,
                "break_end": break_end,
                "break_minutes": break_minutes,
                "location_name": location_name,
                "timbratura_location": timbratura_location,
                "timbratura_lat": timbratura_lat,
                "timbratura_lon": timbratura_lon,
                "timbratura_radius": timbratura_radius,
                "gps_mode": gps_mode,
                "gps_timbratura_location": gps_timbratura_location,
            })
        
        return jsonify({"turni": turni})
    
    except Exception as e:
        app.logger.exception(f"Errore in api_user_turni: {str(e)}")
        return jsonify({"error": f"Errore: {str(e)}", "turni": []}), 500


@app.get("/magazzino")
@login_required
def magazzino_home() -> ResponseReturnValue:
    guard = _magazzino_only()
    if guard is not None:
        return guard

    db = get_db()
    display_name = session.get('user_display') or session.get('user_name') or session.get('user')
    primary_name = session.get('user_name') or display_name or session.get('user')
    initials = session.get('user_initials') or compute_initials(primary_name or "")
    header_clock = datetime.now().strftime("%d/%m/%Y | %H:%M")
    user_role = session.get('user_role', 'user')
    
    # Company logo
    settings = get_company_settings(db)
    company_logo = settings.get('logo_url') if settings else None

    return render_template(
        "magazzino.html",
        user_name=primary_name,
        user_display=display_name,
        user_initials=initials,
        user_role=user_role,
        header_clock=header_clock,
        company_logo=company_logo,
    )


WAREHOUSE_ACTIVITIES_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS warehouse_activities (
    id INT AUTO_INCREMENT PRIMARY KEY,
    project_code VARCHAR(128) NOT NULL,
    activity_label VARCHAR(255) NOT NULL,
    note TEXT,
    username VARCHAR(190) DEFAULT NULL,
    created_ts BIGINT NOT NULL,
    INDEX idx_warehouse_project (project_code),
    INDEX idx_warehouse_created (created_ts)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""


WAREHOUSE_ACTIVITIES_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS warehouse_activities (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_code TEXT NOT NULL,
    activity_label TEXT NOT NULL,
    note TEXT,
    username TEXT,
    created_ts INTEGER NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_warehouse_project ON warehouse_activities(project_code);
CREATE INDEX IF NOT EXISTS idx_warehouse_created ON warehouse_activities(created_ts);
"""


# Tabella per sessioni di lavoro magazzino (con timer)
WAREHOUSE_SESSIONS_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS warehouse_sessions (
    id INT AUTO_INCREMENT PRIMARY KEY,
    project_code VARCHAR(128) NOT NULL,
    activity_label VARCHAR(255) NOT NULL,
    elapsed_ms BIGINT NOT NULL DEFAULT 0,
    start_ts BIGINT DEFAULT NULL,
    end_ts BIGINT DEFAULT NULL,
    note TEXT,
    username VARCHAR(190) DEFAULT NULL,
    created_ts BIGINT NOT NULL,
    INDEX idx_wh_sessions_project (project_code),
    INDEX idx_wh_sessions_created (created_ts),
    INDEX idx_wh_sessions_user (username)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""


WAREHOUSE_SESSIONS_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS warehouse_sessions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_code TEXT NOT NULL,
    activity_label TEXT NOT NULL,
    elapsed_ms INTEGER NOT NULL DEFAULT 0,
    start_ts INTEGER DEFAULT NULL,
    end_ts INTEGER DEFAULT NULL,
    note TEXT,
    username TEXT,
    created_ts INTEGER NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_wh_sessions_project ON warehouse_sessions(project_code);
CREATE INDEX IF NOT EXISTS idx_wh_sessions_created ON warehouse_sessions(created_ts);
CREATE INDEX IF NOT EXISTS idx_wh_sessions_user ON warehouse_sessions(username);
"""


# Progetti manuali magazzino (persistenza cross-device per utente)
WAREHOUSE_MANUAL_PROJECTS_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS warehouse_manual_projects (
    id INT AUTO_INCREMENT PRIMARY KEY,
    project_code VARCHAR(128) NOT NULL,
    name VARCHAR(255) DEFAULT NULL,
    username VARCHAR(190) NOT NULL,
    created_ts BIGINT NOT NULL,
    UNIQUE KEY uniq_wh_manual_user_code (username, project_code),
    INDEX idx_wh_manual_user (username)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""


WAREHOUSE_MANUAL_PROJECTS_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS warehouse_manual_projects (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_code TEXT NOT NULL,
    name TEXT,
    username TEXT NOT NULL,
    created_ts INTEGER NOT NULL,
    UNIQUE(username, project_code)
);
CREATE INDEX IF NOT EXISTS idx_wh_manual_user ON warehouse_manual_projects(username);
"""


# ============================================================
# TABELLE TIMBRATURA
# ============================================================

TIMBRATURE_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS timbrature (
    id INT AUTO_INCREMENT PRIMARY KEY,
    username VARCHAR(190) NOT NULL,
    tipo VARCHAR(50) NOT NULL,
    data DATE NOT NULL,
    ora TIME NOT NULL,
    created_ts BIGINT NOT NULL,
    method VARCHAR(20) DEFAULT NULL,
    gps_lat DECIMAL(10,8) DEFAULT NULL,
    gps_lon DECIMAL(11,8) DEFAULT NULL,
    location_name VARCHAR(255) DEFAULT NULL,
    INDEX idx_timbrature_user_date (username, data),
    INDEX idx_timbrature_date (data)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

TIMBRATURE_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS timbrature (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT NOT NULL,
    tipo TEXT NOT NULL,
    data TEXT NOT NULL,
    ora TEXT NOT NULL,
    created_ts INTEGER NOT NULL,
    method TEXT DEFAULT NULL,
    gps_lat REAL DEFAULT NULL,
    gps_lon REAL DEFAULT NULL,
    location_name TEXT DEFAULT NULL
);
CREATE INDEX IF NOT EXISTS idx_timbrature_user_date ON timbrature(username, data);
CREATE INDEX IF NOT EXISTS idx_timbrature_date ON timbrature(data);
"""


def ensure_timbrature_table(db: DatabaseLike) -> None:
    statement = TIMBRATURE_TABLE_MYSQL if DB_VENDOR == "mysql" else TIMBRATURE_TABLE_SQLITE
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass
    
    # Migrazione: aggiungi colonna ora_mod se non esiste
    if DB_VENDOR == "mysql":
        try:
            db.execute("ALTER TABLE timbrature ADD COLUMN ora_mod TIME DEFAULT NULL")
            db.commit()
        except Exception:
            pass  # Colonna già esistente
    else:
        try:
            db.execute("ALTER TABLE timbrature ADD COLUMN ora_mod TEXT")
            db.commit()
        except Exception:
            pass
    
    # Migrazione: aggiungi colonne GPS se non esistono
    new_columns = [
        ("method", "VARCHAR(20) DEFAULT NULL" if DB_VENDOR == "mysql" else "TEXT DEFAULT NULL"),
        ("gps_lat", "DECIMAL(10,8) DEFAULT NULL" if DB_VENDOR == "mysql" else "REAL DEFAULT NULL"),
        ("gps_lon", "DECIMAL(11,8) DEFAULT NULL" if DB_VENDOR == "mysql" else "REAL DEFAULT NULL"),
        ("location_name", "VARCHAR(255) DEFAULT NULL" if DB_VENDOR == "mysql" else "TEXT DEFAULT NULL"),
    ]
    for col_name, col_type in new_columns:
        try:
            db.execute(f"ALTER TABLE timbrature ADD COLUMN {col_name} {col_type}")
            db.commit()
        except Exception:
            pass  # Colonna già esistente


def ensure_warehouse_activities_table(db: DatabaseLike) -> None:
    statement = (
        WAREHOUSE_ACTIVITIES_TABLE_MYSQL if DB_VENDOR == "mysql" else WAREHOUSE_ACTIVITIES_TABLE_SQLITE
    )
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass


def ensure_warehouse_sessions_table(db: DatabaseLike) -> None:
    statement = (
        WAREHOUSE_SESSIONS_TABLE_MYSQL if DB_VENDOR == "mysql" else WAREHOUSE_SESSIONS_TABLE_SQLITE
    )
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass
    
    # Migrazione: aggiungi colonne start_ts e end_ts se non esistono
    try:
        if DB_VENDOR == "mysql":
            db.execute("ALTER TABLE warehouse_sessions ADD COLUMN start_ts BIGINT DEFAULT NULL")
            db.commit()
    except Exception:
        pass
    try:
        if DB_VENDOR == "mysql":
            db.execute("ALTER TABLE warehouse_sessions ADD COLUMN end_ts BIGINT DEFAULT NULL")
            db.commit()
    except Exception:
        pass
    
    # SQLite: verifica se le colonne esistono
    if DB_VENDOR != "mysql":
        try:
            cursor = db.execute("PRAGMA table_info(warehouse_sessions)")
            columns = {row[1] for row in cursor.fetchall()}
            if "start_ts" not in columns:
                db.execute("ALTER TABLE warehouse_sessions ADD COLUMN start_ts INTEGER DEFAULT NULL")
            if "end_ts" not in columns:
                db.execute("ALTER TABLE warehouse_sessions ADD COLUMN end_ts INTEGER DEFAULT NULL")
            if "intervals" not in columns:
                db.execute("ALTER TABLE warehouse_sessions ADD COLUMN intervals TEXT DEFAULT NULL")
            db.commit()
        except Exception:
            pass
    else:
        # MySQL: aggiungi colonna intervals se non esiste
        try:
            db.execute("ALTER TABLE warehouse_sessions ADD COLUMN intervals TEXT DEFAULT NULL")
            db.commit()
        except Exception:
            pass


def ensure_warehouse_manual_projects_table(db: DatabaseLike) -> None:
    statement = (
        WAREHOUSE_MANUAL_PROJECTS_TABLE_MYSQL
        if DB_VENDOR == "mysql"
        else WAREHOUSE_MANUAL_PROJECTS_TABLE_SQLITE
    )
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass


@app.get("/api/magazzino/projects/today")
@login_required
def api_magazzino_projects_today() -> ResponseReturnValue:
    """Recupera progetti attivi per la data specificata.

    Secondo la documentazione API Rentman (oas.json):
    - I progetti (projects) NON hanno un campo status diretto
    - Lo STATUS è nel SUBPROJECT (/subprojects) come riferimento "/statuses/ID"
    - Ogni progetto ha almeno un subproject

    Logica: recuperiamo i subprojects, filtriamo per status e data,
    poi raggruppiamo per progetto padre.
    """
    guard = _magazzino_only()
    if guard is not None:
        return guard

    requested = _normalize_date(request.args.get("date"))
    target_date = requested or datetime.now().date().isoformat()
    debug = _is_truthy(request.args.get("debug"))

    rentman_info: Dict[str, Any] = {"available": False, "used": False}

    # Keywords per filtrare gli status
    allowed_status_keywords: Tuple[str, ...] = (
        "confermat",
        "confirm",
        "in location",
        "locat",
        "pronto",
        "ready",
    )
    cancelled_keywords: Tuple[str, ...] = ("annull", "cancel")

    def normalize_status(value: str) -> str:
        cleaned = str(value or "").strip().lower()
        cleaned = cleaned.replace("_", " ")
        return " ".join(cleaned.split())

    def is_status_allowed(status_name: str) -> bool:
        normalized = normalize_status(status_name)
        if any(kw in normalized for kw in cancelled_keywords):
            return False
        if any(kw in normalized for kw in allowed_status_keywords):
            return True
        return False

    def subproject_matches_date(subproject: Mapping[str, Any], target: date) -> bool:
        date_fields = [
            ("equipment_period_from", "equipment_period_to"),
            ("usageperiod_start", "usageperiod_end"),
            ("planperiod_start", "planperiod_end"),
        ]
        for start_key, end_key in date_fields:
            start_val = _parse_date_any(subproject.get(start_key))
            end_val = _parse_date_any(subproject.get(end_key))
            if start_val is not None or end_val is not None:
                if start_val is None:
                    start_val = end_val
                if end_val is None:
                    end_val = start_val
                if start_val and end_val:
                    if end_val < start_val:
                        start_val, end_val = end_val, start_val
                    if start_val <= target <= end_val:
                        return True
        return False

    client = get_rentman_client()
    if client:
        rentman_info["available"] = True
        scan_limit = _coerce_int(os.environ.get("JOBLOG_RENTMAN_SUBPROJECT_SCAN_LIMIT")) or 15000
        scan_limit = max(1000, min(scan_limit, 30000))

        # 1) Recupera status map
        status_map: Dict[int, str] = {}
        try:
            for entry in client.get_project_statuses():
                status_id = entry.get("id")
                if isinstance(status_id, int):
                    status_name = entry.get("displayname") or entry.get("name") or str(status_id)
                    status_map[status_id] = str(status_name)
        except Exception:
            status_map = {}

        # 2) Recupera tutti i subprojects
        subprojects_all: List[Dict[str, Any]] = []
        rentman_error_type: Optional[str] = None
        rentman_error_detail: Optional[str] = None
        try:
            for sub in client.iter_collection("/subprojects", limit_total=scan_limit):
                subprojects_all.append(sub)
        except Exception as exc:
            app.logger.exception("Rentman: errore durante scansione subprojects")
            rentman_error_type = type(exc).__name__
            rentman_error_detail = str(exc)

        if rentman_error_type is None:
            target_date_obj = _parse_date_any(target_date)
            if target_date_obj is None:
                target_date_obj = datetime.now().date()

            # 3) Filtra subprojects per status e data
            valid_subprojects: List[Dict[str, Any]] = []
            projects_seen: Set[str] = set()

            for sub in subprojects_all:
                status_ref = sub.get("status")
                status_id = parse_reference(status_ref)
                status_name = status_map.get(status_id, "") if isinstance(status_id, int) else ""

                if not is_status_allowed(status_name):
                    continue

                if not subproject_matches_date(sub, target_date_obj):
                    continue

                project_ref = sub.get("project")
                if not project_ref:
                    continue

                valid_subprojects.append({
                    "subproject_id": sub.get("id"),
                    "project_ref": project_ref,
                    "status_id": status_id,
                    "status_name": status_name,
                })
                projects_seen.add(project_ref)

            # 4) Recupera dettagli progetti padre
            projects_rentman: List[Dict[str, Any]] = []
            seen_codes: Set[str] = set()

            for project_ref in projects_seen:
                project_id = parse_reference(project_ref)
                if project_id is None:
                    continue

                try:
                    payload = client._request("GET", f"/projects/{project_id}")
                    project_data = payload.get("data", {})

                    code = str(project_data.get("number") or project_data.get("reference") or project_id).strip().upper()
                    if not code or code in seen_codes:
                        continue
                    seen_codes.add(code)

                    name = project_data.get("displayname") or project_data.get("name") or code
                    projects_rentman.append({"code": code, "name": str(name)})
                except Exception:
                    continue

            rentman_info["used"] = True
            projects_rentman.sort(key=lambda item: (str(item.get("code") or ""), str(item.get("name") or "")))

            response: Dict[str, Any] = {
                "ok": True,
                "date": target_date,
                "projects": projects_rentman,
                "count": len(projects_rentman),
                "source": "rentman",
            }
            if debug:
                response["debug"] = {
                    "scan_limit": scan_limit,
                    "subprojects_total": len(subprojects_all),
                    "subprojects_valid": len(valid_subprojects),
                    "status_map_count": len(status_map),
                    "status_map": status_map,
                }
            return jsonify(response)

        if debug:
            if rentman_error_type:
                rentman_info["error"] = rentman_error_type
            if rentman_error_detail:
                rentman_info["error_detail"] = rentman_error_detail[:1200]
    else:
        rentman_info["available"] = False
        rentman_info["reason"] = "client_unavailable"

    # Fallback locale: projects.json (solo se Rentman non è disponibile o ha fallito)
    catalog = load_external_projects()
    projects: List[Dict[str, Any]] = []
    for entry in catalog.values():
        code = str(entry.get("project_code") or "").strip().upper()
        name = str(entry.get("project_name") or code).strip() or code
        activities_any = entry.get("activities")
        activities: List[Any] = activities_any if isinstance(activities_any, list) else []
        active = False
        for act in activities:
            if not isinstance(act, dict):
                continue
            if _activity_matches_date(act.get("plan_start"), act.get("plan_end"), target_date):
                active = True
                break
        if active:
            projects.append({"code": code, "name": name})

    projects.sort(key=lambda item: (str(item.get("code") or ""), str(item.get("name") or "")))
    response_local: Dict[str, Any] = {"ok": True, "date": target_date, "projects": projects, "count": len(projects), "source": "local"}
    if debug:
        response_local["debug"] = {"rentman": rentman_info}
    return jsonify(response_local)


@app.get("/api/magazzino/projects/manual")
@login_required
def api_magazzino_projects_manual_list() -> ResponseReturnValue:
    """Restituisce i progetti manuali salvati per l'utente (persistenti cross-device)."""
    guard = _magazzino_only()
    if guard is not None:
        return guard

    db = get_db()
    ensure_warehouse_manual_projects_table(db)
    username = session.get("user")
    rows = db.execute(
        """
        SELECT project_code AS code, name, created_ts
        FROM warehouse_manual_projects
        WHERE username = ?
        ORDER BY created_ts DESC
        LIMIT 500
        """,
        (username,),
    ).fetchall()
    items = [dict(row) for row in rows] if rows else []
    return jsonify({"ok": True, "items": items})


@app.post("/api/magazzino/projects/manual")
@login_required
def api_magazzino_projects_manual_upsert() -> ResponseReturnValue:
    """Salva o aggiorna un progetto manuale per l'utente corrente."""
    guard = _magazzino_only()
    if guard is not None:
        return guard

    data = request.get_json(silent=True) or {}
    code = _normalize_text(data.get("code"))
    name = _normalize_text(data.get("name"))
    if not code:
        return jsonify({"ok": False, "error": "missing_code"}), 400

    db = get_db()
    ensure_warehouse_manual_projects_table(db)
    username = session.get("user")
    now = now_ms()

    if DB_VENDOR == "mysql":
        db.execute(
            """
            INSERT INTO warehouse_manual_projects(project_code, name, username, created_ts)
            VALUES(%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE name=VALUES(name), created_ts=VALUES(created_ts)
            """,
            (code, name, username, now),
        )
    else:
        db.execute(
            """
            INSERT INTO warehouse_manual_projects(project_code, name, username, created_ts)
            VALUES(?,?,?,?)
            ON CONFLICT(username, project_code) DO UPDATE SET name=excluded.name, created_ts=excluded.created_ts
            """,
            (code, name, username, now),
        )
    try:
        db.commit()
    except Exception:
        pass
    return jsonify({"ok": True, "code": code, "name": name, "created_ts": now})


@app.get("/api/magazzino/projects/lookup")
@login_required
def api_magazzino_projects_lookup() -> ResponseReturnValue:
    """Recupera un progetto per codice (numero) anche fuori data odierna.

    Prova prima Rentman (se configurato), altrimenti il catalogo locale.
    """
    guard = _magazzino_only()
    if guard is not None:
        return guard

    code = _normalize_text(request.args.get("code")).upper()
    if not code:
        return jsonify({"ok": False, "error": "missing_code"}), 400

    project: Optional[Dict[str, str]] = None
    source = "local"
    debug = _is_truthy(request.args.get("debug"))
    rentman_debug: Dict[str, Any] = {}

    client = get_rentman_client()
    if client:
        try:
            # Riduci batch per evitare limiti dimensione risposta Rentman
            batch = 150  # < 300 e payload più piccolo
            max_total = 7500
            scanned = 0
            found = False
            offset = 0

            def norm(val: str) -> str:
                return str(val or "").strip().upper()

            while scanned < max_total:
                payload = client._request(
                    "GET",
                    "/projects",
                    params={
                        "limit": batch,
                        "offset": offset,
                        "fields": "number,reference,displayname,name",
                    },
                )
                data = payload.get("data") if isinstance(payload, dict) else None
                if debug:
                    rentman_debug.setdefault("pages", []).append(len(data) if isinstance(data, list) else None)
                if not isinstance(data, list) or not data:
                    break

                for entry in data:
                    num = norm(entry.get("number") or entry.get("reference"))
                    disp = norm(entry.get("displayname") or entry.get("name"))
                    if num == code or disp == code or code in num:
                        name = entry.get("displayname") or entry.get("name") or num
                        project = {"code": num or code, "name": str(name)}
                        source = "rentman"
                        found = True
                        break

                scanned += len(data)
                if found:
                    break
                if len(data) < batch:
                    break
                offset += batch

            if debug:
                rentman_debug["scanned"] = scanned
                rentman_debug["found"] = bool(project)

        except Exception as exc:
            project = None
            if debug:
                rentman_debug["error"] = True
                rentman_debug["error_type"] = type(exc).__name__
                rentman_debug["error_detail"] = str(exc)[:400]

    if project is None:
        catalog = load_external_projects()
        entry = catalog.get(code)
        if entry:
            name = str(entry.get("project_name") or code).strip() or code
            project = {"code": code, "name": name}
            source = "local"

    if project is None:
        resp: Dict[str, Any] = {"ok": False, "error": "not_found"}
        if debug:
            resp["rentman_debug"] = rentman_debug
        return jsonify(resp), 404

    resp: Dict[str, Any] = {"ok": True, "project": project, "source": source}
    if debug:
        resp["rentman_debug"] = rentman_debug
    return jsonify(resp)


@app.get("/api/magazzino/activities")
@login_required
def api_magazzino_activities_list() -> ResponseReturnValue:
    guard = _magazzino_only()
    if guard is not None:
        return guard

    project_code = _normalize_text(request.args.get("project_code")).upper()
    if not project_code:
        return jsonify({"ok": False, "error": "missing_project_code"}), 400

    db = get_db()
    ensure_warehouse_activities_table(db)
    rows = db.execute(
        """
        SELECT id, project_code, activity_label, note, username, created_ts
        FROM warehouse_activities
        WHERE project_code = ?
        ORDER BY created_ts DESC
        LIMIT 200
        """,
        (project_code,),
    ).fetchall()
    items = [dict(row) for row in rows] if rows else []
    return jsonify({"ok": True, "project_code": project_code, "items": items})


@app.post("/api/magazzino/activities")
@login_required
def api_magazzino_activities_create() -> ResponseReturnValue:
    guard = _magazzino_only()
    if guard is not None:
        return guard

    data = request.get_json(silent=True) or {}
    project_code = _normalize_text(data.get("project_code")).upper()
    activity_label = _normalize_text(data.get("activity_label"))
    note = _normalize_text(data.get("note"))
    if not project_code:
        return jsonify({"ok": False, "error": "missing_project_code"}), 400
    if not activity_label:
        return jsonify({"ok": False, "error": "missing_activity_label"}), 400

    db = get_db()
    ensure_warehouse_activities_table(db)
    now = now_ms()
    username = session.get("user")
    db.execute(
        """
        INSERT INTO warehouse_activities(project_code, activity_label, note, username, created_ts)
        VALUES(?,?,?,?,?)
        """,
        (project_code, activity_label, note or None, username, now),
    )
    try:
        db.commit()
    except Exception:
        pass
    return jsonify({"ok": True, "created_ts": now})


# ============== Sessioni Magazzino (Timer) ==============

@app.get("/api/magazzino/sessions")
@login_required
def api_magazzino_sessions_list() -> ResponseReturnValue:
    """Lista sessioni di lavoro magazzino di oggi (solo utente corrente)."""
    guard = _magazzino_only()
    if guard is not None:
        return guard

    import json as json_module
    
    project_code = _normalize_text(request.args.get("project_code")).upper()
    current_user = session.get("user")

    db = get_db()
    ensure_warehouse_sessions_table(db)

    # Filtra sessioni di oggi per l'utente corrente
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_start_ms = int(today_start.timestamp() * 1000)

    # Filtra sempre per utente corrente, opzionalmente anche per progetto
    if project_code:
        rows = db.execute(
            """
            SELECT id, project_code, activity_label, elapsed_ms, start_ts, end_ts, intervals, note, username, created_ts
            FROM warehouse_sessions
            WHERE project_code = ? AND username = ? AND created_ts >= ?
            ORDER BY created_ts DESC
            LIMIT 100
            """,
            (project_code, current_user, today_start_ms),
        ).fetchall()
    else:
        rows = db.execute(
            """
            SELECT id, project_code, activity_label, elapsed_ms, start_ts, end_ts, intervals, note, username, created_ts
            FROM warehouse_sessions
            WHERE username = ? AND created_ts >= ?
            ORDER BY created_ts DESC
            LIMIT 100
            """,
            (current_user, today_start_ms),
        ).fetchall()
    
    items = []
    for row in (rows or []):
        item = dict(row)
        # Parsa intervals da JSON
        if item.get("intervals"):
            try:
                item["intervals"] = json_module.loads(item["intervals"])
            except:
                item["intervals"] = []
        else:
            # Se non ci sono intervals ma c'è start/end, crea un array con un singolo intervallo
            if item.get("start_ts") and item.get("end_ts"):
                item["intervals"] = [{"start": item["start_ts"], "end": item["end_ts"]}]
            else:
                item["intervals"] = []
        items.append(item)
    
    return jsonify({"ok": True, "items": items})


@app.post("/api/magazzino/sessions")
@login_required
def api_magazzino_sessions_create() -> ResponseReturnValue:
    """Salva una nuova sessione di lavoro magazzino."""
    guard = _magazzino_only()
    if guard is not None:
        return guard

    data = request.get_json(silent=True) or {}
    project_code = _normalize_text(data.get("project_code")).upper()
    activity_label = _normalize_text(data.get("activity_label"))
    elapsed_ms = _coerce_int(data.get("elapsed_ms"))
    start_ts = _coerce_int(data.get("start_ts"))
    end_ts = _coerce_int(data.get("end_ts"))
    note = _normalize_text(data.get("note"))

    if not project_code:
        return jsonify({"ok": False, "error": "missing_project_code"}), 400
    if not activity_label:
        return jsonify({"ok": False, "error": "missing_activity_label"}), 400
    if elapsed_ms is None or elapsed_ms < 0:
        elapsed_ms = 0

    db = get_db()
    ensure_warehouse_sessions_table(db)
    now = now_ms()
    username = session.get("user")

    # Salva il primo intervallo
    import json as json_module
    intervals = json_module.dumps([{"start": start_ts, "end": end_ts}]) if start_ts and end_ts else None

    db.execute(
        """
        INSERT INTO warehouse_sessions(project_code, activity_label, elapsed_ms, start_ts, end_ts, intervals, note, username, created_ts)
        VALUES(?,?,?,?,?,?,?,?,?)
        """,
        (project_code, activity_label, elapsed_ms, start_ts, end_ts, intervals, note or None, username, now),
    )
    try:
        db.commit()
    except Exception:
        pass
    return jsonify({"ok": True, "created_ts": now, "elapsed_ms": elapsed_ms, "start_ts": start_ts, "end_ts": end_ts})


@app.put("/api/magazzino/sessions/<int:session_id>")
@login_required
def api_magazzino_sessions_update(session_id: int) -> ResponseReturnValue:
    """Aggiorna una sessione esistente (aggiunge tempo o modifica orari)."""
    guard = _magazzino_only()
    if guard is not None:
        return guard

    import json as json_module
    
    data = request.get_json(silent=True) or {}
    add_elapsed_ms = _coerce_int(data.get("add_elapsed_ms"))  # Tempo da aggiungere
    new_start_ts = _coerce_int(data.get("start_ts"))  # Inizio nuovo intervallo
    new_end_ts = _coerce_int(data.get("end_ts"))  # Fine nuovo intervallo
    interval_start = _coerce_int(data.get("interval_start"))  # Inizio del nuovo intervallo di lavoro

    db = get_db()
    ensure_warehouse_sessions_table(db)
    username = session.get("user")

    # Verifica che la sessione esista e appartenga all'utente
    row = db.execute(
        "SELECT id, elapsed_ms, start_ts, end_ts, intervals FROM warehouse_sessions WHERE id = ? AND username = ?",
        (session_id, username)
    ).fetchone()
    
    if not row:
        return jsonify({"ok": False, "error": "session_not_found"}), 404

    current_elapsed = row["elapsed_ms"] or 0
    current_start = row["start_ts"]
    current_intervals = []
    
    # Carica intervalli esistenti
    if row["intervals"]:
        try:
            current_intervals = json_module.loads(row["intervals"])
        except:
            current_intervals = []
    
    # Se non ci sono intervalli ma c'è start_ts/end_ts, crea il primo
    if not current_intervals and current_start and row["end_ts"]:
        current_intervals = [{"start": current_start, "end": row["end_ts"]}]

    # Calcola nuovi valori
    new_elapsed = current_elapsed
    if add_elapsed_ms and add_elapsed_ms > 0:
        new_elapsed = current_elapsed + add_elapsed_ms
        # Aggiungi nuovo intervallo
        if interval_start and new_end_ts:
            current_intervals.append({"start": interval_start, "end": new_end_ts})
        elif new_end_ts:
            # Usa end_ts - elapsed come inizio approssimativo
            current_intervals.append({"start": new_end_ts - add_elapsed_ms, "end": new_end_ts})

    # Se vengono passati nuovi orari senza add_elapsed (modifica manuale)
    final_start = current_start  # Mantieni l'inizio originale
    final_end = new_end_ts if new_end_ts else row["end_ts"]
    
    # Se modifica manuale degli orari (senza add_elapsed)
    if not add_elapsed_ms and (new_start_ts is not None or new_end_ts is not None):
        final_start = new_start_ts if new_start_ts is not None else current_start
        final_end = new_end_ts if new_end_ts is not None else row["end_ts"]

    intervals_json = json_module.dumps(current_intervals) if current_intervals else None

    db.execute(
        """
        UPDATE warehouse_sessions 
        SET elapsed_ms = ?, start_ts = ?, end_ts = ?, intervals = ?
        WHERE id = ? AND username = ?
        """,
        (new_elapsed, final_start, final_end, intervals_json, session_id, username)
    )
    try:
        db.commit()
    except Exception:
        pass

    return jsonify({
        "ok": True, 
        "id": session_id,
        "elapsed_ms": new_elapsed, 
        "start_ts": final_start, 
        "end_ts": final_end,
        "intervals": current_intervals
    })


@app.route("/api/activities", methods=["GET", "POST"])
@login_required
def api_activities():
    if request.method == "GET":
        db = get_db()
        rows = db.execute(
            "SELECT activity_id, label FROM activities ORDER BY sort_order, label"
        ).fetchall()
        return jsonify({"activities": [dict(row) for row in rows]})

    # POST - create new activity
    data = request.get_json(silent=True) or {}
    label = str(data.get("label") or "").strip()
    if not label:
        return jsonify({"ok": False, "error": "missing_label"}), 400

    raw_id = data.get("activity_id")
    requested_id = ""
    if isinstance(raw_id, str) and raw_id.strip():
        requested_id = _normalize_activity_id(raw_id)
        if not requested_id:
            return jsonify({"ok": False, "error": "invalid_activity_id"}), 400

    db = get_db()
    project_code = session.get('supervisor_project_code', '')
    if not project_code:
        return jsonify({"ok": False, "error": "no_active_project"}), 409
    project_name = session.get('supervisor_project_name') or project_code
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"

    if requested_id:
        existing = db.execute(
            f"SELECT 1 FROM activities WHERE activity_id={placeholder} AND project_code={placeholder}",
            (requested_id, project_code),
        ).fetchone()
        if existing is not None:
            return jsonify({"ok": False, "error": "activity_id_in_use"}), 409
        activity_id = requested_id
    else:
        activity_id = _generate_activity_id(db, label)

    order_row = db.execute(f"SELECT COALESCE(MAX(sort_order), 0) AS max_order FROM activities WHERE project_code={placeholder}", (project_code,)).fetchone()
    max_order = 0
    if order_row is not None:
        value = row_value(order_row, "max_order")
        if value is None:
            value = row_value(order_row, 0)
        if isinstance(value, (int, float)):
            max_order = int(value)
    sort_order = max_order + 1

    plan_start = _normalize_datetime(data.get("plan_start"))
    plan_end = _normalize_datetime(data.get("plan_end"))
    planned_members = _coerce_int(data.get("planned_members"))
    if planned_members is None or planned_members < 0:
        planned_members = 0
    notes_raw = data.get("notes")
    notes = str(notes_raw).strip() if isinstance(notes_raw, str) and notes_raw.strip() else None
    planned_duration_ms = compute_planned_duration_ms(plan_start, plan_end, planned_members)

    db.execute(
        f"""
        INSERT INTO activities(
            activity_id, project_code, label, sort_order, plan_start, plan_end,
            planned_members, planned_duration_ms, notes
        ) VALUES({placeholder},{placeholder},{placeholder},{placeholder},{placeholder},{placeholder},{placeholder},{placeholder},{placeholder})
        """,
        (
            activity_id,
            project_code,
            label,
            sort_order,
            plan_start,
            plan_end,
            planned_members,
            planned_duration_ms,
            notes,
        ),
    )

    meta = load_activity_meta(db)
    meta[str(activity_id)] = {
        "plan_start": plan_start,
        "plan_end": plan_end,
        "planned_members": planned_members,
        "planned_duration_ms": planned_duration_ms,
        "actual_runtime_ms": 0,
    }
    save_activity_meta(db, meta)

    now = now_ms()
    db.execute(
        f"INSERT INTO event_log(ts, kind, details, project_code) VALUES({placeholder},{placeholder},{placeholder},{placeholder})",
        (
            now,
            "create_activity",
            json.dumps({"activity_id": activity_id, "label": label}),
            project_code,
        ),
    )

    db.commit()

    payload = {
        "activity_id": activity_id,
        "label": label,
        "plan_start": plan_start,
        "plan_end": plan_end,
        "planned_members": planned_members,
        "planned_duration_ms": planned_duration_ms,
        "notes": notes,
        "sort_order": sort_order,
    }

    return (
        jsonify(
            {
                "ok": True,
                "activity": payload,
                "project": {"code": project_code, "name": project_name},
            }
        ),
        201,
    )


@app.get("/api/push/status")
@login_required
def api_push_status():
    settings = get_webpush_settings()
    if not settings:
        return jsonify({"enabled": False})

    db = get_db()
    username = session.get("user")
    subscribed = False
    if username:
        row = db.execute(
            "SELECT 1 FROM push_subscriptions WHERE username=? LIMIT 1",
            (username,),
        ).fetchone()
        subscribed = row is not None

    return jsonify(
        {
            "enabled": True,
            "publicKey": settings["vapid_public"],
            "subscribed": subscribed,
        }
    )


@app.post("/api/push/subscribe")
@login_required
def api_push_subscribe():
    settings = get_webpush_settings()
    if not settings:
        return jsonify({"ok": False, "error": "push_not_configured"}), 400

    data = request.get_json(silent=True) or {}
    endpoint_raw = data.get("endpoint")
    if not isinstance(endpoint_raw, str):
        return jsonify({"ok": False, "error": "invalid_subscription"}), 400
    endpoint = endpoint_raw.strip()

    keys = data.get("keys") or {}
    p256dh_raw = keys.get("p256dh")
    auth_raw = keys.get("auth")
    if not isinstance(p256dh_raw, str) or not isinstance(auth_raw, str):
        return jsonify({"ok": False, "error": "invalid_subscription"}), 400
    p256dh = p256dh_raw.strip()
    auth_key = auth_raw.strip()

    if not endpoint or not p256dh or not auth_key:
        return jsonify({"ok": False, "error": "invalid_subscription"}), 400

    encoding_raw = data.get("contentEncoding") or data.get("content_encoding")
    encoding = (
        encoding_raw.strip()
        if isinstance(encoding_raw, str) and encoding_raw.strip()
        else "aes128gcm"
    )

    expiration_raw = data.get("expirationTime")
    expiration_time: Optional[int]
    if isinstance(expiration_raw, (int, float)):
        expiration_time = int(expiration_raw)
    elif isinstance(expiration_raw, str) and expiration_raw.isdigit():
        expiration_time = int(expiration_raw)
    else:
        expiration_time = None

    user_agent = data.get("userAgent")
    if not isinstance(user_agent, str) or not user_agent.strip():
        user_agent = request.headers.get("User-Agent")
    if isinstance(user_agent, str):
        user_agent = user_agent[:255]

    username = session.get("user") or "anonymous"
    now = now_ms()

    db = get_db()
    params = (
        username,
        endpoint,
        p256dh,
        auth_key,
        encoding,
        user_agent,
        expiration_time,
        now,
        now,
    )

    if DB_VENDOR == "mysql":
        db.execute(
            """
            INSERT INTO push_subscriptions(
                username, endpoint, p256dh, auth, content_encoding, user_agent, expiration_time, created_ts, updated_ts
            ) VALUES(?,?,?,?,?,?,?,?,?)
            ON DUPLICATE KEY UPDATE
                username=VALUES(username),
                p256dh=VALUES(p256dh),
                auth=VALUES(auth),
                content_encoding=VALUES(content_encoding),
                user_agent=VALUES(user_agent),
                expiration_time=VALUES(expiration_time),
                updated_ts=VALUES(updated_ts)
            """,
            params,
        )
    else:
        db.execute(
            """
            INSERT INTO push_subscriptions(
                username, endpoint, p256dh, auth, content_encoding, user_agent, expiration_time, created_ts, updated_ts
            ) VALUES(?,?,?,?,?,?,?,?,?)
            ON CONFLICT(endpoint) DO UPDATE SET
                username=excluded.username,
                p256dh=excluded.p256dh,
                auth=excluded.auth,
                content_encoding=excluded.content_encoding,
                user_agent=excluded.user_agent,
                expiration_time=excluded.expiration_time,
                updated_ts=excluded.updated_ts
            """,
            params,
        )

    db.commit()
    return jsonify({"ok": True})


@app.post("/api/push/test")
@login_required
def api_push_test():
    settings = get_webpush_settings()
    if not settings:
        return jsonify({"ok": False, "error": "push_not_configured"}), 400

    db = get_db()
    username = session.get("user")
    if not username:
        return jsonify({"ok": False, "error": "missing_user"}), 400

    rows = db.execute(
        "SELECT endpoint, p256dh, auth, content_encoding FROM push_subscriptions WHERE username=?",
        (username,),
    ).fetchall()

    if not rows:
        return jsonify({"ok": False, "error": "no_subscription"}), 404

    subscriptions = [dict(row) for row in rows]
    invalid_endpoints = set()
    delivered = 0

    payload = {
        "title": "JobLog",
        "body": "Notifica di prova inviata con successo",
        "data": {
            "notification_type": "test_message",
            "issued_at": datetime.now(timezone.utc).isoformat(),
        },
    }

    for sub in subscriptions:
        endpoint = sub.get("endpoint") or ""
        if not endpoint or endpoint in invalid_endpoints:
            continue
        key_p256dh = sub.get("p256dh")
        key_auth = sub.get("auth")
        if not key_p256dh or not key_auth:
            invalid_endpoints.add(endpoint)
            continue

        subscription_info = {
            "endpoint": endpoint,
            "keys": {
                "p256dh": key_p256dh,
                "auth": key_auth,
            },
        }
        encoding = sub.get("content_encoding") or "aes128gcm"

        try:
            webpush(
                subscription_info=subscription_info,
                data=json.dumps(payload),
                vapid_private_key=settings["vapid_private"],
                vapid_claims={"sub": settings["subject"]},
                ttl=60,
                content_encoding=encoding,
            )
            delivered += 1
            record_push_notification(
                db,
                kind="test_message",
                title=payload.get("title", "JobLog"),
                body=payload.get("body"),
                payload=payload,
                activity_id=None,
                username=username,
            )
        except WebPushException as exc:
            status = getattr(exc.response, "status_code", None)
            app.logger.warning("WebPush test fallita (%s): %s", status, exc)
            if status in (404, 410):
                invalid_endpoints.add(endpoint)
        except Exception as exc:  # pragma: no cover - logging best effort
            app.logger.exception("Errore generico nell'invio della notifica di prova", exc_info=exc)

    if invalid_endpoints:
        for endpoint in invalid_endpoints:
            remove_push_subscription(db, endpoint)

    db.commit()
    return jsonify({"ok": True, "delivered": delivered, "invalid": list(invalid_endpoints)})


@app.post("/api/push/unsubscribe")
@login_required
def api_push_unsubscribe():
    data = request.get_json(silent=True) or {}
    endpoint = str(data.get("endpoint") or "").strip()
    if not endpoint:
        return jsonify({"ok": False, "error": "invalid_endpoint"}), 400

    db = get_db()
    username = session.get("user")
    if username:
        db.execute(
            "DELETE FROM push_subscriptions WHERE endpoint=? AND username=?",
            (endpoint, username),
        )
    else:
        db.execute("DELETE FROM push_subscriptions WHERE endpoint=?", (endpoint,))

    db.commit()
    return jsonify({"ok": True})


@app.get("/api/push/notifications")
@login_required
def api_push_notifications():
    username = session.get("user")
    if not username:
        return jsonify({"items": []})

    limit_arg = request.args.get("limit", default="20")
    parsed_limit: Optional[int]
    if isinstance(limit_arg, str) and limit_arg.strip().lower() in {"all", "tutti"}:
        parsed_limit = None
    else:
        try:
            parsed_limit = int(limit_arg)
        except (TypeError, ValueError):
            parsed_limit = 20
        else:
            if parsed_limit <= 0:
                parsed_limit = None

    db = get_db()
    ensure_push_notification_read_column(db)
    items = fetch_recent_push_notifications(db, username=username, limit=parsed_limit)
    return jsonify({"items": items})


@app.post("/api/push/notifications/<int:notification_id>/read")
@login_required
def api_push_notification_mark_read(notification_id: int) -> ResponseReturnValue:
    """Marca una notifica come letta."""
    username = session.get("user")
    if not username:
        return jsonify({"error": "Non autenticato"}), 401
    
    db = get_db()
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Verifica che la notifica appartenga all'utente
    row = db.execute(
        f"SELECT id FROM push_notification_log WHERE id = {placeholder} AND username = {placeholder}",
        (notification_id, username)
    ).fetchone()
    
    if not row:
        return jsonify({"error": "Notifica non trovata"}), 404
    
    # Marca come letta
    db.execute(
        f"UPDATE push_notification_log SET read_at = {placeholder} WHERE id = {placeholder}",
        (now_ms(), notification_id)
    )
    db.commit()
    
    return jsonify({"ok": True, "read_at": now_ms()})


@app.post("/api/push/notifications/read-all")
@login_required
def api_push_notifications_mark_all_read() -> ResponseReturnValue:
    """Marca tutte le notifiche dell'utente come lette."""
    username = session.get("user")
    if not username:
        return jsonify({"error": "Non autenticato"}), 401
    
    db = get_db()
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Marca tutte come lette
    db.execute(
        f"UPDATE push_notification_log SET read_at = {placeholder} WHERE username = {placeholder} AND read_at IS NULL",
        (now_ms(), username)
    )
    db.commit()
    
    return jsonify({"ok": True})


@app.get("/api/state")
@login_required
def api_state():
    db = get_db()
    now = now_ms()
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"

    # Ogni supervisor vede il proprio progetto (dalla sessione)
    project_code = session.get('supervisor_project_code')
    project_name = session.get('supervisor_project_name') or project_code

    if not project_code:
        return jsonify(
            {
                "team": [],
                "activities": [],
                "allPaused": True,
                "timestamp": now,
                "project": None,
            }
        )

    activity_rows = db.execute(
        f"SELECT activity_id, label FROM activities WHERE project_code = {placeholder} ORDER BY sort_order, label",
        (project_code,)
    ).fetchall()

    activity_meta = load_activity_meta(db)

    activity_map: Dict[str, Dict[str, Any]] = {}
    for row in activity_rows:
        activity_id = row["activity_id"]
        activity_key = str(activity_id)
        meta_entry = activity_meta.get(activity_key)
        if not isinstance(meta_entry, dict):
            meta_entry = {}
            activity_meta[activity_key] = meta_entry

        activity_map[activity_id] = {
            "activity_id": activity_id,
            "label": row["label"],
            "members": [],
            "plan_start": meta_entry.get("plan_start"),
            "plan_end": meta_entry.get("plan_end"),
            "planned_members": meta_entry.get("planned_members"),
            "planned_duration_ms": meta_entry.get("planned_duration_ms"),
            "actual_runtime_ms": meta_entry.get("actual_runtime_ms", 0),
        }

    members = db.execute(
        f"SELECT * FROM member_state WHERE project_code = {placeholder} ORDER BY member_name",
        (project_code,)
    ).fetchall()

    team: List[Dict[str, Any]] = []
    active_members: List[Dict[str, Any]] = []
    paused_keys = {
        row["member_key"]
        for row in db.execute(
            f"SELECT member_key FROM member_state WHERE project_code = {placeholder} AND running={placeholder} AND pause_start IS NOT NULL",
            (project_code, RUN_STATE_PAUSED,)
        ).fetchall()
    }
    for row in members:
        running_state = int(row["running"])
        last_start_ts = row["entered_ts"] or row["start_ts"]
        member = {
            "member_key": row["member_key"],
            "member_name": row["member_name"],
            "activity_id": row["activity_id"],
            "running": running_state == RUN_STATE_RUNNING,
            "running_state": running_state,
            "elapsed": compute_elapsed(row, now),
            "paused": row["member_key"] in paused_keys,
            "last_start_ts": last_start_ts,
        }
        if row["activity_id"] and row["activity_id"] in activity_map:
            activity_map[row["activity_id"]]["members"].append(member)
            active_members.append(member)
        else:
            team.append(member)

    meta_dirty = False
    for activity_id, activity in activity_map.items():
        activity_key = str(activity_id)
        activity["members"].sort(key=lambda m: m["member_name"])
        meta_entry = activity_meta.get(activity_key)
        if not isinstance(meta_entry, dict):
            meta_entry = {}
            activity_meta[activity_key] = meta_entry
        stored_value = _coerce_int(meta_entry.get("planned_members"))
        activity["planned_members"] = stored_value

        plan_start_meta = meta_entry.get("plan_start")
        plan_end_meta = meta_entry.get("plan_end")
        planned_duration_ms = _coerce_int(meta_entry.get("planned_duration_ms"))
        if planned_duration_ms is None:
            computed_duration = compute_planned_duration_ms(
                plan_start_meta,
                plan_end_meta,
                stored_value,
            )
            if computed_duration is not None:
                planned_duration_ms = computed_duration
                meta_entry["planned_duration_ms"] = computed_duration
                meta_dirty = True
        activity["planned_duration_ms"] = planned_duration_ms

    if meta_dirty:
        save_activity_meta(db, activity_meta)

    all_paused = not any(m["running"] for m in team + active_members)
    
    # Verifica se il progetto del supervisor è sincronizzato con quello globale
    project_info = {
        "code": project_code,
        "name": project_name or project_code,
    }

    return jsonify(
        {
            "team": team,
            "activities": list(activity_map.values()),
            "allPaused": all_paused,
            "timestamp": now,
            "project": project_info,
        }
    )


@app.get("/api/events")
@login_required
def api_events():
    db = get_db()
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    project_code = session.get('supervisor_project_code')
    if not project_code:
        return jsonify({"events": []})

    activity_labels = {
        row["activity_id"]: row["label"]
        for row in db.execute(
            f"SELECT activity_id, label FROM activities WHERE project_code = {placeholder}",
            (project_code,)
        )
    }

    rows = db.execute(
        f"SELECT id, ts, kind, member_key, details FROM event_log WHERE project_code = {placeholder} ORDER BY ts DESC LIMIT 25",
        (project_code,)
    ).fetchall()

    events: List[Dict[str, Any]] = []
    for row in rows:
        details_raw = row["details"]
        details: Dict[str, Any]
        if details_raw:
            try:
                details = json.loads(details_raw)
            except json.JSONDecodeError:
                details = {}
        else:
            details = {}

        summary = describe_event(row["kind"], details, activity_labels)
        events.append(
            {
                "id": row["id"],
                "timestamp": row["ts"],
                "kind": row["kind"],
                "summary": summary,
            }
        )

    return jsonify({"events": events})


@app.get("/api/project/attachments")
@login_required
def api_project_attachments():
    db = get_db()
    project_code = get_app_state(db, "project_code")
    if not project_code:
        return jsonify({"project": None, "attachments": []})

    project_name = get_app_state(db, "project_name") or project_code
    exhaustive = request.args.get("mode") == "deep"
    attachments = fetch_project_attachments(project_code, exhaustive=exhaustive)
    return jsonify({"project": {"code": project_code, "name": project_name}, "attachments": attachments})


@app.get("/api/project/materials")
@login_required
def api_project_materials():
    db = get_db()
    project_code = get_app_state(db, "project_code")
    if not project_code:
        return jsonify({"project": None, "materials": [], "folders": [], "equipment_checks": {}, "from_cache": False})

    project_name = get_app_state(db, "project_name") or project_code
    mode = (request.args.get("mode") or "").strip().lower()
    refresh_requested = mode == "refresh"

    cached_payload: Optional[Dict[str, Any]] = None
    if not refresh_requested:
        cached_payload = load_project_materials_cache(db, project_code)

    if cached_payload:
        equipment_checks = fetch_equipment_checks(db, project_code)
        return jsonify(
            {
                "project": {"code": project_code, "name": project_name},
                "materials": cached_payload.get("items", []),
                "folders": cached_payload.get("folders", []),
                "equipment_checks": equipment_checks,
                "from_cache": True,
                "updated_ts": cached_payload.get("updated_ts"),
            }
        )

    materials_payload = fetch_project_materials(project_code)
    items = materials_payload.get("items", [])
    folders = materials_payload.get("folders", [])
    saved_ts = save_project_materials_cache(
        db,
        project_code,
        project_name,
        items=items,
        folders=folders,
    )
    try:
        db.commit()
    except Exception:
        app.logger.exception("Impossibile salvare la cache materiali per il progetto %s", project_code)
    equipment_checks = fetch_equipment_checks(db, project_code)
    return jsonify(
        {
            "project": {"code": project_code, "name": project_name},
            "materials": items,
            "folders": folders,
            "equipment_checks": equipment_checks,
            "from_cache": False,
            "updated_ts": saved_ts,
        }
    )


@app.get("/api/project/equipment/checks")
@login_required
def api_project_equipment_checks():
    db = get_db()
    project_code = get_app_state(db, "project_code")
    if not project_code:
        return jsonify({"project": None, "checks": {}})

    project_name = get_app_state(db, "project_name") or project_code
    checks = fetch_equipment_checks(db, project_code)
    return jsonify({"project": {"code": project_code, "name": project_name}, "checks": checks})


@app.post("/api/project/equipment/checks")
@login_required
def api_update_equipment_check():
    data = request.get_json(silent=True) or {}
    item_key = (data.get("item_key") or "").strip()
    if not item_key:
        return jsonify({"ok": False, "error": "missing_item_key"}), 400
    if "checked" not in data:
        return jsonify({"ok": False, "error": "missing_checked_flag"}), 400

    checked = bool(data.get("checked"))

    db = get_db()
    project_code = get_app_state(db, "project_code")
    if not project_code:
        return jsonify({"ok": False, "error": "no_active_project"}), 400

    username = session.get("user")
    timestamp = persist_equipment_check(
        db,
        project_code=project_code,
        item_key=item_key,
        checked=checked,
        username=username,
    )
    db.commit()

    return jsonify({"ok": True, "checked": checked, "timestamp": timestamp})


@app.route("/api/project/local-equipment", methods=["GET", "POST"])
@login_required
def api_local_equipment():
    db = get_db()
    project_code = get_app_state(db, "project_code")
    if not project_code:
        return jsonify({"ok": False, "error": "no_active_project"}), 400

    ensure_local_equipment_table(db)

    if request.method == "GET":
        rows = db.execute(
            "SELECT id, name, quantity, notes, group_name, created_ts, updated_ts "
            "FROM local_equipment WHERE project_code = ? ORDER BY created_ts DESC",
            (project_code,),
        ).fetchall()
        items = []
        for row in rows:
            if isinstance(row, Mapping):
                items.append(dict(row))
            elif isinstance(row, Sequence):
                items.append({
                    "id": row[0],
                    "name": row[1],
                    "quantity": row[2],
                    "notes": row[3],
                    "group_name": row[4],
                    "created_ts": row[5],
                    "updated_ts": row[6],
                })
        return jsonify({"ok": True, "items": items})

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "missing_name"}), 400

    quantity = int(data.get("quantity", 1))
    notes = (data.get("notes") or "").strip() or None
    group_name = (data.get("group_name") or "").strip() or "Attrezzature extra"
    now = int(time.time() * 1000)

    db.execute(
        "INSERT INTO local_equipment (project_code, name, quantity, notes, group_name, created_ts, updated_ts) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        (project_code, name, quantity, notes, group_name, now, now),
    )
    new_id = _last_insert_id(db)
    db.commit()

    return jsonify({
        "ok": True,
        "item": {
            "id": new_id,
            "name": name,
            "quantity": quantity,
            "notes": notes,
            "group_name": group_name,
            "created_ts": now,
            "updated_ts": now,
        },
    })


@app.delete("/api/project/local-equipment/<int:item_id>")
@login_required
def api_delete_local_equipment(item_id: int):
    db = get_db()
    project_code = get_app_state(db, "project_code")
    if not project_code:
        return jsonify({"ok": False, "error": "no_active_project"}), 400

    ensure_local_equipment_table(db)
    db.execute(
        "DELETE FROM local_equipment WHERE id = ? AND project_code = ?",
        (item_id, project_code),
    )
    db.commit()
    return jsonify({"ok": True})


# ──────────────────────────────────────────────────────────────────────────────
# Foto progetto
# ──────────────────────────────────────────────────────────────────────────────

@app.route("/api/project/photos", methods=["GET", "POST"])
@login_required
def api_project_photos():
    db = get_db()
    project_code = get_app_state(db, "project_code")
    if not project_code:
        return jsonify({"ok": False, "error": "no_active_project"}), 400

    ensure_project_photos_table(db)

    if request.method == "GET":
        rows = db.execute(
            "SELECT id, filename, original_name, mime_type, file_size, caption, created_ts "
            "FROM project_photos WHERE project_code = ? ORDER BY created_ts DESC",
            (project_code,),
        ).fetchall()
        items = []
        for row in rows:
            if isinstance(row, Mapping):
                items.append(dict(row))
            elif isinstance(row, Sequence):
                items.append({
                    "id": row[0],
                    "filename": row[1],
                    "original_name": row[2],
                    "mime_type": row[3],
                    "file_size": row[4],
                    "caption": row[5],
                    "created_ts": row[6],
                })
        return jsonify({"ok": True, "items": items, "project_code": project_code})

    # POST - Upload nuova foto
    if "photo" not in request.files:
        return jsonify({"ok": False, "error": "no_file"}), 400

    file = request.files["photo"]
    if file.filename == "":
        return jsonify({"ok": False, "error": "no_filename"}), 400

    if not allowed_photo_file(file.filename):
        return jsonify({"ok": False, "error": "invalid_file_type"}), 400

    # Leggi il file in memoria per verificare la dimensione
    file_data = file.read()
    if len(file_data) > MAX_PHOTO_SIZE:
        return jsonify({"ok": False, "error": "file_too_large"}), 400

    # Genera nome file univoco
    original_name = file.filename
    ext = original_name.rsplit(".", 1)[1].lower() if "." in original_name else "jpg"
    unique_filename = f"{project_code}_{int(time.time() * 1000)}_{os.urandom(4).hex()}.{ext}"

    # Salva il file
    file_path = os.path.join(PHOTOS_UPLOAD_FOLDER, unique_filename)
    with open(file_path, "wb") as f:
        f.write(file_data)

    # Salva nel database
    caption = request.form.get("caption", "").strip() or None
    mime_type = file.content_type or f"image/{ext}"
    now = int(time.time() * 1000)

    db.execute(
        "INSERT INTO project_photos (project_code, filename, original_name, mime_type, file_size, caption, created_ts) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        (project_code, unique_filename, original_name, mime_type, len(file_data), caption, now),
    )
    new_id = _last_insert_id(db)
    db.commit()

    return jsonify({
        "ok": True,
        "item": {
            "id": new_id,
            "filename": unique_filename,
            "original_name": original_name,
            "mime_type": mime_type,
            "file_size": len(file_data),
            "caption": caption,
            "created_ts": now,
        },
    })


@app.get("/api/project/photos/<filename>")
@login_required
def api_get_photo(filename: str):
    """Serve una foto dal filesystem"""
    file_path = os.path.join(PHOTOS_UPLOAD_FOLDER, filename)
    if not os.path.exists(file_path):
        return jsonify({"ok": False, "error": "not_found"}), 404

    # Sicurezza: verifica che il file appartenga al progetto attivo
    db = get_db()
    project_code = get_app_state(db, "project_code")
    ensure_project_photos_table(db)

    row = db.execute(
        "SELECT id FROM project_photos WHERE filename = ? AND project_code = ?",
        (filename, project_code),
    ).fetchone()

    if not row:
        return jsonify({"ok": False, "error": "not_authorized"}), 403

    return send_from_directory(PHOTOS_UPLOAD_FOLDER, filename)


@app.delete("/api/project/photos/<int:photo_id>")
@login_required
def api_delete_photo(photo_id: int):
    db = get_db()
    project_code = get_app_state(db, "project_code")
    if not project_code:
        return jsonify({"ok": False, "error": "no_active_project"}), 400

    ensure_project_photos_table(db)

    # Recupera il filename per eliminare il file
    row = db.execute(
        "SELECT filename FROM project_photos WHERE id = ? AND project_code = ?",
        (photo_id, project_code),
    ).fetchone()

    if not row:
        return jsonify({"ok": False, "error": "not_found"}), 404

    filename = row["filename"] if isinstance(row, Mapping) else row[0]
    file_path = os.path.join(PHOTOS_UPLOAD_FOLDER, filename)

    # Elimina dal database
    db.execute(
        "DELETE FROM project_photos WHERE id = ? AND project_code = ?",
        (photo_id, project_code),
    )
    db.commit()

    # Elimina il file dal filesystem
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except OSError as e:
        app.logger.warning("Impossibile eliminare file foto %s: %s", file_path, e)

    return jsonify({"ok": True})


@app.post("/api/load_project")
@login_required
def api_load_project():
    data = request.get_json(silent=True) or {}
    project_code = (data.get("project_code") or "").strip().upper()
    project_date = (data.get("project_date") or "").strip()
    if not project_code:
        return jsonify({"ok": False, "error": "missing_project_code"}), 400
    if not project_date:
        return jsonify({"ok": False, "error": "missing_project_date"}), 400

    db = get_db()
    if has_active_member_sessions(db):
        return (
            jsonify(
                {
                    "ok": False,
                    "error": "active_sessions_present",
                    "message": "Sono presenti operatori con attività in corso o in pausa. Termina le sessioni prima di ricaricare il progetto.",
                }
            ),
            409,
        )

    plan = mock_fetch_project(project_code, project_date)
    if plan is None:
        clear_project_state(db)
        db.commit()
        # Pulisce anche la sessione del supervisor
        session.pop('supervisor_project_code', None)
        session.pop('supervisor_project_name', None)
        return jsonify({"ok": False, "error": "project_not_found"}), 404

    apply_project_plan(db, plan)
    db.commit()
    
    # Salva il progetto nella sessione del supervisor (individuale)
    session['supervisor_project_code'] = plan["project_code"]
    session['supervisor_project_name'] = plan.get("project_name")
    
    # Salva anche nel database per persistenza dopo logout/login
    username = session.get('user')
    if username:
        placeholder = "%s" if DB_VENDOR == "mysql" else "?"
        try:
            db.execute(
                f"UPDATE app_users SET current_project_code={placeholder}, current_project_name={placeholder} WHERE username={placeholder}",
                (plan["project_code"], plan.get("project_name"), username)
            )
            db.commit()
            app.logger.info("Progetto %s salvato per utente %s", plan["project_code"], username)
        except Exception as e:
            app.logger.warning("Impossibile salvare progetto corrente per %s: %s", username, e)

    return jsonify(
        {
            "ok": True,
            "project": {
                "code": plan["project_code"],
                "name": plan.get("project_name"),
            },
        }
    )


@app.post("/api/move")
@login_required
def api_move():
    data = request.get_json(silent=True) or {}
    member_key = (data.get("member_key") or "").strip()
    member_name = (data.get("member_name") or "").strip()
    activity_id = data.get("activity_id")

    if isinstance(activity_id, str):
        activity_id = activity_id.strip()
    if activity_id == "":
        activity_id = None

    if not member_key or not member_name:
        return jsonify({"ok": False, "error": "invalid_payload"}), 400

    project_code = session.get('supervisor_project_code')
    if not project_code:
        return jsonify({"ok": False, "error": "no_project_selected"}), 400

    db = get_db()
    now = now_ms()
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"

    if activity_id:
        exists = db.execute(
            f"SELECT 1 FROM activities WHERE activity_id={placeholder} AND project_code={placeholder}",
            (activity_id, project_code),
        ).fetchone()
        if exists is None:
            return jsonify({"ok": False, "error": "unknown_activity"}), 400

    existing = db.execute(
        f"SELECT * FROM member_state WHERE member_key={placeholder} AND project_code={placeholder}",
        (member_key, project_code),
    ).fetchone()

    if existing is None:
        db.execute(
            f"""
            INSERT INTO member_state(
                member_key, project_code, member_name, activity_id, running, start_ts, elapsed_cached, pause_start, entered_ts
            ) VALUES({placeholder},{placeholder},{placeholder},{placeholder},{placeholder},{placeholder},{placeholder},{placeholder},{placeholder})
            """,
            (member_key, project_code, member_name, None, 0, None, 0, None, None),
        )
        existing = db.execute(
            f"SELECT * FROM member_state WHERE member_key={placeholder} AND project_code={placeholder}",
            (member_key, project_code),
        ).fetchone()
    else:
        db.execute(
            f"UPDATE member_state SET member_name={placeholder} WHERE member_key={placeholder} AND project_code={placeholder}",
            (member_name, member_key, project_code),
        )

    if existing is None:
        app.logger.error("Member state insert fallita per %s", member_key)
        return jsonify({"ok": False, "error": "member_state_error"}), 500

    previous_activity = existing["activity_id"]
    previous_entered_ts = row_value(existing, "entered_ts")
    prev_elapsed = compute_elapsed(existing, now)
    normalized_previous = str(previous_activity) if previous_activity else None
    normalized_target = str(activity_id) if activity_id else None
    same_activity = normalized_previous is not None and normalized_previous == normalized_target

    running = RUN_STATE_RUNNING if activity_id else RUN_STATE_PAUSED
    start_ts = now if running else None
    reset_elapsed = bool(activity_id) and not same_activity
    elapsed_cached = 0 if reset_elapsed else prev_elapsed
    next_entered_ts = previous_entered_ts
    if activity_id and not same_activity:
        next_entered_ts = now
    if not activity_id:
        next_entered_ts = None

    activity_meta = load_activity_meta(db)
    meta_changed = False
    auto_closed_previous = False
    if normalized_previous and normalized_previous != normalized_target:
        if prev_elapsed > 0:
            meta_changed = increment_activity_runtime(activity_meta, normalized_previous, prev_elapsed)
        auto_closed_previous = True

    db.execute(
        f"""
        UPDATE member_state
        SET activity_id={placeholder}, running={placeholder}, start_ts={placeholder}, elapsed_cached={placeholder}, pause_start=NULL, entered_ts={placeholder}
        WHERE member_key={placeholder} AND project_code={placeholder}
        """,
        (activity_id, running, start_ts, elapsed_cached, next_entered_ts, member_key, project_code),
    )

    if meta_changed:
        save_activity_meta(db, activity_meta)

    if auto_closed_previous:
        # Calcola il tempo totale partendo dall'ultimo move verso questa attività
        activity_start_ts = find_last_move_ts(db, member_key, normalized_previous)
        if activity_start_ts is None:
            activity_start_ts = previous_entered_ts or existing["start_ts"]

        total_ms = 0
        if activity_start_ts:
            total_ms = max(0, now - activity_start_ts)
        
        pause_ms = max(0, total_ms - prev_elapsed)
        
        finish_payload = {
            "member_name": existing["member_name"],
            "activity_id": normalized_previous,
            "duration_ms": prev_elapsed,
            "total_ms": total_ms,
            "pause_ms": pause_ms,
            "auto_close": True,
            "project_code": project_code,
        }
        db.execute(
            f"INSERT INTO event_log(ts, kind, member_key, details, project_code) VALUES({placeholder},{placeholder},{placeholder},{placeholder},{placeholder})",
            (now, "finish_activity", member_key, json.dumps(finish_payload), project_code),
        )

    move_details = {
        "from": previous_activity,
        "to": activity_id,
        "member_name": member_name,
        "duration_ms": prev_elapsed,
        "project_code": project_code,
    }
    db.execute(
        f"INSERT INTO event_log(ts, kind, member_key, details, project_code) VALUES({placeholder},{placeholder},{placeholder},{placeholder},{placeholder})",
        (now, "move", member_key, json.dumps(move_details), project_code),
    )

    db.commit()
    return jsonify({"ok": True})


@app.post("/api/start_activity")
@login_required
def api_start_activity():
    """Avvia i timer per tutti i membri di una specifica attività."""
    data = request.get_json(silent=True) or {}
    activity_id = (data.get("activity_id") or "").strip()
    
    if not activity_id:
        return jsonify({"ok": False, "error": "missing_activity_id"}), 400
    
    now = now_ms()
    db = get_db()
    project_code = session.get("supervisor_project_code", "")
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"

    # Verifica che l'attività esista per questo progetto
    activity_exists = db.execute(
        f"SELECT 1 FROM activities WHERE activity_id={placeholder} AND project_code={placeholder}",
        (activity_id, project_code),
    ).fetchone()
    
    if not activity_exists:
        return jsonify({"ok": False, "error": "activity_not_found"}), 404

    # Trova tutti i membri assegnati a questa attività con timer non avviato
    rows = db.execute(
        f"SELECT member_key FROM member_state WHERE activity_id={placeholder} AND running={placeholder} AND project_code={placeholder}",
        (activity_id, RUN_STATE_PAUSED, project_code),
    ).fetchall()

    if not rows:
        return jsonify({"ok": True, "affected": 0})

    affected = 0
    for row in rows:
        db.execute(
            f"UPDATE member_state SET running={placeholder}, start_ts={placeholder}, pause_start=NULL WHERE member_key={placeholder} AND project_code={placeholder}",
            (RUN_STATE_RUNNING, now, row["member_key"], project_code),
        )
        affected += 1

    db.execute(
        f"INSERT INTO event_log(ts, kind, details, project_code) VALUES({placeholder},{placeholder},{placeholder},{placeholder})",
        (now, "start_activity", json.dumps({"activity_id": activity_id, "affected": affected}), project_code),
    )

    db.commit()
    return jsonify({"ok": True, "affected": affected})


@app.post("/api/start_member")
@login_required
def api_start_member():
    """Avvia il timer per un singolo membro."""
    data = request.get_json(silent=True) or {}
    member_key = (data.get("member_key") or "").strip()
    
    if not member_key:
        return jsonify({"ok": False, "error": "missing_member_key"}), 400
    
    now = now_ms()
    db = get_db()
    project_code = session.get("supervisor_project_code", "")
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"

    # Verifica che il membro esista e abbia un'attività assegnata
    member = db.execute(
        f"SELECT member_key, member_name, activity_id, running FROM member_state WHERE member_key={placeholder} AND project_code={placeholder}",
        (member_key, project_code),
    ).fetchone()
    
    if not member:
        return jsonify({"ok": False, "error": "member_not_found"}), 404
    
    if not member["activity_id"]:
        return jsonify({"ok": False, "error": "no_activity_assigned"}), 400
    
    if member["running"] == RUN_STATE_RUNNING:
        return jsonify({"ok": False, "error": "already_running"}), 400

    # Avvia il timer
    db.execute(
        f"UPDATE member_state SET running={placeholder}, start_ts={placeholder}, pause_start=NULL WHERE member_key={placeholder} AND project_code={placeholder}",
        (RUN_STATE_RUNNING, now, member_key, project_code),
    )

    db.execute(
        f"INSERT INTO event_log(ts, kind, details, project_code) VALUES({placeholder},{placeholder},{placeholder},{placeholder})",
        (now, "start_member", json.dumps({"member_key": member_key}), project_code),
    )

    db.commit()
    return jsonify({"ok": True})


@app.post("/api/start_all")
@login_required
def api_start_all():
    """Avvia i timer per tutti i membri che hanno un'attività assegnata."""
    now = now_ms()
    db = get_db()
    project_code = session.get("supervisor_project_code", "")
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"

    # Trova tutti i membri con activity_id assegnato ma non in esecuzione
    rows = db.execute(
        f"SELECT member_key FROM member_state WHERE activity_id IS NOT NULL AND running={placeholder} AND project_code={placeholder}",
        (RUN_STATE_PAUSED, project_code),
    ).fetchall()

    if not rows:
        return jsonify({"ok": True, "affected": 0})

    affected = 0
    for row in rows:
        db.execute(
            f"UPDATE member_state SET running={placeholder}, start_ts={placeholder}, pause_start=NULL WHERE member_key={placeholder} AND project_code={placeholder}",
            (RUN_STATE_RUNNING, now, row["member_key"], project_code),
        )
        affected += 1

    db.execute(
        f"INSERT INTO event_log(ts, kind, details, project_code) VALUES({placeholder},{placeholder},{placeholder},{placeholder})",
        (now, "start_all", json.dumps({"affected": affected}), project_code),
    )

    db.commit()
    return jsonify({"ok": True, "affected": affected})


@app.post("/api/pause_all")
@login_required
def api_pause_all():
    now = now_ms()
    db = get_db()
    project_code = session.get("supervisor_project_code", "")
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"

    rows = db.execute(
        f"SELECT member_key, start_ts, elapsed_cached FROM member_state WHERE running={placeholder} AND project_code={placeholder}",
        (RUN_STATE_RUNNING, project_code),
    ).fetchall()

    for row in rows:
        start_ts = row["start_ts"] or now
        elapsed = (row["elapsed_cached"] or 0) + max(0, now - start_ts)
        db.execute(
            f"""
            UPDATE member_state
            SET running={placeholder}, start_ts=NULL, elapsed_cached={placeholder}, pause_start={placeholder}
            WHERE member_key={placeholder} AND project_code={placeholder}
            """,
            (RUN_STATE_PAUSED, elapsed, now, row["member_key"], project_code),
        )

    if rows:
        db.execute(
            f"INSERT INTO event_log(ts, kind, details, project_code) VALUES({placeholder},{placeholder},{placeholder},{placeholder})",
            (now, "pause_all", json.dumps({"affected": len(rows)}), project_code),
        )

    db.commit()
    return jsonify({"ok": True})


@app.post("/api/resume_all")
@login_required
def api_resume_all():
    now = now_ms()
    db = get_db()
    project_code = session.get("supervisor_project_code", "")
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"

    rows = db.execute(
        f"SELECT member_key FROM member_state WHERE running={placeholder} AND pause_start IS NOT NULL AND project_code={placeholder}",
        (RUN_STATE_PAUSED, project_code),
    ).fetchall()

    for row in rows:
        db.execute(
            f"UPDATE member_state SET running={placeholder}, start_ts={placeholder}, pause_start=NULL WHERE member_key={placeholder} AND project_code={placeholder}",
            (RUN_STATE_RUNNING, now, row["member_key"], project_code),
        )

    if rows:
        db.execute(
            f"INSERT INTO event_log(ts, kind, details, project_code) VALUES({placeholder},{placeholder},{placeholder},{placeholder})",
            (now, "resume_all", json.dumps({"affected": len(rows)}), project_code),
        )

    db.commit()
    return jsonify({"ok": True})


@app.post("/api/finish_all")
@login_required
def api_finish_all():
    now = now_ms()
    db = get_db()
    project_code = session.get("supervisor_project_code", "")
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"

    rows = db.execute(
        f"SELECT * FROM member_state WHERE activity_id IS NOT NULL AND project_code={placeholder}",
        (project_code,),
    ).fetchall()

    activity_meta = load_activity_meta(db)
    meta_changed = False

    affected = 0
    for row in rows:
        elapsed = compute_elapsed(row, now)
        if row["activity_id"] and elapsed > 0:
            meta_changed |= increment_activity_runtime(activity_meta, str(row["activity_id"]), elapsed)
        
        activity_start_ts = find_last_move_ts(db, row["member_key"], str(row["activity_id"]))
        if activity_start_ts is None:
            activity_start_ts = row["entered_ts"] or row["start_ts"]

        total_ms = 0
        if activity_start_ts:
            total_ms = max(0, now - activity_start_ts)
        
        pause_ms = max(0, total_ms - elapsed)
        
        db.execute(
            f"""
            UPDATE member_state
            SET activity_id=NULL, running={placeholder}, start_ts=NULL, elapsed_cached=0, pause_start=NULL, entered_ts=NULL
            WHERE member_key={placeholder} AND project_code={placeholder}
            """,
            (RUN_STATE_FINISHED, row["member_key"], project_code),
        )
        db.execute(
            f"INSERT INTO event_log(ts, kind, member_key, details, project_code) VALUES({placeholder},{placeholder},{placeholder},{placeholder},{placeholder})",
            (
                now,
                "finish_activity",
                row["member_key"],
                json.dumps(
                    {
                        "member_name": row["member_name"],
                        "activity_id": row["activity_id"],
                        "duration_ms": elapsed,
                        "total_ms": total_ms,
                        "pause_ms": pause_ms,
                    }
                ),
                project_code,
            ),
        )
        affected += 1

    if affected:
        db.execute(
            f"INSERT INTO event_log(ts, kind, details, project_code) VALUES({placeholder},{placeholder},{placeholder},{placeholder})",
            (
                now,
                "finish_all",
                json.dumps({"affected": affected}),
                project_code,
            ),
        )

    if meta_changed:
        save_activity_meta(db, activity_meta)

    db.commit()
    return jsonify({"ok": True, "affected": affected})


@app.post("/api/member/pause")
@login_required
def api_member_pause():
    data = request.get_json(silent=True) or {}
    member_key = (data.get("member_key") or "").strip()
    if not member_key:
        return jsonify({"ok": False, "error": "missing_member_key"}), 400

    db = get_db()
    project_code = session.get("supervisor_project_code", "")
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    member = fetch_member(db, member_key, project_code)
    if member is None:
        return jsonify({"ok": False, "error": "member_not_found"}), 404

    if not member["activity_id"]:
        return jsonify({"ok": False, "error": "member_not_assigned"}), 400

    if member["pause_start"] is not None:
        return jsonify({"ok": True, "already_paused": True})

    if member["running"] != RUN_STATE_RUNNING:
        return jsonify({"ok": False, "error": "member_not_running"}), 400

    now = now_ms()
    elapsed = compute_elapsed(member, now)

    db.execute(
        f"""
        UPDATE member_state
        SET running={placeholder}, start_ts=NULL, elapsed_cached={placeholder}, pause_start={placeholder}
        WHERE member_key={placeholder} AND project_code={placeholder}
        """,
        (RUN_STATE_PAUSED, elapsed, now, member_key, project_code),
    )

    db.execute(
        f"INSERT INTO event_log(ts, kind, member_key, details, project_code) VALUES({placeholder},{placeholder},{placeholder},{placeholder},{placeholder})",
        (
            now,
            "pause_member",
            member_key,
            json.dumps(
                {
                    "member_name": member["member_name"],
                    "activity_id": member["activity_id"],
                    "duration_ms": elapsed,
                }
            ),
            project_code,
        ),
    )

    # Genera timbratura inizio_pausa per l'operatore
    member_username = get_username_from_member_key(db, member_key)
    if member_username:
        create_supervisor_pause_timbratura(
            db,
            username=member_username,
            tipo='inizio_pausa',
            member_name=member["member_name"],
            supervisor_username=session.get("user")
        )

    db.commit()
    return jsonify({"ok": True})


@app.post("/api/member/resume")
@login_required
def api_member_resume():
    data = request.get_json(silent=True) or {}
    member_key = (data.get("member_key") or "").strip()
    if not member_key:
        return jsonify({"ok": False, "error": "missing_member_key"}), 400

    db = get_db()
    project_code = session.get("supervisor_project_code", "")
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    member = fetch_member(db, member_key, project_code)
    if member is None:
        return jsonify({"ok": False, "error": "member_not_found"}), 404

    if not member["activity_id"]:
        return jsonify({"ok": False, "error": "member_not_assigned"}), 400

    if member["running"] == RUN_STATE_RUNNING:
        return jsonify({"ok": True, "already_running": True})

    if member["pause_start"] is None:
        return jsonify({"ok": False, "error": "member_not_paused"}), 400

    now = now_ms()

    db.execute(
        f"UPDATE member_state SET running={placeholder}, start_ts={placeholder}, pause_start=NULL WHERE member_key={placeholder} AND project_code={placeholder}",
        (RUN_STATE_RUNNING, now, member_key, project_code),
    )

    db.execute(
        f"INSERT INTO event_log(ts, kind, member_key, details, project_code) VALUES({placeholder},{placeholder},{placeholder},{placeholder},{placeholder})",
        (
            now,
            "resume_member",
            member_key,
            json.dumps(
                {
                    "member_name": member["member_name"],
                    "activity_id": member["activity_id"],
                }
            ),
            project_code,
        ),
    )

    # Genera timbratura fine_pausa per l'operatore
    member_username = get_username_from_member_key(db, member_key)
    if member_username:
        create_supervisor_pause_timbratura(
            db,
            username=member_username,
            tipo='fine_pausa',
            member_name=member["member_name"],
            supervisor_username=session.get("user")
        )

    db.commit()
    return jsonify({"ok": True})


@app.post("/api/member/finish")
@login_required
def api_member_finish():
    data = request.get_json(silent=True) or {}
    member_key = (data.get("member_key") or "").strip()
    if not member_key:
        return jsonify({"ok": False, "error": "missing_member_key"}), 400

    db = get_db()
    project_code = session.get("supervisor_project_code", "")
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    member = fetch_member(db, member_key, project_code)
    if member is None:
        return jsonify({"ok": False, "error": "member_not_found"}), 404

    if not member["activity_id"]:
        return jsonify({"ok": False, "error": "member_not_assigned"}), 400

    now = now_ms()

    elapsed = compute_elapsed(member, now)
    activity_meta = load_activity_meta(db)
    meta_changed = False
    if member["activity_id"] and elapsed > 0:
        meta_changed = increment_activity_runtime(activity_meta, str(member["activity_id"]), elapsed)

    activity_start_ts = find_last_move_ts(db, member_key, str(member["activity_id"]))
    if activity_start_ts is None:
        activity_start_ts = member["entered_ts"] or member["start_ts"]

    total_ms = 0
    if activity_start_ts:
        total_ms = max(0, now - activity_start_ts)
    pause_ms = max(0, total_ms - elapsed)

    db.execute(
        f"""
        UPDATE member_state
        SET activity_id=NULL, running={placeholder}, start_ts=NULL, elapsed_cached=0, pause_start=NULL, entered_ts=NULL
        WHERE member_key={placeholder} AND project_code={placeholder}
        """,
        (RUN_STATE_FINISHED, member_key, project_code),
    )

    db.execute(
        f"INSERT INTO event_log(ts, kind, member_key, details, project_code) VALUES({placeholder},{placeholder},{placeholder},{placeholder},{placeholder})",
        (
            now,
            "finish_activity",
            member_key,
            json.dumps(
                {
                    "member_name": member["member_name"],
                    "activity_id": member["activity_id"],
                    "duration_ms": elapsed,
                    "total_ms": total_ms,
                    "pause_ms": pause_ms,
                }
            ),
            project_code,
        ),
    )

    if meta_changed:
        save_activity_meta(db, activity_meta)

    db.commit()
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
#  API - Gestione operatori nel progetto (capo squadra)
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/project/available-operators")
@login_required
def api_project_available_operators():
    """Lista operatori disponibili (non già nel progetto) dalla tabella crew_members."""
    db = get_db()
    project_code = session.get("supervisor_project_code", "")
    
    if not project_code:
        return jsonify({"ok": False, "error": "no_project"}), 400
    
    ensure_crew_members_table(db)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Ottieni gli operatori già nel progetto
    existing_keys = set()
    existing_rows = db.execute(
        f"SELECT member_key FROM member_state WHERE project_code = {placeholder}",
        (project_code,)
    ).fetchall()
    for row in existing_rows:
        key = row["member_key"] if isinstance(row, dict) else row[0]
        existing_keys.add(key)
    
    # Ottieni tutti gli operatori attivi da crew_members
    if DB_VENDOR == "mysql":
        crew_rows = db.execute(
            "SELECT rentman_id, name FROM crew_members WHERE is_active = 1 ORDER BY name"
        ).fetchall()
    else:
        crew_rows = db.execute(
            "SELECT rentman_id, name FROM crew_members WHERE is_active = 1 ORDER BY name"
        ).fetchall()
    
    available = []
    for row in crew_rows:
        rentman_id = row["rentman_id"] if isinstance(row, dict) else row[0]
        name = row["name"] if isinstance(row, dict) else row[1]
        member_key = f"rentman-crew-{rentman_id}"
        
        # Escludi operatori già nel progetto
        if member_key not in existing_keys:
            available.append({
                "key": member_key,
                "name": name,
                "rentman_id": rentman_id
            })
    
    return jsonify({"ok": True, "operators": available})


@app.post("/api/project/add-operator")
@login_required
def api_project_add_operator():
    """Aggiunge un operatore al progetto corrente."""
    data = request.get_json(silent=True) or {}
    
    # Supporta sia l'aggiunta da crew_members (rentman_id) che un nuovo operatore manuale (name)
    rentman_id = data.get("rentman_id")
    name = (data.get("name") or "").strip()
    
    if not rentman_id and not name:
        return jsonify({"ok": False, "error": "missing_data"}), 400
    
    db = get_db()
    project_code = session.get("supervisor_project_code", "")
    
    if not project_code:
        return jsonify({"ok": False, "error": "no_project"}), 400
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    now = now_ms()
    
    if rentman_id:
        # Aggiunta da crew_members
        member_key = f"rentman-crew-{rentman_id}"
        
        # Verifica che non esista già nel progetto
        existing = db.execute(
            f"SELECT 1 FROM member_state WHERE member_key = {placeholder} AND project_code = {placeholder}",
            (member_key, project_code)
        ).fetchone()
        
        if existing:
            return jsonify({"ok": False, "error": "already_in_project"}), 400
        
        # Ottieni il nome da crew_members
        if DB_VENDOR == "mysql":
            crew = db.execute(
                "SELECT name FROM crew_members WHERE rentman_id = %s", (rentman_id,)
            ).fetchone()
        else:
            crew = db.execute(
                "SELECT name FROM crew_members WHERE rentman_id = ?", (rentman_id,)
            ).fetchone()
        
        if not crew:
            return jsonify({"ok": False, "error": "operator_not_found"}), 404
        
        member_name = crew["name"] if isinstance(crew, dict) else crew[0]
    else:
        # Operatore manuale - genera una chiave unica
        base_key = f"local-{name.lower().replace(' ', '-')}"
        member_key = base_key
        counter = 1
        
        while True:
            existing = db.execute(
                f"SELECT 1 FROM member_state WHERE member_key = {placeholder} AND project_code = {placeholder}",
                (member_key, project_code)
            ).fetchone()
            if not existing:
                break
            member_key = f"{base_key}-{counter}"
            counter += 1
        
        member_name = name
    
    # Inserisci l'operatore nel progetto
    db.execute(
        f"""
        INSERT INTO member_state (member_key, project_code, member_name, activity_id, running, start_ts, elapsed_cached, pause_start, entered_ts)
        VALUES ({placeholder}, {placeholder}, {placeholder}, NULL, {placeholder}, NULL, 0, NULL, NULL)
        """,
        (member_key, project_code, member_name, RUN_STATE_PAUSED)
    )
    
    # Log evento
    db.execute(
        f"INSERT INTO event_log(ts, kind, details, project_code) VALUES({placeholder},{placeholder},{placeholder},{placeholder})",
        (now, "add_operator", json.dumps({"member_key": member_key, "member_name": member_name}), project_code)
    )
    
    db.commit()
    
    return jsonify({
        "ok": True,
        "member": {
            "key": member_key,
            "name": member_name
        }
    })


@app.post("/api/project/remove-operator")
@login_required
def api_project_remove_operator():
    """Rimuove un operatore dal progetto corrente."""
    data = request.get_json(silent=True) or {}
    member_key = (data.get("member_key") or "").strip()
    
    if not member_key:
        return jsonify({"ok": False, "error": "missing_member_key"}), 400
    
    db = get_db()
    project_code = session.get("supervisor_project_code", "")
    
    if not project_code:
        return jsonify({"ok": False, "error": "no_project"}), 400
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    now = now_ms()
    
    # Verifica che l'operatore esista e non sia in attività
    member = db.execute(
        f"SELECT member_name, activity_id, running FROM member_state WHERE member_key = {placeholder} AND project_code = {placeholder}",
        (member_key, project_code)
    ).fetchone()
    
    if not member:
        return jsonify({"ok": False, "error": "member_not_found"}), 404
    
    member_name = member["member_name"] if isinstance(member, dict) else member[0]
    activity_id = member["activity_id"] if isinstance(member, dict) else member[1]
    running = member["running"] if isinstance(member, dict) else member[2]
    
    # Non permettere la rimozione se l'operatore ha un timer attivo
    if running == RUN_STATE_RUNNING:
        return jsonify({"ok": False, "error": "member_running", "message": "Ferma il timer prima di rimuovere l'operatore"}), 400
    
    # Rimuovi l'operatore
    db.execute(
        f"DELETE FROM member_state WHERE member_key = {placeholder} AND project_code = {placeholder}",
        (member_key, project_code)
    )
    
    # Log evento
    db.execute(
        f"INSERT INTO event_log(ts, kind, details, project_code) VALUES({placeholder},{placeholder},{placeholder},{placeholder})",
        (now, "remove_operator", json.dumps({"member_key": member_key, "member_name": member_name}), project_code)
    )
    
    db.commit()
    
    return jsonify({"ok": True})


@app.get("/api/export")
@login_required
def api_export():
    """Esporta i dati delle attività in formato Excel o CSV."""
    export_format = request.args.get("format", "excel").lower()
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    project_filter = request.args.get("project_code", "").strip()

    db = get_db()
    
    # Ottieni informazioni progetto corrente
    raw_project_code = get_app_state(db, "project_code")
    project_name = get_app_state(db, "project_name") or raw_project_code or ""

    if not raw_project_code:
        return jsonify({"ok": False, "error": "no_active_project"}), 400

    project_code = str(raw_project_code)
    project_name = str(project_name or project_code)

    # Se c'è un filtro progetto e non corrisponde al progetto attivo, errore
    if project_filter and project_filter != project_code:
        return jsonify({"ok": False, "error": "project_mismatch"}), 400

    start_date_filter = parse_iso_date(start_date)
    end_date_filter = parse_iso_date(end_date)

    session_rows = build_session_rows(
        db,
        start_date=start_date_filter,
        end_date=end_date_filter,
    )

    export_data = []
    for session_row in session_rows:
        start_dt = datetime.fromtimestamp(session_row["start_ts"] / 1000, tz=timezone.utc)
        end_dt = datetime.fromtimestamp(session_row["end_ts"] / 1000, tz=timezone.utc)
        status_label = "Completato" if session_row["status"] == "completed" else "In corso"
        export_data.append(
            {
                "operatore": session_row["member_name"],
                "attivita": session_row["activity_label"],
                "data_inizio": start_dt.strftime("%d/%m/%Y"),
                "ora_inizio": start_dt.strftime("%H:%M:%S"),
                "data_fine": end_dt.strftime("%d/%m/%Y") if session_row["status"] == "completed" else "In corso",
                "ora_fine": end_dt.strftime("%H:%M:%S") if session_row["status"] == "completed" else "-",
                "durata_netta": format_duration_ms(session_row["net_ms"]) or "00:00:00",
                "tempo_pausa": format_duration_ms(session_row["pause_ms"]) or "00:00:00",
                "num_pause": str(session_row["pause_count"]),
                "stato": status_label,
            }
        )
    
    app.logger.info("Export: generati %s record per l'export", len(export_data))

    # Genera file in base al formato
    if export_format == "csv":
        return generate_csv_export(export_data, project_code, project_name)
    else:
        return generate_excel_export(export_data, project_code, project_name)


def generate_excel_export(data: List[Dict[str, Any]], project_code: str, project_name: str):
    """Genera un file Excel con template professionale."""
    wb = Workbook()
    ws_raw = wb.active
    if ws_raw is None:  # pragma: no cover - openpyxl should always provide an active sheet
        ws_raw = wb.create_sheet(title="Report Attività")
    ws: Worksheet = cast(Worksheet, ws_raw)
    ws.title = "Report Attività"

    # Stili
    header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    title_font = Font(name="Calibri", size=18, bold=True, color="1E293B")
    title_alignment = Alignment(horizontal="left", vertical="center")
    
    cell_font = Font(name="Calibri", size=11)
    cell_alignment = Alignment(horizontal="left", vertical="center")
    
    border_thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # Titolo report
    ws["A1"] = f"🔷 JobLOG - Report Attività"
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    
    # Info progetto
    ws["A2"] = f"Progetto: {project_code} - {project_name or project_code}"
    ws.merge_cells("A2:J2")
    project_cell = ws["A2"]
    project_cell.font = Font(name="Calibri", size=12, color="64748B")
    project_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Data generazione
    now = datetime.now()
    ws["A3"] = f"Generato il: {now.strftime('%d/%m/%Y alle %H:%M')}"
    ws.merge_cells("A3:J3")
    date_cell = ws["A3"]
    date_cell.font = Font(name="Calibri", size=10, color="94A3B8")
    date_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Riga vuota
    ws.append([])

    # Header colonne
    headers = ["Operatore", "Attività", "Data Inizio", "Ora Inizio", "Data Fine", "Ora Fine", "Durata Netta", "Tempo Pausa", "N° Pause", "Stato"]
    ws.append(headers)
    
    header_row = ws.max_row
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_thin

    # Dati
    for row_data in data:
        ws.append([
            row_data["operatore"],
            row_data["attivita"],
            row_data["data_inizio"],
            row_data["ora_inizio"],
            row_data["data_fine"],
            row_data["ora_fine"],
            row_data["durata_netta"],
            row_data["tempo_pausa"],
            row_data["num_pause"],
            row_data["stato"],
        ])
        
        row_num = ws.max_row
        for col_num in range(1, 11):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border_thin
            
            # Alternating row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    # Totale sessioni
    total_row = ws.max_row + 2
    total_cell_ref = f"A{total_row}"
    ws[total_cell_ref] = f"Totale Sessioni: {len(data)}"
    ws.merge_cells(f"A{total_row}:I{total_row}")
    total_cell = ws[total_cell_ref]
    total_cell.font = Font(name="Calibri", size=12, bold=True, color="1E293B")
    total_cell.alignment = Alignment(horizontal="right", vertical="center")

    # Auto-fit colonne
    column_widths = {
        "A": 20,  # Operatore
        "B": 30,  # Attività
        "C": 12,  # Data Inizio
        "D": 11,  # Ora Inizio
        "E": 12,  # Data Fine
        "F": 11,  # Ora Fine
        "G": 14,  # Durata Netta
        "H": 13,  # Tempo Pausa
        "I": 10,  # N° Pause
        "J": 12,  # Stato
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Salva in memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"joblog_report_{project_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def generate_csv_export(data: List[Dict[str, Any]], project_code: str, project_name: str):
    """Genera un file CSV con encoding UTF-8 BOM."""
    output = io.StringIO()
    
    # UTF-8 BOM per compatibilità Excel
    output.write("\ufeff")
    
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    
    # Header informativo
    writer.writerow([f"JobLOG - Report Attività"])
    writer.writerow([f"Progetto: {project_code} - {project_name or project_code}"])
    writer.writerow([f"Generato il: {datetime.now().strftime('%d/%m/%Y alle %H:%M')}"])
    writer.writerow([])  # Riga vuota
    
    # Header colonne
    writer.writerow(["Operatore", "Attività", "Data Inizio", "Ora Inizio", "Data Fine", "Ora Fine", "Durata Netta", "Tempo Pausa", "N° Pause", "Stato"])
    
    # Dati
    for row in data:
        writer.writerow([
            row["operatore"],
            row["attivita"],
            row["data_inizio"],
            row["ora_inizio"],
            row["data_fine"],
            row["ora_fine"],
            row["durata_netta"],
            row["tempo_pausa"],
            row["num_pause"],
            row["stato"],
        ])
    
    # Totale
    writer.writerow([])
    writer.writerow(["", "", "", "", "", "", "", "", f"Totale Sessioni: {len(data)}", ""])

    # Prepara per download
    output.seek(0)
    bytes_output = io.BytesIO(output.getvalue().encode("utf-8"))
    bytes_output.seek(0)

    filename = f"joblog_report_{project_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return send_file(
        bytes_output,
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename,
    )


@app.get("/admin")
@app.get("/admin/dashboard")
@login_required
def admin_dashboard_page() -> ResponseReturnValue:
    if not is_admin_or_supervisor():
        abort(403)

    display_name = session.get("user_display") or session.get("user_name") or session.get("user")
    primary_name = session.get("user_name") or display_name or session.get("user")
    initials = session.get("user_initials") or compute_initials(primary_name or "")

    return render_template(
        "admin_dashboard.html",
        user_name=primary_name,
        user_display=display_name,
        user_initials=initials,
        is_admin=bool(session.get("is_admin")),
    )


@app.get("/admin/sessions")
@login_required
def admin_sessions_page() -> ResponseReturnValue:
    if not is_admin_or_supervisor():
        abort(403)

    display_name = session.get('user_display') or session.get('user_name') or session.get('user')
    primary_name = session.get('user_name') or display_name or session.get('user')
    initials = session.get('user_initials') or compute_initials(primary_name or "")

    return render_template(
        "admin_sessions.html",
        user_name=primary_name,
        user_display=display_name,
        user_initials=initials,
        is_admin=bool(session.get("is_admin")),
    )


@app.get("/api/admin/magazzino/summary")
@login_required
def api_admin_magazzino_summary() -> ResponseReturnValue:
    """Riepilogo ore magazzino per progetto (range date)."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403

    # Supporto range date: date_start/date_end oppure date singola (retrocompatibilità)
    date_start = parse_iso_date(request.args.get("date_start")) or parse_iso_date(request.args.get("date")) or datetime.now().date()
    date_end = parse_iso_date(request.args.get("date_end")) or date_start
    
    # Assicura che date_end >= date_start
    if date_end < date_start:
        date_start, date_end = date_end, date_start

    start_dt = datetime.combine(date_start, datetime.min.time())
    end_dt = datetime.combine(date_end, datetime.min.time()) + timedelta(days=1)
    start_ms = int(start_dt.timestamp() * 1000)
    end_ms = int(end_dt.timestamp() * 1000)

    db = get_db()
    ensure_warehouse_sessions_table(db)

    rows = db.execute(
        """
        SELECT project_code, COUNT(*) AS sessions, COALESCE(SUM(elapsed_ms), 0) AS total_ms
        FROM warehouse_sessions
        WHERE created_ts >= ? AND created_ts < ?
        GROUP BY project_code
        ORDER BY total_ms DESC, project_code ASC
        """,
        (start_ms, end_ms),
    ).fetchall()

    team_sessions = build_session_rows(
        db,
        start_date=date_start,
        end_date=date_end,
    )
    team_total_ms = sum(_coerce_int(s.get("net_ms")) or 0 for s in team_sessions)
    team_total_sessions = len(team_sessions)

    items: List[Dict[str, Any]] = []
    total_ms = 0
    total_sessions = 0
    for row in rows or []:
        ms_value = _coerce_int(row["total_ms"]) or 0
        sessions_count = _coerce_int(row["sessions"]) or 0
        total_ms += ms_value
        total_sessions += sessions_count
        items.append(
            {
                "project_code": row["project_code"],
                "sessions": sessions_count,
                "total_ms": ms_value,
            }
        )

    return jsonify(
        {
            "ok": True,
            "date": date_start.isoformat() if date_start == date_end else f"{date_start.isoformat()} - {date_end.isoformat()}",
            "items": items,
            "total_ms": total_ms,
            "total_sessions": total_sessions,
            "team_total_ms": team_total_ms,
            "team_total_sessions": team_total_sessions,
        }
    )


# ═══════════════════════════════════════════════════════════════════
# DEBUG: SIMULATED DATE API
# ═══════════════════════════════════════════════════════════════════

@app.route("/api/admin/simulated-date", methods=["GET", "POST", "DELETE"], endpoint="api_admin_simulated_date")
@login_required
def api_admin_simulated_date() -> ResponseReturnValue:
    """
    GET: Mostra la data simulata attuale
    POST: Imposta una data simulata (es: {"date": "2026-01-10"})
    DELETE: Rimuove la data simulata (usa data reale)
    """
    global SIMULATED_DATE
    
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403
    
    if request.method == "GET":
        return jsonify({
            "simulated_date": SIMULATED_DATE,
            "real_date": datetime.now().strftime("%Y-%m-%d"),
            "effective_date": get_simulated_today().strftime("%Y-%m-%d")
        })
    
    if request.method == "DELETE":
        old_date = SIMULATED_DATE
        SIMULATED_DATE = None
        app.logger.info(f"Data simulata rimossa (era: {old_date})")
        return jsonify({
            "success": True,
            "message": "Data simulata rimossa, usando data reale",
            "effective_date": get_simulated_today().strftime("%Y-%m-%d")
        })
    
    # POST: imposta nuova data simulata
    data = request.get_json() or {}
    new_date = data.get("date")
    
    if not new_date:
        return jsonify({"error": "Specifica 'date' nel formato YYYY-MM-DD"}), 400
    
    try:
        # Valida il formato
        datetime.strptime(new_date, "%Y-%m-%d")
        SIMULATED_DATE = new_date
        app.logger.info(f"Data simulata impostata a: {SIMULATED_DATE}")
        return jsonify({
            "success": True,
            "message": f"Data simulata impostata a {new_date}",
            "simulated_date": SIMULATED_DATE,
            "effective_date": get_simulated_today().strftime("%Y-%m-%d")
        })
    except ValueError:
        return jsonify({"error": "Formato data non valido, usa YYYY-MM-DD"}), 400


@app.get("/api/admin/open-sessions")
@login_required
def api_admin_open_sessions() -> ResponseReturnValue:
    """Restituisce le sessioni aperte (timer in corso) per tutti i progetti."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403

    db = get_db()
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    now = int(datetime.now(tz=timezone.utc).timestamp() * 1000)

    # Ottieni tutti i membri con sessioni attive (running o in pausa con attività)
    rows = db.execute(
        f"""
        SELECT ms.member_key, ms.member_name, ms.activity_id, ms.running, 
               ms.start_ts, ms.elapsed_cached, ms.pause_start, ms.entered_ts,
               ms.project_code,
               a.label AS activity_label,
               a.notes AS activity_notes
        FROM member_state ms
        LEFT JOIN activities a ON ms.activity_id = a.activity_id AND ms.project_code = a.project_code
        WHERE ms.activity_id IS NOT NULL
        ORDER BY ms.project_code, ms.member_name
        """
    ).fetchall()

    # Conta le pause per ogni membro/attività dal log
    pause_counts: Dict[str, int] = {}
    pause_rows = db.execute(
        """
        SELECT member_key, COUNT(*) as pause_count
        FROM event_log
        WHERE kind = 'pause_member'
        GROUP BY member_key
        """
    ).fetchall()
    for pr in pause_rows:
        pause_counts[pr["member_key"]] = pr["pause_count"]

    open_sessions = []
    for row in rows:
        elapsed = compute_elapsed(row, now)
        start_ts = row["entered_ts"] or row["start_ts"]
        running_state = int(row["running"])
        member_key = row["member_key"]
        
        open_sessions.append({
            "member_key": member_key,
            "member_name": row["member_name"],
            "activity_id": row["activity_id"],
            "activity_label": row["activity_label"] or row["activity_id"],
            "project_code": row["project_code"],
            "running": running_state == RUN_STATE_RUNNING,
            "paused": running_state == RUN_STATE_PAUSED or row["pause_start"] is not None,
            "start_ts": start_ts,
            "elapsed_ms": elapsed,
            "pause_count": pause_counts.get(member_key, 0),
            "notes": row["activity_notes"] or "",
        })

    return jsonify({
        "ok": True,
        "open_sessions": open_sessions,
        "count": len(open_sessions),
    })


@app.route("/api/admin/day-sessions", methods=["GET"], endpoint="api_admin_day_sessions")
@login_required
def api_admin_day_sessions() -> ResponseReturnValue:
    """Restituisce sessioni di squadra e magazzino per il range di date indicato."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403

    # Supporto range date: date_start/date_end oppure date singola (retrocompatibilità)
    date_start = parse_iso_date(request.args.get("date_start")) or parse_iso_date(request.args.get("date")) or datetime.now().date()
    date_end = parse_iso_date(request.args.get("date_end")) or date_start
    project_filter = request.args.get("project") or None
    
    # Assicura che date_end >= date_start
    if date_end < date_start:
        date_start, date_end = date_end, date_start

    start_dt = datetime.combine(date_start, datetime.min.time())
    end_dt = datetime.combine(date_end, datetime.min.time()) + timedelta(days=1)
    start_ms = int(start_dt.timestamp() * 1000)
    end_ms = int(end_dt.timestamp() * 1000)

    db = get_db()
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"

    team_sessions = build_session_rows(
        db,
        start_date=date_start,
        end_date=date_end,
        project_filter=project_filter,
    )

    ensure_warehouse_sessions_table(db)
    
    # Query magazzino con filtro progetto opzionale
    if project_filter:
        wh_rows = db.execute(
            f"""
            SELECT project_code, activity_label, elapsed_ms, username, created_ts, note, start_ts, end_ts
            FROM warehouse_sessions
            WHERE created_ts >= {placeholder} AND created_ts < {placeholder} AND project_code = {placeholder}
            ORDER BY created_ts DESC
            LIMIT 500
            """,
            (start_ms, end_ms, project_filter),
        ).fetchall()
    else:
        wh_rows = db.execute(
            f"""
            SELECT project_code, activity_label, elapsed_ms, username, created_ts, note, start_ts, end_ts
            FROM warehouse_sessions
            WHERE created_ts >= {placeholder} AND created_ts < {placeholder}
            ORDER BY created_ts DESC
            LIMIT 500
            """,
            (start_ms, end_ms),
        ).fetchall()

    magazzino_sessions = [
        {
            "project_code": row["project_code"],
            "activity_label": row["activity_label"],
            "elapsed_ms": _coerce_int(row["elapsed_ms"]) or 0,
            "username": row["username"],
            "created_ts": _coerce_int(row["created_ts"]) or 0,
            "note": row["note"] or "",
            "start_ts": _coerce_int(row["start_ts"]) if row["start_ts"] else None,
            "end_ts": _coerce_int(row["end_ts"]) if row["end_ts"] else None,
        }
        for row in wh_rows or []
    ]

    # Calcola totali
    team_total_ms = sum(_coerce_int(s.get("net_ms")) or 0 for s in team_sessions)
    magazzino_total_ms = sum(_coerce_int(s.get("elapsed_ms")) or 0 for s in magazzino_sessions)
    combined_total_ms = team_total_ms + magazzino_total_ms

    # Calcola ore pianificate per il progetto da Rentman (filtrate per data)
    planned_total_ms = 0
    activity_breakdown = []
    if project_filter:
        # Recupera ore pianificate da rentman_plannings per la data selezionata
        planned_rows = db.execute(
            f"SELECT SUM(hours_planned) as total_hours FROM rentman_plannings WHERE project_code = {placeholder} AND planning_date = {placeholder} AND is_obsolete = 0",
            (project_filter, date_start.isoformat())
        ).fetchone()
        if planned_rows and planned_rows["total_hours"]:
            # Converti ore in millisecondi
            planned_total_ms = int(float(planned_rows["total_hours"]) * 3600000)
        
        # Calcola distribuzione ore per attività (squadra + magazzino)
        activity_hours: Dict[str, int] = {}
        
        # Aggiungi ore squadra
        for s in team_sessions:
            if s.get("status") == "completed":
                act_label = s.get("activity_label") or s.get("activity_id") or "Altro"
                activity_hours[act_label] = activity_hours.get(act_label, 0) + (_coerce_int(s.get("net_ms")) or 0)
        
        # Aggiungi ore magazzino
        for s in magazzino_sessions:
            act_label = s.get("activity_label") or "Magazzino"
            activity_hours[act_label] = activity_hours.get(act_label, 0) + (_coerce_int(s.get("elapsed_ms")) or 0)
        
        # Converti in lista per il frontend (usa combined_total_ms per percentuali)
        for label, ms in sorted(activity_hours.items(), key=lambda x: -x[1]):
            activity_breakdown.append({
                "label": label,
                "ms": ms,
                "percent": round((ms / combined_total_ms * 100) if combined_total_ms > 0 else 0, 1)
            })

    return jsonify(
        {
            "ok": True,
            "date": date_start.isoformat() if date_start == date_end else f"{date_start.isoformat()} - {date_end.isoformat()}",
            "project_filter": project_filter,
            "team_sessions": team_sessions,
            "magazzino_sessions": magazzino_sessions,
            "team_total_ms": team_total_ms,
            "magazzino_total_ms": magazzino_total_ms,
            "combined_total_ms": combined_total_ms,
            "planned_total_ms": planned_total_ms,
            "activity_breakdown": activity_breakdown,
        }
    )


@app.get("/api/admin/projects-list")
@login_required
def api_admin_projects_list() -> ResponseReturnValue:
    """Restituisce la lista dei progetti con sessioni recenti."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403

    db = get_db()
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Ottieni progetti da event_log (ultimi 30 giorni)
    thirty_days_ago = int((datetime.now() - timedelta(days=30)).timestamp() * 1000)
    
    projects_set = set()
    
    # Da event_log
    rows = db.execute(
        f"SELECT DISTINCT JSON_EXTRACT(details, '$.project_code') as project_code FROM event_log WHERE ts > {placeholder} AND details IS NOT NULL",
        (thirty_days_ago,)
    ).fetchall()
    for row in rows:
        pc = row["project_code"]
        if pc and pc != "null" and pc.strip('"'):
            projects_set.add(pc.strip('"'))
    
    # Da warehouse_sessions
    wh_rows = db.execute(
        f"SELECT DISTINCT project_code FROM warehouse_sessions WHERE created_ts > {placeholder}",
        (thirty_days_ago,)
    ).fetchall()
    for row in wh_rows:
        if row["project_code"]:
            projects_set.add(row["project_code"])
    
    # Da activities
    act_rows = db.execute("SELECT DISTINCT project_code FROM activities").fetchall()
    for row in act_rows:
        if row["project_code"]:
            projects_set.add(row["project_code"])
    
    projects = sorted([p for p in projects_set if p])
    
    return jsonify({
        "ok": True,
        "projects": projects,
    })


@app.route(
    "/api/admin/day-sessions/export.xlsx",
    methods=["GET"],
    endpoint="api_admin_day_sessions_export_xlsx",
)
@login_required
def api_admin_day_sessions_export_xlsx() -> ResponseReturnValue:
    """Esporta in Excel le sessioni (Squadra + Magazzino) per il range di date indicato."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403

    # Supporto range date: date_start/date_end oppure date singola (retrocompatibilità)
    date_start = parse_iso_date(request.args.get("date_start")) or parse_iso_date(request.args.get("date")) or datetime.now().date()
    date_end = parse_iso_date(request.args.get("date_end")) or date_start
    
    # Assicura che date_end >= date_start
    if date_end < date_start:
        date_start, date_end = date_end, date_start

    start_dt = datetime.combine(date_start, datetime.min.time())
    end_dt = datetime.combine(date_end, datetime.min.time()) + timedelta(days=1)
    start_ms = int(start_dt.timestamp() * 1000)
    end_ms = int(end_dt.timestamp() * 1000)

    db = get_db()

    team_sessions = build_session_rows(
        db,
        start_date=date_start,
        end_date=date_end,
    )

    ensure_warehouse_sessions_table(db)
    wh_rows = db.execute(
        """
        SELECT project_code, activity_label, elapsed_ms, username, created_ts
        FROM warehouse_sessions
        WHERE created_ts >= ? AND created_ts < ?
        ORDER BY created_ts DESC
        LIMIT 2000
        """,
        (start_ms, end_ms),
    ).fetchall()

    merged_rows: List[Dict[str, Any]] = []
    for s in team_sessions or []:
        # Estrai data dalla sessione (usa timezone locale)
        start_ts = s.get("start_ts") or s.get("end_ts")
        if start_ts:
            try:
                dt = datetime.fromtimestamp(start_ts / 1000)  # Timezone locale
                date_str = dt.strftime("%d/%m/%Y")
            except Exception:
                date_str = ""
        else:
            date_str = ""
        merged_rows.append(
            {
                "date": date_str,
                "source": "Squadra",
                "project_code": s.get("project_code") or "",
                "user": s.get("member_name") or s.get("member_key") or "",
                "activity": s.get("activity_label") or s.get("activity_id") or "",
                "duration_ms": _coerce_int(s.get("net_ms")) or 0,
                "sort_ts": start_ts or 0,
            }
        )

    for row in wh_rows or []:
        # Estrai data dalla sessione magazzino (usa timezone locale)
        created_ts = _coerce_int(row.get("created_ts")) or 0
        if created_ts:
            try:
                dt = datetime.fromtimestamp(created_ts / 1000)  # Timezone locale
                date_str = dt.strftime("%d/%m/%Y")
            except Exception:
                date_str = ""
        else:
            date_str = ""
        merged_rows.append(
            {
                "date": date_str,
                "source": "Magazzino",
                "project_code": row.get("project_code") or "",
                "user": row.get("username") or "",
                "activity": row.get("activity_label") or "",
                "duration_ms": _coerce_int(row.get("elapsed_ms")) or 0,
                "sort_ts": created_ts,
            }
        )
    
    # Ordina per data/timestamp (più recenti prima)
    merged_rows.sort(key=lambda x: x.get("sort_ts", 0), reverse=True)

    wb = Workbook()
    ws_raw = wb.active
    if ws_raw is None:  # pragma: no cover
        ws_raw = wb.create_sheet(title="Sessioni")
    ws: Worksheet = cast(Worksheet, ws_raw)
    ws.title = "Sessioni"

    title_font = Font(name="Calibri", size=16, bold=True, color="1E293B")
    subtitle_font = Font(name="Calibri", size=11, color="64748B")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Calibri", size=11)
    cell_alignment = Alignment(horizontal="left", vertical="center")
    border_thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    ws["A1"] = "JobLog - Export sessioni"
    ws.merge_cells("A1:F1")
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws["A2"] = f"Data: {date_start.strftime('%d/%m/%Y')}" if date_start == date_end else f"Periodo: {date_start.strftime('%d/%m/%Y')} - {date_end.strftime('%d/%m/%Y')}"
    ws.merge_cells("A2:F2")
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.append([])

    headers = ["Data", "Fonte", "Progetto", "Utente", "Attività", "Ore"]
    ws.append(headers)
    header_row = ws.max_row
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_thin

    for item in merged_rows:
        ws.append(
            [
                item["date"],
                item["source"],
                str(item["project_code"] or ""),
                str(item["user"] or ""),
                str(item["activity"] or ""),
                format_duration_ms(item["duration_ms"]) or "00:00:00",
            ]
        )
        row_num = ws.max_row
        for col_num in range(1, 7):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border_thin
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    ws.column_dimensions[get_column_letter(1)].width = 12
    ws.column_dimensions[get_column_letter(2)].width = 12
    ws.column_dimensions[get_column_letter(3)].width = 14
    ws.column_dimensions[get_column_letter(4)].width = 22
    ws.column_dimensions[get_column_letter(5)].width = 42
    ws.column_dimensions[get_column_letter(6)].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"sessioni_{date_start.isoformat()}.xlsx" if date_start == date_end else f"sessioni_{date_start.isoformat()}_{date_end.isoformat()}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.get("/api/admin/sessions")
@login_required
def api_admin_sessions():
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403

    db = get_db()

    start_date = parse_iso_date(request.args.get("start_date", ""))
    end_date = parse_iso_date(request.args.get("end_date", ""))
    member_filter = request.args.get("member")
    activity_filter = request.args.get("activity_id")
    search_term = (request.args.get("search") or "").strip().lower()

    limit_param = request.args.get("limit")
    try:
        limit = int(limit_param) if limit_param else 200
    except (TypeError, ValueError):
        limit = 200
    limit = max(50, min(limit, 2000))

    # --- Sessioni squadra ---
    team_sessions = build_session_rows(
        db,
        start_date=start_date,
        end_date=end_date,
        member_filter=member_filter,
        activity_filter=activity_filter,
    )

    # --- Sessioni magazzino ---
    ensure_warehouse_sessions_table(db)
    wh_conditions: List[str] = []
    wh_params: List[Any] = []
    if start_date:
        start_dt = datetime.combine(start_date, datetime.min.time())
        wh_conditions.append("created_ts >= ?")
        wh_params.append(int(start_dt.timestamp() * 1000))
    if end_date:
        end_dt = datetime.combine(end_date, datetime.min.time()) + timedelta(days=1)
        wh_conditions.append("created_ts < ?")
        wh_params.append(int(end_dt.timestamp() * 1000))
    if member_filter:
        wh_conditions.append("username = ?")
        wh_params.append(member_filter)
    wh_where = (" WHERE " + " AND ".join(wh_conditions)) if wh_conditions else ""
    wh_rows = db.execute(
        f"""
        SELECT id, project_code, activity_label, elapsed_ms, username, created_ts
        FROM warehouse_sessions
        {wh_where}
        ORDER BY created_ts DESC
        LIMIT 1000
        """,
        tuple(wh_params),
    ).fetchall()

    # Unisci e filtra
    all_sessions: List[Dict[str, Any]] = []

    for item in team_sessions:
        all_sessions.append({**item, "_source": "Squadra", "_sort_ts": item.get("end_ts") or item.get("start_ts") or 0})

    for row in wh_rows or []:
        created_ts = _coerce_int(row["created_ts"]) or 0
        elapsed_ms = _coerce_int(row["elapsed_ms"]) or 0
        all_sessions.append({
            "_source": "Magazzino",
            "_sort_ts": created_ts,
            "member_key": row["username"] or "",
            "member_name": row["username"] or "",
            "activity_id": "",
            "activity_label": row["activity_label"] or "",
            "project_code": row["project_code"] or "",
            "start_ts": created_ts - elapsed_ms if elapsed_ms else created_ts,
            "end_ts": created_ts,
            "status": "completed",
            "net_ms": elapsed_ms,
            "pause_ms": 0,
            "pause_count": 0,
            "auto_closed": False,
            "override_id": None,
            "manual_entry": False,
            "note": "",
            "wh_id": row["id"],
        })

    # Filtra per search_term
    if search_term:
        filtered: List[Dict[str, Any]] = []
        for item in all_sessions:
            haystacks = [
                item.get("member_name", ""),
                item.get("member_key", ""),
                item.get("activity_label", ""),
                item.get("activity_id", ""),
                item.get("project_code", ""),
            ]
            if any(search_term in str(value).lower() for value in haystacks):
                filtered.append(item)
        all_sessions = filtered

    # Ordina per timestamp decrescente
    all_sessions.sort(key=lambda x: x.get("_sort_ts") or 0, reverse=True)
    all_sessions = all_sessions[:limit]

    payload = []
    for item in all_sessions:
        start_ts = item.get("start_ts") or 0
        end_ts = item.get("end_ts") or 0
        start_dt = datetime.fromtimestamp(start_ts / 1000, tz=timezone.utc) if start_ts else None
        end_dt = datetime.fromtimestamp(end_ts / 1000, tz=timezone.utc) if end_ts else None
        payload.append(
            {
                "source": item.get("_source", "Squadra"),
                "member_key": item.get("member_key", ""),
                "member_name": item.get("member_name", ""),
                "activity_id": item.get("activity_id", ""),
                "activity_label": item.get("activity_label", ""),
                "project_code": item.get("project_code", ""),
                "start_ts": start_ts,
                "end_ts": end_ts,
                "start_iso": start_dt.isoformat() if start_dt else "",
                "end_iso": end_dt.isoformat() if end_dt else "",
                "status": item.get("status", "completed"),
                "net_ms": item.get("net_ms", 0),
                "pause_ms": item.get("pause_ms", 0),
                "pause_count": item.get("pause_count", 0),
                "net_hms": format_duration_ms(item.get("net_ms")) or "00:00:00",
                "pause_hms": format_duration_ms(item.get("pause_ms")) or "00:00:00",
                "auto_closed": item.get("auto_closed", False),
                "override_id": item.get("override_id"),
                "manual_entry": bool(item.get("manual_entry")),
                "note": item.get("note") or "",
                "source_member_key": item.get("source_member_key"),
                "source_activity_id": item.get("source_activity_id"),
                "source_start_ts": item.get("source_start_ts"),
                "editable": item.get("_source") == "Squadra",
                "wh_id": item.get("wh_id"),
            }
        )

    return jsonify(
        {
            "sessions": payload,
            "count": len(payload),
            "limit": limit,
            "generated_at": now_ms(),
        }
    )


def _admin_or_supervisor() -> Optional[ResponseReturnValue]:
    """Guard per endpoint accessibili a admin e supervisor."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403
    return None


def _admin_only() -> Optional[ResponseReturnValue]:
    """Guard per endpoint accessibili solo ad admin."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403
    return None


def _json_error(message: str, status: int = 400) -> ResponseReturnValue:
    return jsonify({"error": message}), status


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


@app.post("/api/admin/sessions/save")
@login_required
def api_admin_sessions_save() -> ResponseReturnValue:
    guard = _admin_or_supervisor()
    if guard is not None:
        return guard

    data = request.get_json(silent=True) or {}
    source = _normalize_text(data.get("source")) or "Squadra"
    project_code = _normalize_text(data.get("project_code")) or ""
    member_key = _normalize_text(data.get("member_key"))
    member_name = _normalize_text(data.get("member_name")) or member_key or "Operatore"
    activity_id = _normalize_text(data.get("activity_id"))
    activity_label = _normalize_text(data.get("activity_label")) or activity_id or "Attività"
    if not member_key:
        return _json_error("member_key (ID operatore / Username) è obbligatorio")
    if not activity_label:
        return _json_error("activity_label (Descrizione attività) è obbligatoria")

    start_ts = _normalize_epoch_ms(data.get("start_ts"))
    if start_ts is None:
        return _json_error("start_ts non valido")

    end_ts_raw = data.get("end_ts")
    end_ts_value = _normalize_epoch_ms(end_ts_raw) if end_ts_raw is not None else None
    if end_ts_value is not None and end_ts_value < start_ts:
        return _json_error("end_ts deve essere successivo a start_ts")

    net_ms = _normalize_epoch_ms(data.get("net_ms"))
    if net_ms is None:
        if end_ts_value is not None:
            net_ms = max(0, end_ts_value - start_ts)
        else:
            net_ms = 0

    pause_ms = _normalize_epoch_ms(data.get("pause_ms")) or 0
    pause_count = max(0, _coerce_int(data.get("pause_count")) or 0)
    note = _normalize_text(data.get("note"))
    override_id = _coerce_int(data.get("override_id"))

    db = get_db()
    user = session.get("user") or "admin"
    now = now_ms()

    # Sessione MAGAZZINO: salva in warehouse_sessions
    if source == "Magazzino":
        ensure_warehouse_sessions_table(db)
        created_ts = end_ts_value if end_ts_value else start_ts + net_ms
        db.execute(
            """
            INSERT INTO warehouse_sessions(project_code, activity_label, elapsed_ms, username, created_ts)
            VALUES(?,?,?,?,?)
            """,
            (project_code, activity_label, net_ms, member_key, created_ts),
        )
        db.commit()
        return jsonify({"ok": True, "source": "Magazzino"})

    # Sessione SQUADRA: salva in activity_session_overrides
    if not activity_id:
        activity_id = activity_label

    source_member_key = _normalize_text(data.get("source_member_key")) or None
    source_activity_id = _normalize_text(data.get("source_activity_id")) or None
    source_start_ts = data.get("source_start_ts")
    source_start_ms = (
        _normalize_epoch_ms(source_start_ts) if source_start_ts is not None else None
    )

    manual_entry = bool(data.get("manual_entry"))
    if not manual_entry and not (source_member_key and source_activity_id and source_start_ms):
        manual_entry = True

    status = "completed" if end_ts_value is not None else "running"
    end_ts_final = end_ts_value if end_ts_value is not None else start_ts

    ensure_session_override_table(db)

    params = (
        member_key,
        member_name,
        activity_id,
        activity_label,
        project_code,
        start_ts,
        end_ts_final,
        net_ms,
        pause_ms,
        pause_count,
        status,
        source_member_key,
        source_activity_id,
        source_start_ms,
        1 if manual_entry else 0,
        note or None,
        user,
        user,
        now,
        now,
    )

    if override_id:
        existing = db.execute(
            "SELECT id FROM activity_session_overrides WHERE id=?",
            (override_id,),
        ).fetchone()
        if not existing:
            return _json_error("override non trovato", 404)
        db.execute(
            """
            UPDATE activity_session_overrides
            SET member_key=?, member_name=?, activity_id=?, activity_label=?, project_code=?,
                start_ts=?, end_ts=?, net_ms=?, pause_ms=?, pause_count=?, status=?,
                source_member_key=?, source_activity_id=?, source_start_ts=?,
                manual_entry=?, note=?, updated_by=?, updated_ts=?
            WHERE id=?
            """,
            (
                member_key,
                member_name,
                activity_id,
                activity_label,
                project_code,
                start_ts,
                end_ts_final,
                net_ms,
                pause_ms,
                pause_count,
                status,
                source_member_key,
                source_activity_id,
                source_start_ms,
                1 if manual_entry else 0,
                note or None,
                user,
                now,
                override_id,
            ),
        )
    else:
        db.execute(
            """
            INSERT INTO activity_session_overrides(
                member_key, member_name, activity_id, activity_label, project_code,
                start_ts, end_ts, net_ms, pause_ms, pause_count, status,
                source_member_key, source_activity_id, source_start_ts,
                manual_entry, note, created_by, updated_by, created_ts, updated_ts
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            params,
        )
        override_id = _last_insert_id(db)

    db.commit()

    if not override_id:
        return jsonify({"ok": True})

    row = db.execute(
        "SELECT * FROM activity_session_overrides WHERE id=?",
        (override_id,),
    ).fetchone()
    if not row:
        return jsonify({"ok": True})

    payload = _override_row_to_session(dict(row))
    return jsonify({"ok": True, "session": payload})


@app.delete("/api/admin/sessions/<int:override_id>")
@login_required
def api_admin_sessions_delete(override_id: int) -> ResponseReturnValue:
    guard = _admin_only()
    if guard is not None:
        return guard

    db = get_db()
    ensure_session_override_table(db)
    existing = db.execute(
        "SELECT id FROM activity_session_overrides WHERE id=?",
        (override_id,),
    ).fetchone()
    if not existing:
        return _json_error("override non trovato", 404)

    db.execute("DELETE FROM activity_session_overrides WHERE id=?", (override_id,))
    db.commit()
    return jsonify({"ok": True})


@app.post("/api/_reset")
@login_required
def api_reset():
    db = get_db()
    seed_demo_data(db)
    db.commit()
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
#  ANALISI ATTIVITÀ CROSS-PROGETTO
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/admin/activity-analysis")
@login_required
def admin_activity_analysis_page() -> ResponseReturnValue:
    """Pagina analisi attività cross-progetto."""
    if not is_admin_or_supervisor():
        abort(403)

    display_name = session.get("user_display") or session.get("user_name") or session.get("user")
    primary_name = session.get("user_name") or display_name or session.get("user")
    initials = session.get("user_initials") or compute_initials(primary_name or "")

    return render_template(
        "admin_activity_analysis.html",
        user_name=primary_name,
        user_display=display_name,
        user_initials=initials,
        is_admin=bool(session.get("is_admin")),
    )


@app.get("/api/admin/activity-analysis")
@login_required
def api_admin_activity_analysis() -> ResponseReturnValue:
    """API per analisi cross-progetto delle attività."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403

    mode = request.args.get("mode", "list")  # 'list' o 'analysis'
    date_start = parse_iso_date(request.args.get("date_start")) or (datetime.now().date() - timedelta(days=30))
    date_end = parse_iso_date(request.args.get("date_end")) or datetime.now().date()

    if date_end < date_start:
        date_start, date_end = date_end, date_start

    start_dt = datetime.combine(date_start, datetime.min.time())
    end_dt = datetime.combine(date_end, datetime.min.time()) + timedelta(days=1)
    start_ms = int(start_dt.timestamp() * 1000)
    end_ms = int(end_dt.timestamp() * 1000)

    db = get_db()

    # Raccogli sessioni da squadra
    team_sessions = build_session_rows(db, start_date=date_start, end_date=date_end)

    # Raccogli sessioni magazzino
    ensure_warehouse_sessions_table(db)
    wh_rows = db.execute(
        """
        SELECT project_code, activity_label, elapsed_ms, username, created_ts
        FROM warehouse_sessions
        WHERE created_ts >= ? AND created_ts < ?
        """,
        (start_ms, end_ms),
    ).fetchall()

    # Unisci tutte le sessioni
    all_sessions: List[Dict[str, Any]] = []

    for s in team_sessions:
        all_sessions.append({
            "project": s.get("project_code") or "N/A",
            "activity": s.get("activity_label") or s.get("activity_id") or "N/A",
            "duration_ms": _coerce_int(s.get("net_ms")) or 0,
            "source": "Squadra"
        })

    for row in wh_rows or []:
        all_sessions.append({
            "project": row["project_code"] or "N/A",
            "activity": row["activity_label"] or "N/A",
            "duration_ms": _coerce_int(row["elapsed_ms"]) or 0,
            "source": "Magazzino"
        })

    if mode == "list":
        # Restituisci lista attività con statistiche aggregate
        activity_stats: Dict[str, Dict[str, Any]] = {}
        for s in all_sessions:
            act = s["activity"]
            if act not in activity_stats:
                activity_stats[act] = {
                    "name": act,
                    "total_ms": 0,
                    "sessions": 0,
                    "projects": set()
                }
            activity_stats[act]["total_ms"] += s["duration_ms"]
            activity_stats[act]["sessions"] += 1
            activity_stats[act]["projects"].add(s["project"])

        activities = sorted(
            [
                {
                    "name": v["name"],
                    "total_ms": v["total_ms"],
                    "sessions": v["sessions"],
                    "projects": len(v["projects"])
                }
                for v in activity_stats.values()
            ],
            key=lambda x: x["total_ms"],
            reverse=True
        )

        return jsonify({
            "ok": True,
            "activities": activities,
            "date_start": date_start.isoformat(),
            "date_end": date_end.isoformat()
        })

    # mode == 'analysis': analisi dettagliata per attività selezionate
    selected_activities = request.args.getlist("activity")
    if not selected_activities:
        return jsonify({"error": "Nessuna attività selezionata"}), 400

    # Filtra sessioni per attività selezionate
    filtered = [s for s in all_sessions if s["activity"] in selected_activities]

    # Raggruppa per progetto e attività
    matrix: Dict[str, Dict[str, Dict[str, Any]]] = {}
    projects_set: Set[str] = set()
    details: List[Dict[str, Any]] = []

    # Raccogli dati per matrice
    grouped: Dict[str, Dict[str, List[int]]] = {}
    for s in filtered:
        proj = s["project"]
        act = s["activity"]
        projects_set.add(proj)

        if proj not in grouped:
            grouped[proj] = {}
        if act not in grouped[proj]:
            grouped[proj][act] = []
        grouped[proj][act].append(s["duration_ms"])

    # Calcola statistiche per ogni combinazione progetto/attività
    total_ms = 0
    total_sessions = 0

    for proj, acts in grouped.items():
        if proj not in matrix:
            matrix[proj] = {}
        for act, durations in acts.items():
            n = len(durations)
            tot = sum(durations)
            avg = tot / n if n > 0 else 0
            min_d = min(durations) if durations else 0
            max_d = max(durations) if durations else 0

            # Calcola varianza percentuale (coefficiente di variazione)
            if avg > 0 and n > 1:
                variance = sum((d - avg) ** 2 for d in durations) / n
                std_dev = variance ** 0.5
                cv_pct = (std_dev / avg) * 100
            else:
                cv_pct = 0

            matrix[proj][act] = {
                "total_ms": tot,
                "sessions": n,
                "avg_ms": int(avg),
                "min_ms": min_d,
                "max_ms": max_d
            }

            details.append({
                "project": proj,
                "activity": act,
                "sessions": n,
                "total_ms": tot,
                "avg_ms": int(avg),
                "min_ms": min_d,
                "max_ms": max_d,
                "variance_pct": round(cv_pct, 1)
            })

            total_ms += tot
            total_sessions += n

    # Ordina dettagli per ore totali decrescenti
    details.sort(key=lambda x: x["total_ms"], reverse=True)

    return jsonify({
        "ok": True,
        "projects": sorted(projects_set),
        "activities": selected_activities,
        "matrix": matrix,
        "details": details,
        "total_ms": total_ms,
        "total_sessions": total_sessions,
        "date_start": date_start.isoformat(),
        "date_end": date_end.isoformat()
    })


@app.get("/api/admin/activity-analysis/export.xlsx")
@login_required
def api_admin_activity_analysis_export() -> ResponseReturnValue:
    """Esporta analisi attività in Excel."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403

    date_start = parse_iso_date(request.args.get("date_start")) or (datetime.now().date() - timedelta(days=30))
    date_end = parse_iso_date(request.args.get("date_end")) or datetime.now().date()
    selected_activities = request.args.getlist("activity")

    if not selected_activities:
        return jsonify({"error": "Nessuna attività selezionata"}), 400

    # Riusa la logica dell'API
    if date_end < date_start:
        date_start, date_end = date_end, date_start

    start_dt = datetime.combine(date_start, datetime.min.time())
    end_dt = datetime.combine(date_end, datetime.min.time()) + timedelta(days=1)
    start_ms = int(start_dt.timestamp() * 1000)
    end_ms = int(end_dt.timestamp() * 1000)

    db = get_db()
    team_sessions = build_session_rows(db, start_date=date_start, end_date=date_end)

    ensure_warehouse_sessions_table(db)
    wh_rows = db.execute(
        """
        SELECT project_code, activity_label, elapsed_ms
        FROM warehouse_sessions
        WHERE created_ts >= ? AND created_ts < ?
        """,
        (start_ms, end_ms),
    ).fetchall()

    all_sessions: List[Dict[str, Any]] = []
    for s in team_sessions:
        all_sessions.append({
            "project": s.get("project_code") or "N/A",
            "activity": s.get("activity_label") or "N/A",
            "duration_ms": _coerce_int(s.get("net_ms")) or 0,
        })
    for row in wh_rows or []:
        all_sessions.append({
            "project": row["project_code"] or "N/A",
            "activity": row["activity_label"] or "N/A",
            "duration_ms": _coerce_int(row["elapsed_ms"]) or 0,
        })

    filtered = [s for s in all_sessions if s["activity"] in selected_activities]

    # Raggruppa e calcola statistiche
    grouped: Dict[str, Dict[str, List[int]]] = {}
    for s in filtered:
        proj, act = s["project"], s["activity"]
        if proj not in grouped:
            grouped[proj] = {}
        if act not in grouped[proj]:
            grouped[proj][act] = []
        grouped[proj][act].append(s["duration_ms"])

    rows_data: List[Dict[str, Any]] = []
    for proj, acts in grouped.items():
        for act, durations in acts.items():
            n = len(durations)
            tot = sum(durations)
            avg = tot / n if n > 0 else 0
            rows_data.append({
                "project": proj,
                "activity": act,
                "sessions": n,
                "total_ms": tot,
                "avg_ms": int(avg),
                "min_ms": min(durations) if durations else 0,
                "max_ms": max(durations) if durations else 0,
            })

    rows_data.sort(key=lambda x: x["total_ms"], reverse=True)

    # Genera Excel
    wb = Workbook()
    ws_raw = wb.active
    if ws_raw is None:
        ws_raw = wb.create_sheet(title="Analisi Attività")
    ws: Worksheet = cast(Worksheet, ws_raw)
    ws.title = "Analisi Attività"

    title_font = Font(name="Calibri", size=16, bold=True, color="1E293B")
    subtitle_font = Font(name="Calibri", size=11, color="64748B")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Calibri", size=11)
    cell_alignment = Alignment(horizontal="left", vertical="center")
    border_thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    ws["A1"] = "🔬 JobLog - Analisi Attività Cross-Progetto"
    ws.merge_cells("A1:G1")
    ws["A1"].font = title_font

    period_str = f"Periodo: {date_start.strftime('%d/%m/%Y')} - {date_end.strftime('%d/%m/%Y')}"
    ws["A2"] = period_str
    ws.merge_cells("A2:G2")
    ws["A2"].font = subtitle_font

    ws["A3"] = f"Attività analizzate: {', '.join(selected_activities)}"
    ws.merge_cells("A3:G3")
    ws["A3"].font = subtitle_font

    ws.append([])

    headers = ["Progetto", "Attività", "Sessioni", "Ore Totali", "Media", "Min", "Max"]
    ws.append(headers)
    header_row = ws.max_row
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_thin

    for item in rows_data:
        ws.append([
            item["project"],
            item["activity"],
            item["sessions"],
            format_duration_ms(item["total_ms"]) or "00:00:00",
            format_duration_ms(item["avg_ms"]) or "00:00:00",
            format_duration_ms(item["min_ms"]) or "00:00:00",
            format_duration_ms(item["max_ms"]) or "00:00:00",
        ])
        row_num = ws.max_row
        for col_num in range(1, 8):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border_thin
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for col_letter, width in {"A": 18, "B": 30, "C": 12, "D": 14, "E": 12, "F": 12, "G": 12}.items():
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"analisi_attivita_{date_start.isoformat()}_{date_end.isoformat()}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# Registrazione lazy del worker: si assicura che il thread sia attivo al primo accesso
@app.before_request
def _ensure_notification_worker() -> None:
    if _NOTIFICATION_THREAD is None or not _NOTIFICATION_THREAD.is_alive():
        start_notification_worker()


atexit.register(stop_notification_worker)


# Registrazione lazy del worker CedolinoWeb per retry timbrate
@app.before_request
def _ensure_cedolino_retry_worker() -> None:
    if _CEDOLINO_RETRY_THREAD is None or not _CEDOLINO_RETRY_THREAD.is_alive():
        start_cedolino_retry_worker()


# ═══════════════════════════════════════════════════════════════════════════════
#  CREW MEMBERS - DATABASE (Operatori Rentman)
# ═══════════════════════════════════════════════════════════════════════════════

CREW_MEMBERS_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS crew_members (
    id INT AUTO_INCREMENT PRIMARY KEY,
    rentman_id INT NOT NULL UNIQUE,
    name VARCHAR(255) NOT NULL,
    external_id VARCHAR(255) DEFAULT NULL,
    external_group_id VARCHAR(255) DEFAULT NULL,
    group_id INT DEFAULT NULL COMMENT 'FK a user_groups per sede GPS',
    email VARCHAR(255) DEFAULT NULL,
    phone VARCHAR(50) DEFAULT NULL,
    is_active TINYINT(1) DEFAULT 1,
    created_ts BIGINT NOT NULL,
    updated_ts BIGINT NOT NULL,
    INDEX idx_crew_external (external_id),
    INDEX idx_crew_external_group (external_group_id),
    INDEX idx_crew_group (group_id),
    INDEX idx_crew_active (is_active)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

CREW_MEMBERS_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS crew_members (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    rentman_id INTEGER NOT NULL UNIQUE,
    name TEXT NOT NULL,
    external_id TEXT DEFAULT NULL,
    external_group_id TEXT DEFAULT NULL,
    group_id INTEGER DEFAULT NULL,
    email TEXT DEFAULT NULL,
    phone TEXT DEFAULT NULL,
    is_active INTEGER DEFAULT 1,
    created_ts INTEGER NOT NULL,
    updated_ts INTEGER NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_crew_external ON crew_members(external_id);
CREATE INDEX IF NOT EXISTS idx_crew_external_group ON crew_members(external_group_id);
CREATE INDEX IF NOT EXISTS idx_crew_group ON crew_members(group_id);
CREATE INDEX IF NOT EXISTS idx_crew_active ON crew_members(is_active);
"""


def ensure_crew_members_table(db: DatabaseLike) -> None:
    """Crea la tabella crew_members se non esiste."""
    statement = (
        CREW_MEMBERS_TABLE_MYSQL if DB_VENDOR == "mysql" else CREW_MEMBERS_TABLE_SQLITE
    )
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        try:
            cursor = db.execute(sql)
            try:
                cursor.close()
            except AttributeError:
                pass
        except Exception:
            pass  # Tabella/indice già esistente
    
    # Migrazione: aggiunge colonna external_group_id se non esiste
    try:
        if DB_VENDOR == "mysql":
            # Verifica se la colonna esiste già
            check = db.execute(
                "SELECT COUNT(*) FROM information_schema.columns WHERE table_name='crew_members' AND column_name='external_group_id'"
            ).fetchone()
            col_exists = (check[0] if check else 0) > 0
            if not col_exists:
                db.execute("ALTER TABLE crew_members ADD COLUMN external_group_id VARCHAR(255) DEFAULT NULL AFTER external_id")
                db.commit()
                app.logger.info("Colonna external_group_id aggiunta a crew_members")
        else:
            # SQLite: verifica tramite PRAGMA
            cols = db.execute("PRAGMA table_info(crew_members)").fetchall()
            col_names = [c[1] for c in cols]
            if "external_group_id" not in col_names:
                db.execute("ALTER TABLE crew_members ADD COLUMN external_group_id TEXT DEFAULT NULL")
                db.commit()
                app.logger.info("Colonna external_group_id aggiunta a crew_members")
    except Exception as e:
        app.logger.warning("Migrazione external_group_id: %s", e)
    
    # Migrazione: aggiunge colonna timbratura_override per eccezioni per operatore
    try:
        if DB_VENDOR == "mysql":
            check = db.execute(
                "SELECT COUNT(*) FROM information_schema.columns WHERE table_name='crew_members' AND column_name='timbratura_override'"
            ).fetchone()
            col_exists = (check[0] if check else 0) > 0
            if not col_exists:
                db.execute("ALTER TABLE crew_members ADD COLUMN timbratura_override TEXT DEFAULT NULL")
                db.commit()
                app.logger.info("Colonna timbratura_override aggiunta a crew_members")
        else:
            cols = db.execute("PRAGMA table_info(crew_members)").fetchall()
            col_names = [c[1] for c in cols]
            if "timbratura_override" not in col_names:
                db.execute("ALTER TABLE crew_members ADD COLUMN timbratura_override TEXT DEFAULT NULL")
                db.commit()
                app.logger.info("Colonna timbratura_override aggiunta a crew_members")
    except Exception as e:
        app.logger.warning("Migrazione timbratura_override: %s", e)
    
    # Migrazione: aggiunge colonna group_id per collegamento a user_groups (sede GPS)
    try:
        if DB_VENDOR == "mysql":
            check = db.execute(
                "SELECT COUNT(*) FROM information_schema.columns WHERE table_name='crew_members' AND column_name='group_id'"
            ).fetchone()
            col_exists = (check[0] if check else 0) > 0
            if not col_exists:
                db.execute("ALTER TABLE crew_members ADD COLUMN group_id INT DEFAULT NULL COMMENT 'FK a user_groups per sede GPS'")
                db.execute("CREATE INDEX idx_crew_group ON crew_members(group_id)")
                db.commit()
                app.logger.info("Colonna group_id aggiunta a crew_members")
        else:
            cols = db.execute("PRAGMA table_info(crew_members)").fetchall()
            col_names = [c[1] for c in cols]
            if "group_id" not in col_names:
                db.execute("ALTER TABLE crew_members ADD COLUMN group_id INTEGER DEFAULT NULL")
                db.execute("CREATE INDEX IF NOT EXISTS idx_crew_group ON crew_members(group_id)")
                db.commit()
                app.logger.info("Colonna group_id aggiunta a crew_members")
    except Exception as e:
        app.logger.warning("Migrazione group_id: %s", e)


def sync_crew_member_from_rentman(db: DatabaseLike, rentman_id: int, name: str) -> None:
    """Sincronizza un operatore da Rentman nel database locale (insert or update name)."""
    now = now_ms()
    if DB_VENDOR == "mysql":
        existing = db.execute(
            "SELECT id FROM crew_members WHERE rentman_id = %s", (rentman_id,)
        ).fetchone()
        if existing:
            db.execute(
                "UPDATE crew_members SET name = %s, updated_ts = %s WHERE rentman_id = %s",
                (name, now, rentman_id)
            )
        else:
            db.execute(
                "INSERT INTO crew_members (rentman_id, name, created_ts, updated_ts) VALUES (%s, %s, %s, %s)",
                (rentman_id, name, now, now)
            )
    else:
        existing = db.execute(
            "SELECT id FROM crew_members WHERE rentman_id = ?", (rentman_id,)
        ).fetchone()
        if existing:
            db.execute(
                "UPDATE crew_members SET name = ?, updated_ts = ? WHERE rentman_id = ?",
                (name, now, rentman_id)
            )
        else:
            db.execute(
                "INSERT INTO crew_members (rentman_id, name, created_ts, updated_ts) VALUES (?, ?, ?, ?)",
                (rentman_id, name, now, now)
            )


# ═══════════════════════════════════════════════════════════════════════════════
#  CEDOLINO WEB - Funzioni di integrazione
# ═══════════════════════════════════════════════════════════════════════════════

def _get_pending_overtime_request_id(db: DatabaseLike, username: str, date_str: str) -> Optional[int]:
    """
    Verifica se esiste una richiesta di Extra Turno pending per un utente in una data specifica.
    
    Args:
        db: connessione database
        username: username dell'utente
        date_str: data in formato YYYY-MM-DD
    
    Returns:
        ID della richiesta Extra Turno pending, o None se non esiste
    """
    try:
        # Ottieni l'ID del tipo "Extra Turno"
        overtime_type_id = get_overtime_request_type_id(db)
        placeholder = "%s" if DB_VENDOR == "mysql" else "?"
        
        app.logger.info(f"Checking pending overtime for user={username}, date={date_str}, type_id={overtime_type_id}")
        
        row = db.execute(f"""
            SELECT id FROM user_requests 
            WHERE username = {placeholder} 
              AND request_type_id = {placeholder} 
              AND date_from = {placeholder} 
              AND status = 'pending'
            ORDER BY created_ts DESC
            LIMIT 1
        """, (username, overtime_type_id, date_str)).fetchone()
        
        if row:
            result_id = row['id'] if isinstance(row, Mapping) else row[0]
            app.logger.info(f"Found pending overtime request id={result_id} for {username} on {date_str}")
            return result_id
        
        app.logger.info(f"No pending overtime found for {username} on {date_str}")
        return None
    except Exception as e:
        app.logger.warning(f"Errore verifica Extra Turno pending: {e}")
        return None


def ensure_cedolino_timbrature_table(db: DatabaseLike) -> None:
    """Crea la tabella cedolino_timbrature se non esiste."""
    statement = (
        CEDOLINO_TIMBRATURE_TABLE_MYSQL if DB_VENDOR == "mysql" else CEDOLINO_TIMBRATURE_TABLE_SQLITE
    )
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        try:
            cursor = db.execute(sql)
            try:
                cursor.close()
            except AttributeError:
                pass
        except Exception:
            pass
    
    # Migrazioni: aggiungi colonne se non esistono
    if DB_VENDOR == "mysql":
        migrations = [
            "ALTER TABLE cedolino_timbrature ADD COLUMN username VARCHAR(190) DEFAULT NULL",
            "ALTER TABLE cedolino_timbrature ADD COLUMN ora_originale TIME DEFAULT NULL",
            "ALTER TABLE cedolino_timbrature ADD COLUMN ora_modificata TIME DEFAULT NULL",
            "ALTER TABLE cedolino_timbrature ADD COLUMN data_riferimento DATE DEFAULT NULL",
            "ALTER TABLE cedolino_timbrature ADD COLUMN overtime_request_id INT DEFAULT NULL COMMENT 'ID richiesta straordinario collegata'",
            "CREATE INDEX idx_cedolino_overtime ON cedolino_timbrature(overtime_request_id)",
        ]
        for migration in migrations:
            try:
                db.execute(migration)
                db.commit()
            except Exception:
                pass  # Colonna già esistente
    else:
        migrations = [
            "ALTER TABLE cedolino_timbrature ADD COLUMN username TEXT",
            "ALTER TABLE cedolino_timbrature ADD COLUMN ora_originale TEXT",
            "ALTER TABLE cedolino_timbrature ADD COLUMN ora_modificata TEXT",
            "ALTER TABLE cedolino_timbrature ADD COLUMN data_riferimento TEXT",
            "ALTER TABLE cedolino_timbrature ADD COLUMN overtime_request_id INTEGER DEFAULT NULL",
        ]
        for migration in migrations:
            try:
                db.execute(migration)
                db.commit()
            except Exception:
                pass
        # Indice separato per SQLite
        try:
            db.execute("CREATE INDEX IF NOT EXISTS idx_cedolino_overtime ON cedolino_timbrature(overtime_request_id)")
            db.commit()
        except Exception:
            pass


def get_cedolino_settings() -> Optional[Dict[str, Any]]:
    """
    Restituisce le impostazioni CedolinoWeb dal config.json.
    La sincronizzazione è attiva solo se:
    1. config.json ha cedolino_web.enabled = true
    2. company_settings.custom_settings.cedolino_sync_enabled = true (se configurato)
    """
    config = load_config()
    section = config.get("cedolino_web")
    if not section or not isinstance(section, dict):
        return None
    if not section.get("enabled"):
        return None
    
    # Verifica anche l'impostazione nel database (se presente)
    try:
        db = get_db()
        settings = get_company_settings(db)
        custom_settings = settings.get("custom_settings", {})
        # Se il flag esiste nel database e è False, disabilita la sincronizzazione
        if "cedolino_sync_enabled" in custom_settings:
            if not custom_settings.get("cedolino_sync_enabled"):
                return None
    except Exception as e:
        # Se c'è un errore nel database, usa solo il config.json
        app.logger.warning(f"Errore verifica cedolino_sync_enabled da DB: {e}")
    
    return section


def get_external_id_for_member(db: DatabaseLike, member_key: str) -> Optional[str]:
    """
    Recupera l'external_id (ID CedolinoWeb) per un operatore dato il member_key.
    Il member_key ha formato 'rentman-crew-{rentman_id}'.
    """
    if not member_key:
        return None
    
    # Estrai rentman_id dal member_key
    if not member_key.startswith("rentman-crew-"):
        app.logger.warning("CedolinoWeb: member_key non valido (formato atteso: rentman-crew-ID): %s", member_key)
        return None
    
    try:
        rentman_id = int(member_key.replace("rentman-crew-", ""))
    except ValueError:
        app.logger.warning("CedolinoWeb: impossibile estrarre rentman_id da %s", member_key)
        return None
    
    # Cerca l'external_id nella tabella crew_members
    if DB_VENDOR == "mysql":
        row = db.execute(
            "SELECT external_id FROM crew_members WHERE rentman_id = %s",
            (rentman_id,)
        ).fetchone()
    else:
        row = db.execute(
            "SELECT external_id FROM crew_members WHERE rentman_id = ?",
            (rentman_id,)
        ).fetchone()
    
    if not row:
        app.logger.debug("CedolinoWeb: nessun operatore trovato per rentman_id %s", rentman_id)
        return None
    
    external_id = row["external_id"] if isinstance(row, dict) else row[0]
    return external_id if external_id else None


def get_username_from_member_key(db: DatabaseLike, member_key: str) -> Optional[str]:
    """
    Recupera lo username dell'utente associato a un member_key.
    Il member_key ha formato 'rentman-crew-{rentman_id}' dove rentman_id è l'ID
    dell'assignment Rentman (non il crew_id).
    
    Il mapping avviene in due passi:
    1. Cerca il crew_id dalla tabella rentman_plannings tramite rentman_id
    2. Cerca lo username dalla tabella app_users tramite rentman_crew_id = crew_id
    
    Args:
        db: connessione database
        member_key: chiave operatore (es. rentman-crew-1903)
    
    Returns:
        username se trovato, None altrimenti
    """
    if not member_key:
        return None
    
    # Estrai rentman_id (assignment_id) dal member_key
    if not member_key.startswith("rentman-crew-"):
        app.logger.debug("get_username_from_member_key: member_key non valido (formato atteso: rentman-crew-ID): %s", member_key)
        return None
    
    try:
        rentman_id = int(member_key.replace("rentman-crew-", ""))
    except ValueError:
        app.logger.warning("get_username_from_member_key: impossibile estrarre rentman_id da %s", member_key)
        return None
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Step 1: Cerca il crew_id dalla tabella rentman_plannings tramite rentman_id
    planning_row = db.execute(
        f"SELECT crew_id FROM rentman_plannings WHERE rentman_id = {placeholder} LIMIT 1",
        (rentman_id,)
    ).fetchone()
    
    if not planning_row:
        app.logger.debug("get_username_from_member_key: nessun planning trovato per rentman_id %s", rentman_id)
        return None
    
    crew_id = planning_row["crew_id"] if isinstance(planning_row, dict) else planning_row[0]
    if not crew_id:
        app.logger.debug("get_username_from_member_key: crew_id nullo per rentman_id %s", rentman_id)
        return None
    
    # Step 2: Cerca lo username dalla tabella app_users tramite rentman_crew_id = crew_id
    user_row = db.execute(
        f"SELECT username FROM app_users WHERE rentman_crew_id = {placeholder} AND is_active = 1",
        (crew_id,)
    ).fetchone()
    
    if not user_row:
        app.logger.debug("get_username_from_member_key: nessun utente trovato per crew_id %s (rentman_id=%s)", crew_id, rentman_id)
        return None
    
    username = user_row["username"] if isinstance(user_row, dict) else user_row[0]
    app.logger.debug("get_username_from_member_key: trovato username '%s' per member_key '%s' (rentman_id=%s, crew_id=%s)", 
                     username, member_key, rentman_id, crew_id)
    return username


def create_supervisor_pause_timbratura(
    db: DatabaseLike,
    username: str,
    tipo: str,
    member_name: Optional[str] = None,
    supervisor_username: Optional[str] = None
) -> bool:
    """
    Genera una timbratura di pausa (inizio_pausa o fine_pausa) quando il supervisor
    mette in pausa o riprende un operatore.
    
    Args:
        db: connessione database
        username: username dell'operatore
        tipo: 'inizio_pausa' o 'fine_pausa'
        member_name: nome visualizzato dell'operatore (per logging)
        supervisor_username: username del supervisor che ha generato la timbratura
    
    Returns:
        True se la timbratura è stata creata, False altrimenti
    """
    if not username:
        return False
    
    if tipo not in ('inizio_pausa', 'fine_pausa'):
        app.logger.error("create_supervisor_pause_timbratura: tipo non valido: %s", tipo)
        return False
    
    # Data e ora correnti
    today = datetime.now().strftime("%Y-%m-%d")
    ora = datetime.now().strftime("%H:%M:%S")
    created_ts = now_ms()
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Calcola ora_mod in base alle regole
    ora_mod = ora
    
    # Recupera le regole di timbratura dell'utente
    try:
        rules = get_user_timbratura_rules(db, username)
    except Exception as e:
        app.logger.warning(f"Errore recupero regole timbratura: {e}, uso default")
        rules = {
            'pausa_blocco_minimo_minuti': 30,
            'pausa_incremento_minuti': 15,
            'pausa_tolleranza_minuti': 5
        }
    
    # Per fine_pausa, calcola ora_mod basato sulla durata della pausa
    if tipo == 'fine_pausa':
        try:
            # Recupera l'ora di inizio pausa (l'ultima non chiusa)
            inizio_pausa_row = db.execute(
                f"""SELECT ora, ora_mod FROM timbrature 
                   WHERE username = {placeholder} AND data = {placeholder} AND tipo = 'inizio_pausa'
                   ORDER BY created_ts DESC LIMIT 1""",
                (username, today)
            ).fetchone()
            
            if inizio_pausa_row:
                inizio_ora_mod = inizio_pausa_row['ora_mod'] if isinstance(inizio_pausa_row, dict) else inizio_pausa_row[1]
                if not inizio_ora_mod:
                    inizio_ora_mod = inizio_pausa_row['ora'] if isinstance(inizio_pausa_row, dict) else inizio_pausa_row[0]
                
                # Formatta l'ora di inizio pausa
                if hasattr(inizio_ora_mod, 'strftime'):
                    inizio_str = inizio_ora_mod.strftime("%H:%M")
                elif hasattr(inizio_ora_mod, 'total_seconds'):
                    # È un timedelta (MySQL TIME restituisce timedelta)
                    total_sec = int(inizio_ora_mod.total_seconds())
                    inizio_str = f"{total_sec // 3600:02d}:{(total_sec % 3600) // 60:02d}"
                else:
                    inizio_str = str(inizio_ora_mod)[:5]
                
                # Calcola durata modificata usando le regole pause
                durata_mod = calcola_pausa_mod(inizio_str, ora[:5], rules)
                
                # Calcola ora_mod di fine_pausa = inizio_pausa_mod + durata_mod
                inizio_parts = inizio_str.split(':')
                inizio_min = int(inizio_parts[0]) * 60 + int(inizio_parts[1])
                fine_mod_min = inizio_min + durata_mod
                
                h = fine_mod_min // 60
                m = fine_mod_min % 60
                ora_mod = f"{h:02d}:{m:02d}:00"
                
                # Calcola durata effettiva per il log
                ora_parts = ora[:5].split(':')
                ora_min = int(ora_parts[0]) * 60 + int(ora_parts[1])
                durata_effettiva = ora_min - inizio_min
                
                app.logger.info(
                    f"Pausa supervisor: {inizio_str} -> {ora[:5]} (durata effettiva {durata_effettiva} min, durata mod {durata_mod} min, fine mod {ora_mod})"
                )
        except Exception as e:
            app.logger.warning(f"Errore calcolo ora_mod pausa: {e}, uso ora originale")
            ora_mod = ora
    
    # Inserisce la timbratura di pausa
    try:
        db.execute(
            f"""
            INSERT INTO timbrature (username, tipo, data, ora, ora_mod, created_ts, method, created_by)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
            """,
            (username, tipo, today, ora, ora_mod, created_ts, "supervisor", supervisor_username)
        )
        
        app.logger.info(
            "Timbratura %s generata da supervisor per %s (%s) alle %s (ora_mod: %s)",
            tipo, username, member_name or username, ora, ora_mod
        )
        
        # Invia a CedolinoWeb
        timeframe_id = TIMEFRAME_INIZIO_PAUSA if tipo == 'inizio_pausa' else TIMEFRAME_FINE_PAUSA
        
        # Recupera display_name per CedolinoWeb
        user_row = db.execute(
            f"SELECT display_name FROM app_users WHERE username = {placeholder}",
            (username,)
        ).fetchone()
        display_name = username
        if user_row:
            display_name = (user_row['display_name'] if isinstance(user_row, dict) else user_row[0]) or username
        
        send_timbrata_utente(
            db,
            username=username,
            member_name=display_name,
            timeframe_id=timeframe_id,
            data_riferimento=today,
            ora_originale=ora,
            ora_modificata=ora_mod,
        )
        
        return True
        
    except Exception as e:
        app.logger.error("create_supervisor_pause_timbratura: errore inserimento timbratura: %s", e)
        return False


def get_external_id_for_username(db: DatabaseLike, username: str, return_reason: bool = False):
    """
    Recupera l'external_id (ID CedolinoWeb) per un utente dato il suo username.
    
    L'external_id viene cercato SOLO nella tabella app_users.
    
    Args:
        db: connessione database
        username: username dell'utente
        return_reason: se True, ritorna anche il motivo in caso di errore
    
    Returns:
        Se return_reason=False: external_id o None
        Se return_reason=True: (external_id, reason) dove reason spiega il problema
    """
    if not username:
        if return_reason:
            return None, "Username non fornito"
        return None
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Recupera external_id dalla tabella app_users
    user_row = db.execute(
        f"SELECT external_id FROM app_users WHERE username = {placeholder}",
        (username,)
    ).fetchone()
    
    if not user_row:
        app.logger.debug("CedolinoWeb: utente %s non trovato", username)
        if return_reason:
            return None, "Utente non trovato nel database"
        return None
    
    # Estrai il valore
    if isinstance(user_row, dict):
        external_id = user_row.get('external_id')
    else:
        external_id = user_row[0]
    
    if external_id:
        app.logger.debug("CedolinoWeb: utente %s ha external_id: %s", username, external_id)
        if return_reason:
            return external_id, ""
        return external_id
    
    # Nessun external_id configurato
    app.logger.debug("CedolinoWeb: utente %s non ha external_id configurato", username)
    if return_reason:
        return None, "Utente non ha l'ID Esterno CedolinoWeb configurato. Vai in Gestione Utenti e inserisci l'ID Esterno."
    return None


def get_external_group_id_for_username(db: DatabaseLike, username: str) -> Optional[str]:
    """
    Recupera l'external_group_id (Gruppo ID CedolinoWeb) per un utente dato il suo username.
    
    Ordine di ricerca:
    1. Prima cerca external_group_id direttamente in app_users (override utente)
    2. Se non trovato, cerca cedolino_group_id dalla tabella user_groups tramite group_id
    """
    if not username:
        return None
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Prima cerca external_group_id diretto dall'utente E il group_id per fallback
    user_row = db.execute(
        f"SELECT external_group_id, group_id FROM app_users WHERE username = {placeholder}",
        (username,)
    ).fetchone()
    
    if not user_row:
        return None
    
    # Estrai i valori
    if isinstance(user_row, dict):
        external_group_id = user_row.get('external_group_id')
        group_id = user_row.get('group_id')
    else:
        external_group_id = user_row[0]
        group_id = user_row[1]
    
    # Se l'utente ha un external_group_id diretto, usalo
    if external_group_id:
        app.logger.info(f"CedolinoWeb gruppo: utente {username} ha external_group_id diretto: {external_group_id}")
        return external_group_id
    
    # Altrimenti cerca il cedolino_group_id dal gruppo associato
    if group_id:
        group_row = db.execute(
            f"SELECT cedolino_group_id FROM user_groups WHERE id = {placeholder}",
            (group_id,)
        ).fetchone()
        
        if group_row:
            if isinstance(group_row, dict):
                cedolino_group_id = group_row.get('cedolino_group_id')
            else:
                cedolino_group_id = group_row[0]
            
            if cedolino_group_id:
                app.logger.info(f"CedolinoWeb gruppo: utente {username} usa cedolino_group_id dal gruppo: {cedolino_group_id}")
                return cedolino_group_id
    
    app.logger.warning(f"CedolinoWeb gruppo: utente {username} non ha gruppo_id associato")
    return None


# ═══════════════════════════════════════════════════════════════════
# TEST PAYLOAD - Imposta a True per vedere il payload completo nei log
# ═══════════════════════════════════════════════════════════════════
CEDOLINO_TEST_PAYLOAD = False  # <-- Cambia a True per attivare


def call_cedolino_webservice(
    external_id: str,
    timeframe_id: int,
    data_riferimento: str,
    data_originale: str,
    data_modificata: str,
    endpoint: Optional[str] = None,
    gruppo_id: Optional[str] = None
) -> Tuple[bool, Optional[str], str]:
    """
    Chiama il webservice CedolinoWeb per registrare una timbrata.
    
    Args:
        external_id: ID esterno dell'operatore (codice_utente e assunzione_id)
        timeframe_id: tipo di timbrata (1=inizio, 3=pausa, 4=fine pausa, 8=fine)
        data_riferimento: data di riferimento (formato YYYY-MM-DD)
        data_originale: data/ora originale (formato YYYY-MM-DD HH:MM:SS)
        data_modificata: data/ora modificata (formato YYYY-MM-DD HH:MM:SS)
        endpoint: URL del webservice (default: CEDOLINO_WEB_ENDPOINT)
        gruppo_id: ID del gruppo esterno (default: NULL)
    
    Returns:
        Tuple (success: bool, error_message: Optional[str], request_url: str)
    """
    import requests
    
    if not endpoint:
        endpoint = CEDOLINO_WEB_ENDPOINT
    
    params = {
        "data_riferimento": data_riferimento,
        "data_originale": data_originale,
        "data_modificata": data_modificata,
        "codice_utente": external_id,
        "codice_terminale": CEDOLINO_CODICE_TERMINALE,
        "timeframe_id": str(timeframe_id),
        "assunzione_id": external_id,
        "terminale_id": "NULL",
        "gruppo_id": gruppo_id if gruppo_id else "NULL",
        "turno_id": "NULL",
        "note": "",
        "validata": "true",
    }
    
    # Costruisci URL completo per debug
    from urllib.parse import urlencode
    full_url = f"{endpoint}?{urlencode(params)}"
    
    # ═══════════════════════════════════════════════════════════════════
    # TEST PAYLOAD LOG - Mostra payload dettagliato
    # ═══════════════════════════════════════════════════════════════════
    if CEDOLINO_TEST_PAYLOAD:
        app.logger.info("=" * 80)
        app.logger.info("🧪 CEDOLINO TEST PAYLOAD 🧪")
        app.logger.info("=" * 80)
        app.logger.info("ENDPOINT: %s", endpoint)
        app.logger.info("-" * 40)
        app.logger.info("PAYLOAD PARAMETERS:")
        for key, value in params.items():
            app.logger.info("  %-20s : %s", key, value)
        app.logger.info("-" * 40)
        app.logger.info("FULL URL: %s", full_url)
        app.logger.info("=" * 80)
    else:
        app.logger.info("CedolinoWeb REQUEST URL: %s", full_url)
    
    try:
        app.logger.info(
            "CedolinoWeb: invio timbrata per %s, timeframe=%s, originale=%s, modificata=%s",
            external_id, timeframe_id, data_originale, data_modificata
        )
        response = requests.get(endpoint, params=params, timeout=30)
        
        # Log della risposta
        app.logger.info("CedolinoWeb RESPONSE: status=%s, body=%s", response.status_code, response.text[:500] if response.text else "(vuoto)")
        
        if response.status_code == 200:
            app.logger.info("CedolinoWeb: timbrata registrata con successo per %s", external_id)
            return True, None, full_url
        else:
            error_msg = f"HTTP {response.status_code}: {response.text[:200]}"
            app.logger.warning("CedolinoWeb: errore HTTP - %s", error_msg)
            return False, error_msg, full_url
            
    except requests.Timeout:
        error_msg = "Timeout durante la connessione"
        app.logger.warning("CedolinoWeb: %s", error_msg)
        return False, error_msg, full_url
    except requests.RequestException as e:
        error_msg = f"Errore di rete: {str(e)}"
        app.logger.warning("CedolinoWeb: %s", error_msg)
        return False, error_msg, full_url
    except Exception as e:
        error_msg = f"Errore imprevisto: {str(e)}"
        app.logger.exception("CedolinoWeb: errore imprevisto")
        return False, error_msg, full_url


def send_timbrata(
    db: DatabaseLike,
    member_key: str,
    member_name: str,
    timeframe_id: int,
    timestamp_ms: Optional[int] = None,
    project_code: Optional[str] = None,
    activity_id: Optional[str] = None,
) -> Tuple[bool, Optional[str], Optional[str]]:
    """
    Registra una timbrata operatore e tenta l'invio a CedolinoWeb.
    Per le timbrature operatori usa lo stesso orario per originale e modificato.
    
    Args:
        db: connessione database
        member_key: chiave operatore (es. rentman-crew-1656)
        member_name: nome operatore
        timeframe_id: tipo timbrata (1, 3, 4, 8)
        timestamp_ms: timestamp (default: now)
        project_code: codice progetto corrente
        activity_id: ID attività corrente
    
    Returns:
        Tuple (success: bool, external_id: Optional[str], error: Optional[str])
        - success=True se timbrata inviata o CedolinoWeb disabilitato
        - external_id=None se operatore non ha ID esterno
        - error=messaggio se fallito
    """
    settings = get_cedolino_settings()
    
    # Se CedolinoWeb non è configurato/abilitato, ritorna OK silenziosamente
    if not settings:
        return True, None, None
    
    if timestamp_ms is None:
        timestamp_ms = now_ms()
    
    # Recupera external_id
    external_id = get_external_id_for_member(db, member_key)
    if not external_id:
        # Operatore senza ID esterno - blocca l'operazione
        return False, None, "Operatore senza ID esterno CedolinoWeb"
    
    # Calcola data_riferimento e ora
    dt = datetime.fromtimestamp(timestamp_ms / 1000.0)
    data_riferimento = dt.strftime("%Y-%m-%d")
    ora = dt.strftime("%H:%M:%S")
    data_ora_completa = f"{data_riferimento} {ora}"
    
    # Assicurati che la tabella esista
    ensure_cedolino_timbrature_table(db)
    
    now = now_ms()
    
    # Salva la timbrata nel database
    if DB_VENDOR == "mysql":
        db.execute(
            """
            INSERT INTO cedolino_timbrature 
            (member_key, member_name, username, external_id, timeframe_id, timestamp_ms, 
             data_riferimento, ora_originale, ora_modificata, project_code, activity_id, created_ts)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (member_key, member_name, None, external_id, timeframe_id, timestamp_ms,
             data_riferimento, ora, ora, project_code, activity_id, now)
        )
    else:
        db.execute(
            """
            INSERT INTO cedolino_timbrature 
            (member_key, member_name, username, external_id, timeframe_id, timestamp_ms, 
             data_riferimento, ora_originale, ora_modificata, project_code, activity_id, created_ts)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (member_key, member_name, None, external_id, timeframe_id, timestamp_ms,
             data_riferimento, ora, ora, project_code, activity_id, now)
        )
    
    # Recupera l'ID appena inserito
    timbrata_id = _last_insert_id(db)
    
    # Tenta l'invio al webservice
    endpoint = settings.get("endpoint") or CEDOLINO_WEB_ENDPOINT
    success, error, _url = call_cedolino_webservice(
        external_id, timeframe_id, data_riferimento, data_ora_completa, data_ora_completa, endpoint
    )
    
    # Aggiorna lo stato di sincronizzazione
    if success:
        if DB_VENDOR == "mysql":
            db.execute(
                "UPDATE cedolino_timbrature SET synced_ts = %s WHERE id = %s",
                (now_ms(), timbrata_id)
            )
        else:
            db.execute(
                "UPDATE cedolino_timbrature SET synced_ts = ? WHERE id = ?",
                (now_ms(), timbrata_id)
            )
    else:
        if DB_VENDOR == "mysql":
            db.execute(
                "UPDATE cedolino_timbrature SET sync_error = %s, sync_attempts = 1 WHERE id = %s",
                (error, timbrata_id)
            )
        else:
            db.execute(
                "UPDATE cedolino_timbrature SET sync_error = ?, sync_attempts = 1 WHERE id = ?",
                (error, timbrata_id)
            )
    
    return success, external_id, error


def send_timbrata_utente(
    db: DatabaseLike,
    username: str,
    member_name: str,
    timeframe_id: int,
    data_riferimento: str,
    ora_originale: str,
    ora_modificata: str,
    overtime_request_id: Optional[int] = None,
) -> Tuple[bool, Optional[str], Optional[str], Optional[str]]:
    """
    Registra una timbrata utente e tenta l'invio a CedolinoWeb.
    Usa ora_originale e ora_modificata dalla tabella timbrature.
    
    Se overtime_request_id è specificato, la timbrata viene salvata ma NON sincronizzata
    fino a quando la richiesta di straordinario non viene revisionata.
    
    Args:
        db: connessione database
        username: username dell'utente
        member_name: nome operatore/utente
        timeframe_id: tipo timbrata (1, 3, 4, 8)
        data_riferimento: data della timbrata (YYYY-MM-DD)
        ora_originale: orario reale della timbrata (HH:MM:SS)
        ora_modificata: orario modificato/arrotondato (HH:MM:SS)
        overtime_request_id: ID richiesta straordinario (se presente, blocca sincronizzazione)
    
    Returns:
        Tuple (success: bool, external_id: Optional[str], error: Optional[str], request_url: Optional[str])
        - success=True se timbrata inviata o CedolinoWeb disabilitato o bloccata per straordinario
        - external_id=None se operatore non ha ID esterno
        - error=messaggio se fallito
        - request_url=URL completo della chiamata (per debug)
    """
    settings = get_cedolino_settings()
    
    # Se CedolinoWeb non è configurato/abilitato, ritorna OK silenziosamente
    if not settings:
        return True, None, None, None
    
    # Recupera external_id dall'username con motivo dettagliato
    external_id, reason = get_external_id_for_username(db, username, return_reason=True)
    if not external_id:
        # Utente senza ID esterno - blocca l'operazione
        return False, None, reason or "Utente senza ID esterno CedolinoWeb", None
    
    # Recupera external_group_id dall'username
    external_group_id = get_external_group_id_for_username(db, username)
    
    # Assicurati che ora_modificata abbia valore (fallback a ora_originale)
    if not ora_modificata:
        ora_modificata = ora_originale
    
    # Componi data_originale e data_modificata
    data_originale = f"{data_riferimento} {ora_originale}"
    data_modificata = f"{data_riferimento} {ora_modificata}"
    
    # Timestamp in millisecondi
    try:
        dt = datetime.strptime(f"{data_riferimento} {ora_originale}", "%Y-%m-%d %H:%M:%S")
        timestamp_ms = int(dt.timestamp() * 1000)
    except ValueError:
        # Prova con formato senza secondi
        try:
            dt = datetime.strptime(f"{data_riferimento} {ora_originale}", "%Y-%m-%d %H:%M")
            timestamp_ms = int(dt.timestamp() * 1000)
        except ValueError:
            timestamp_ms = now_ms()
    
    # Assicurati che la tabella esista
    ensure_cedolino_timbrature_table(db)
    
    now = now_ms()
    
    # Salva la timbrata nel database (con overtime_request_id se presente)
    if DB_VENDOR == "mysql":
        db.execute(
            """
            INSERT INTO cedolino_timbrature 
            (member_key, member_name, username, external_id, timeframe_id, timestamp_ms, 
             data_riferimento, ora_originale, ora_modificata, project_code, activity_id, 
             overtime_request_id, created_ts)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (None, member_name, username, external_id, timeframe_id, timestamp_ms,
             data_riferimento, ora_originale, ora_modificata, None, None, overtime_request_id, now)
        )
    else:
        db.execute(
            """
            INSERT INTO cedolino_timbrature 
            (member_key, member_name, username, external_id, timeframe_id, timestamp_ms, 
             data_riferimento, ora_originale, ora_modificata, project_code, activity_id, 
             overtime_request_id, created_ts)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (None, member_name, username, external_id, timeframe_id, timestamp_ms,
             data_riferimento, ora_originale, ora_modificata, None, None, overtime_request_id, now)
        )
    
    # Recupera l'ID appena inserito
    timbrata_id = _last_insert_id(db)
    
    # Se c'è un overtime_request_id, NON sincronizzare ora - sarà fatto dopo la revisione
    if overtime_request_id:
        app.logger.info(
            "CedolinoWeb: timbrata %s per %s bloccata in attesa revisione straordinario (request_id=%s)",
            timbrata_id, username, overtime_request_id
        )
        # Salva un messaggio informativo nell'errore
        if DB_VENDOR == "mysql":
            db.execute(
                "UPDATE cedolino_timbrature SET sync_error = %s WHERE id = %s",
                ("In attesa revisione straordinario", timbrata_id)
            )
        else:
            db.execute(
                "UPDATE cedolino_timbrature SET sync_error = ? WHERE id = ?",
                ("In attesa revisione straordinario", timbrata_id)
            )
        # Ritorna success=True perché la timbrata è stata salvata correttamente
        return True, external_id, None, None
    
    # Tenta l'invio al webservice
    endpoint = settings.get("endpoint") or CEDOLINO_WEB_ENDPOINT
    success, error, request_url = call_cedolino_webservice(
        external_id, timeframe_id, data_riferimento, data_originale, data_modificata, endpoint, external_group_id
    )
    
    # Aggiorna lo stato di sincronizzazione
    if success:
        if DB_VENDOR == "mysql":
            db.execute(
                "UPDATE cedolino_timbrature SET synced_ts = %s WHERE id = %s",
                (now_ms(), timbrata_id)
            )
        else:
            db.execute(
                "UPDATE cedolino_timbrature SET synced_ts = ? WHERE id = ?",
                (now_ms(), timbrata_id)
            )
    else:
        if DB_VENDOR == "mysql":
            db.execute(
                "UPDATE cedolino_timbrature SET sync_error = %s, sync_attempts = 1 WHERE id = %s",
                (error, timbrata_id)
            )
        else:
            db.execute(
                "UPDATE cedolino_timbrature SET sync_error = ?, sync_attempts = 1 WHERE id = ?",
                (error, timbrata_id)
            )
    
    return success, external_id, error, request_url


def retry_pending_timbrature(db: DatabaseLike, max_attempts: int = 5) -> int:
    """
    Ritenta l'invio delle timbrate non sincronizzate.
    Esclude le timbrature bloccate per straordinario in attesa di revisione.
    
    Args:
        db: connessione database
        max_attempts: numero massimo di tentativi
    
    Returns:
        Numero di timbrate sincronizzate con successo
    """
    # Assicura che la tabella esista con tutte le colonne (migrazioni)
    ensure_cedolino_timbrature_table(db)
    
    settings = get_cedolino_settings()
    if not settings:
        return 0
    
    endpoint = settings.get("endpoint") or CEDOLINO_WEB_ENDPOINT
    
    # Recupera timbrate non sincronizzate con tentativi < max
    # IMPORTANTE: Esclude TUTTE quelle con overtime_request_id - vengono gestite
    # esclusivamente da _sync_overtime_blocked_timbrature dopo la revisione dell'admin
    if DB_VENDOR == "mysql":
        rows = db.execute(
            """
            SELECT ct.id, ct.external_id, ct.timeframe_id, ct.data_riferimento, 
                   ct.ora_originale, ct.ora_modificata, ct.sync_attempts, ct.username
            FROM cedolino_timbrature ct
            WHERE ct.synced_ts IS NULL 
              AND ct.sync_attempts < %s
              AND ct.overtime_request_id IS NULL
            ORDER BY ct.created_ts ASC
            LIMIT 50
            """,
            (max_attempts,)
        ).fetchall()
    else:
        rows = db.execute(
            """
            SELECT ct.id, ct.external_id, ct.timeframe_id, ct.data_riferimento, 
                   ct.ora_originale, ct.ora_modificata, ct.sync_attempts, ct.username
            FROM cedolino_timbrature ct
            WHERE ct.synced_ts IS NULL 
              AND ct.sync_attempts < ?
              AND ct.overtime_request_id IS NULL
            ORDER BY ct.created_ts ASC
            LIMIT 50
            """,
            (max_attempts,)
        ).fetchall()
    
    synced_count = 0
    for row in rows:
        timbrata_id = row["id"] if isinstance(row, dict) else row[0]
        external_id = row["external_id"] if isinstance(row, dict) else row[1]
        timeframe_id = row["timeframe_id"] if isinstance(row, dict) else row[2]
        data_rif = row["data_riferimento"] if isinstance(row, dict) else row[3]
        ora_orig = row["ora_originale"] if isinstance(row, dict) else row[4]
        ora_mod = row["ora_modificata"] if isinstance(row, dict) else row[5]
        attempts = row["sync_attempts"] if isinstance(row, dict) else row[6]
        
        # Formatta data_riferimento come stringa se necessario
        if hasattr(data_rif, 'strftime'):
            data_riferimento = data_rif.strftime("%Y-%m-%d")
        else:
            data_riferimento = str(data_rif)
        
        # Formatta ora come stringa se necessario
        if hasattr(ora_orig, 'strftime'):
            ora_originale = ora_orig.strftime("%H:%M:%S")
        else:
            ora_originale = str(ora_orig)
        
        if hasattr(ora_mod, 'strftime'):
            ora_modificata = ora_mod.strftime("%H:%M:%S")
        else:
            ora_modificata = str(ora_mod) if ora_mod else ora_originale
        
        # Componi data_originale e data_modificata
        data_originale = f"{data_riferimento} {ora_originale}"
        data_modificata = f"{data_riferimento} {ora_modificata}"
        
        success, error, _url = call_cedolino_webservice(
            external_id, timeframe_id, data_riferimento, data_originale, data_modificata, endpoint
        )
        
        if success:
            if DB_VENDOR == "mysql":
                db.execute(
                    "UPDATE cedolino_timbrature SET synced_ts = %s, sync_error = NULL WHERE id = %s",
                    (now_ms(), timbrata_id)
                )
            else:
                db.execute(
                    "UPDATE cedolino_timbrature SET synced_ts = ?, sync_error = NULL WHERE id = ?",
                    (now_ms(), timbrata_id)
                )
            synced_count += 1
        else:
            if DB_VENDOR == "mysql":
                db.execute(
                    "UPDATE cedolino_timbrature SET sync_error = %s, sync_attempts = %s WHERE id = %s",
                    (error, attempts + 1, timbrata_id)
                )
            else:
                db.execute(
                    "UPDATE cedolino_timbrature SET sync_error = ?, sync_attempts = ? WHERE id = ?",
                    (error, attempts + 1, timbrata_id)
                )
    
    if rows:
        db.commit()
    
    return synced_count


def _cedolino_retry_worker(stop_event: Event) -> None:
    """Worker thread per ritentare l'invio delle timbrate CedolinoWeb non sincronizzate."""
    app.logger.info(
        "CedolinoWeb retry worker: avviato (intervallo %ss)", CEDOLINO_RETRY_INTERVAL_SECONDS
    )
    
    while not stop_event.is_set():
        try:
            with app.app_context():
                settings = get_cedolino_settings()
                if not settings:
                    app.logger.debug("CedolinoWeb retry worker: integrazione disabilitata")
                else:
                    max_attempts = settings.get("max_retry_attempts", 10)
                    db = get_db()
                    synced = retry_pending_timbrature(db, max_attempts)
                    if synced > 0:
                        app.logger.info("CedolinoWeb retry worker: sincronizzate %s timbrate", synced)
        except Exception as exc:
            app.logger.exception("CedolinoWeb retry worker: errore", exc_info=exc)
        finally:
            stop_event.wait(CEDOLINO_RETRY_INTERVAL_SECONDS)


def start_cedolino_retry_worker() -> None:
    """Avvia il worker per i retry CedolinoWeb."""
    global _CEDOLINO_RETRY_THREAD, _CEDOLINO_RETRY_STOP
    
    if _CEDOLINO_RETRY_THREAD and _CEDOLINO_RETRY_THREAD.is_alive():
        return
    
    _CEDOLINO_RETRY_STOP = Event()
    _CEDOLINO_RETRY_THREAD = Thread(
        target=_cedolino_retry_worker,
        args=(_CEDOLINO_RETRY_STOP,),
        name="joblog-cedolino-retry",
        daemon=True
    )
    _CEDOLINO_RETRY_THREAD.start()
    app.logger.info("CedolinoWeb retry worker: thread avviato")


def stop_cedolino_retry_worker() -> None:
    """Ferma il worker per i retry CedolinoWeb."""
    global _CEDOLINO_RETRY_THREAD, _CEDOLINO_RETRY_STOP
    
    stop_event = _CEDOLINO_RETRY_STOP
    thread = _CEDOLINO_RETRY_THREAD
    
    if stop_event is not None:
        stop_event.set()
    
    if thread and thread.is_alive():
        thread.join(timeout=5)
    
    _CEDOLINO_RETRY_THREAD = None
    _CEDOLINO_RETRY_STOP = None


atexit.register(stop_cedolino_retry_worker)


# ═══════════════════════════════════════════════════════════════════════════════
#  PIANIFICAZIONI RENTMAN - DATABASE
# ═══════════════════════════════════════════════════════════════════════════════

RENTMAN_PLANNINGS_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS rentman_plannings (
    id INT AUTO_INCREMENT PRIMARY KEY,
    rentman_id INT NOT NULL,
    planning_date DATE NOT NULL,
    crew_id INT,
    crew_name VARCHAR(255),
    function_id INT,
    function_name VARCHAR(255),
    project_id INT,
    project_name VARCHAR(500),
    project_code VARCHAR(128),
    subproject_id INT,
    location_id INT,
    location_name VARCHAR(500),
    location_address VARCHAR(500),
    location_lat DECIMAL(12,8) DEFAULT NULL COMMENT 'Latitudine location',
    location_lon DECIMAL(12,8) DEFAULT NULL COMMENT 'Longitudine location',
    timbratura_gps_mode VARCHAR(20) DEFAULT 'group',
    gps_timbratura_location VARCHAR(255),
    plan_start DATETIME,
    plan_end DATETIME,
    break_start TIME DEFAULT NULL COMMENT 'Inizio pausa',
    break_end TIME DEFAULT NULL COMMENT 'Fine pausa',
    break_minutes INT DEFAULT NULL COMMENT 'Durata pausa in minuti',
    hours_planned DECIMAL(10,2),
    hours_registered DECIMAL(10,2),
    remark TEXT,
    remark_planner TEXT,
    is_leader TINYINT(1) DEFAULT 0,
    transport VARCHAR(50),
    sent_to_webservice TINYINT(1) DEFAULT 0,
    sent_ts BIGINT DEFAULT NULL,
    webservice_response TEXT,
    is_obsolete TINYINT(1) DEFAULT 0 COMMENT 'Turno rimosso da Rentman',
    created_ts BIGINT NOT NULL,
    updated_ts BIGINT NOT NULL,
    UNIQUE KEY uniq_rentman_planning (rentman_id, planning_date),
    INDEX idx_planning_date (planning_date),
    INDEX idx_planning_crew (crew_id),
    INDEX idx_planning_project (project_code),
    INDEX idx_planning_sent (sent_to_webservice),
    INDEX idx_planning_obsolete (is_obsolete)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

RENTMAN_PLANNINGS_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS rentman_plannings (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    rentman_id INTEGER NOT NULL,
    planning_date TEXT NOT NULL,
    crew_id INTEGER,
    crew_name TEXT,
    function_id INTEGER,
    function_name TEXT,
    project_id INTEGER,
    project_name TEXT,
    project_code TEXT,
    subproject_id INTEGER,
    location_id INTEGER,
    location_name TEXT,
    location_address TEXT,
    location_lat REAL DEFAULT NULL,
    location_lon REAL DEFAULT NULL,
    timbratura_gps_mode TEXT DEFAULT 'group',
    gps_timbratura_location TEXT,
    plan_start TEXT,
    plan_end TEXT,
    break_start TEXT DEFAULT NULL,
    break_end TEXT DEFAULT NULL,
    break_minutes INTEGER DEFAULT NULL,
    hours_planned REAL,
    hours_registered REAL,
    remark TEXT,
    remark_planner TEXT,
    is_leader INTEGER DEFAULT 0,
    transport TEXT,
    sent_to_webservice INTEGER DEFAULT 0,
    sent_ts INTEGER DEFAULT NULL,
    webservice_response TEXT,
    is_obsolete INTEGER DEFAULT 0,
    created_ts INTEGER NOT NULL,
    updated_ts INTEGER NOT NULL,
    UNIQUE(rentman_id, planning_date)
);
CREATE INDEX IF NOT EXISTS idx_planning_date ON rentman_plannings(planning_date);
CREATE INDEX IF NOT EXISTS idx_planning_crew ON rentman_plannings(crew_id);
CREATE INDEX IF NOT EXISTS idx_planning_project ON rentman_plannings(project_code);
CREATE INDEX IF NOT EXISTS idx_planning_sent ON rentman_plannings(sent_to_webservice);
CREATE INDEX IF NOT EXISTS idx_planning_obsolete ON rentman_plannings(is_obsolete);
"""


# ═══════════════════════════════════════════════════════════════════════════════
#  REQUEST TYPES - TIPOLOGIE RICHIESTE (ferie, permessi, rimborsi, ecc.)
# ═══════════════════════════════════════════════════════════════════════════════

REQUEST_TYPES_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS request_types (
    id INT AUTO_INCREMENT PRIMARY KEY,
    name VARCHAR(100) NOT NULL,
    value_type ENUM('hours', 'days', 'amount', 'km', 'minutes', 'timbratura') NOT NULL,
    external_id VARCHAR(100),
    abbreviation VARCHAR(10),
    description TEXT,
    active TINYINT(1) DEFAULT 1,
    sort_order INT DEFAULT 0,
    is_giustificativo TINYINT(1) DEFAULT 0,
    created_ts BIGINT NOT NULL DEFAULT 0,
    updated_ts BIGINT NOT NULL DEFAULT 0,
    INDEX idx_request_type_active (active),
    INDEX idx_request_type_value (value_type)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

REQUEST_TYPES_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS request_types (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    value_type TEXT NOT NULL CHECK(value_type IN ('hours', 'days', 'amount', 'km', 'minutes', 'timbratura')),
    external_id TEXT,
    abbreviation TEXT,
    description TEXT,
    active INTEGER DEFAULT 1,
    sort_order INTEGER DEFAULT 0,
    is_giustificativo INTEGER DEFAULT 0,
    created_ts INTEGER NOT NULL,
    updated_ts INTEGER NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_request_type_active ON request_types(active);
CREATE INDEX IF NOT EXISTS idx_request_type_value ON request_types(value_type);
"""

# Tabella per le richieste degli utenti
USER_REQUESTS_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS user_requests (
    id INT AUTO_INCREMENT PRIMARY KEY,
    user_id INT NOT NULL,
    username VARCHAR(100) NOT NULL,
    request_type_id INT NOT NULL,
    date_from DATE NOT NULL,
    date_to DATE,
    value_amount DECIMAL(10,2) NOT NULL,
    notes TEXT,
    cdc VARCHAR(100),
    attachment_path VARCHAR(500),
    status ENUM('pending', 'approved', 'rejected') DEFAULT 'pending',
    reviewed_by VARCHAR(100),
    reviewed_ts BIGINT,
    review_notes TEXT,
    created_ts BIGINT NOT NULL DEFAULT 0,
    updated_ts BIGINT NOT NULL DEFAULT 0,
    INDEX idx_request_user (user_id),
    INDEX idx_request_username (username),
    INDEX idx_request_status (status),
    INDEX idx_request_date (date_from),
    INDEX idx_request_type (request_type_id),
    FOREIGN KEY (request_type_id) REFERENCES request_types(id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

USER_REQUESTS_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS user_requests (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    username TEXT NOT NULL,
    request_type_id INTEGER NOT NULL,
    date_from TEXT NOT NULL,
    date_to TEXT,
    value_amount REAL NOT NULL,
    notes TEXT,
    cdc TEXT,
    attachment_path TEXT,
    status TEXT DEFAULT 'pending' CHECK(status IN ('pending', 'approved', 'rejected')),
    reviewed_by TEXT,
    reviewed_ts INTEGER,
    review_notes TEXT,
    created_ts INTEGER NOT NULL,
    updated_ts INTEGER NOT NULL,
    FOREIGN KEY (request_type_id) REFERENCES request_types(id)
);
CREATE INDEX IF NOT EXISTS idx_request_user ON user_requests(user_id);
CREATE INDEX IF NOT EXISTS idx_request_username ON user_requests(username);
CREATE INDEX IF NOT EXISTS idx_request_status ON user_requests(status);
CREATE INDEX IF NOT EXISTS idx_request_date ON user_requests(date_from);
CREATE INDEX IF NOT EXISTS idx_request_type ON user_requests(request_type_id);
"""

# Tabella per i documenti aziendali (circolari, comunicazioni, buste paga)
USER_DOCUMENTS_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS user_documents (
    id INT AUTO_INCREMENT PRIMARY KEY,
    category ENUM('circolare', 'comunicazione', 'busta_paga') NOT NULL,
    title VARCHAR(255) NOT NULL,
    description TEXT,
    file_path VARCHAR(500),
    file_name VARCHAR(255),
    target_users JSON DEFAULT NULL,
    target_all BOOLEAN DEFAULT TRUE,
    created_by VARCHAR(100) NOT NULL,
    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    INDEX idx_doc_category (category),
    INDEX idx_doc_created (created_at)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

USER_DOCUMENTS_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS user_documents (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    category TEXT NOT NULL CHECK(category IN ('circolare', 'comunicazione', 'busta_paga')),
    title TEXT NOT NULL,
    description TEXT,
    file_path TEXT,
    file_name TEXT,
    target_users TEXT DEFAULT NULL,
    target_all INTEGER DEFAULT 1,
    created_by TEXT NOT NULL,
    created_at TEXT DEFAULT CURRENT_TIMESTAMP
);
CREATE INDEX IF NOT EXISTS idx_doc_category ON user_documents(category);
CREATE INDEX IF NOT EXISTS idx_doc_created ON user_documents(created_at);
"""

# Tabella per tracciare lettura documenti
USER_DOCUMENTS_READ_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS user_documents_read (
    id INT AUTO_INCREMENT PRIMARY KEY,
    document_id INT NOT NULL,
    username VARCHAR(100) NOT NULL,
    read_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    UNIQUE KEY unique_read (document_id, username),
    FOREIGN KEY (document_id) REFERENCES user_documents(id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

USER_DOCUMENTS_READ_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS user_documents_read (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    document_id INTEGER NOT NULL,
    username TEXT NOT NULL,
    read_at TEXT DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(document_id, username),
    FOREIGN KEY (document_id) REFERENCES user_documents(id) ON DELETE CASCADE
)
"""

# ═══════════════════════════════════════════════════════════════════════════════
# STRAORDINARI (OVERTIME) - Tabelle e funzioni
# ═══════════════════════════════════════════════════════════════════════════════

OVERTIME_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS overtime_requests (
    id INT AUTO_INCREMENT PRIMARY KEY,
    username VARCHAR(100) NOT NULL,
    date DATE NOT NULL,
    session_id INT,
    planning_id INT,
    shift_source ENUM('rentman', 'manual', 'none') DEFAULT 'none',
    planned_start TIME,
    planned_end TIME,
    actual_start TIME,
    actual_end TIME,
    extra_minutes_before INT DEFAULT 0,
    extra_minutes_after INT DEFAULT 0,
    total_extra_minutes INT NOT NULL,
    overtime_type ENUM('before_shift', 'after_shift', 'both', 'extra_day') DEFAULT 'after_shift',
    notes TEXT,
    status ENUM('pending', 'approved', 'rejected') DEFAULT 'pending',
    reviewed_by VARCHAR(100),
    reviewed_ts BIGINT,
    review_notes TEXT,
    created_ts BIGINT NOT NULL,
    updated_ts BIGINT NOT NULL,
    INDEX idx_overtime_user (username),
    INDEX idx_overtime_date (date),
    INDEX idx_overtime_status (status),
    INDEX idx_overtime_session (session_id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

OVERTIME_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS overtime_requests (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT NOT NULL,
    date TEXT NOT NULL,
    session_id INTEGER,
    planning_id INTEGER,
    shift_source TEXT DEFAULT 'none' CHECK(shift_source IN ('rentman', 'manual', 'none')),
    planned_start TEXT,
    planned_end TEXT,
    actual_start TEXT,
    actual_end TEXT,
    extra_minutes_before INTEGER DEFAULT 0,
    extra_minutes_after INTEGER DEFAULT 0,
    total_extra_minutes INTEGER NOT NULL,
    overtime_type TEXT DEFAULT 'after_shift' CHECK(overtime_type IN ('before_shift', 'after_shift', 'both', 'extra_day')),
    notes TEXT,
    status TEXT DEFAULT 'pending' CHECK(status IN ('pending', 'approved', 'rejected')),
    reviewed_by TEXT,
    reviewed_ts INTEGER,
    review_notes TEXT,
    created_ts INTEGER NOT NULL,
    updated_ts INTEGER NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_overtime_user ON overtime_requests(username);
CREATE INDEX IF NOT EXISTS idx_overtime_date ON overtime_requests(date);
CREATE INDEX IF NOT EXISTS idx_overtime_status ON overtime_requests(status);
CREATE INDEX IF NOT EXISTS idx_overtime_session ON overtime_requests(session_id);
"""


def ensure_overtime_table(db: DatabaseLike) -> None:
    """Crea la tabella overtime_requests se non esiste."""
    statement = (
        OVERTIME_TABLE_MYSQL if DB_VENDOR == "mysql" else OVERTIME_TABLE_SQLITE
    )
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass
    db.commit()


def _ensure_overtime_request_type(db: DatabaseLike) -> int:
    """
    Assicura che esista il tipo richiesta 'Extra Turno' e ritorna il suo ID.
    Questo tipo viene usato per le richieste di Extra Turno automatiche e manuali.
    
    Nota: Migra automaticamente il vecchio nome 'Straordinario' a 'Extra Turno'.
    """
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Prima cerca se esiste con il nuovo nome
    row = db.execute(
        f"SELECT id FROM request_types WHERE name = {placeholder}",
        ("Extra Turno",)
    ).fetchone()
    
    if row:
        return row["id"] if isinstance(row, Mapping) else row[0]
    
    # Cerca se esiste con il vecchio nome e rinominalo
    old_row = db.execute(
        f"SELECT id FROM request_types WHERE name = {placeholder}",
        ("Straordinario",)
    ).fetchone()
    
    if old_row:
        # Rinomina da "Straordinario" a "Extra Turno"
        old_id = old_row["id"] if isinstance(old_row, Mapping) else old_row[0]
        db.execute(f"""
            UPDATE request_types 
            SET name = {placeholder}, description = {placeholder}
            WHERE id = {placeholder}
        """, ("Extra Turno", "Richiesta di riconoscimento ore di Extra Turno", old_id))
        db.commit()
        app.logger.info(f"Migrato tipo richiesta 'Straordinario' → 'Extra Turno' (id={old_id})")
        return old_id
    
    # Crea il tipo se non esiste
    db.execute(f"""
        INSERT INTO request_types (name, value_type, description, active, sort_order)
        VALUES ({placeholder}, {placeholder}, {placeholder}, 1, 100)
    """, ("Extra Turno", "minutes", "Richiesta di riconoscimento ore di Extra Turno"))
    db.commit()
    
    # Recupera l'ID appena creato
    row = db.execute(
        f"SELECT id FROM request_types WHERE name = {placeholder}",
        ("Extra Turno",)
    ).fetchone()
    
    return row["id"] if isinstance(row, Mapping) else row[0]


def _ensure_missed_punch_request_type(db: DatabaseLike) -> int:
    """
    Assicura che esista il tipo richiesta 'Mancata Timbratura' e ritorna il suo ID.
    """
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Cerca se esiste già
    row = db.execute(
        f"SELECT id FROM request_types WHERE name = {placeholder}",
        ("Mancata Timbratura",)
    ).fetchone()
    
    if row:
        return row["id"] if isinstance(row, Mapping) else row[0]
    
    # Crea il tipo se non esiste
    db.execute(f"""
        INSERT INTO request_types (name, value_type, description, active, sort_order)
        VALUES ({placeholder}, {placeholder}, {placeholder}, 1, 4)
    """, ("Mancata Timbratura", "timbratura", "Richiesta di inserimento timbratura mancante"))
    db.commit()
    
    # Recupera l'ID appena creato
    row = db.execute(
        f"SELECT id FROM request_types WHERE name = {placeholder}",
        ("Mancata Timbratura",)
    ).fetchone()
    
    return row["id"] if isinstance(row, Mapping) else row[0]


def get_overtime_request_type_id(db: DatabaseLike) -> int:
    """Ritorna l'ID del tipo richiesta 'Extra Turno', creandolo se necessario."""
    ensure_request_types_table(db)
    return _ensure_overtime_request_type(db)


def ensure_request_types_table(db: DatabaseLike) -> None:
    """Crea la tabella request_types se non esiste."""
    statement = (
        REQUEST_TYPES_TABLE_MYSQL if DB_VENDOR == "mysql" else REQUEST_TYPES_TABLE_SQLITE
    )
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass
    
    # Migrazione MySQL: aggiungi DEFAULT alle colonne timestamp se mancante
    if DB_VENDOR == "mysql":
        try:
            db.execute("ALTER TABLE request_types MODIFY COLUMN created_ts BIGINT NOT NULL DEFAULT 0")
            db.commit()
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE request_types MODIFY COLUMN updated_ts BIGINT NOT NULL DEFAULT 0")
            db.commit()
        except Exception:
            pass
        # Poi aggiungi 'minutes' e 'timbratura' all'ENUM
        try:
            db.execute("""
                ALTER TABLE request_types 
                MODIFY COLUMN value_type ENUM('hours', 'days', 'amount', 'km', 'minutes', 'timbratura') NOT NULL
            """)
            db.commit()
            app.logger.info("Migrazione ENUM value_type completata con successo")
        except Exception as e:
            app.logger.warning(f"Migrazione ENUM value_type: {e}")
        
        # Migrazione: aggiungi colonna abbreviation se non esiste
        try:
            db.execute("ALTER TABLE request_types ADD COLUMN abbreviation VARCHAR(10)")
            db.commit()
            app.logger.info("Migrazione: aggiunta colonna abbreviation a request_types")
        except Exception:
            pass  # Colonna già esiste
        
        # Migrazione: aggiungi colonna external_id se non esiste
        try:
            db.execute("ALTER TABLE request_types ADD COLUMN external_id VARCHAR(100)")
            db.commit()
            app.logger.info("Migrazione: aggiunta colonna external_id a request_types")
        except Exception:
            pass  # Colonna già esiste
        
        # Migrazione: aggiungi colonna is_giustificativo se non esiste
        try:
            db.execute("ALTER TABLE request_types ADD COLUMN is_giustificativo TINYINT(1) DEFAULT 0")
            db.commit()
            app.logger.info("Migrazione: aggiunta colonna is_giustificativo a request_types")
        except Exception:
            pass  # Colonna già esiste
    
    # Assicura che esista il tipo "Extra Turno" per le richieste automatiche
    _ensure_overtime_request_type(db)
    
    # Assicura che esista il tipo "Mancata Timbratura"
    _ensure_missed_punch_request_type(db)


def ensure_user_requests_table(db: DatabaseLike) -> None:
    """Crea la tabella user_requests se non esiste e aggiunge colonne mancanti."""
    statement = (
        USER_REQUESTS_TABLE_MYSQL if DB_VENDOR == "mysql" else USER_REQUESTS_TABLE_SQLITE
    )
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass
    
    # Migrazione MySQL: aggiungi DEFAULT alle colonne timestamp se mancante
    if DB_VENDOR == "mysql":
        try:
            db.execute("ALTER TABLE user_requests MODIFY COLUMN created_ts BIGINT NOT NULL DEFAULT 0")
            db.commit()
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE user_requests MODIFY COLUMN updated_ts BIGINT NOT NULL DEFAULT 0")
            db.commit()
        except Exception:
            pass
    
    # Aggiungi colonne mancanti se la tabella esisteva già
    try:
        if DB_VENDOR == "mysql":
            db.execute("ALTER TABLE user_requests ADD COLUMN cdc VARCHAR(100)")
            db.commit()
    except Exception:
        pass  # Colonna già esiste
    
    try:
        if DB_VENDOR == "mysql":
            db.execute("ALTER TABLE user_requests ADD COLUMN attachment_path VARCHAR(500)")
            db.commit()
    except Exception:
        pass  # Colonna già esiste
    
    # Migrazione: aggiungi colonna tratte per rimborsi km
    try:
        if DB_VENDOR == "mysql":
            db.execute("ALTER TABLE user_requests ADD COLUMN tratte JSON DEFAULT NULL")
        else:
            db.execute("ALTER TABLE user_requests ADD COLUMN tratte TEXT DEFAULT NULL")
        db.commit()
    except Exception:
        pass  # Colonna già esiste
    
    # Migrazione: aggiungi colonna extra_data per dati aggiuntivi (es. straordinari)
    try:
        if DB_VENDOR == "mysql":
            db.execute("ALTER TABLE user_requests ADD COLUMN extra_data JSON DEFAULT NULL")
        else:
            db.execute("ALTER TABLE user_requests ADD COLUMN extra_data TEXT DEFAULT NULL")
        db.commit()
    except Exception:
        pass  # Colonna già esiste


def ensure_user_documents_table(db: DatabaseLike) -> None:
    """Crea le tabelle user_documents e user_documents_read se non esistono."""
    # Tabella documenti
    statement = (
        USER_DOCUMENTS_TABLE_MYSQL if DB_VENDOR == "mysql" else USER_DOCUMENTS_TABLE_SQLITE
    )
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass
    
    # Aggiungi colonna notified_at se non esiste
    existing = _get_existing_columns(db, "user_documents")
    if "notified_at" not in existing:
        col_type = "BIGINT" if DB_VENDOR == "mysql" else "INTEGER"
        try:
            db.execute(f"ALTER TABLE user_documents ADD COLUMN notified_at {col_type} DEFAULT NULL")
            db.commit()
        except Exception:
            pass
    
    # Tabella letture
    statement = (
        USER_DOCUMENTS_READ_TABLE_MYSQL if DB_VENDOR == "mysql" else USER_DOCUMENTS_READ_TABLE_SQLITE
    )
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass


def ensure_rentman_plannings_table(db: DatabaseLike) -> None:
    """Crea la tabella rentman_plannings se non esiste."""
    statement = (
        RENTMAN_PLANNINGS_TABLE_MYSQL if DB_VENDOR == "mysql" else RENTMAN_PLANNINGS_TABLE_SQLITE
    )
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass
    
    # Migrazione: aggiungi colonne location se non esistono
    if DB_VENDOR == "mysql":
        try:
            # Verifica se la colonna esiste già
            cursor = db.execute("SHOW COLUMNS FROM rentman_plannings LIKE 'location_id'")
            if not cursor.fetchone():
                db.execute("ALTER TABLE rentman_plannings ADD COLUMN subproject_id INT AFTER project_code")
                db.execute("ALTER TABLE rentman_plannings ADD COLUMN location_id INT AFTER subproject_id")
                db.execute("ALTER TABLE rentman_plannings ADD COLUMN location_name VARCHAR(500) AFTER location_id")
                db.execute("ALTER TABLE rentman_plannings ADD COLUMN location_address VARCHAR(500) AFTER location_name")
                db.commit()
                app.logger.info("Migrazione rentman_plannings: aggiunte colonne location")
        except Exception as e:
            app.logger.warning(f"Migrazione rentman_plannings location: {e}")
        
        # Migrazione: aggiungi colonne plan_start e plan_end se non esistono
        try:
            cursor = db.execute("SHOW COLUMNS FROM rentman_plannings LIKE 'plan_start'")
            if not cursor.fetchone():
                db.execute("ALTER TABLE rentman_plannings ADD COLUMN plan_start DATETIME DEFAULT NULL AFTER location_address")
                db.execute("ALTER TABLE rentman_plannings ADD COLUMN plan_end DATETIME DEFAULT NULL AFTER plan_start")
                db.commit()
                app.logger.info("Migrazione rentman_plannings: aggiunte colonne plan_start e plan_end")
        except Exception as e:
            app.logger.warning(f"Migrazione rentman_plannings plan_start/plan_end: {e}")
        
        # Migrazione: aggiungi colonne pausa se non esistono
        try:
            cursor = db.execute("SHOW COLUMNS FROM rentman_plannings LIKE 'break_start'")
            if not cursor.fetchone():
                db.execute("ALTER TABLE rentman_plannings ADD COLUMN break_start TIME DEFAULT NULL AFTER plan_end")
                db.execute("ALTER TABLE rentman_plannings ADD COLUMN break_end TIME DEFAULT NULL AFTER break_start")
                db.execute("ALTER TABLE rentman_plannings ADD COLUMN break_minutes INT DEFAULT NULL AFTER break_end")
                db.commit()
                app.logger.info("Migrazione rentman_plannings: aggiunte colonne pausa")
        except Exception as e:
            app.logger.warning(f"Migrazione rentman_plannings pausa: {e}")
        
        # Migrazione: aggiungi colonna timbratura_gps_mode se non esiste
        try:
            cursor = db.execute("SHOW COLUMNS FROM rentman_plannings LIKE 'timbratura_gps_mode'")
            if not cursor.fetchone():
                db.execute("ALTER TABLE rentman_plannings ADD COLUMN timbratura_gps_mode VARCHAR(20) DEFAULT 'group' AFTER location_address")
                db.commit()
                app.logger.info("Migrazione rentman_plannings: aggiunta colonna timbratura_gps_mode")
        except Exception as e:
            app.logger.warning(f"Migrazione rentman_plannings timbratura_gps_mode: {e}")
        
        # Migrazione: aggiungi colonna gps_timbratura_location se non esiste
        try:
            cursor = db.execute("SHOW COLUMNS FROM rentman_plannings LIKE 'gps_timbratura_location'")
            if not cursor.fetchone():
                db.execute("ALTER TABLE rentman_plannings ADD COLUMN gps_timbratura_location VARCHAR(255) AFTER timbratura_gps_mode")
                db.commit()
                app.logger.info("Migrazione rentman_plannings: aggiunta colonna gps_timbratura_location")
        except Exception as e:
            app.logger.warning(f"Migrazione rentman_plannings gps_timbratura_location: {e}")
        # Migrazione: aggiungi colonna is_obsolete se non esiste
        try:
            cursor = db.execute("SHOW COLUMNS FROM rentman_plannings LIKE 'is_obsolete'")
            if not cursor.fetchone():
                db.execute("ALTER TABLE rentman_plannings ADD COLUMN is_obsolete TINYINT(1) DEFAULT 0 AFTER webservice_response")
                db.execute("CREATE INDEX idx_planning_obsolete ON rentman_plannings(is_obsolete)")
                db.commit()
                app.logger.info("Migrazione rentman_plannings: aggiunta colonna is_obsolete")
        except Exception as e:
            app.logger.warning(f"Migrazione rentman_plannings is_obsolete: {e}")
        
        # Migrazione: aggiungi colonne location_lat e location_lon se non esistono
        try:
            cursor = db.execute("SHOW COLUMNS FROM rentman_plannings LIKE 'location_lat'")
            if not cursor.fetchone():
                db.execute("ALTER TABLE rentman_plannings ADD COLUMN location_lat DECIMAL(12,8) DEFAULT NULL AFTER location_address")
                db.execute("ALTER TABLE rentman_plannings ADD COLUMN location_lon DECIMAL(12,8) DEFAULT NULL AFTER location_lat")
                db.commit()
                app.logger.info("Migrazione rentman_plannings: aggiunte colonne location_lat e location_lon")
        except Exception as e:
            app.logger.warning(f"Migrazione rentman_plannings location_lat/lon: {e}")
        
        # Migrazione: aggiungi colonna remark_planner se non esiste
        try:
            cursor = db.execute("SHOW COLUMNS FROM rentman_plannings LIKE 'remark_planner'")
            if not cursor.fetchone():
                db.execute("ALTER TABLE rentman_plannings ADD COLUMN remark_planner TEXT AFTER remark")
                db.commit()
                app.logger.info("Migrazione rentman_plannings: aggiunta colonna remark_planner")
        except Exception as e:
            app.logger.warning(f"Migrazione rentman_plannings remark_planner: {e}")
    else:
        # SQLite migrations
        migrations_sqlite = [
            "ALTER TABLE rentman_plannings ADD COLUMN break_start TEXT DEFAULT NULL",
            "ALTER TABLE rentman_plannings ADD COLUMN break_end TEXT DEFAULT NULL",
            "ALTER TABLE rentman_plannings ADD COLUMN break_minutes INTEGER DEFAULT NULL",
            "ALTER TABLE rentman_plannings ADD COLUMN timbratura_gps_mode TEXT DEFAULT 'group'",
            "ALTER TABLE rentman_plannings ADD COLUMN gps_timbratura_location TEXT DEFAULT NULL",
            "ALTER TABLE rentman_plannings ADD COLUMN is_obsolete INTEGER DEFAULT 0",
            "ALTER TABLE rentman_plannings ADD COLUMN location_lat REAL DEFAULT NULL",
            "ALTER TABLE rentman_plannings ADD COLUMN location_lon REAL DEFAULT NULL",
            "ALTER TABLE rentman_plannings ADD COLUMN remark_planner TEXT DEFAULT NULL",
        ]
        for migration in migrations_sqlite:
            try:
                db.execute(migration)
                db.commit()
            except Exception:
                pass  # Colonna già esistente


# ═══════════════════════════════════════════════════════════════════════════════
#  PIANIFICAZIONI RENTMAN - ROUTES
# ═══════════════════════════════════════════════════════════════════════════════


def get_joblog_hours_for_date(db: DatabaseLike, target_date: str) -> Dict[str, float]:
    """
    Calcola le ore registrate in JobLog per ogni operatore in una data specifica.
    
    Returns:
        Dict con member_name.lower() come chiave e ore totali come valore.
    """
    try:
        from datetime import datetime as dt_parse
        # Parse target date
        target_dt = dt_parse.strptime(target_date, "%Y-%m-%d").date()
    except ValueError:
        app.logger.warning(f"Data non valida per JobLog hours: {target_date}")
        return {}
    
    # Calcola timestamp inizio e fine giornata
    start_of_day = dt_parse.combine(target_dt, dt_parse.min.time())
    end_of_day = dt_parse.combine(target_dt, dt_parse.max.time())
    start_ts = int(start_of_day.timestamp() * 1000)
    end_ts = int(end_of_day.timestamp() * 1000)
    
    # Query per eventi finish_activity con duration_ms
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    query = f"""
        SELECT el.ts, el.member_key, el.details, ms.member_name
        FROM event_log el
        LEFT JOIN member_state ms ON el.member_key = ms.member_key AND el.project_code = ms.project_code
        WHERE el.kind = 'finish_activity'
        AND el.ts >= {placeholder} AND el.ts <= {placeholder}
    """
    
    try:
        rows = db.execute(query, (start_ts, end_ts)).fetchall()
    except Exception as exc:
        app.logger.error(f"Errore query JobLog hours: {exc}")
        return {}
    
    # Accumula ore per operatore
    hours_by_member: Dict[str, float] = {}
    
    for row in rows:
        try:
            details = json.loads(row["details"]) if row["details"] else {}
        except json.JSONDecodeError:
            continue
        
        duration_ms = details.get("duration_ms", 0)
        if not duration_ms:
            continue
        
        # Usa member_name dalla tabella member_state o dai dettagli evento
        member_name = row["member_name"] or details.get("member_name") or row["member_key"]
        if not member_name:
            continue
        
        # Normalizza il nome (lowercase) per il matching
        member_name_lower = member_name.strip().lower()
        
        # Converti ms in ore
        hours = duration_ms / (1000 * 60 * 60)
        
        if member_name_lower in hours_by_member:
            hours_by_member[member_name_lower] += hours
        else:
            hours_by_member[member_name_lower] = hours
    
    return hours_by_member


def match_crew_name_to_joblog(crew_name: str, joblog_hours: Dict[str, float]) -> Optional[float]:
    """
    Cerca di matchare un nome operatore Rentman con i dati JobLog.
    Usa matching esatto e fuzzy sul nome.
    """
    if not crew_name:
        return None
    
    crew_lower = crew_name.strip().lower()
    
    # 1. Match esatto
    if crew_lower in joblog_hours:
        return joblog_hours[crew_lower]
    
    # 2. Prova a confrontare parti del nome
    crew_parts = set(crew_lower.split())
    
    for member_name, hours in joblog_hours.items():
        member_parts = set(member_name.split())
        
        # Se almeno 2 parole coincidono, o se una parola è contenuta nell'altra
        common = crew_parts & member_parts
        if len(common) >= 2:
            return hours
        
        # Se il nome completo è contenuto
        if crew_lower in member_name or member_name in crew_lower:
            return hours
        
        # Match su cognome (tipicamente la seconda parola)
        crew_words = crew_lower.split()
        member_words = member_name.split()
        
        # Se hanno almeno 2 parole e la seconda (cognome) coincide
        if len(crew_words) >= 2 and len(member_words) >= 2:
            if crew_words[1] == member_words[1]:
                return hours
    
    return None


@app.get("/admin/rentman-planning")
@login_required
def admin_rentman_planning_page() -> ResponseReturnValue:
    """Pagina pianificazioni Rentman."""
    if not is_admin_or_supervisor():
        abort(403)

    display_name = session.get("user_display") or session.get("user_name") or session.get("user")
    primary_name = session.get("user_name") or display_name or session.get("user")
    initials = session.get("user_initials") or compute_initials(primary_name or "")

    return render_template(
        "admin_rentman_planning.html",
        user_name=primary_name,
        user_display=display_name,
        user_initials=initials,
        is_admin=bool(session.get("is_admin")),
    )


@app.get("/admin/locations")
@login_required
def admin_locations_page() -> ResponseReturnValue:
    """Pagina gestione location GPS."""
    if not is_admin_or_supervisor():
        abort(403)

    display_name = session.get("user_display") or session.get("user_name") or session.get("user")
    primary_name = session.get("user_name") or display_name or session.get("user")
    initials = session.get("user_initials") or compute_initials(primary_name or "")

    return render_template(
        "admin_locations.html",
        user_name=primary_name,
        user_display=display_name,
        user_initials=initials,
        is_admin=bool(session.get("is_admin")),
    )


@app.get("/api/admin/locations")
@login_required
def api_admin_locations_list() -> ResponseReturnValue:
    """Restituisce tutte le location da Rentman con stato coordinate."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "Accesso negato"}), 403

    db = get_db()
    ensure_rentman_plannings_table(db)
    ensure_location_cache_table(db)

    # Recupera tutte le location univoche da Rentman (location_name non NULL)
    locations_rows = db.execute("""
        SELECT DISTINCT location_name, location_address
        FROM rentman_plannings
        WHERE location_name IS NOT NULL AND location_name != ''
        ORDER BY location_name
    """).fetchall()

    locations = []
    for row in locations_rows:
        location_name = row["location_name"] if isinstance(row, dict) else row[0]
        location_address = row["location_address"] if isinstance(row, dict) else row[1]

        # Controlla se ha coordinate in cache (ora restituisce anche il raggio)
        cached = get_location_cache(db, location_name)

        locations.append({
            "name": location_name,
            "address": location_address or "",
            "has_cache": cached is not None,
            "latitude": cached[0] if cached else None,
            "longitude": cached[1] if cached else None,
            "radius_meters": cached[2] if cached else 300,
        })

    return jsonify({"locations": locations})


@app.post("/api/admin/locations/<location_name>")
@login_required
def api_admin_locations_save(location_name: str) -> ResponseReturnValue:
    """Salva le coordinate e il raggio per una location."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "Accesso negato"}), 403

    data = request.get_json() or {}
    latitude = data.get("latitude")
    longitude = data.get("longitude")
    radius_meters = data.get("radius_meters", 300)

    if latitude is None or longitude is None:
        return jsonify({"error": "Latitudine e longitudine richieste"}), 400

    try:
        latitude = float(latitude)
        longitude = float(longitude)
        radius_meters = int(radius_meters) if radius_meters else 300
    except (ValueError, TypeError):
        return jsonify({"error": "Coordinate non valide"}), 400

    db = get_db()
    save_location_cache(db, location_name, latitude, longitude, radius_meters=radius_meters)

    return jsonify({"ok": True, "message": f"Coordinate salvate per {location_name} (raggio: {radius_meters}m)"})


@app.get("/api/admin/rentman-planning")
@login_required
def api_admin_rentman_planning() -> ResponseReturnValue:
    """API per recuperare pianificazioni Rentman per una data."""
    app.logger.warning("🔴🔴🔴 INIZIO /api/admin/rentman-planning")
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403

    target_date = request.args.get("date")
    if not target_date:
        target_date = datetime.now().date().isoformat()
    
    app.logger.warning(f"🔴 API rentman-planning: target_date={target_date}")

    try:
        from rentman_client import RentmanClient, RentmanError
        client = RentmanClient()
    except Exception as exc:
        app.logger.warning("Rentman client non disponibile: %s", exc)
        return jsonify({"error": "Rentman non configurato", "details": str(exc)}), 500

    try:
        plannings = client.get_crew_plannings_by_date(target_date)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except RentmanError as exc:
        return jsonify({"error": "Errore Rentman", "details": str(exc)}), 500

    app.logger.info(f"📥 RENTMAN: Ricevuti {len(plannings)} turni per {target_date}")
    for idx, p in enumerate(plannings):
        crew_name = p.get("displayname", "?")
        start = p.get("planperiod_start", "?")
        end = p.get("planperiod_end", "?")
        app.logger.info(f"  [{idx}] {crew_name}: {start} - {end}")

    # DEBUG: Log tutti i campi della prima pianificazione per capire la struttura
    if plannings:
        app.logger.info("Rentman planning fields: %s", list(plannings[0].keys()))
        app.logger.info("Rentman first planning data: %s", plannings[0])

    # Calcola ore JobLog per questa data
    db = get_db()
    joblog_hours = get_joblog_hours_for_date(db, target_date)

    # Arricchisci i dati con info su crew e progetto
    # Cache per evitare chiamate duplicate
    crew_cache: Dict[int, Dict[str, Any]] = {}
    function_cache: Dict[int, Dict[str, Any]] = {}
    project_cache: Dict[int, Dict[str, Any]] = {}
    subproject_cache: Dict[int, Dict[str, Any]] = {}
    contact_cache: Dict[int, Dict[str, Any]] = {}
    
    # Pre-carica mapping rentman_crew_id -> group_id da app_users
    ensure_user_groups_table(db)
    
    # Carica utenti JobLog con il loro rentman_crew_id e group_id
    crew_group_map: Dict[int, int] = {}  # rentman_crew_id -> group_id
    if DB_VENDOR == "mysql":
        user_rows = db.execute("SELECT rentman_crew_id, group_id FROM app_users WHERE rentman_crew_id IS NOT NULL AND group_id IS NOT NULL").fetchall()
    else:
        user_rows = db.execute("SELECT rentman_crew_id, group_id FROM app_users WHERE rentman_crew_id IS NOT NULL AND group_id IS NOT NULL").fetchall()
    for row in user_rows:
        rentman_crew_id = row["rentman_crew_id"] if isinstance(row, Mapping) else row[0]
        group_id = row["group_id"] if isinstance(row, Mapping) else row[1]
        if rentman_crew_id and group_id:
            crew_group_map[rentman_crew_id] = group_id
    
    app.logger.info("crew_group_map (da app_users): %s", crew_group_map)
    
    # Carica tutti i gruppi con la loro sede GPS
    group_gps_map: Dict[int, str] = {}  # group_id -> gps_location_name
    if DB_VENDOR == "mysql":
        group_rows = db.execute("SELECT id, gps_location_name FROM user_groups WHERE gps_location_name IS NOT NULL AND gps_location_name != ''").fetchall()
    else:
        group_rows = db.execute("SELECT id, gps_location_name FROM user_groups WHERE gps_location_name IS NOT NULL AND gps_location_name != ''").fetchall()
    for row in group_rows:
        group_id = row["id"] if isinstance(row, Mapping) else row[0]
        gps_loc = row["gps_location_name"] if isinstance(row, Mapping) else row[1]
        if group_id and gps_loc:
            group_gps_map[group_id] = gps_loc
    
    app.logger.info("group_gps_map: %s", group_gps_map)

    results = []
    for planning in plannings:
        # Estrai ID dal riferimento (es. "/crew/123" -> 123)
        crew_ref = planning.get("crewmember", "")
        crew_id = None
        if crew_ref and "/" in crew_ref:
            try:
                crew_id = int(crew_ref.split("/")[-1])
            except (ValueError, IndexError):
                pass

        function_ref = planning.get("function", "")
        function_id = None
        if function_ref and "/" in function_ref:
            try:
                function_id = int(function_ref.split("/")[-1])
            except (ValueError, IndexError):
                pass

        # Recupera dettagli crew
        crew_name = planning.get("displayname", "")
        if crew_id and crew_id not in crew_cache:
            crew_data = client.get_crew_member(crew_id)
            if crew_data:
                crew_cache[crew_id] = crew_data
        if crew_id and crew_id in crew_cache:
            cd = crew_cache[crew_id]
            crew_name = cd.get("displayname") or f"{cd.get('firstname', '')} {cd.get('lastname', '')}".strip() or crew_name

        # Recupera dettagli funzione, progetto e location
        project_name = ""
        project_code = ""
        project_id = None
        function_name = ""
        subproject_id = None
        location_id = None
        location_name = ""
        location_address = ""
        location_lat = None
        location_lon = None
        
        if function_id and function_id not in function_cache:
            func_data = client.get_project_function(function_id)
            if func_data:
                function_cache[function_id] = func_data
        if function_id and function_id in function_cache:
            fd = function_cache[function_id]
            function_name = fd.get("name") or fd.get("displayname") or ""
            
            # Estrai subproject dalla funzione
            subproject_ref = fd.get("subproject", "")
            if subproject_ref and "/" in subproject_ref:
                try:
                    subproject_id = int(subproject_ref.split("/")[-1])
                    if subproject_id not in subproject_cache:
                        subproj_data = client.get_subproject(subproject_id)
                        if subproj_data:
                            subproject_cache[subproject_id] = subproj_data
                    if subproject_id in subproject_cache:
                        sp = subproject_cache[subproject_id]
                        # Estrai location dal subproject (è un riferimento a /contacts/ID)
                        location_ref = sp.get("location", "")
                        if location_ref and "/" in location_ref:
                            try:
                                location_id = int(location_ref.split("/")[-1])
                                if location_id not in contact_cache:
                                    contact_data = client.get_contact(location_id)
                                    if contact_data:
                                        contact_cache[location_id] = contact_data
                                if location_id in contact_cache:
                                    ct = contact_cache[location_id]
                                    location_name = ct.get("displayname") or ct.get("name") or ""
                                    # Costruisci indirizzo completo dai campi visit_*
                                    addr_parts = []
                                    if ct.get("visit_street"):
                                        addr_parts.append(ct.get("visit_street"))
                                        if ct.get("visit_number") and ct.get("visit_number") != "S.n.":
                                            addr_parts[-1] += " " + ct.get("visit_number")
                                    if ct.get("visit_postalcode"):
                                        addr_parts.append(ct.get("visit_postalcode"))
                                    if ct.get("visit_city"):
                                        addr_parts.append(ct.get("visit_city"))
                                    if ct.get("visit_state"):
                                        addr_parts.append(f"({ct.get('visit_state')})")
                                    location_address = ", ".join(addr_parts) if addr_parts else ""
                                    # Recupera coordinate se disponibili (non null)
                                    lat = ct.get("latitude")
                                    lon = ct.get("longitude")
                                    if lat is not None and lon is not None:
                                        try:
                                            location_lat = float(lat)
                                            location_lon = float(lon)
                                        except (ValueError, TypeError):
                                            pass
                                    elif location_address:
                                        # Rentman non fornisce coordinate, prova geocoding dall'indirizzo
                                        coords = geocode_address(location_address)
                                        if coords:
                                            location_lat, location_lon = coords
                            except (ValueError, IndexError):
                                pass
                except (ValueError, IndexError):
                    pass
            
            # Estrai progetto dalla funzione
            project_ref = fd.get("project", "")
            if project_ref and "/" in project_ref:
                try:
                    project_id = int(project_ref.split("/")[-1])
                    if project_id not in project_cache:
                        proj_data = client.get_project(project_id)
                        if proj_data:
                            project_cache[project_id] = proj_data
                    if project_id in project_cache:
                        pd = project_cache[project_id]
                        project_name = pd.get("name") or pd.get("displayname") or ""
                        project_code = pd.get("number") or pd.get("reference") or ""
                        
                        # Fallback: se non abbiamo location dal subproject, prova dal progetto
                        if not location_id:
                            proj_location_ref = pd.get("location", "")
                            if proj_location_ref and "/" in proj_location_ref:
                                try:
                                    location_id = int(proj_location_ref.split("/")[-1])
                                    if location_id not in contact_cache:
                                        contact_data = client.get_contact(location_id)
                                        if contact_data:
                                            contact_cache[location_id] = contact_data
                                    if location_id in contact_cache:
                                        ct = contact_cache[location_id]
                                        location_name = ct.get("displayname") or ct.get("name") or ""
                                        # Costruisci indirizzo completo
                                        addr_parts = []
                                        if ct.get("visit_street"):
                                            addr_parts.append(ct.get("visit_street"))
                                            if ct.get("visit_number") and ct.get("visit_number") != "S.n.":
                                                addr_parts[-1] += " " + str(ct.get("visit_number"))
                                        if ct.get("visit_postalcode"):
                                            addr_parts.append(ct.get("visit_postalcode"))
                                        if ct.get("visit_city"):
                                            addr_parts.append(ct.get("visit_city"))
                                        if ct.get("visit_state"):
                                            addr_parts.append(f"({ct.get('visit_state')})")
                                        location_address = ", ".join(addr_parts) if addr_parts else ""
                                        # Coordinate
                                        lat = ct.get("latitude")
                                        lon = ct.get("longitude")
                                        if lat is not None and lon is not None:
                                            try:
                                                location_lat = float(lat)
                                                location_lon = float(lon)
                                            except (ValueError, TypeError):
                                                pass
                                        elif location_address:
                                            # Geocoding dall'indirizzo
                                            coords = geocode_address(location_address)
                                            if coords:
                                                location_lat, location_lon = coords
                                except (ValueError, IndexError):
                                    pass
                except (ValueError, IndexError):
                    pass

        # CONTROLLO CACHE GLOBALE: se le coordinate sono salvate nella cache, usa quelle
        location_radius = 300
        if location_name:
            ensure_location_cache_table(db)
            cached_coords = get_location_cache(db, location_name, location_id)
            if cached_coords:
                location_lat, location_lon, location_radius = cached_coords
                app.logger.info(f"✅ Location '{location_name}' (id={location_id}): usando coordinate dalla cache globale: {location_lat}, {location_lon}, raggio={location_radius}m")
            else:
                app.logger.info(f"⚠️ Location '{location_name}' (id={location_id}): nessuna cache, usando coordinate da Rentman (lat={location_lat}, lon={location_lon})")
        
        # Calcola ore JobLog per questo operatore
        joblog_registered = match_crew_name_to_joblog(crew_name, joblog_hours)
        
        # Calcola la pausa dalla differenza tra durata turno e ore pianificate
        # Rentman non fornisce un campo specifico per la pausa, ma la include già
        # nelle hours_planned (che sono al netto della pausa)
        break_minutes = None
        plan_start_str = planning.get("planperiod_start")
        plan_end_str = planning.get("planperiod_end")
        hours_planned_raw = planning.get("hours_planned")
        
        if plan_start_str and plan_end_str and hours_planned_raw is not None:
            try:
                from datetime import datetime as dt_parse
                # Parse datetime ISO - rimuovi timezone per semplificare
                start_clean = plan_start_str.split("+")[0].replace("Z", "")
                end_clean = plan_end_str.split("+")[0].replace("Z", "")
                
                start_dt = dt_parse.fromisoformat(start_clean)
                end_dt = dt_parse.fromisoformat(end_clean)
                
                # Calcola durata totale in minuti
                total_duration_minutes = (end_dt - start_dt).total_seconds() / 60
                
                # hours_planned è in secondi, converti in minuti
                hours_planned_val = float(hours_planned_raw)
                if hours_planned_val > 100:  # È in secondi
                    planned_minutes = hours_planned_val / 60
                else:  # È già in ore
                    planned_minutes = hours_planned_val * 60
                
                # La pausa è la differenza
                calculated_break = total_duration_minutes - planned_minutes
                
                if calculated_break > 0:
                    break_minutes = int(round(calculated_break))
            except Exception as e:
                app.logger.warning(f"Errore calcolo pausa: {e}")
        
        # Recupera sede GPS di timbratura dal gruppo dell'operatore
        gps_timbratura_location = None
        if crew_id and crew_id in crew_group_map:
            group_id = crew_group_map[crew_id]
            if group_id in group_gps_map:
                gps_timbratura_location = group_gps_map[group_id]
        
        # CONTROLLO CACHE GLOBALE: se le coordinate sono salvate nella cache, usa quelle
        location_radius = 300
        if location_name:
            ensure_location_cache_table(db)
            cached_coords = get_location_cache(db, location_name, location_id)
            if cached_coords:
                location_lat, location_lon, location_radius = cached_coords
                app.logger.info(f"✅ Location '{location_name}' (id={location_id}): usando coordinate dalla cache globale: {location_lat}, {location_lon}, raggio={location_radius}m")
            else:
                app.logger.info(f"⚠️ Location '{location_name}' (id={location_id}): nessuna cache, usando coordinate da Rentman (lat={location_lat}, lon={location_lon})")
        
        app.logger.warning(f"🔴 API /api/admin/rentman-planning: {crew_name} - location={location_name}, coords=({location_lat}, {location_lon})")
        
        results.append({
            "id": planning.get("id"),
            "rentman_id": planning.get("id"),  # Aggiungi rentman_id esplicitamente
            "crew_id": crew_id,
            "crew_name": crew_name,
            "function_name": function_name,
            "project_name": project_name,
            "project_code": project_code,
            "project_id": project_id,
            "subproject_id": subproject_id,
            "location_id": location_id,
            "location_name": location_name,
            "location_address": location_address,
            "location_lat": location_lat,
            "location_lon": location_lon,
            "gps_timbratura_location": gps_timbratura_location,
            "plan_start": planning.get("planperiod_start"),
            "plan_end": planning.get("planperiod_end"),
            "break_minutes": break_minutes,
            "hours_planned": planning.get("hours_planned"),
            "hours_registered": round(joblog_registered, 2) if joblog_registered is not None else 0,
            "hours_rentman": planning.get("hours_registered"),  # Mantieni valore originale Rentman per riferimento
            "remark": planning.get("remark", ""),
            "remark_planner": planning.get("remark_planner", ""),  # Nota al pianificatore
            "is_leader": planning.get("project_leader", False),
            "transport": planning.get("transport", ""),
        })

    # Ordina per orario di inizio e poi per nome crew
    results.sort(key=lambda x: (x.get("start") or "", x.get("crew_name") or ""))

    # Merge con dati salvati nel DB per preservare sent_to_webservice e rilevare modifiche
    db = get_db()
    ensure_rentman_plannings_table(db)
    
    if DB_VENDOR == "mysql":
        saved_rows = db.execute(
            "SELECT rentman_id, sent_to_webservice, plan_start, plan_end, project_name, sent_ts, break_start, break_end, break_minutes, gps_timbratura_location, timbratura_gps_mode FROM rentman_plannings WHERE planning_date = %s",
            (target_date,)
        ).fetchall()
    else:
        saved_rows = db.execute(
            "SELECT rentman_id, sent_to_webservice, plan_start, plan_end, project_name, sent_ts, break_start, break_end, break_minutes, gps_timbratura_location, timbratura_gps_mode FROM rentman_plannings WHERE planning_date = ?",
            (target_date,)
        ).fetchall()
    
    # Crea mappa rentman_id -> {sent, old_start, old_end, old_project, sent_ts, break_*, gps}
    saved_map = {}
    for row in saved_rows:
        if isinstance(row, Mapping):
            saved_map[row["rentman_id"]] = {
                "sent": bool(row["sent_to_webservice"]),
                "old_start": row["plan_start"],
                "old_end": row["plan_end"],
                "old_project": row["project_name"],
                "sent_ts": row["sent_ts"],
                "break_start": row.get("break_start"),
                "break_end": row.get("break_end"),
                "break_minutes": row.get("break_minutes"),
                "gps_timbratura_location": row.get("gps_timbratura_location"),
                "timbratura_gps_mode": row.get("timbratura_gps_mode"),
            }
        else:
            saved_map[row[0]] = {
                "sent": bool(row[1]),
                "old_start": row[2],
                "old_end": row[3],
                "old_project": row[4],
                "sent_ts": row[5],
                "break_start": row[6] if len(row) > 6 else None,
                "break_end": row[7] if len(row) > 7 else None,
                "break_minutes": row[8] if len(row) > 8 else None,
                "gps_timbratura_location": row[9] if len(row) > 9 else None,
                "timbratura_gps_mode": row[10] if len(row) > 10 else None,
            }
    
    # Arricchisci i risultati con info invio e modifiche
    for r in results:
        rentman_id = r.get("id")
        if rentman_id in saved_map:
            saved = saved_map[rentman_id]
            r["sent_to_webservice"] = saved["sent"]
            
            # Recupera pausa salvata se non presente da Rentman
            if not r.get("break_start") and saved.get("break_start"):
                r["break_start"] = saved["break_start"]
            if not r.get("break_end") and saved.get("break_end"):
                r["break_end"] = saved["break_end"]
            if not r.get("break_minutes") and saved.get("break_minutes"):
                r["break_minutes"] = saved["break_minutes"]
            
            # Preserva coordinate manuali se Rentman non le ha (null)
            rentman_lat = r.get("location_lat")
            rentman_lon = r.get("location_lon")
            saved_lat = saved.get("location_lat")
            saved_lon = saved.get("location_lon")
            rentman_has_coords = rentman_lat is not None and rentman_lon is not None
            saved_has_coords = saved_lat is not None and saved_lon is not None
            
            if not rentman_has_coords and saved_has_coords:
                r["location_lat"] = saved["location_lat"]
                r["location_lon"] = saved["location_lon"]
            
            # Preserva gps_timbratura_location e mode dal DB se non calcolato
            if saved.get("gps_timbratura_location") and not r.get("gps_timbratura_location"):
                r["gps_timbratura_location"] = saved["gps_timbratura_location"]
            if saved.get("timbratura_gps_mode"):
                r["timbratura_gps_mode"] = saved["timbratura_gps_mode"]
            
            # Aggiungi timestamp invio formattato
            if saved["sent_ts"]:
                try:
                    ts_val = saved["sent_ts"]
                    if isinstance(ts_val, (int, float)):
                        # sent_ts è in millisecondi (es: 1735634845125)
                        ts_sec = ts_val / 1000 if ts_val > 1e12 else ts_val
                        r["sent_ts"] = datetime.fromtimestamp(ts_sec).strftime("%d/%m/%Y %H:%M:%S")
                    elif isinstance(ts_val, str):
                        # Potrebbe essere ISO string (es: 2025-12-31T08:47:25.125Z)
                        ts_clean = ts_val.replace("Z", "+00:00")
                        dt = datetime.fromisoformat(ts_clean)
                        r["sent_ts"] = dt.strftime("%d/%m/%Y %H:%M:%S")
                    elif hasattr(ts_val, 'strftime'):
                        # È già un datetime
                        r["sent_ts"] = ts_val.strftime("%d/%m/%Y %H:%M:%S")
                    else:
                        r["sent_ts"] = str(ts_val)  # Fallback: converti a stringa
                except Exception:
                    r["sent_ts"] = None
            else:
                r["sent_ts"] = None
            
            # Rileva se è stato modificato rispetto all'ultimo invio
            if saved["sent"]:
                # Confronta orari e progetto
                # Gestisci sia stringhe che datetime objects
                new_start_raw = r.get("start", "")
                new_start = str(new_start_raw)[:16] if new_start_raw else ""
                old_start_raw = saved["old_start"]
                old_start = str(old_start_raw)[:16] if old_start_raw else ""
                new_end_raw = r.get("end", "")
                new_end = str(new_end_raw)[:16] if new_end_raw else ""
                old_end_raw = saved["old_end"]
                old_end = str(old_end_raw)[:16] if old_end_raw else ""
                
                is_modified = (
                    new_start != old_start or 
                    new_end != old_end or 
                    r.get("project_name", "") != (saved["old_project"] or "")
                )
                r["is_modified"] = is_modified
            else:
                r["is_modified"] = False
        else:
            r["sent_to_webservice"] = False
            r["is_modified"] = False
            r["sent_ts"] = None

    return jsonify({
        "ok": True,
        "date": target_date,
        "count": len(results),
        "plannings": results,
    })


@app.post("/api/admin/rentman/planning/update-break")
@login_required
def api_admin_rentman_planning_update_break() -> ResponseReturnValue:
    """Aggiorna solo i campi pausa di una pianificazione."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403

    data = request.get_json() or {}
    rentman_id = data.get("rentman_id")
    target_date = data.get("date")
    break_start = data.get("break_start")
    break_end = data.get("break_end")
    break_minutes = data.get("break_minutes")

    app.logger.info(f"update-break chiamato con: rentman_id={rentman_id}, date={target_date}")

    if not rentman_id or not target_date:
        return jsonify({"error": "rentman_id e date richiesti"}), 400

    db = get_db()
    ensure_rentman_plannings_table(db)

    # Verifica che il record esista - cerca anche per id diretto (potrebbe essere l'id del DB)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    existing = db.execute(
        f"SELECT id, rentman_id FROM rentman_plannings WHERE (rentman_id = {placeholder} OR id = {placeholder}) AND planning_date = {placeholder}",
        (rentman_id, rentman_id, target_date)
    ).fetchone()

    if not existing:
        app.logger.warning(f"Pianificazione non trovata per rentman_id={rentman_id}, date={target_date}")
        return jsonify({"error": "Pianificazione non trovata. Salva prima le pianificazioni."}), 404

    record_id = existing['id'] if isinstance(existing, dict) else existing[0]
    actual_rentman_id = existing['rentman_id'] if isinstance(existing, dict) else existing[1]
    app.logger.info(f"Trovato record: id={record_id}, rentman_id={actual_rentman_id}")

    # Aggiorna solo i campi pausa
    now_ms = int(time.time() * 1000)
    
    db.execute(f"""
        UPDATE rentman_plannings 
        SET break_start = {placeholder}, break_end = {placeholder}, break_minutes = {placeholder}, updated_ts = {placeholder}
        WHERE id = {placeholder}
    """, (break_start, break_end, break_minutes, now_ms, record_id))

    db.commit()

    app.logger.info(f"Pausa aggiornata per id={record_id}: start={break_start}, end={break_end}, minutes={break_minutes}")

    return jsonify({"success": True, "message": "Pausa aggiornata"})


@app.post("/api/admin/rentman/planning/update-gps-mode")
@login_required
def api_admin_rentman_planning_update_gps_mode() -> ResponseReturnValue:
    """Aggiorna la modalità GPS per la timbratura (sede gruppo vs location progetto)."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403

    data = request.get_json() or {}
    planning_id = data.get("id")
    rentman_id = data.get("rentman_id")
    gps_mode = data.get("timbratura_gps_mode")

    app.logger.info(f"update-gps-mode: id={planning_id}, rentman_id={rentman_id}, mode={gps_mode}")

    if not planning_id and not rentman_id:
        return jsonify({"error": "ID pianificazione richiesto"}), 400

    if gps_mode not in ("group", "location"):
        return jsonify({"error": "Modalità GPS non valida. Usa 'group' o 'location'"}), 400

    db = get_db()
    now_ms = int(time.time() * 1000)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"

    # Converti in int se possibile
    try:
        if planning_id:
            planning_id = int(planning_id)
        if rentman_id:
            rentman_id = int(rentman_id)
    except (ValueError, TypeError):
        pass

    # Trova il record - prova prima con id, poi con rentman_id
    existing = None
    
    if planning_id:
        existing = db.execute(
            f"SELECT id FROM rentman_plannings WHERE id = {placeholder}",
            (planning_id,)
        ).fetchone()
        app.logger.info(f"Cercato per id={planning_id}: {'trovato' if existing else 'non trovato'}")
    
    if not existing and rentman_id:
        existing = db.execute(
            f"SELECT id FROM rentman_plannings WHERE rentman_id = {placeholder}",
            (rentman_id,)
        ).fetchone()
        app.logger.info(f"Cercato per rentman_id={rentman_id}: {'trovato' if existing else 'non trovato'}")
    
    if not existing and planning_id:
        # Prova planning_id come rentman_id
        existing = db.execute(
            f"SELECT id FROM rentman_plannings WHERE rentman_id = {placeholder}",
            (planning_id,)
        ).fetchone()
        app.logger.info(f"Cercato per planning_id come rentman_id={planning_id}: {'trovato' if existing else 'non trovato'}")

    if not existing:
        app.logger.warning(f"Pianificazione non trovata: id={planning_id}, rentman_id={rentman_id}")
        return jsonify({"error": "Pianificazione non trovata"}), 404

    # Recupera l'ID del record trovato
    record_id = existing["id"] if isinstance(existing, dict) else existing[0]
    
    # Aggiorna
    db.execute(f"""
        UPDATE rentman_plannings 
        SET timbratura_gps_mode = {placeholder}, updated_ts = {placeholder}
        WHERE id = {placeholder}
    """, (gps_mode, now_ms, record_id))
    
    db.commit()

    mode_label = "Sede Gruppo" if gps_mode == "group" else "Location Progetto"
    app.logger.info(f"✅ Modalità GPS aggiornata per id={record_id}: {mode_label}")

    # Verifica che il salvataggio sia andato a buon fine
    verify = db.execute(f"SELECT timbratura_gps_mode FROM rentman_plannings WHERE id = {placeholder}", (record_id,)).fetchone()
    saved_mode = verify['timbratura_gps_mode'] if isinstance(verify, dict) else (verify[0] if verify else None)
    app.logger.info(f"🔍 Verifica dopo save: mode nel DB = {saved_mode}")

    return jsonify({"success": True, "message": f"Modalità GPS impostata a: {mode_label}"})


@app.post("/api/admin/rentman-planning/save")
@login_required
def api_admin_rentman_planning_save() -> ResponseReturnValue:
    """Salva le pianificazioni Rentman nel database locale."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403

    data = request.get_json() or {}
    target_date = data.get("date")
    plannings = data.get("plannings", [])

    if not target_date:
        return jsonify({"error": "Data richiesta"}), 400

    if not plannings:
        return jsonify({"error": "Nessuna pianificazione da salvare"}), 400

    db = get_db()
    ensure_rentman_plannings_table(db)
    ensure_crew_members_table(db)  # Assicura che la tabella operatori esista

    now_ms = int(time.time() * 1000)
    saved = 0
    updated = 0
    synced_crews = set()  # Track già sincronizzati per evitare duplicati

    def parse_iso_datetime(dt_str: str | None) -> str | None:
        """Converte datetime ISO in formato MySQL compatibile."""
        if not dt_str:
            return None
        try:
            # Prova a parsare datetime ISO con timezone
            from datetime import datetime as dt_parse
            # Rimuovi timezone se presente e converti
            if '+' in dt_str:
                dt_str = dt_str.split('+')[0]
            elif dt_str.endswith('Z'):
                dt_str = dt_str[:-1]
            # Converti in formato MySQL
            parsed = dt_parse.fromisoformat(dt_str)
            return parsed.strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            return dt_str

    for p in plannings:
        # IMPORTANTE: usa rentman_id se presente, altrimenti id
        # Quando i dati vengono dal merge frontend, 'id' è l'ID del DB
        rentman_id = p.get("rentman_id") or p.get("id")

        if not rentman_id:
            continue

        # Sincronizza operatore nel database locale (una volta sola per crew_id)
        crew_id = p.get("crew_id")
        crew_name = p.get("crew_name")
        if crew_id and crew_id not in synced_crews:
            sync_crew_member_from_rentman(db, crew_id, crew_name or f"Crew {crew_id}")
            synced_crews.add(crew_id)
            # NON fare continue qui! Devo comunque salvare la pianificazione

        # Converti hours da secondi a ore se necessario
        hours_planned = p.get("hours_planned")
        if hours_planned and float(hours_planned) > 100:
            hours_planned = float(hours_planned) / 3600

        hours_registered = p.get("hours_registered")
        if hours_registered and float(hours_registered) > 100:
            hours_registered = float(hours_registered) / 3600

        # Converti datetime in formato MySQL compatibile
        plan_start = parse_iso_datetime(p.get("plan_start"))
        plan_end = parse_iso_datetime(p.get("plan_end"))

        # Estrai project_id dalla funzione (se disponibile)
        project_id = p.get("project_id")
        function_id = p.get("function_id")

        # Check if exists
        if DB_VENDOR == "mysql":
            existing = db.execute(
                "SELECT id, sent_to_webservice, plan_start, plan_end FROM rentman_plannings WHERE rentman_id = %s AND planning_date = %s",
                (rentman_id, target_date)
            ).fetchone()
        else:
            existing = db.execute(
                "SELECT id, sent_to_webservice, plan_start, plan_end FROM rentman_plannings WHERE rentman_id = ? AND planning_date = ?",
                (rentman_id, target_date)
            ).fetchone()

        if existing:
            # Update existing record (but preserve sent status and timbratura_gps_mode!)
            app.logger.warning(f"🔴 SAVE UPDATE DB: rentman_id={rentman_id}, location_name={p.get('location_name')}, lat={p.get('location_lat')}, lon={p.get('location_lon')}")
            
            # Nota: preserva le note esistenti se quelle da Rentman sono vuote
            new_remark = p.get("remark") or None
            new_remark_planner = p.get("remark_planner") or None
            
            if DB_VENDOR == "mysql":
                db.execute("""
                    UPDATE rentman_plannings SET
                        crew_id = %s, crew_name = %s, function_id = %s, function_name = %s,
                        project_id = %s, project_name = %s, project_code = %s,
                        subproject_id = %s, location_id = %s, location_name = %s, location_address = %s,
                        location_lat = %s, location_lon = %s,
                        gps_timbratura_location = %s,
                        plan_start = %s, plan_end = %s, 
                        break_start = %s, break_end = %s, break_minutes = %s,
                        hours_planned = %s, hours_registered = %s,
                        remark = COALESCE(%s, remark), remark_planner = COALESCE(%s, remark_planner), 
                        is_leader = %s, transport = %s, updated_ts = %s,
                        is_obsolete = 0
                    WHERE rentman_id = %s AND planning_date = %s
                """, (
                    p.get("crew_id"), p.get("crew_name"), function_id, p.get("function_name"),
                    project_id, p.get("project_name"), p.get("project_code"),
                    p.get("subproject_id"), p.get("location_id"), p.get("location_name"), p.get("location_address"),
                    p.get("location_lat"), p.get("location_lon"),
                    p.get("gps_timbratura_location"),
                    plan_start, plan_end,
                    p.get("break_start"), p.get("break_end"), p.get("break_minutes"),
                    hours_planned, hours_registered,
                    new_remark, new_remark_planner, 1 if p.get("is_leader") else 0, p.get("transport"), now_ms,
                    rentman_id, target_date
                ))
            else:
                db.execute("""
                    UPDATE rentman_plannings SET
                        crew_id = ?, crew_name = ?, function_id = ?, function_name = ?,
                        project_id = ?, project_name = ?, project_code = ?,
                        subproject_id = ?, location_id = ?, location_name = ?, location_address = ?,
                        location_lat = ?, location_lon = ?,
                        gps_timbratura_location = ?,
                        plan_start = ?, plan_end = ?,
                        break_start = ?, break_end = ?, break_minutes = ?,
                        hours_planned = ?, hours_registered = ?,
                        remark = COALESCE(?, remark), remark_planner = COALESCE(?, remark_planner), 
                        is_leader = ?, transport = ?, updated_ts = ?,
                        is_obsolete = 0
                    WHERE rentman_id = ? AND planning_date = ?
                """, (
                    p.get("crew_id"), p.get("crew_name"), function_id, p.get("function_name"),
                    project_id, p.get("project_name"), p.get("project_code"),
                    p.get("subproject_id"), p.get("location_id"), p.get("location_name"), p.get("location_address"),
                    p.get("location_lat"), p.get("location_lon"),
                    p.get("gps_timbratura_location"),
                    plan_start, plan_end,
                    p.get("break_start"), p.get("break_end"), p.get("break_minutes"),
                    hours_planned, hours_registered,
                    new_remark, new_remark_planner, 1 if p.get("is_leader") else 0, p.get("transport"), now_ms,
                    rentman_id, target_date
                ))
            updated += 1
        else:
            # Insert new record
            app.logger.warning(f"🔴 SAVE INSERT DB: rentman_id={rentman_id}, location_name={p.get('location_name')}, lat={p.get('location_lat')}, lon={p.get('location_lon')}")
            if DB_VENDOR == "mysql":
                db.execute("""
                    INSERT INTO rentman_plannings (
                        rentman_id, planning_date, crew_id, crew_name, function_id, function_name,
                        project_id, project_name, project_code, subproject_id, location_id, location_name, location_address,
                        location_lat, location_lon,
                        gps_timbratura_location, timbratura_gps_mode,
                        plan_start, plan_end, break_start, break_end, break_minutes,
                        hours_planned, hours_registered, remark, remark_planner, is_leader, transport,
                        sent_to_webservice, created_ts, updated_ts
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 0, %s, %s)
                """, (
                    rentman_id, target_date, p.get("crew_id"), p.get("crew_name"),
                    function_id, p.get("function_name"), project_id, p.get("project_name"),
                    p.get("project_code"), p.get("subproject_id"), p.get("location_id"), p.get("location_name"), p.get("location_address"),
                    p.get("location_lat"), p.get("location_lon"),
                    p.get("gps_timbratura_location"), p.get("timbratura_gps_mode") or "group",
                    plan_start, plan_end, p.get("break_start"), p.get("break_end"), p.get("break_minutes"),
                    hours_planned, hours_registered, p.get("remark"), p.get("remark_planner"),
                    1 if p.get("is_leader") else 0, p.get("transport"), now_ms, now_ms
                ))
            else:
                db.execute("""
                    INSERT INTO rentman_plannings (
                        rentman_id, planning_date, crew_id, crew_name, function_id, function_name,
                        project_id, project_name, project_code, subproject_id, location_id, location_name, location_address,
                        location_lat, location_lon,
                        gps_timbratura_location, timbratura_gps_mode,
                        plan_start, plan_end, break_start, break_end, break_minutes,
                        hours_planned, hours_registered, remark, remark_planner, is_leader, transport,
                        sent_to_webservice, created_ts, updated_ts
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
                """, (
                    rentman_id, target_date, p.get("crew_id"), p.get("crew_name"),
                    function_id, p.get("function_name"), project_id, p.get("project_name"),
                    p.get("project_code"), p.get("subproject_id"), p.get("location_id"), p.get("location_name"), p.get("location_address"),
                    p.get("location_lat"), p.get("location_lon"),
                    p.get("gps_timbratura_location"), p.get("timbratura_gps_mode") or "group",
                    plan_start, plan_end, p.get("break_start"), p.get("break_end"), p.get("break_minutes"),
                    hours_planned, hours_registered, p.get("remark"), p.get("remark_planner"),
                    1 if p.get("is_leader") else 0, p.get("transport"), now_ms, now_ms
                ))
            saved += 1

    # Raccogli tutti i rentman_id ricevuti dalla sincronizzazione
    # IMPORTANTE: il frontend invia sia 'rentman_id' (originale Rentman) che 'id' (può essere DB id o Rentman id)
    # Preferisci 'rentman_id' se presente, altrimenti usa 'id'
    synced_rentman_ids = []
    for p in plannings:
        rid = p.get("rentman_id") or p.get("id")
        if rid:
            synced_rentman_ids.append(rid)
    app.logger.info(f"🔴 synced_rentman_ids raccolti: {synced_rentman_ids}")
    
    # Rimuovi i turni NON PIÙ INVIATI (sent_to_webservice = 0) che non sono nella sincronizzazione
    # Marca come obsoleti i turni già inviati (sent_to_webservice = 1) che non sono più in Rentman
    removed = 0
    marked_obsolete = 0
    
    if synced_rentman_ids:
        if DB_VENDOR == "mysql":
            placeholders = ",".join(["%s"] * len(synced_rentman_ids))
            # Rimuovi turni non più presenti che NON sono stati ancora inviati
            result = db.execute(f"""
                DELETE FROM rentman_plannings 
                WHERE planning_date = %s 
                  AND sent_to_webservice = 0
                  AND rentman_id NOT IN ({placeholders})
            """, (target_date, *synced_rentman_ids))
            removed = result.rowcount if hasattr(result, 'rowcount') else 0
            
            # Marca come obsoleti i turni già inviati che non sono più in Rentman
            result_obsolete = db.execute(f"""
                UPDATE rentman_plannings
                SET is_obsolete = 1, updated_ts = %s
                WHERE planning_date = %s 
                  AND sent_to_webservice = 1
                  AND rentman_id NOT IN ({placeholders})
            """, (now_ms, target_date, *synced_rentman_ids))
            marked_obsolete = result_obsolete.rowcount if hasattr(result_obsolete, 'rowcount') else 0
        else:
            placeholders = ",".join(["?"] * len(synced_rentman_ids))
            result = db.execute(f"""
                DELETE FROM rentman_plannings 
                WHERE planning_date = ? 
                  AND sent_to_webservice = 0
                  AND rentman_id NOT IN ({placeholders})
            """, (target_date, *synced_rentman_ids))
            removed = result.rowcount if hasattr(result, 'rowcount') else 0
            
            # Marca come obsoleti i turni già inviati che non sono più in Rentman
            result_obsolete = db.execute(f"""
                UPDATE rentman_plannings
                SET is_obsolete = 1, updated_ts = ?
                WHERE planning_date = ? 
                  AND sent_to_webservice = 1
                  AND rentman_id NOT IN ({placeholders})
            """, (now_ms, target_date, *synced_rentman_ids))
            marked_obsolete = result_obsolete.rowcount if hasattr(result_obsolete, 'rowcount') else 0
    
    db.commit()

    return jsonify({
        "ok": True,
        "saved": saved,
        "updated": updated,
        "removed": removed,
        "marked_obsolete": marked_obsolete,
        "total": saved + updated,
    })


@app.get("/api/admin/rentman-planning/saved")
@login_required
def api_admin_rentman_planning_saved() -> ResponseReturnValue:
    """Recupera le pianificazioni salvate dal database locale."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403

    target_date = request.args.get("date")
    if not target_date:
        target_date = datetime.now().date().isoformat()

    db = get_db()
    ensure_rentman_plannings_table(db)

    if DB_VENDOR == "mysql":
        rows = db.execute(
            "SELECT * FROM rentman_plannings WHERE planning_date = %s AND is_obsolete = 0 ORDER BY plan_start, crew_name",
            (target_date,)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM rentman_plannings WHERE planning_date = ? AND is_obsolete = 0 ORDER BY plan_start, crew_name",
            (target_date,)
        ).fetchall()

    plannings = []
    for row in rows:
        if isinstance(row, Mapping):
            plannings.append(dict(row))
        else:
            # SQLite row to dict - ordine colonne deve corrispondere alla tabella
            cols = ["id", "rentman_id", "planning_date", "crew_id", "crew_name", "function_id",
                    "function_name", "project_id", "project_name", "project_code", 
                    "subproject_id", "location_id", "location_name", "location_address",
                    "location_lat", "location_lon",
                    "timbratura_gps_mode", "gps_timbratura_location",
                    "plan_start", "plan_end", "break_start", "break_end", "break_minutes",
                    "hours_planned", "hours_registered", "remark", "remark_planner", "is_leader",
                    "transport", "sent_to_webservice", "sent_ts", "webservice_response",
                    "created_ts", "updated_ts", "is_obsolete"]
            plannings.append(dict(zip(cols, row)))

    # DEBUG: Log GPS mode per ogni pianificazione
    for p in plannings:
        app.logger.info(f"📍 SAVED API: {p.get('crew_name')} - gps_mode={p.get('timbratura_gps_mode')}, gps_loc={p.get('gps_timbratura_location')}, lat={p.get('location_lat')}, lon={p.get('location_lon')}")

    return jsonify({
        "ok": True,
        "date": target_date,
        "count": len(plannings),
        "plannings": plannings,
    })


@app.post("/api/admin/rentman-planning/send")
@login_required
def api_admin_rentman_planning_send() -> ResponseReturnValue:
    """Invia le pianificazioni al webservice esterno e notifica gli utenti."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403

    data = request.get_json() or {}
    planning_ids = data.get("ids", [])  # Lista di ID locali da inviare
    plannings = data.get("plannings", [])  # O lista di pianificazioni con rentman_id
    force_resend = data.get("force_resend", False)  # Se True, invia notifiche anche per già inviati

    db = get_db()
    ensure_rentman_plannings_table(db)

    rows = []
    
    if planning_ids:
        # Recupera le pianificazioni per ID locale
        placeholders = ",".join(["%s" if DB_VENDOR == "mysql" else "?"] * len(planning_ids))
        query = f"SELECT * FROM rentman_plannings WHERE id IN ({placeholders})"
        rows = db.execute(query, planning_ids).fetchall()
    elif plannings:
        # Recupera per rentman_id (se esistono nel DB)
        rentman_ids = [p.get("rentman_id") or p.get("id") for p in plannings if p.get("rentman_id") or p.get("id")]
        if rentman_ids:
            placeholders = ",".join(["%s" if DB_VENDOR == "mysql" else "?"] * len(rentman_ids))
            query = f"SELECT * FROM rentman_plannings WHERE rentman_id IN ({placeholders})"
            rows = db.execute(query, rentman_ids).fetchall()

    if not rows and not plannings:
        return jsonify({"error": "Nessuna pianificazione selezionata"}), 400

    # TODO: Implementare l'invio al webservice
    # Per ora segna come inviate
    now_ms_val = int(time.time() * 1000)
    sent_count = 0
    errors = []
    
    # Raccogli info per notifiche push
    users_to_notify: Dict[int, List[Dict[str, Any]]] = {}  # crew_id -> list of turni

    # Se abbiamo righe dal DB, processa quelle
    if rows:
        for row in rows:
            local_id = row["id"] if isinstance(row, Mapping) else row[0]
            crew_id = row["crew_id"] if isinstance(row, Mapping) else row[3]
            planning_date = row["planning_date"] if isinstance(row, Mapping) else row[2]
            project_name = row["project_name"] if isinstance(row, Mapping) else row[8]
            
            # Controlla se era già stato inviato
            was_sent = row["sent_to_webservice"] if isinstance(row, Mapping) else row[17]
            
            try:
                # TODO: Chiamata al webservice qui
                # response = requests.post(webservice_url, json=payload)
                # webservice_response = response.text
                
                webservice_response = "OK - Placeholder (webservice non configurato)"
                
                # Aggiorna il record come inviato
                if DB_VENDOR == "mysql":
                    db.execute("""
                        UPDATE rentman_plannings 
                        SET sent_to_webservice = 1, sent_ts = %s, webservice_response = %s, updated_ts = %s
                        WHERE id = %s
                    """, (now_ms_val, webservice_response, now_ms_val, local_id))
                else:
                    db.execute("""
                        UPDATE rentman_plannings 
                        SET sent_to_webservice = 1, sent_ts = ?, webservice_response = ?, updated_ts = ?
                        WHERE id = ?
                    """, (now_ms_val, webservice_response, now_ms_val, local_id))
                
                sent_count += 1
                
                # Aggiungi alla lista per notifica:
                # - Se non era già inviato (primo invio)
                # - OPPURE se force_resend=True (reinvio manuale o turno modificato)
                should_notify = not was_sent or force_resend
                if should_notify and crew_id:
                    if crew_id not in users_to_notify:
                        users_to_notify[crew_id] = []
                    users_to_notify[crew_id].append({
                        "date": str(planning_date)[:10] if planning_date else "",
                        "project": project_name or "Progetto",
                        "is_update": was_sent  # Per differenziare il messaggio notifica
                    })
                
            except Exception as exc:
                app.logger.error("Errore invio pianificazione %s: %s", local_id, exc)
                errors.append({"id": local_id, "error": str(exc)})
    else:
        # Se non abbiamo righe dal DB ma abbiamo pianificazioni raw, segna come "inviate" ma non salvate
        # Questo è un caso limite: l'utente vuole inviare senza salvare nel DB locale
        for p in plannings:
            try:
                # TODO: Chiamata al webservice qui
                webservice_response = "OK - Placeholder (webservice non configurato, non salvato nel DB)"
                sent_count += 1
            except Exception as exc:
                errors.append({"id": p.get("id"), "error": str(exc)})

    db.commit()
    
    # Invia notifiche push agli utenti
    notifications_sent = 0
    if users_to_notify:
        notifications_sent = _send_turni_notifications(db, users_to_notify)

    return jsonify({
        "ok": True,
        "sent": sent_count,
        "notifications_sent": notifications_sent,
        "errors": errors,
    })


def _send_turni_notifications(db: DatabaseLike, users_to_notify: Dict[int, List[Dict[str, Any]]]) -> int:
    """Invia notifiche push agli utenti per i nuovi turni pubblicati."""
    app.logger.info("_send_turni_notifications chiamata con %d crew_id", len(users_to_notify))
    
    settings = get_webpush_settings()
    if not settings:
        app.logger.info("Notifiche push non configurate, skip invio turni")
        return 0
    
    notifications_sent = 0
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    for crew_id, turni in users_to_notify.items():
        app.logger.info("Cerco utente per crew_id=%s", crew_id)
        
        # Trova l'utente associato a questo crew_id
        user_row = db.execute(
            f"SELECT username FROM app_users WHERE rentman_crew_id = {placeholder}",
            (crew_id,)
        ).fetchone()
        
        if not user_row:
            app.logger.warning("Nessun utente trovato per crew_id=%s", crew_id)
            continue
        
        username = user_row['username'] if isinstance(user_row, dict) else user_row[0]
        app.logger.info("Trovato utente %s per crew_id=%s", username, crew_id)
        
        # Recupera le subscription push dell'utente
        subscriptions = db.execute(
            f"SELECT endpoint, p256dh, auth FROM push_subscriptions WHERE username = {placeholder}",
            (username,)
        ).fetchall()
        
        if not subscriptions:
            app.logger.warning("Nessuna subscription push per utente %s", username)
            continue
        
        app.logger.info("Trovate %d subscription per %s", len(subscriptions), username)
        
        # Prepara il messaggio
        if len(turni) == 1:
            title = "📅 Nuovo turno pubblicato"
            body = f"{turni[0]['date']} - {turni[0]['project']}"
        else:
            title = f"📅 {len(turni)} nuovi turni pubblicati"
            body = f"Hai {len(turni)} nuovi turni. Apri l'app per i dettagli."
        
        payload = {
            "title": title,
            "body": body,
            "icon": "/static/icons/icon-192.png",
            "badge": "/static/icons/badge-72.png",
            "tag": "turni-update",
            "data": {
                "url": "/user/turni",
                "type": "turni_published"
            }
        }
        
        # Invia a tutte le subscription dell'utente
        for sub in subscriptions:
            endpoint = sub['endpoint'] if isinstance(sub, dict) else sub[0]
            p256dh = sub['p256dh'] if isinstance(sub, dict) else sub[1]
            auth = sub['auth'] if isinstance(sub, dict) else sub[2]
            
            subscription_info = {
                "endpoint": endpoint,
                "keys": {
                    "p256dh": p256dh,
                    "auth": auth
                }
            }
            
            try:
                webpush(
                    subscription_info=subscription_info,
                    data=json.dumps(payload),
                    vapid_private_key=settings["vapid_private"],
                    vapid_claims={"sub": settings["subject"]},
                    ttl=86400,  # 24 ore
                )
                notifications_sent += 1
                app.logger.info("Notifica turno inviata a %s", username)
                    
            except WebPushException as e:
                app.logger.warning("Errore invio notifica turno a %s: %s", username, e)
                # Rimuovi subscription se non valida
                if e.response and e.response.status_code in {404, 410}:
                    remove_push_subscription(db, endpoint)
            except Exception as e:
                app.logger.error("Errore generico invio notifica turno: %s", e)
        
        # Salva la notifica nel log (una volta per utente, dopo aver provato tutte le subscription)
        if notifications_sent > 0:
            try:
                record_push_notification(
                    db,
                    kind="turni_published",
                    title=title,
                    body=body,
                    payload=payload,
                    username=username,
                )
                app.logger.info("Notifica turno salvata nel log per %s", username)
            except Exception as e:
                app.logger.error("Errore salvataggio notifica nel log: %s", e)
    
    return notifications_sent


def _send_document_notifications(
    db: DatabaseLike,
    category: str,
    title: str,
    target_all: bool,
    target_users_json: str,
    doc_id: int = None
) -> int:
    """Invia notifiche push per un nuovo documento caricato."""
    settings = get_webpush_settings()
    if not settings:
        app.logger.info("Notifiche push non configurate, skip invio documento")
        return 0
    
    notifications_sent = 0
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Etichette categoria per il messaggio
    category_labels = {
        "circolare": "📋 Nuova Circolare",
        "comunicazione": "📢 Nuova Comunicazione",
        "busta_paga": "💰 Nuovo Cedolino"
    }
    
    notification_title = category_labels.get(category, "📄 Nuovo Documento")
    notification_body = title
    
    # Costruisci URL con documento specifico se disponibile
    target_url = "/user/documents"
    if doc_id:
        target_url = f"/user/documents?doc={doc_id}"
    
    payload = {
        "title": notification_title,
        "body": notification_body,
        "icon": "/static/icons/icon-192x192.png",
        "badge": "/static/icons/icon-72x72.png",
        "tag": f"document-{category}",
        "data": {
            "url": target_url,
            "type": "new_document",
            "doc_id": doc_id
        }
    }
    
    # Determina i destinatari
    target_usernames = []
    
    if target_all:
        # Tutti gli operatori (role = 'user')
        rows = db.execute(
            f"SELECT username FROM app_users WHERE role = {placeholder} AND is_active = 1",
            ("user",)
        ).fetchall()
        target_usernames = [r['username'] if isinstance(r, dict) else r[0] for r in rows]
    else:
        # Utenti specifici dal JSON
        try:
            target_usernames = json.loads(target_users_json) if target_users_json else []
        except:
            target_usernames = []
    
    app.logger.info("Invio notifica documento a %d utenti", len(target_usernames))
    
    for username in target_usernames:
        # Recupera le subscription push dell'utente
        subscriptions = db.execute(
            f"SELECT endpoint, p256dh, auth FROM push_subscriptions WHERE username = {placeholder}",
            (username,)
        ).fetchall()
        
        if not subscriptions:
            continue
        
        user_notified = False
        
        # Invia a tutte le subscription dell'utente
        for sub in subscriptions:
            endpoint = sub['endpoint'] if isinstance(sub, dict) else sub[0]
            p256dh = sub['p256dh'] if isinstance(sub, dict) else sub[1]
            auth = sub['auth'] if isinstance(sub, dict) else sub[2]
            
            subscription_info = {
                "endpoint": endpoint,
                "keys": {
                    "p256dh": p256dh,
                    "auth": auth
                }
            }
            
            try:
                webpush(
                    subscription_info=subscription_info,
                    data=json.dumps(payload),
                    vapid_private_key=settings["vapid_private"],
                    vapid_claims={"sub": settings["subject"]},
                    ttl=86400,
                )
                notifications_sent += 1
                user_notified = True
                app.logger.info("Notifica documento inviata a %s", username)
                    
            except WebPushException as e:
                app.logger.warning("Errore invio notifica documento a %s: %s", username, e)
                if e.response and e.response.status_code in {404, 410}:
                    remove_push_subscription(db, endpoint)
            except Exception as e:
                app.logger.error("Errore generico invio notifica documento: %s", e)
        
        # Salva nel log
        if user_notified:
            try:
                record_push_notification(
                    db,
                    kind="new_document",
                    title=notification_title,
                    body=notification_body,
                    payload=payload,
                    username=username,
                )
            except Exception as e:
                app.logger.error("Errore salvataggio notifica documento nel log: %s", e)
    
    app.logger.info("Inviate %d notifiche documento totali", notifications_sent)
    return notifications_sent


# ═══════════════════════════════════════════════════════════════════════════════
#  QR CODE TIMBRATURA - GiQR LEVEL 4 LITE
# ═══════════════════════════════════════════════════════════════════════════════

QR_DEVICE_ID = os.environ.get("QR_DEVICE_ID", "WebApp-001")
QR_REFRESH_SECONDS = int(os.environ.get("QR_REFRESH_SECONDS", "10"))


def _generate_giqr_payload() -> dict:
    """Genera payload GiQR Level 4 Lite con firma."""
    now = datetime.now()
    
    dt_str = now.strftime("%Y%m%d%H%M%S")
    nonce = random.randint(0, 65535)
    
    # Firma semplice (algoritmo dal codice originale)
    sig = (
        now.hour * 97 +
        now.minute * 59 +
        now.second * 31 +
        now.year +
        now.month +
        now.day +
        sum(QR_DEVICE_ID.encode())
    ) % 100000
    
    return {
        "v": 1,
        "dt": dt_str,
        "dev": QR_DEVICE_ID,
        "n": nonce,
        "sig": sig
    }


@app.route("/api/qr-timbratura", methods=["GET"])
@login_required
def api_qr_timbratura():
    """Genera QR code dinamico per timbratura in formato base64 PNG."""
    payload = _generate_giqr_payload()
    
    # JSON → bytes → Base64
    raw_json = json.dumps(payload).encode("utf-8")
    b64_payload = base64.b64encode(raw_json).decode("utf-8")
    
    # Genera QR code
    qr = qrcode.QRCode(box_size=8, border=2)
    qr.add_data(b64_payload)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Converti in base64 per invio al frontend
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    img_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    
    return jsonify({
        "ok": True,
        "image": f"data:image/png;base64,{img_base64}",
        "payload": payload,
        "refresh_seconds": QR_REFRESH_SECONDS,
        "timestamp": datetime.now().isoformat()
    })


@app.route("/qr-timbratura")
@login_required
def qr_timbratura_page():
    """Pagina standalone per QR timbratura (schermo intero)."""
    return render_template(
        "qr_timbratura.html",
        refresh_seconds=QR_REFRESH_SECONDS,
        device_id=QR_DEVICE_ID
    )


def _validate_giqr_payload(payload: dict) -> Tuple[bool, str]:
    """
    Valida un payload GiQR.
    Ritorna (is_valid, error_message).
    """
    if not payload:
        return False, "Payload vuoto"
    
    required_fields = ["v", "dt", "dev", "n", "sig"]
    for field in required_fields:
        if field not in payload:
            return False, f"Campo mancante: {field}"
    
    # Verifica versione
    if payload.get("v") != 1:
        return False, "Versione QR non supportata"
    
    # Verifica timestamp (non più vecchio di 60 secondi)
    dt_str = payload.get("dt", "")
    try:
        qr_time = datetime.strptime(dt_str, "%Y%m%d%H%M%S")
        now = datetime.now()
        age_seconds = abs((now - qr_time).total_seconds())
        if age_seconds > 60:
            return False, f"QR scaduto ({int(age_seconds)} secondi)"
    except ValueError:
        return False, "Timestamp non valido"
    
    # Verifica firma
    dev = payload.get("dev", "")
    sig_expected = (
        qr_time.hour * 97 +
        qr_time.minute * 59 +
        qr_time.second * 31 +
        qr_time.year +
        qr_time.month +
        qr_time.day +
        sum(dev.encode())
    ) % 100000
    
    if payload.get("sig") != sig_expected:
        return False, "Firma QR non valida"
    
    return True, ""


@app.post("/api/timbratura/validate-qr")
@login_required
def api_timbratura_validate_qr():
    """Valida un QR code scansionato per la timbratura."""
    data = request.get_json()
    if not data:
        return jsonify({"valid": False, "error": "Dati mancanti"}), 400
    
    qr_data = data.get("qr_data", "")
    
    # Decodifica Base64 → JSON
    try:
        decoded = base64.b64decode(qr_data).decode("utf-8")
        payload = json.loads(decoded)
    except Exception:
        return jsonify({"valid": False, "error": "QR non valido"}), 400
    
    # Valida payload
    is_valid, error = _validate_giqr_payload(payload)
    
    if not is_valid:
        return jsonify({"valid": False, "error": error}), 400
    
    # Genera un token temporaneo di validazione (valido 5 minuti)
    validation_token = secrets.token_urlsafe(32)
    expires = now_ms() + (5 * 60 * 1000)  # 5 minuti
    
    # Salva in sessione
    session["timbratura_token"] = validation_token
    session["timbratura_token_expires"] = expires
    session["timbratura_method"] = "qr"
    session["timbratura_location"] = None  # QR non ha location
    session["timbratura_gps_lat"] = None
    session["timbratura_gps_lon"] = None
    
    return jsonify({
        "valid": True,
        "token": validation_token,
        "expires_in": 300,
        "device": payload.get("dev", "")
    })


# ═══════════════════════════════════════════════════════════════════════════════
#  TIMBRATURA GPS - Geolocalizzazione
# ═══════════════════════════════════════════════════════════════════════════════

def get_timbratura_config() -> Dict[str, Any]:
    """Restituisce la configurazione timbratura.
    
    Priorità:
    1. Database (company_settings.custom_settings.timbratura)
    2. config.json (fallback per retrocompatibilità)
    3. Valori di default
    """
    default_config = {
        "qr_enabled": True,
        "gps_enabled": False,
        "gps_locations": [],
        "gps_max_accuracy_meters": 50
    }
    
    # Prova prima dal database
    try:
        db = get_db()
        company_settings = get_company_settings(db)
        custom_settings = company_settings.get("custom_settings", {})
        
        if isinstance(custom_settings, str):
            custom_settings = json.loads(custom_settings) if custom_settings else {}
        
        if "timbratura" in custom_settings:
            db_config = custom_settings["timbratura"]
            # Merge con i default
            return {
                "qr_enabled": db_config.get("qr_enabled", True),
                "gps_enabled": db_config.get("gps_enabled", False),
                "gps_locations": db_config.get("gps_locations", []),
                "gps_max_accuracy_meters": db_config.get("gps_max_accuracy_meters", 50)
            }
    except Exception as e:
        app.logger.warning(f"Errore lettura config timbratura da DB: {e}")
    
    # Fallback: config.json (per retrocompatibilità)
    config = load_config()
    if "timbratura" in config:
        file_config = config["timbratura"]
        return {
            "qr_enabled": file_config.get("qr_enabled", True),
            "gps_enabled": file_config.get("gps_enabled", False),
            "gps_locations": file_config.get("gps_locations", []),
            "gps_max_accuracy_meters": file_config.get("gps_max_accuracy_meters", 50)
        }
    
    return default_config


def get_user_timbratura_config(member_key: str = None) -> Dict[str, Any]:
    """
    Restituisce la configurazione timbratura effettiva per un utente.
    
    Se l'utente ha delle eccezioni configurate (timbratura_override),
    queste DISABILITANO i metodi per quell'operatore.
    
    La logica è:
    - Config aziendale definisce i metodi disponibili per tutti
    - Le eccezioni operatore DISABILITANO metodi specifici
    - Se tutti i metodi sono disabilitati = timbratura diretta
    """
    base_config = get_timbratura_config()
    
    if not member_key:
        return base_config
    
    # Cerca le eccezioni per questo operatore
    try:
        db = get_db()
        ensure_crew_members_table(db)
        
        # Cerca prima per corrispondenza esatta, poi per corrispondenza parziale (nome inizia con)
        row = None
        member_key_lower = member_key.lower().strip()
        
        if DB_VENDOR == "mysql":
            # Prima cerca corrispondenza esatta
            row = db.execute(
                "SELECT timbratura_override FROM crew_members WHERE rentman_id = %s OR LOWER(name) = %s LIMIT 1",
                (member_key, member_key_lower)
            ).fetchone()
            
            # Se non trovato, cerca per nome che inizia con member_key (es. "angelo" -> "Angelo Ruggieri")
            if not row:
                row = db.execute(
                    "SELECT timbratura_override FROM crew_members WHERE LOWER(name) LIKE %s AND timbratura_override IS NOT NULL LIMIT 1",
                    (member_key_lower + '%',)
                ).fetchone()
        else:
            row = db.execute(
                "SELECT timbratura_override FROM crew_members WHERE rentman_id = ? OR LOWER(name) = ? LIMIT 1",
                (member_key, member_key_lower)
            ).fetchone()
            
            if not row:
                row = db.execute(
                    "SELECT timbratura_override FROM crew_members WHERE LOWER(name) LIKE ? AND timbratura_override IS NOT NULL LIMIT 1",
                    (member_key_lower + '%',)
                ).fetchone()
        
        if row and row[0]:
            override = row[0]
            if isinstance(override, str):
                override = json.loads(override)
            
            app.logger.info(f"Trovata eccezione timbratura per '{member_key}': {override}")
            
            # Le eccezioni DISABILITANO i metodi (override con False)
            if "qr_disabled" in override and override.get("qr_disabled"):
                base_config["qr_enabled"] = False
            if "gps_disabled" in override and override.get("gps_disabled"):
                base_config["gps_enabled"] = False
                
    except Exception as e:
        app.logger.warning(f"Errore lettura eccezioni timbratura per {member_key}: {e}")
    
    return base_config


def haversine_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """
    Calcola la distanza in metri tra due coordinate usando la formula di Haversine.
    """
    import math
    R = 6371000  # Raggio della Terra in metri
    
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    
    a = math.sin(delta_phi / 2) ** 2 + \
        math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    
    return R * c


@app.get("/api/timbratura/config")
@login_required
def api_timbratura_config():
    """Restituisce le opzioni di timbratura disponibili per l'utente corrente."""
    # Recupera il nome utente dalla sessione per verificare eventuali eccezioni
    # user_name contiene il nome completo che dovrebbe corrispondere a crew_members.name
    user_full_name = session.get("user_name") or session.get("user_display") or session.get("user")
    timb_config = get_user_timbratura_config(user_full_name)
    
    app.logger.debug(f"Timbratura config per utente '{user_full_name}': qr={timb_config.get('qr_enabled')}, gps={timb_config.get('gps_enabled')}")
    
    # Includi le coordinate delle locations per il calcolo della distanza lato client
    locations = []
    for loc in timb_config.get("gps_locations", []):
        locations.append({
            "name": loc.get("name", "Sede"),
            "latitude": loc.get("latitude"),
            "longitude": loc.get("longitude"),
            "radius_meters": loc.get("radius_meters", 300)
        })
    
    return jsonify({
        "qr_enabled": timb_config.get("qr_enabled", True),
        "gps_enabled": timb_config.get("gps_enabled", False),
        "gps_locations": locations,
        "gps_max_accuracy_meters": timb_config.get("gps_max_accuracy_meters", 50)
    })


@app.post("/api/timbratura/validate-gps")
@login_required
def api_timbratura_validate_gps():
    """
    Valida la posizione GPS dell'utente per la timbratura.
    Verifica che l'utente si trovi entro il raggio di una delle sedi configurate.
    Se il turno dell'utente ha una sede specifica, valida solo contro quella sede.
    """
    data = request.get_json()
    if not data:
        return jsonify({"valid": False, "error": "Dati mancanti"}), 400
    
    latitude = data.get("latitude")
    longitude = data.get("longitude")
    accuracy = data.get("accuracy", 999999)  # Accuratezza in metri
    shift_location_name = data.get("shift_location_name")
    # Coordinate della sede inviate dal frontend (calcolate dal backend /api/user/turno-oggi)
    shift_location_lat = data.get("shift_location_lat")
    shift_location_lon = data.get("shift_location_lon")
    
    app.logger.info(f"VALIDATE-GPS: shift_location_name={shift_location_name}, lat={latitude}, lon={longitude}, accuracy={accuracy}")
    app.logger.info(f"VALIDATE-GPS: shift_location_lat={shift_location_lat}, shift_location_lon={shift_location_lon}")
    
    if latitude is None or longitude is None:
        return jsonify({"valid": False, "error": "Coordinate GPS mancanti"}), 400
    
    try:
        latitude = float(latitude)
        longitude = float(longitude)
        accuracy = float(accuracy)
    except (TypeError, ValueError):
        return jsonify({"valid": False, "error": "Coordinate non valide"}), 400
    
    # Recupera il nome utente dalla sessione per verificare eventuali eccezioni
    username = session.get("user")
    user_full_name = session.get("user_name") or session.get("user_display") or username
    timb_config = get_user_timbratura_config(user_full_name)
    
    # Verifica se GPS è abilitato (considerando eccezioni utente)
    if not timb_config.get("gps_enabled", False):
        return jsonify({"valid": False, "error": "Timbratura GPS non abilitata"}), 400
    
    # Verifica accuratezza del GPS
    max_accuracy = timb_config.get("gps_max_accuracy_meters", 50)
    if accuracy > max_accuracy:
        return jsonify({
            "valid": False, 
            "error": f"Precisione GPS insufficiente ({int(accuracy)}m). Richiesta: max {max_accuracy}m. Prova all'aperto o attendi qualche secondo."
        }), 400
    
    locations = timb_config.get("gps_locations", [])
    if not locations:
        return jsonify({"valid": False, "error": "Nessuna sede configurata per timbratura GPS"}), 400
    
    # Debug logging
    app.logger.info(f"GPS Validate - Input: lat={latitude}, lon={longitude}, acc={accuracy}")
    app.logger.info(f"GPS Validate - Locations raw: {locations}")
    for loc in locations:
        app.logger.info(f"GPS Validate - Location '{loc.get('name')}': lat={loc.get('latitude')} ({type(loc.get('latitude'))}), lon={loc.get('longitude')} ({type(loc.get('longitude'))})")
    
    # Verifica se il turno dell'utente ha una sede specifica
    shift_location_name = data.get('shift_location_name')  # Sede specifica passata dal frontend
    app.logger.info(f"VALIDATE-GPS: Ricevuto shift_location_name dal frontend: '{shift_location_name}'")
    
    if not shift_location_name and username:
        # Prova a recuperare dal turno odierno
        db = get_db()
        ensure_employee_shifts_table(db)
        day_of_week = datetime.now().weekday()
        placeholder = "%s" if DB_VENDOR == "mysql" else "?"
        shift_row = db.execute(f"""
            SELECT location_name FROM employee_shifts
            WHERE username = {placeholder} AND day_of_week = {placeholder} AND is_active = 1
        """, (username, day_of_week)).fetchone()
        if shift_row:
            shift_location_name = shift_row['location_name'] if isinstance(shift_row, dict) else shift_row[0]
    
    # Se il turno ha una sede specifica, filtra le locations per usare solo quella
    if shift_location_name:
        filtered_locations = [loc for loc in locations if loc.get('name') == shift_location_name]
        if filtered_locations:
            locations = filtered_locations
            app.logger.info(f"VALIDATE-GPS: Sede specifica '{shift_location_name}' trovata! Filtrando locations. Rimaste: {len(locations)}")
        else:
            app.logger.warning(f"VALIDATE-GPS: Sede '{shift_location_name}' NON trovata nelle locations disponibili: {[l.get('name') for l in locations]}")
            # Se la sede non è trovata in configurazione ma abbiamo le coordinate dal backend, usale
            if shift_location_lat is not None and shift_location_lon is not None:
                app.logger.info(f"VALIDATE-GPS: Usando coordinate ricevute dal frontend: lat={shift_location_lat}, lon={shift_location_lon}")
                # Crea una location virtuale con le coordinate ricevute (convertendo in float)
                try:
                    virtual_lat = float(shift_location_lat)
                    virtual_lon = float(shift_location_lon)
                except (ValueError, TypeError):
                    app.logger.error(f"VALIDATE-GPS: Errore conversione coordinate: lat={shift_location_lat}, lon={shift_location_lon}")
                    return jsonify({"valid": False, "error": "Coordinate sede non valide"}), 400
                
                # Recupera il raggio dalla cache se disponibile
                virtual_radius = 300
                db = get_db()
                cached = get_location_cache(db, shift_location_name)
                if cached:
                    virtual_radius = cached[2]  # Il terzo elemento è il raggio
                    app.logger.info(f"VALIDATE-GPS: Usando raggio dalla cache: {virtual_radius}m")
                
                virtual_location = {
                    "name": shift_location_name or "Sede",
                    "latitude": virtual_lat,
                    "longitude": virtual_lon,
                    "radius_meters": virtual_radius
                }
                locations = [virtual_location]
    
    # Verifica se l'utente è entro il raggio di una delle sedi
    matched_location = None
    min_distance = float('inf')
    
    for loc in locations:
        loc_lat = loc.get("latitude")
        loc_lon = loc.get("longitude")
        loc_radius = loc.get("radius_meters", 300)
        
        if loc_lat is None or loc_lon is None:
            continue
        
        distance = haversine_distance(latitude, longitude, loc_lat, loc_lon)
        
        if distance < min_distance:
            min_distance = distance
        
        # Considera anche l'accuratezza del GPS nell'errore
        effective_distance = distance - accuracy  # Distanza minima possibile
        
        if effective_distance <= loc_radius:
            matched_location = loc
            break
    
    if not matched_location:
        return jsonify({
            "valid": False, 
            "error": f"Non sei in una sede autorizzata. Distanza minima: {int(min_distance)}m",
            "distance": int(min_distance)
        }), 400
    
    # GPS valido! Genera token come per QR
    validation_token = secrets.token_urlsafe(32)
    expires = now_ms() + (5 * 60 * 1000)  # 5 minuti
    
    # Salva in sessione (stesso meccanismo del QR)
    session["timbratura_token"] = validation_token
    session["timbratura_token_expires"] = expires
    session["timbratura_method"] = "gps"
    session["timbratura_location"] = matched_location.get("name", "Sede")
    session["timbratura_gps_lat"] = latitude
    session["timbratura_gps_lon"] = longitude
    
    app.logger.info(f"GPS validato per {session.get('user')}: {matched_location.get('name')} (distanza: {int(min_distance)}m, accuracy: {int(accuracy)}m)")
    
    return jsonify({
        "valid": True,
        "token": validation_token,
        "expires_in": 300,
        "location": matched_location.get("name", "Sede"),
        "distance": int(min_distance),
        "accuracy": int(accuracy)
    })


# ═══════════════════════════════════════════════════════════════════════════════
#  PLANNING GRUPPO - VISTA SETTIMANALE
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/admin/group-planning")
@app.get("/admin/group-planning/<int:group_id>")
@login_required
def admin_group_planning_page(group_id: Optional[int] = None) -> ResponseReturnValue:
    """Pagina planning settimanale per gruppo."""
    if not session.get("is_admin") and session.get("role") != "supervisor":
        abort(403)

    display_name = session.get("user_display") or session.get("user_name") or session.get("user")
    primary_name = session.get("user_name") or display_name or session.get("user")
    initials = session.get("user_initials") or compute_initials(primary_name or "")

    db = get_db()
    ensure_user_groups_table(db)

    # Recupera tutti i gruppi attivi
    groups = db.execute("SELECT id, name FROM user_groups WHERE is_active = 1 ORDER BY name").fetchall()
    groups_list = [{"id": g["id"] if isinstance(g, dict) else g[0], 
                    "name": g["name"] if isinstance(g, dict) else g[1]} for g in groups]

    # Se non specificato, usa il primo gruppo
    if group_id is None and groups_list:
        group_id = groups_list[0]["id"]
    
    group_name = "Nessun Gruppo"
    for g in groups_list:
        if g["id"] == group_id:
            group_name = g["name"]
            break

    return render_template(
        "admin_group_planning.html",
        user_name=primary_name,
        user_display=display_name,
        user_initials=initials,
        is_admin=session.get("is_admin"),
        groups=groups_list,
        group_id=group_id,
        group_name=group_name,
    )


@app.get("/api/admin/group-planning/<int:group_id>")
@login_required
def api_admin_group_planning(group_id: int) -> ResponseReturnValue:
    """API per recuperare il planning settimanale di un gruppo."""
    if not session.get("is_admin") and session.get("role") != "supervisor":
        return jsonify({"error": "forbidden"}), 403

    start_date = request.args.get("start")
    end_date = request.args.get("end")

    if not start_date or not end_date:
        return jsonify({"error": "Date start e end richieste"}), 400

    db = get_db()
    ensure_user_groups_table(db)
    ensure_rentman_plannings_table(db)

    placeholder = "%s" if DB_VENDOR == "mysql" else "?"

    # Recupera gli utenti del gruppo
    users = db.execute(f"""
        SELECT u.username, u.display_name, u.full_name, u.rentman_crew_id
        FROM app_users u
        WHERE u.group_id = {placeholder} AND u.is_active = 1
        ORDER BY u.display_name, u.username
    """, (group_id,)).fetchall()

    users_list = []
    crew_ids = []
    username_by_crew = {}

    for u in users:
        username = u["username"] if isinstance(u, dict) else u[0]
        display_name = u["display_name"] if isinstance(u, dict) else u[1]
        crew_id = u["rentman_crew_id"] if isinstance(u, dict) else u[3]
        
        users_list.append({
            "username": username,
            "display_name": display_name,
            "crew_id": crew_id
        })
        
        if crew_id:
            crew_ids.append(crew_id)
            username_by_crew[crew_id] = username

    shifts_list = []

    # Se ci sono crew_id, recupera i turni Rentman
    if crew_ids:
        placeholders = ",".join([placeholder] * len(crew_ids))
        
        plannings = db.execute(f"""
            SELECT crew_id, planning_date, project_code, project_name, function_name,
                   plan_start, plan_end, hours_planned, remark, is_leader, transport,
                   break_start, break_end, break_minutes, location_name, location_id
            FROM rentman_plannings
            WHERE crew_id IN ({placeholders})
              AND is_obsolete = 0
              AND planning_date >= {placeholder}
              AND planning_date <= {placeholder}
            ORDER BY planning_date, plan_start
        """, (*crew_ids, start_date, end_date)).fetchall()

        for p in plannings:
            crew_id = p["crew_id"] if isinstance(p, dict) else p[0]
            username = username_by_crew.get(crew_id)
            if not username:
                continue

            planning_date = p["planning_date"] if isinstance(p, dict) else p[1]
            plan_start = p["plan_start"] if isinstance(p, dict) else p[5]
            plan_end = p["plan_end"] if isinstance(p, dict) else p[6]
            hours_planned = p["hours_planned"] if isinstance(p, dict) else p[7]
            location_name = p["location_name"] if isinstance(p, dict) else p[14]
            location_id = p["location_id"] if isinstance(p, dict) else p[15]
            
            # Coordinate dalla cache globale
            location_lat, location_lon, location_radius = None, None, 300
            if location_name:
                ensure_location_cache_table(db)
                cached_coords = get_location_cache(db, location_name, location_id)
                if cached_coords:
                    location_lat, location_lon, location_radius = cached_coords
                    app.logger.info(f"✅ Group-planning: Location '{location_name}' (id={location_id}) - usando coordinate dalla cache: {location_lat}, {location_lon}, raggio={location_radius}m")

            # Normalizza data
            if hasattr(planning_date, 'isoformat'):
                date_str = planning_date.isoformat()
            else:
                date_str = str(planning_date)[:10]

            # Formatta orari
            start_str = ""
            end_str = ""
            if plan_start:
                if hasattr(plan_start, 'strftime'):
                    start_str = plan_start.strftime("%H:%M")
                else:
                    s = str(plan_start)
                    start_str = s[11:16] if len(s) > 11 else s[:5]
            if plan_end:
                if hasattr(plan_end, 'strftime'):
                    end_str = plan_end.strftime("%H:%M")
                else:
                    s = str(plan_end)
                    end_str = s[11:16] if len(s) > 11 else s[:5]

            shifts_list.append({
                "username": username,
                "date": date_str,
                "type": "shift",
                "project_code": p["project_code"] if isinstance(p, dict) else p[2],
                "project_name": p["project_name"] if isinstance(p, dict) else p[3],
                "function": p["function_name"] if isinstance(p, dict) else p[4],
                "start": start_str,
                "end": end_str,
                "hours": float(hours_planned or 0),
                "note": p["remark"] if isinstance(p, dict) else p[8],
                "is_leader": bool(p["is_leader"] if isinstance(p, dict) else p[9]),
                "location_name": location_name,
                "location_lat": location_lat,
                "location_lon": location_lon,
            })

    # Aggiungere le ferie/permessi approvati (solo tipo 1=Ferie e 3=Permesso)
    if users_list:
        usernames = [u["username"] for u in users_list]
        placeholders = ",".join([placeholder] * len(usernames))
        
        requests = db.execute(f"""
            SELECT username, request_type_id, date_from, date_to, notes, extra_data
            FROM user_requests
            WHERE username IN ({placeholders})
              AND status = 'approved'
              AND request_type_id IN (1, 3)
              AND date_from <= {placeholder}
              AND date_to >= {placeholder}
            ORDER BY date_from
        """, (*usernames, end_date, start_date)).fetchall()
        
        for req in requests:
            username = req["username"] if isinstance(req, dict) else req[0]
            request_type_id = req["request_type_id"] if isinstance(req, dict) else req[1]
            date_from = req["date_from"] if isinstance(req, dict) else req[2]
            date_to = req["date_to"] if isinstance(req, dict) else req[3]
            notes = req["notes"] if isinstance(req, dict) else req[4]
            extra_data_str = req["extra_data"] if isinstance(req, dict) else req[5]
            
            # Normalizza le date
            if hasattr(date_from, 'isoformat'):
                date_from_str = date_from.isoformat()
            else:
                date_from_str = str(date_from)[:10]
            
            if hasattr(date_to, 'isoformat'):
                date_to_str = date_to.isoformat()
            else:
                date_to_str = str(date_to)[:10]
            
            # Leggi il tipo di richiesta
            req_type = db.execute(
                f"SELECT name FROM request_types WHERE id = {placeholder}",
                (request_type_id,)
            ).fetchone()
            
            type_name = req_type["name"] if req_type and isinstance(req_type, dict) else (req_type[0] if req_type else "Permesso")
            
            # Estrai orari da extra_data per i permessi
            time_start = "00:00"
            time_end = "24:00"
            hours = 0
            extra_data = None
            
            if request_type_id == 3 and extra_data_str:  # Permesso
                try:
                    import json
                    # Parse extra_data se è stringa
                    if isinstance(extra_data_str, str):
                        extra = json.loads(extra_data_str)
                    else:
                        extra = extra_data_str
                    
                    time_start = extra.get("time_start", "00:00")
                    time_end = extra.get("time_end", "24:00")
                    extra_data = {"time_start": time_start, "time_end": time_end}
                    
                    # Calcola ore
                    from datetime import datetime as dt
                    start_time = dt.strptime(time_start, "%H:%M")
                    end_time = dt.strptime(time_end, "%H:%M")
                    delta = end_time - start_time
                    hours = delta.total_seconds() / 3600
                except Exception as e:
                    app.logger.error(f"Errore parsing extra_data per permesso: {e}")
                    pass
            
            # Aggiungi una card per ogni giorno della feria
            current_date = date_from
            while current_date <= date_to:
                date_str = current_date.isoformat() if hasattr(current_date, 'isoformat') else str(current_date)[:10]
                
                shift_item = {
                    "username": username,
                    "date": date_str,
                    "type": "vacation",
                    "project_code": type_name,
                    "project_name": type_name,
                    "function": type_name,
                    "start": time_start,
                    "end": time_end,
                    "hours": hours,
                    "note": notes or "",
                    "is_leader": False,
                }
                
                # Aggiungi extra_data solo se presente
                if extra_data:
                    shift_item["extra_data"] = extra_data
                
                shifts_list.append(shift_item)
                
                # Incrementa di un giorno
                from datetime import timedelta
                current_date = current_date + timedelta(days=1)

    return jsonify({
        "users": users_list,
        "shifts": shifts_list
    })


# ═══════════════════════════════════════════════════════════════════════════════
#  GESTIONE GRUPPI UTENTI - ADMIN UI
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/admin/groups")
@login_required
def admin_groups_page() -> ResponseReturnValue:
    """Pagina gestione gruppi utenti (solo admin)."""
    if not session.get("is_admin"):
        abort(403)

    display_name = session.get("user_display") or session.get("user_name") or session.get("user")
    primary_name = session.get("user_name") or display_name or session.get("user")
    initials = session.get("user_initials") or compute_initials(primary_name or "")

    # Verifica se CedolinoWeb è attivo (config.json + database settings)
    cedolino_sync_enabled = get_cedolino_settings() is not None
    app.logger.info(f"admin_groups_page: cedolino_sync_enabled={cedolino_sync_enabled}")

    # Recupera le location GPS configurate
    timbratura_config = get_timbratura_config()
    gps_locations = timbratura_config.get("gps_locations", [])

    return render_template(
        "admin_groups.html",
        user_name=primary_name,
        user_display=display_name,
        user_initials=initials,
        is_admin=True,
        cedolino_sync_enabled=cedolino_sync_enabled,
        gps_locations=gps_locations,
    )


@app.get("/api/admin/groups")
@login_required
def api_admin_groups_list() -> ResponseReturnValue:
    """Lista tutti i gruppi."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403

    db = get_db()
    ensure_user_groups_table(db)
    
    rows = db.execute("""
        SELECT id, name, description, cedolino_group_id, gps_location_name, is_active, created_ts, updated_ts
        FROM user_groups
        ORDER BY name ASC
    """).fetchall()

    groups = []
    for row in rows:
        groups.append({
            "id": row["id"] if isinstance(row, dict) else row[0],
            "name": row["name"] if isinstance(row, dict) else row[1],
            "description": row["description"] if isinstance(row, dict) else row[2],
            "cedolino_group_id": row["cedolino_group_id"] if isinstance(row, dict) else row[3],
            "gps_location_name": row["gps_location_name"] if isinstance(row, dict) else row[4],
            "is_active": bool(row["is_active"] if isinstance(row, dict) else row[5]),
            "created_ts": row["created_ts"] if isinstance(row, dict) else row[6],
            "updated_ts": row["updated_ts"] if isinstance(row, dict) else row[7],
        })

    return jsonify({"groups": groups})


@app.post("/api/admin/groups")
@login_required
def api_admin_groups_create() -> ResponseReturnValue:
    """Crea un nuovo gruppo."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403

    data = request.get_json()
    if not data:
        return jsonify({"error": "Dati non validi"}), 400

    name = (data.get("name") or "").strip()
    description = (data.get("description") or "").strip() or None
    cedolino_group_id = (data.get("cedolino_group_id") or "").strip() or None
    gps_location_name = (data.get("gps_location_name") or "").strip() or None
    is_active = data.get("is_active", True)

    if not name:
        return jsonify({"error": "Nome gruppo richiesto"}), 400

    db = get_db()
    ensure_user_groups_table(db)

    # Verifica se esiste già
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    existing = db.execute(f"SELECT id FROM user_groups WHERE name = {placeholder}", (name,)).fetchone()

    if existing:
        return jsonify({"error": f"Il gruppo '{name}' esiste già"}), 409

    now = now_ms()

    if DB_VENDOR == "mysql":
        db.execute("""
            INSERT INTO user_groups (name, description, cedolino_group_id, gps_location_name, is_active, created_ts, updated_ts)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (name, description, cedolino_group_id, gps_location_name, 1 if is_active else 0, now, now))
    else:
        db.execute("""
            INSERT INTO user_groups (name, description, cedolino_group_id, gps_location_name, is_active, created_ts, updated_ts)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (name, description, cedolino_group_id, gps_location_name, 1 if is_active else 0, now, now))
    
    new_id = _last_insert_id(db)
    db.commit()

    app.logger.info("Admin %s ha creato gruppo '%s' (id=%s)", session.get("user"), name, new_id)
    return jsonify({"ok": True, "id": new_id, "name": name}), 201


@app.put("/api/admin/groups/<int:group_id>")
@login_required
def api_admin_groups_update(group_id: int) -> ResponseReturnValue:
    """Modifica un gruppo esistente."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403

    data = request.get_json()
    if not data:
        return jsonify({"error": "Dati non validi"}), 400

    db = get_db()
    ensure_user_groups_table(db)

    # Verifica che il gruppo esista
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    existing = db.execute(f"SELECT id FROM user_groups WHERE id = {placeholder}", (group_id,)).fetchone()

    if not existing:
        return jsonify({"error": f"Gruppo con id={group_id} non trovato"}), 404

    # Prepara i campi da aggiornare
    updates = []
    params = []

    if "name" in data:
        name = (data["name"] or "").strip()
        if name:
            # Verifica unicità nome
            check = db.execute(
                f"SELECT id FROM user_groups WHERE name = {placeholder} AND id != {placeholder}",
                (name, group_id)
            ).fetchone()
            if check:
                return jsonify({"error": f"Il nome '{name}' è già utilizzato da un altro gruppo"}), 409
            updates.append("name = " + placeholder)
            params.append(name)

    if "description" in data:
        description = (data["description"] or "").strip() or None
        updates.append("description = " + placeholder)
        params.append(description)

    if "cedolino_group_id" in data:
        cedolino_group_id = (data["cedolino_group_id"] or "").strip() or None
        updates.append("cedolino_group_id = " + placeholder)
        params.append(cedolino_group_id)

    if "gps_location_name" in data:
        gps_location_name = (data["gps_location_name"] or "").strip() or None
        updates.append("gps_location_name = " + placeholder)
        params.append(gps_location_name)

    if "is_active" in data:
        is_active = data["is_active"]
        updates.append("is_active = " + placeholder)
        params.append(1 if is_active else 0)

    if not updates:
        return jsonify({"error": "Nessun campo da aggiornare"}), 400

    updates.append("updated_ts = " + placeholder)
    params.append(now_ms())
    params.append(group_id)

    sql = f"UPDATE user_groups SET {', '.join(updates)} WHERE id = {placeholder}"
    db.execute(sql, tuple(params))
    db.commit()

    app.logger.info("Admin %s ha modificato gruppo id=%s", session.get("user"), group_id)
    return jsonify({"ok": True})


@app.delete("/api/admin/groups/<int:group_id>")
@login_required
def api_admin_groups_delete(group_id: int) -> ResponseReturnValue:
    """Elimina un gruppo."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403

    db = get_db()
    ensure_user_groups_table(db)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"

    # Verifica che il gruppo esista
    existing = db.execute(f"SELECT id, name FROM user_groups WHERE id = {placeholder}", (group_id,)).fetchone()
    if not existing:
        return jsonify({"error": f"Gruppo con id={group_id} non trovato"}), 404

    group_name = existing["name"] if isinstance(existing, dict) else existing[1]

    # Verifica se ci sono utenti associati
    users_count = db.execute(
        f"SELECT COUNT(*) as cnt FROM app_users WHERE group_id = {placeholder}",
        (group_id,)
    ).fetchone()
    count = users_count["cnt"] if isinstance(users_count, dict) else users_count[0]
    
    if count > 0:
        return jsonify({"error": f"Impossibile eliminare: {count} utente/i ancora associato/i a questo gruppo"}), 400

    db.execute(f"DELETE FROM user_groups WHERE id = {placeholder}", (group_id,))
    db.commit()

    app.logger.info("Admin %s ha eliminato gruppo '%s' (id=%s)", session.get("user"), group_name, group_id)
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
#  GESTIONE UTENTI - ADMIN UI
# ═══════════════════════════════════════════════════════════════════════════════

VALID_USER_ROLES = {"user", "supervisor", "admin", "magazzino"}


@app.get("/admin/users")
@login_required
def admin_users_page() -> ResponseReturnValue:
    """Pagina gestione utenti (solo admin)."""
    if not session.get("is_admin"):
        abort(403)

    display_name = session.get("user_display") or session.get("user_name") or session.get("user")
    primary_name = session.get("user_name") or display_name or session.get("user")
    initials = session.get("user_initials") or compute_initials(primary_name or "")

    # Verifica se CedolinoWeb è attivo (config.json + database settings)
    cedolino_sync_enabled = get_cedolino_settings() is not None

    return render_template(
        "admin_users.html",
        user_name=primary_name,
        user_display=display_name,
        user_initials=initials,
        is_admin=True,
        cedolino_sync_enabled=cedolino_sync_enabled,
    )


@app.get("/api/admin/users")
@login_required
def api_admin_users_list() -> ResponseReturnValue:
    """Lista tutti gli utenti."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403

    db = get_db()
    ensure_user_groups_table(db)
    
    rows = db.execute("""
        SELECT u.username, u.display_name, u.full_name, u.role, u.is_active, 
               u.created_ts, u.updated_ts, u.rentman_crew_id, c.name as crew_name,
               u.external_id, u.external_group_id, u.group_id, g.name as group_name,
               g.cedolino_group_id as group_cedolino_id
        FROM app_users u
        LEFT JOIN crew_members c ON u.rentman_crew_id = c.rentman_id
        LEFT JOIN user_groups g ON u.group_id = g.id
        ORDER BY u.username ASC
    """).fetchall()

    users = []
    for row in rows:
        external_id = row["external_id"] if isinstance(row, dict) else row[9]
        
        users.append({
            "username": row["username"] if isinstance(row, dict) else row[0],
            "display_name": row["display_name"] if isinstance(row, dict) else row[1],
            "full_name": row["full_name"] if isinstance(row, dict) else row[2],
            "role": row["role"] if isinstance(row, dict) else row[3],
            "is_active": bool(row["is_active"] if isinstance(row, dict) else row[4]),
            "created_ts": row["created_ts"] if isinstance(row, dict) else row[5],
            "updated_ts": row["updated_ts"] if isinstance(row, dict) else row[6],
            "rentman_crew_id": row["rentman_crew_id"] if isinstance(row, dict) else row[7],
            "crew_name": row["crew_name"] if isinstance(row, dict) else row[8],
            "external_id": external_id,
            "external_group_id": row["external_group_id"] if isinstance(row, dict) else row[10],
            "group_id": row["group_id"] if isinstance(row, dict) else row[11],
            "group_name": row["group_name"] if isinstance(row, dict) else row[12],
            "group_cedolino_id": row["group_cedolino_id"] if isinstance(row, dict) else row[13],
        })

    return jsonify({"users": users})


@app.post("/api/admin/users")
@login_required
def api_admin_users_create() -> ResponseReturnValue:
    """Crea un nuovo utente."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403

    data = request.get_json()
    if not data:
        return jsonify({"error": "Dati non validi"}), 400

    username = (data.get("username") or "").strip().lower()
    password = data.get("password", "")
    display_name = (data.get("display_name") or "").strip()
    full_name = (data.get("full_name") or "").strip() or None
    role = (data.get("role") or "user").strip().lower()
    is_active = data.get("is_active", True)
    rentman_crew_id = data.get("rentman_crew_id")
    if rentman_crew_id is not None:
        rentman_crew_id = int(rentman_crew_id) if rentman_crew_id else None

    if not username:
        return jsonify({"error": "Username richiesto"}), 400
    if not password:
        return jsonify({"error": "Password richiesta"}), 400
    if not display_name:
        display_name = username
    if role not in VALID_USER_ROLES:
        return jsonify({"error": f"Ruolo non valido. Valori ammessi: {', '.join(sorted(VALID_USER_ROLES))}"}), 400

    db = get_db()

    # Verifica se esiste già
    if DB_VENDOR == "mysql":
        existing = db.execute("SELECT username FROM app_users WHERE username = %s", (username,)).fetchone()
    else:
        existing = db.execute("SELECT username FROM app_users WHERE username = ?", (username,)).fetchone()

    if existing:
        return jsonify({"error": f"L'utente '{username}' esiste già"}), 409

    now = now_ms()
    password_hashed = hash_password(password)

    if DB_VENDOR == "mysql":
        db.execute("""
            INSERT INTO app_users (username, password_hash, display_name, full_name, role, is_active, rentman_crew_id, created_ts, updated_ts)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (username, password_hashed, display_name, full_name, role, 1 if is_active else 0, rentman_crew_id, now, now))
    else:
        db.execute("""
            INSERT INTO app_users (username, password_hash, display_name, full_name, role, is_active, rentman_crew_id, created_ts, updated_ts)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (username, password_hashed, display_name, full_name, role, 1 if is_active else 0, rentman_crew_id, now, now))
    db.commit()

    app.logger.info("Admin %s ha creato utente %s con ruolo %s", session.get("user"), username, role)
    return jsonify({"ok": True, "username": username}), 201


@app.put("/api/admin/users/<username>")
@login_required
def api_admin_users_update(username: str) -> ResponseReturnValue:
    """Modifica un utente esistente."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403

    username = username.strip().lower()
    data = request.get_json()
    app.logger.info("PUT /api/admin/users/%s - payload: %s", username, data)
    if not data:
        return jsonify({"error": "Dati non validi"}), 400

    db = get_db()

    # Verifica che l'utente esista
    if DB_VENDOR == "mysql":
        existing = db.execute("SELECT username FROM app_users WHERE username = %s", (username,)).fetchone()
    else:
        existing = db.execute("SELECT username FROM app_users WHERE username = ?", (username,)).fetchone()

    if not existing:
        return jsonify({"error": f"Utente '{username}' non trovato"}), 404

    # Prepara i campi da aggiornare
    updates = []
    params = []

    if "display_name" in data:
        display_name = (data["display_name"] or "").strip()
        if display_name:
            updates.append("display_name = " + ("%s" if DB_VENDOR == "mysql" else "?"))
            params.append(display_name)

    if "full_name" in data:
        full_name = (data["full_name"] or "").strip() or None
        updates.append("full_name = " + ("%s" if DB_VENDOR == "mysql" else "?"))
        params.append(full_name)

    if "role" in data:
        role = (data["role"] or "").strip().lower()
        if role not in VALID_USER_ROLES:
            return jsonify({"error": f"Ruolo non valido. Valori ammessi: {', '.join(sorted(VALID_USER_ROLES))}"}), 400
        # Impedisci di rimuovere il ruolo admin a se stessi
        current_user = session.get("user", "").lower()
        if username == current_user and role != "admin":
            return jsonify({"error": "Non puoi rimuovere il ruolo admin a te stesso"}), 400
        updates.append("role = " + ("%s" if DB_VENDOR == "mysql" else "?"))
        params.append(role)

    if "is_active" in data:
        is_active = data["is_active"]
        # Impedisci di disattivare se stessi
        current_user = session.get("user", "").lower()
        if username == current_user and not is_active:
            return jsonify({"error": "Non puoi disattivare il tuo account"}), 400
        updates.append("is_active = " + ("%s" if DB_VENDOR == "mysql" else "?"))
        params.append(1 if is_active else 0)

    if "password" in data and data["password"]:
        password_hashed = hash_password(data["password"])
        updates.append("password_hash = " + ("%s" if DB_VENDOR == "mysql" else "?"))
        params.append(password_hashed)

    if "rentman_crew_id" in data:
        rentman_crew_id = data.get("rentman_crew_id")
        # Può essere None per rimuovere l'associazione, o un intero
        if rentman_crew_id is not None and rentman_crew_id != "":
            rentman_crew_id = int(rentman_crew_id)
        else:
            rentman_crew_id = None
        updates.append("rentman_crew_id = " + ("%s" if DB_VENDOR == "mysql" else "?"))
        params.append(rentman_crew_id)

    # Gestione external_id per CedolinoWeb (ID diretto utente)
    if "external_id" in data:
        external_id = data.get("external_id")
        # Può essere None/vuoto per rimuovere l'ID, o una stringa
        if external_id is not None and external_id != "":
            external_id = str(external_id).strip()
        else:
            external_id = None
        updates.append("external_id = " + ("%s" if DB_VENDOR == "mysql" else "?"))
        params.append(external_id)

    # Gestione external_group_id per CedolinoWeb (Gruppo ID diretto utente - deprecato, usare group_id)
    if "external_group_id" in data:
        external_group_id = data.get("external_group_id")
        # Può essere None/vuoto per rimuovere l'ID, o una stringa
        if external_group_id is not None and external_group_id != "":
            external_group_id = str(external_group_id).strip()
        else:
            external_group_id = None
        updates.append("external_group_id = " + ("%s" if DB_VENDOR == "mysql" else "?"))
        params.append(external_group_id)

    # Gestione group_id per collegamento a user_groups
    if "group_id" in data:
        group_id = data.get("group_id")
        # Può essere None/vuoto per rimuovere l'associazione, o un intero
        if group_id is not None and group_id != "":
            group_id = int(group_id)
        else:
            group_id = None
        updates.append("group_id = " + ("%s" if DB_VENDOR == "mysql" else "?"))
        params.append(group_id)

    if not updates:
        return jsonify({"error": "Nessun campo da aggiornare"}), 400

    updates.append("updated_ts = " + ("%s" if DB_VENDOR == "mysql" else "?"))
    params.append(now_ms())
    params.append(username)

    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    sql = f"UPDATE app_users SET {', '.join(updates)} WHERE username = {placeholder}"
    app.logger.info("SQL UPDATE: %s con params: %s", sql, params)
    
    cursor = db.execute(sql, tuple(params))
    app.logger.info("Rows affected: %s", cursor.rowcount if hasattr(cursor, 'rowcount') else 'N/A')
    
    db.commit()
    app.logger.info("COMMIT eseguito per utente %s", username)
    
    # Verifica immediata
    if DB_VENDOR == "mysql":
        verify = db.execute("SELECT rentman_crew_id FROM app_users WHERE username = %s", (username,)).fetchone()
        app.logger.info("Verifica dopo commit - rentman_crew_id: %s", verify)

    app.logger.info("Admin %s ha modificato utente %s", session.get("user"), username)
    return jsonify({"ok": True, "username": username})


@app.delete("/api/admin/users/<username>")
@login_required
def api_admin_users_delete(username: str) -> ResponseReturnValue:
    """Elimina un utente."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403

    username = username.strip().lower()

    # Impedisci di eliminare se stessi
    current_user = session.get("user", "").lower()
    if username == current_user:
        return jsonify({"error": "Non puoi eliminare il tuo account"}), 400

    db = get_db()

    # Verifica che l'utente esista
    if DB_VENDOR == "mysql":
        existing = db.execute("SELECT username FROM app_users WHERE username = %s", (username,)).fetchone()
    else:
        existing = db.execute("SELECT username FROM app_users WHERE username = ?", (username,)).fetchone()

    if not existing:
        return jsonify({"error": f"Utente '{username}' non trovato"}), 404

    if DB_VENDOR == "mysql":
        db.execute("DELETE FROM app_users WHERE username = %s", (username,))
    else:
        db.execute("DELETE FROM app_users WHERE username = ?", (username,))
    db.commit()

    app.logger.info("Admin %s ha eliminato utente %s", session.get("user"), username)
    return jsonify({"ok": True, "deleted": username})


# ═══════════════════════════════════════════════════════════════════════════════
#  GESTIONE OPERATORI (Admin) - API
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/operators")
@login_required
def admin_operators() -> ResponseReturnValue:
    """Pagina gestione operatori."""
    if not session.get("is_admin"):
        flash("Accesso non autorizzato", "danger")
        return redirect(url_for("index"))
    
    return render_template("admin_operators.html", is_admin=True)


@app.get("/api/admin/operators")
@login_required
def api_admin_operators_list() -> ResponseReturnValue:
    """Lista tutti gli operatori dal database."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403

    db = get_db()
    ensure_crew_members_table(db)

    cursor = db.execute(
        "SELECT id, rentman_id, name, email, phone, is_active, created_ts, updated_ts, timbratura_override "
        "FROM crew_members ORDER BY name"
    )
    rows = cursor.fetchall()

    operators = []
    for row in rows:
        timbratura_override = row[8] if len(row) > 8 else None
        if timbratura_override and isinstance(timbratura_override, str):
            try:
                timbratura_override = json.loads(timbratura_override)
            except:
                timbratura_override = None
        
        operators.append({
            "id": row[0],
            "rentman_id": row[1],
            "name": row[2],
            "email": row[3],
            "phone": row[4],
            "is_active": bool(row[5]),
            "created_ts": row[6],
            "updated_ts": row[7],
            "timbratura_override": timbratura_override
        })

    return jsonify({"ok": True, "operators": operators})


@app.put("/api/admin/operators/<int:operator_id>")
@login_required
def api_admin_operators_update(operator_id: int) -> ResponseReturnValue:
    """Aggiorna un operatore (email, phone, is_active, timbratura_override)."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403

    data = request.get_json(silent=True) or {}

    db = get_db()
    ensure_crew_members_table(db)

    # Verifica che l'operatore esista
    if DB_VENDOR == "mysql":
        existing = db.execute("SELECT id, name FROM crew_members WHERE id = %s", (operator_id,)).fetchone()
    else:
        existing = db.execute("SELECT id, name FROM crew_members WHERE id = ?", (operator_id,)).fetchone()

    if not existing:
        return jsonify({"error": "Operatore non trovato"}), 404

    # Prepara i campi da aggiornare
    email = data.get("email")
    if email is not None:
        email = email.strip() if email else None

    phone = data.get("phone")
    if phone is not None:
        phone = phone.strip() if phone else None

    is_active = data.get("is_active")
    if is_active is not None:
        is_active = 1 if is_active else 0

    now = now_ms()

    # Costruisci query di aggiornamento dinamica
    updates = []
    params = []

    if "email" in data:
        updates.append("email = " + ("%s" if DB_VENDOR == "mysql" else "?"))
        params.append(email)
    if "phone" in data:
        updates.append("phone = " + ("%s" if DB_VENDOR == "mysql" else "?"))
        params.append(phone)
    if "is_active" in data:
        updates.append("is_active = " + ("%s" if DB_VENDOR == "mysql" else "?"))
        params.append(is_active)
    
    # Gestione eccezioni timbratura (qr_disabled, gps_disabled)
    if "timbratura_override" in data:
        timbratura_override = data.get("timbratura_override")
        if timbratura_override and isinstance(timbratura_override, dict):
            # Verifica se c'è almeno una disabilitazione attiva
            has_override = timbratura_override.get("qr_disabled") or timbratura_override.get("gps_disabled")
            if has_override:
                timbratura_override = json.dumps(timbratura_override)
            else:
                timbratura_override = None
        else:
            timbratura_override = None
        updates.append("timbratura_override = " + ("%s" if DB_VENDOR == "mysql" else "?"))
        params.append(timbratura_override)

    if not updates:
        return jsonify({"error": "Nessun campo da aggiornare"}), 400

    updates.append("updated_ts = " + ("%s" if DB_VENDOR == "mysql" else "?"))
    params.append(now)
    params.append(operator_id)

    sql = f"UPDATE crew_members SET {', '.join(updates)} WHERE id = " + ("%s" if DB_VENDOR == "mysql" else "?")

    db.execute(sql, tuple(params))
    db.commit()

    app.logger.info("Admin %s ha modificato operatore %s (id=%d)", session.get("user"), existing[1], operator_id)
    return jsonify({"ok": True, "id": operator_id})


@app.post("/api/admin/operators/sync")
@login_required
def api_admin_operators_sync() -> ResponseReturnValue:
    """Forza la sincronizzazione degli operatori da Rentman."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403

    try:
        from rentman_client import rentman_request
    except ImportError:
        return jsonify({"error": "Client Rentman non disponibile"}), 500

    # Recupera tutti i crew members da Rentman
    try:
        response = rentman_request("GET", "/crew")
        if not response:
            return jsonify({"error": "Nessuna risposta da Rentman"}), 502
    except Exception as e:
        app.logger.error("Errore sync operatori Rentman: %s", e)
        return jsonify({"error": str(e)}), 500

    db = get_db()
    ensure_crew_members_table(db)

    synced = 0
    crew_data = response.get("data", [])
    for crew in crew_data:
        rentman_id = crew.get("id")
        name = crew.get("displayname") or crew.get("name") or f"Crew {rentman_id}"
        if rentman_id:
            sync_crew_member_from_rentman(db, rentman_id, name)
            synced += 1

    db.commit()

    app.logger.info("Admin %s ha sincronizzato %d operatori da Rentman", session.get("user"), synced)
    return jsonify({"ok": True, "synced": synced})


# ═══════════════════════════════════════════════════════════════════════════════
#  LOCATION CACHE - Gestione coordinate GPS per location Rentman
# ═══════════════════════════════════════════════════════════════════════════════

def ensure_location_cache_table(db: DatabaseLike) -> None:
    """Crea la tabella location_cache per salvare le coordinate GPS delle location."""
    if DB_VENDOR == "mysql":
        db.execute("""
            CREATE TABLE IF NOT EXISTS location_cache (
                id INT AUTO_INCREMENT PRIMARY KEY,
                rentman_location_id INT COMMENT 'ID della location da Rentman',
                location_name VARCHAR(500) NOT NULL COMMENT 'Nome della location Rentman',
                latitude DECIMAL(10,7) NOT NULL COMMENT 'Latitudine GPS',
                longitude DECIMAL(10,7) NOT NULL COMMENT 'Longitudine GPS',
                radius_meters INT DEFAULT 300 COMMENT 'Raggio tolleranza GPS in metri',
                created_ts BIGINT DEFAULT 0,
                updated_ts BIGINT DEFAULT 0,
                INDEX idx_location_name (location_name),
                INDEX idx_rentman_location_id (rentman_location_id)
            ) COMMENT='Cache delle coordinate GPS per le location Rentman'
        """)
        
        # Aggiungi la colonna rentman_location_id se manca
        try:
            db.execute("ALTER TABLE location_cache ADD COLUMN rentman_location_id INT COMMENT 'ID della location da Rentman' AFTER id")
            db.execute("ALTER TABLE location_cache ADD INDEX idx_rentman_location_id (rentman_location_id)")
            app.logger.info("✅ Colonna rentman_location_id aggiunta a location_cache")
        except Exception as e:
            if "Duplicate column" in str(e) or "already exists" in str(e):
                app.logger.info("ℹ️ Colonna rentman_location_id esiste già")
            else:
                app.logger.warning(f"⚠️ Errore aggiunta colonna rentman_location_id: {e}")
        
        # Aggiungi la colonna radius_meters se manca
        try:
            db.execute("ALTER TABLE location_cache ADD COLUMN radius_meters INT DEFAULT 300 COMMENT 'Raggio tolleranza GPS in metri' AFTER longitude")
            app.logger.info("✅ Colonna radius_meters aggiunta a location_cache")
        except Exception as e:
            if "Duplicate column" in str(e) or "already exists" in str(e):
                app.logger.info("ℹ️ Colonna radius_meters esiste già")
            else:
                app.logger.warning(f"⚠️ Errore aggiunta colonna radius_meters: {e}")
    else:
        db.execute("""
            CREATE TABLE IF NOT EXISTS location_cache (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rentman_location_id INTEGER,
                location_name TEXT NOT NULL UNIQUE,
                latitude REAL NOT NULL,
                longitude REAL NOT NULL,
                radius_meters INTEGER DEFAULT 300,
                created_ts INTEGER DEFAULT 0,
                updated_ts INTEGER DEFAULT 0
            )
        """)
    
    db.commit()

def get_location_cache(db: DatabaseLike, location_name: str, rentman_location_id: Optional[int] = None) -> Optional[Tuple[float, float, int]]:
    """Recupera le coordinate GPS e il raggio dalla cache per una location.
    Cerca prima per location_id (se fornito), poi per location_name.
    Ritorna (lat, lon, radius_meters) oppure None."""
    if not location_name:
        return None
    
    ensure_location_cache_table(db)
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Cerca prima per location_id se disponibile (più preciso)
    if rentman_location_id:
        row = db.execute(
            f"SELECT latitude, longitude, COALESCE(radius_meters, 300) as radius_meters FROM location_cache WHERE rentman_location_id = {placeholder}",
            (rentman_location_id,)
        ).fetchone()
        if row:
            lat = row["latitude"] if isinstance(row, dict) else row[0]
            lon = row["longitude"] if isinstance(row, dict) else row[1]
            radius = row["radius_meters"] if isinstance(row, dict) else row[2]
            return (lat, lon, radius)
    
    # Fallback: cerca per location_name
    row = db.execute(
        f"SELECT latitude, longitude, COALESCE(radius_meters, 300) as radius_meters FROM location_cache WHERE location_name = {placeholder}",
        (location_name,)
    ).fetchone()
    
    if row:
        lat = row["latitude"] if isinstance(row, dict) else row[0]
        lon = row["longitude"] if isinstance(row, dict) else row[1]
        radius = row["radius_meters"] if isinstance(row, dict) else row[2]
        return (lat, lon, radius)
    
    return None

def save_location_cache(db: DatabaseLike, location_name: str, latitude: float, longitude: float, rentman_location_id: Optional[int] = None, radius_meters: int = 300) -> None:
    """Salva le coordinate GPS e il raggio nella cache per una location."""
    if not location_name or latitude is None or longitude is None:
        return
    
    ensure_location_cache_table(db)
    
    now = int(datetime.now().timestamp() * 1000)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    try:
        # Prova l'INSERT o l'UPDATE se già esiste
        if DB_VENDOR == "mysql":
            db.execute("""
                INSERT INTO location_cache (rentman_location_id, location_name, latitude, longitude, radius_meters, created_ts, updated_ts)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE rentman_location_id = %s, latitude = %s, longitude = %s, radius_meters = %s, updated_ts = %s
            """, (rentman_location_id, location_name, latitude, longitude, radius_meters, now, now, rentman_location_id, latitude, longitude, radius_meters, now))
        else:
            # SQLite: prova INSERT, altrimenti UPDATE
            db.execute(
                "INSERT OR IGNORE INTO location_cache (rentman_location_id, location_name, latitude, longitude, radius_meters, created_ts, updated_ts) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (rentman_location_id, location_name, latitude, longitude, radius_meters, now, now)
            )
            db.execute(
                "UPDATE location_cache SET rentman_location_id = ?, latitude = ?, longitude = ?, radius_meters = ?, updated_ts = ? WHERE location_name = ?",
                (rentman_location_id, latitude, longitude, radius_meters, now, location_name)
            )
        db.commit()
        app.logger.info(f"Location cache salvata: location_id={rentman_location_id}, name={location_name} ({latitude}, {longitude}), raggio={radius_meters}m")
    except Exception as e:
        app.logger.error(f"Errore salvataggio location cache: {e}")

# ═══════════════════════════════════════════════════════════════════════════════
#  REGOLE TIMBRATURE - ADMIN
# ═══════════════════════════════════════════════════════════════════════════════

def ensure_timbratura_rules_table(db):
    """Crea la tabella timbratura_rules se non esiste."""
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    if DB_VENDOR == "mysql":
        db.execute("""
            CREATE TABLE IF NOT EXISTS timbratura_rules (
                id INT PRIMARY KEY DEFAULT 1,
                anticipo_max_minuti INT DEFAULT 30,
                tolleranza_ritardo_minuti INT DEFAULT 5,
                arrotondamento_ingresso_minuti INT DEFAULT 15,
                arrotondamento_ingresso_tipo VARCHAR(1) DEFAULT '+',
                arrotondamento_uscita_minuti INT DEFAULT 15,
                arrotondamento_uscita_tipo VARCHAR(1) DEFAULT '-',
                pausa_blocco_minimo_minuti INT DEFAULT 30,
                pausa_incremento_minuti INT DEFAULT 15,
                pausa_tolleranza_minuti INT DEFAULT 5,
                updated_ts BIGINT,
                updated_by VARCHAR(100)
            )
        """)
        # Aggiunge colonne se mancanti (per upgrade)
        try:
            db.execute("ALTER TABLE timbratura_rules ADD COLUMN arrotondamento_ingresso_tipo VARCHAR(1) DEFAULT '+'")
            db.commit()
        except:
            pass
        try:
            db.execute("ALTER TABLE timbratura_rules ADD COLUMN arrotondamento_uscita_tipo VARCHAR(1) DEFAULT '-'")
            db.commit()
        except:
            pass
    else:
        db.execute("""
            CREATE TABLE IF NOT EXISTS timbratura_rules (
                id INTEGER PRIMARY KEY DEFAULT 1,
                anticipo_max_minuti INTEGER DEFAULT 30,
                tolleranza_ritardo_minuti INTEGER DEFAULT 5,
                arrotondamento_ingresso_minuti INTEGER DEFAULT 15,
                arrotondamento_ingresso_tipo TEXT DEFAULT '+',
                arrotondamento_uscita_minuti INTEGER DEFAULT 15,
                arrotondamento_uscita_tipo TEXT DEFAULT '-',
                pausa_blocco_minimo_minuti INTEGER DEFAULT 30,
                pausa_incremento_minuti INTEGER DEFAULT 15,
                pausa_tolleranza_minuti INTEGER DEFAULT 5,
                updated_ts INTEGER,
                updated_by TEXT
            )
        """)
        # Aggiunge colonne se mancanti (per upgrade)
        try:
            db.execute("ALTER TABLE timbratura_rules ADD COLUMN arrotondamento_ingresso_tipo TEXT DEFAULT '+'")
            db.commit()
        except:
            pass
        try:
            db.execute("ALTER TABLE timbratura_rules ADD COLUMN arrotondamento_uscita_tipo TEXT DEFAULT '-'")
            db.commit()
        except:
            pass
    db.commit()


# ═══════════════════════════════════════════════════════════════════════════════
#  REGOLE TIMBRATURE PER GRUPPO
# ═══════════════════════════════════════════════════════════════════════════════

GROUP_TIMBRATURA_RULES_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS group_timbratura_rules (
    id INT AUTO_INCREMENT PRIMARY KEY,
    group_id INT NOT NULL,
    rounding_mode ENUM('single', 'daily') NOT NULL DEFAULT 'single' COMMENT 'single=arrotonda singola timbrata, daily=arrotonda totale giornaliero',
    flessibilita_ingresso_minuti INT DEFAULT 30 COMMENT 'Minuti di flessibilità in ingresso rispetto al turno',
    flessibilita_uscita_minuti INT DEFAULT 30 COMMENT 'Minuti di flessibilità in uscita rispetto al turno',
    arrotondamento_giornaliero_minuti INT DEFAULT 15 COMMENT 'Blocco arrotondamento per daily mode',
    arrotondamento_giornaliero_tipo ENUM('floor', 'ceil', 'nearest') DEFAULT 'floor' COMMENT 'floor=in difetto, ceil=in eccesso, nearest=al più vicino',
    oltre_flessibilita_action ENUM('allow', 'warn', 'block') DEFAULT 'allow' COMMENT 'Azione se timbrata oltre flessibilità',
    usa_regole_pausa_standard TINYINT(1) DEFAULT 1 COMMENT 'Se true, usa regole pausa globali',
    is_active TINYINT(1) NOT NULL DEFAULT 1,
    created_ts BIGINT NOT NULL,
    updated_ts BIGINT NOT NULL,
    updated_by VARCHAR(100),
    UNIQUE KEY uk_group_rules (group_id),
    FOREIGN KEY (group_id) REFERENCES user_groups(id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

GROUP_TIMBRATURA_RULES_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS group_timbratura_rules (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    group_id INTEGER NOT NULL UNIQUE,
    rounding_mode TEXT NOT NULL DEFAULT 'single',
    flessibilita_ingresso_minuti INTEGER DEFAULT 30,
    flessibilita_uscita_minuti INTEGER DEFAULT 30,
    arrotondamento_giornaliero_minuti INTEGER DEFAULT 15,
    arrotondamento_giornaliero_tipo TEXT DEFAULT 'floor',
    oltre_flessibilita_action TEXT DEFAULT 'allow',
    usa_regole_pausa_standard INTEGER DEFAULT 1,
    is_active INTEGER NOT NULL DEFAULT 1,
    created_ts INTEGER NOT NULL,
    updated_ts INTEGER NOT NULL,
    updated_by TEXT,
    FOREIGN KEY (group_id) REFERENCES user_groups(id) ON DELETE CASCADE
)
"""


def ensure_group_timbratura_rules_table(db):
    """Crea la tabella group_timbratura_rules se non esiste."""
    try:
        # Prima assicura che user_groups esista
        ensure_user_groups_table(db)
        
        if DB_VENDOR == "mysql":
            db.execute(GROUP_TIMBRATURA_RULES_TABLE_MYSQL)
        else:
            db.execute(GROUP_TIMBRATURA_RULES_TABLE_SQLITE)
        db.commit()
    except Exception as e:
        # Se fallisce per la foreign key, prova senza
        app.logger.warning(f"ensure_group_timbratura_rules_table: {e}")
        try:
            if DB_VENDOR == "mysql":
                # Crea senza foreign key
                db.execute("""
                    CREATE TABLE IF NOT EXISTS group_timbratura_rules (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        group_id INT NOT NULL,
                        rounding_mode ENUM('single', 'daily') NOT NULL DEFAULT 'single',
                        flessibilita_ingresso_minuti INT DEFAULT 30,
                        flessibilita_uscita_minuti INT DEFAULT 30,
                        arrotondamento_giornaliero_minuti INT DEFAULT 15,
                        arrotondamento_giornaliero_tipo ENUM('floor', 'ceil', 'nearest') DEFAULT 'floor',
                        oltre_flessibilita_action ENUM('allow', 'warn', 'block') DEFAULT 'allow',
                        usa_regole_pausa_standard TINYINT(1) DEFAULT 1,
                        is_active TINYINT(1) NOT NULL DEFAULT 1,
                        created_ts BIGINT NOT NULL DEFAULT 0,
                        updated_ts BIGINT NOT NULL DEFAULT 0,
                        updated_by VARCHAR(100),
                        UNIQUE KEY uk_group_rules (group_id)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
                """)
            db.commit()
        except Exception as e2:
            app.logger.error(f"ensure_group_timbratura_rules_table fallback: {e2}")


def get_group_timbratura_rules(db, group_id: int) -> Optional[dict]:
    """
    Ottiene le regole timbrature per un gruppo specifico.
    Ritorna None se il gruppo non ha regole specifiche.
    """
    ensure_group_timbratura_rules_table(db)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    row = db.execute(
        f"SELECT * FROM group_timbratura_rules WHERE group_id = {placeholder} AND is_active = 1",
        (group_id,)
    ).fetchone()
    
    if not row:
        return None
    
    if isinstance(row, dict):
        return row
    
    # Converte tuple a dict
    columns = ['id', 'group_id', 'rounding_mode', 'flessibilita_ingresso_minuti',
               'flessibilita_uscita_minuti', 'arrotondamento_giornaliero_minuti',
               'arrotondamento_giornaliero_tipo', 'oltre_flessibilita_action',
               'usa_regole_pausa_standard', 'is_active', 'created_ts', 'updated_ts', 'updated_by']
    return dict(zip(columns, row))


def get_user_timbratura_rules(db, username: str) -> dict:
    """
    Ottiene le regole timbrature per un utente.
    
    Logica:
    1. Se l'utente appartiene a un gruppo con regole specifiche, usa quelle
    2. Altrimenti usa le regole globali di timbratura_rules
    
    Returns:
        dict con le regole + campo 'source' ('group' o 'global') e 'rounding_mode'
    """
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Trova il gruppo dell'utente
    user_row = db.execute(
        f"SELECT group_id FROM app_users WHERE username = {placeholder}",
        (username,)
    ).fetchone()
    
    group_id = None
    if user_row:
        group_id = user_row['group_id'] if isinstance(user_row, dict) else user_row[0]
    
    # Se ha un gruppo, cerca regole specifiche
    group_rules = None
    if group_id:
        group_rules = get_group_timbratura_rules(db, group_id)
    
    # Ottieni sempre le regole globali come fallback
    global_rules = get_timbratura_rules(db)
    
    if group_rules:
        # Merge: regole gruppo + regole pausa da globali (se usa_regole_pausa_standard)
        result = {
            'source': 'group',
            'group_id': group_id,
            'rounding_mode': group_rules.get('rounding_mode', 'single'),
            'flessibilita_ingresso_minuti': group_rules.get('flessibilita_ingresso_minuti', 30),
            'flessibilita_uscita_minuti': group_rules.get('flessibilita_uscita_minuti', 30),
            'arrotondamento_giornaliero_minuti': group_rules.get('arrotondamento_giornaliero_minuti', 15),
            'arrotondamento_giornaliero_tipo': group_rules.get('arrotondamento_giornaliero_tipo', 'floor'),
            'oltre_flessibilita_action': group_rules.get('oltre_flessibilita_action', 'allow'),
            # Per single mode, usa comunque i valori globali
            'anticipo_max_minuti': global_rules.get('anticipo_max_minuti', 30),
            'tolleranza_ritardo_minuti': global_rules.get('tolleranza_ritardo_minuti', 5),
            'arrotondamento_ingresso_minuti': global_rules.get('arrotondamento_ingresso_minuti', 15),
            'arrotondamento_ingresso_tipo': global_rules.get('arrotondamento_ingresso_tipo', '+'),
            'arrotondamento_uscita_minuti': global_rules.get('arrotondamento_uscita_minuti', 15),
            'arrotondamento_uscita_tipo': global_rules.get('arrotondamento_uscita_tipo', '-'),
        }
        
        # Regole pausa: da globali se usa_regole_pausa_standard
        if group_rules.get('usa_regole_pausa_standard', True):
            result['pausa_blocco_minimo_minuti'] = global_rules.get('pausa_blocco_minimo_minuti', 30)
            result['pausa_incremento_minuti'] = global_rules.get('pausa_incremento_minuti', 15)
            result['pausa_tolleranza_minuti'] = global_rules.get('pausa_tolleranza_minuti', 5)
        else:
            # Prendi la pausa minima dal turno dell'utente per oggi
            from datetime import datetime
            day_of_week = datetime.now().weekday()  # 0=Lunedì, 6=Domenica
            try:
                ensure_employee_shifts_table(db)
                shift_row = db.execute(
                    f"""SELECT break_start, break_end FROM employee_shifts 
                       WHERE username = {placeholder} AND day_of_week = {placeholder} AND is_active = 1
                       LIMIT 1""",
                    (username, day_of_week)
                ).fetchone()
                
                if shift_row:
                    break_start = shift_row['break_start'] if isinstance(shift_row, dict) else shift_row[0]
                    break_end = shift_row['break_end'] if isinstance(shift_row, dict) else shift_row[1]
                    
                    if break_start and break_end:
                        # Converti timedelta in minuti
                        if hasattr(break_start, 'total_seconds'):
                            bs_min = int(break_start.total_seconds()) // 60
                            be_min = int(break_end.total_seconds()) // 60
                        else:
                            # Stringa HH:MM
                            bs_parts = str(break_start)[:5].split(':')
                            be_parts = str(break_end)[:5].split(':')
                            bs_min = int(bs_parts[0]) * 60 + int(bs_parts[1])
                            be_min = int(be_parts[0]) * 60 + int(be_parts[1])
                        
                        pausa_minima_turno = be_min - bs_min
                        if pausa_minima_turno > 0:
                            result['pausa_blocco_minimo_minuti'] = pausa_minima_turno
                            result['pausa_incremento_minuti'] = 0  # Nessun incremento, usa esattamente la pausa del turno
                            result['pausa_tolleranza_minuti'] = 0  # Nessuna tolleranza
                            result['pausa_source'] = 'shift'
                            app.logger.info(f"Pausa da turno per {username}: {pausa_minima_turno} min (day {day_of_week})")
                        else:
                            # Pausa turno = 0, usa default
                            result['pausa_blocco_minimo_minuti'] = global_rules.get('pausa_blocco_minimo_minuti', 30)
                            result['pausa_incremento_minuti'] = global_rules.get('pausa_incremento_minuti', 15)
                            result['pausa_tolleranza_minuti'] = global_rules.get('pausa_tolleranza_minuti', 5)
                    else:
                        # No break definito nel turno, usa globali
                        result['pausa_blocco_minimo_minuti'] = global_rules.get('pausa_blocco_minimo_minuti', 30)
                        result['pausa_incremento_minuti'] = global_rules.get('pausa_incremento_minuti', 15)
                        result['pausa_tolleranza_minuti'] = global_rules.get('pausa_tolleranza_minuti', 5)
                else:
                    # Nessun turno per oggi, usa globali
                    result['pausa_blocco_minimo_minuti'] = global_rules.get('pausa_blocco_minimo_minuti', 30)
                    result['pausa_incremento_minuti'] = global_rules.get('pausa_incremento_minuti', 15)
                    result['pausa_tolleranza_minuti'] = global_rules.get('pausa_tolleranza_minuti', 5)
            except Exception as e:
                app.logger.warning(f"Errore lettura pausa da turno per {username}: {e}")
                result['pausa_blocco_minimo_minuti'] = global_rules.get('pausa_blocco_minimo_minuti', 30)
                result['pausa_incremento_minuti'] = global_rules.get('pausa_incremento_minuti', 15)
                result['pausa_tolleranza_minuti'] = global_rules.get('pausa_tolleranza_minuti', 5)
        
        return result
    else:
        # Nessuna regola gruppo: usa globali con rounding_mode='single'
        return {
            'source': 'global',
            'group_id': None,
            'rounding_mode': 'single',
            'flessibilita_ingresso_minuti': global_rules.get('anticipo_max_minuti', 30),
            'flessibilita_uscita_minuti': 30,
            'arrotondamento_giornaliero_minuti': 15,
            'arrotondamento_giornaliero_tipo': 'floor',
            'oltre_flessibilita_action': 'allow',
            **global_rules
        }


def get_timbratura_rules(db) -> dict:
    """Ottiene le regole timbrature dal database."""
    ensure_timbratura_rules_table(db)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    row = db.execute("SELECT * FROM timbratura_rules WHERE id = 1").fetchone()
    
    if not row:
        # Inserisce valori di default
        db.execute("""
            INSERT INTO timbratura_rules (id, anticipo_max_minuti, tolleranza_ritardo_minuti,
                arrotondamento_ingresso_minuti, arrotondamento_uscita_minuti,
                pausa_blocco_minimo_minuti, pausa_incremento_minuti, pausa_tolleranza_minuti)
            VALUES (1, 30, 5, 15, 15, 30, 15, 5)
        """)
        db.commit()
        row = db.execute("SELECT * FROM timbratura_rules WHERE id = 1").fetchone()
    
    if isinstance(row, dict):
        return row
    
    # Converte tuple a dict
    columns = ['id', 'anticipo_max_minuti', 'tolleranza_ritardo_minuti',
               'arrotondamento_ingresso_minuti', 'arrotondamento_uscita_minuti',
               'pausa_blocco_minimo_minuti', 'pausa_incremento_minuti', 
               'pausa_tolleranza_minuti', 'updated_ts', 'updated_by']
    return dict(zip(columns, row))


@app.get("/admin/timbratura-rules")
@login_required
def admin_timbratura_rules_page() -> ResponseReturnValue:
    """Pagina configurazione regole timbrature (solo admin)."""
    if not session.get("is_admin"):
        abort(403)
    return render_template("admin_timbratura_rules.html", is_admin=True)


@app.get("/admin/group-timbratura-rules")
@login_required
def admin_group_timbratura_rules_page() -> ResponseReturnValue:
    """Pagina configurazione regole timbrature per gruppo (solo admin)."""
    if not session.get("is_admin"):
        abort(403)
    return render_template("admin_group_timbratura_rules.html", is_admin=True)


@app.get("/api/admin/timbratura-rules")
@login_required
def api_get_timbratura_rules():
    """Restituisce le regole timbrature correnti."""
    if not session.get("is_admin"):
        return jsonify({"error": "Non autorizzato"}), 403
    
    db = get_db()
    rules = get_timbratura_rules(db)
    return jsonify(rules)


@app.post("/api/admin/timbratura-rules")
@login_required
def api_save_timbratura_rules():
    """Salva le regole timbrature."""
    if not session.get("is_admin"):
        return jsonify({"error": "Non autorizzato"}), 403
    
    data = request.get_json()
    if not data:
        return jsonify({"error": "Dati mancanti"}), 400
    
    db = get_db()
    ensure_timbratura_rules_table(db)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Valida i dati numerici
    fields = {
        'anticipo_max_minuti': (0, 120),
        'tolleranza_ritardo_minuti': (0, 60),
        'arrotondamento_ingresso_minuti': (1, 60),
        'arrotondamento_uscita_minuti': (1, 60),
        'pausa_blocco_minimo_minuti': (5, 120),
        'pausa_incremento_minuti': (5, 60),
        'pausa_tolleranza_minuti': (0, 30)
    }
    
    # Campi di tipo arrotondamento (+ - ~)
    tipo_fields = ['arrotondamento_ingresso_tipo', 'arrotondamento_uscita_tipo']
    valid_tipos = ['+', '-', '~']
    
    values = {}
    for field, (min_val, max_val) in fields.items():
        val = data.get(field)
        if val is not None:
            val = int(val)
            if val < min_val or val > max_val:
                return jsonify({"error": f"{field} deve essere tra {min_val} e {max_val}"}), 400
            values[field] = val
    
    # Valida e aggiungi campi tipo
    for field in tipo_fields:
        val = data.get(field)
        if val is not None:
            if val not in valid_tipos:
                return jsonify({"error": f"{field} deve essere uno di: +, -, ~"}), 400
            values[field] = val
    
    if not values:
        return jsonify({"error": "Nessun campo da aggiornare"}), 400
    
    # Costruisce query di update
    set_clause = ", ".join([f"{k} = {placeholder}" for k in values.keys()])
    params = list(values.values()) + [now_ms(), session.get('user')]
    
    db.execute(
        f"UPDATE timbratura_rules SET {set_clause}, updated_ts = {placeholder}, updated_by = {placeholder} WHERE id = 1",
        params
    )
    db.commit()
    
    app.logger.info("Admin %s ha aggiornato le regole timbrature: %s", session.get('user'), values)
    return jsonify({"success": True})


# ═══════════════════════════════════════════════════════════════════════════════
#  API REGOLE TIMBRATURE PER GRUPPO
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/admin/group-timbratura-rules")
@login_required
def api_get_all_group_timbratura_rules():
    """Restituisce tutte le regole timbrature per gruppo."""
    if not session.get("is_admin"):
        return jsonify({"error": "Non autorizzato"}), 403
    
    db = get_db()
    ensure_group_timbratura_rules_table(db)
    ensure_user_groups_table(db)
    
    # Ottieni tutti i gruppi con eventuali regole
    rows = db.execute("""
        SELECT g.id, g.name, g.description,
               r.id as rule_id, r.rounding_mode, r.flessibilita_ingresso_minuti,
               r.flessibilita_uscita_minuti, r.arrotondamento_giornaliero_minuti,
               r.arrotondamento_giornaliero_tipo, r.oltre_flessibilita_action,
               r.usa_regole_pausa_standard, r.is_active as rule_active
        FROM user_groups g
        LEFT JOIN group_timbratura_rules r ON g.id = r.group_id
        WHERE g.is_active = 1
        ORDER BY g.name
    """).fetchall()
    
    result = []
    for row in rows:
        if isinstance(row, dict):
            item = {
                'group_id': row['id'],
                'group_name': row['name'],
                'group_description': row['description'],
                'has_rules': row['rule_id'] is not None,
                'rules': None
            }
            if row['rule_id']:
                item['rules'] = {
                    'id': row['rule_id'],
                    'rounding_mode': row['rounding_mode'],
                    'flessibilita_ingresso_minuti': row['flessibilita_ingresso_minuti'],
                    'flessibilita_uscita_minuti': row['flessibilita_uscita_minuti'],
                    'arrotondamento_giornaliero_minuti': row['arrotondamento_giornaliero_minuti'],
                    'arrotondamento_giornaliero_tipo': row['arrotondamento_giornaliero_tipo'],
                    'oltre_flessibilita_action': row['oltre_flessibilita_action'],
                    'usa_regole_pausa_standard': bool(row['usa_regole_pausa_standard']),
                    'is_active': bool(row['rule_active'])
                }
        else:
            item = {
                'group_id': row[0],
                'group_name': row[1],
                'group_description': row[2],
                'has_rules': row[3] is not None,
                'rules': None
            }
            if row[3]:
                item['rules'] = {
                    'id': row[3],
                    'rounding_mode': row[4],
                    'flessibilita_ingresso_minuti': row[5],
                    'flessibilita_uscita_minuti': row[6],
                    'arrotondamento_giornaliero_minuti': row[7],
                    'arrotondamento_giornaliero_tipo': row[8],
                    'oltre_flessibilita_action': row[9],
                    'usa_regole_pausa_standard': bool(row[10]),
                    'is_active': bool(row[11])
                }
        result.append(item)
    
    return jsonify(result)


@app.get("/api/admin/group-timbratura-rules/<int:group_id>")
@login_required
def api_get_group_timbratura_rules(group_id: int):
    """Restituisce le regole timbrature per un gruppo specifico."""
    if not session.get("is_admin"):
        return jsonify({"error": "Non autorizzato"}), 403
    
    db = get_db()
    rules = get_group_timbratura_rules(db, group_id)
    
    if rules:
        return jsonify(rules)
    else:
        # Ritorna valori di default per nuovo inserimento
        return jsonify({
            'group_id': group_id,
            'rounding_mode': 'single',
            'flessibilita_ingresso_minuti': 30,
            'flessibilita_uscita_minuti': 30,
            'arrotondamento_giornaliero_minuti': 15,
            'arrotondamento_giornaliero_tipo': 'floor',
            'oltre_flessibilita_action': 'allow',
            'usa_regole_pausa_standard': True,
            'is_new': True
        })


@app.post("/api/admin/group-timbratura-rules/<int:group_id>")
@login_required
def api_save_group_timbratura_rules(group_id: int):
    """Salva o aggiorna le regole timbrature per un gruppo."""
    if not session.get("is_admin"):
        return jsonify({"error": "Non autorizzato"}), 403
    
    data = request.get_json()
    if not data:
        return jsonify({"error": "Dati mancanti"}), 400
    
    db = get_db()
    ensure_group_timbratura_rules_table(db)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Verifica che il gruppo esista
    group_row = db.execute(
        f"SELECT id FROM user_groups WHERE id = {placeholder}",
        (group_id,)
    ).fetchone()
    if not group_row:
        return jsonify({"error": "Gruppo non trovato"}), 404
    
    # Validazione
    rounding_mode = data.get('rounding_mode', 'single')
    if rounding_mode not in ('single', 'daily'):
        return jsonify({"error": "rounding_mode deve essere 'single' o 'daily'"}), 400
    
    flessibilita_ingresso = int(data.get('flessibilita_ingresso_minuti', 30))
    flessibilita_uscita = int(data.get('flessibilita_uscita_minuti', 30))
    arrot_giornaliero = int(data.get('arrotondamento_giornaliero_minuti', 15))
    
    arrot_tipo = data.get('arrotondamento_giornaliero_tipo', 'floor')
    if arrot_tipo not in ('floor', 'ceil', 'nearest'):
        return jsonify({"error": "arrotondamento_giornaliero_tipo deve essere 'floor', 'ceil' o 'nearest'"}), 400
    
    oltre_action = data.get('oltre_flessibilita_action', 'allow')
    if oltre_action not in ('allow', 'warn', 'block'):
        return jsonify({"error": "oltre_flessibilita_action deve essere 'allow', 'warn' o 'block'"}), 400
    
    usa_pausa_std = 1 if data.get('usa_regole_pausa_standard', True) else 0
    is_active = 1 if data.get('is_active', True) else 0
    
    now = now_ms()
    user = session.get('user')
    
    # Verifica se esiste già una regola per questo gruppo
    existing = db.execute(
        f"SELECT id FROM group_timbratura_rules WHERE group_id = {placeholder}",
        (group_id,)
    ).fetchone()
    
    if existing:
        # Update
        db.execute(f"""
            UPDATE group_timbratura_rules SET
                rounding_mode = {placeholder},
                flessibilita_ingresso_minuti = {placeholder},
                flessibilita_uscita_minuti = {placeholder},
                arrotondamento_giornaliero_minuti = {placeholder},
                arrotondamento_giornaliero_tipo = {placeholder},
                oltre_flessibilita_action = {placeholder},
                usa_regole_pausa_standard = {placeholder},
                is_active = {placeholder},
                updated_ts = {placeholder},
                updated_by = {placeholder}
            WHERE group_id = {placeholder}
        """, (rounding_mode, flessibilita_ingresso, flessibilita_uscita, arrot_giornaliero,
              arrot_tipo, oltre_action, usa_pausa_std, is_active, now, user, group_id))
    else:
        # Insert
        db.execute(f"""
            INSERT INTO group_timbratura_rules 
                (group_id, rounding_mode, flessibilita_ingresso_minuti, flessibilita_uscita_minuti,
                 arrotondamento_giornaliero_minuti, arrotondamento_giornaliero_tipo,
                 oltre_flessibilita_action, usa_regole_pausa_standard, is_active,
                 created_ts, updated_ts, updated_by)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder},
                    {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
        """, (group_id, rounding_mode, flessibilita_ingresso, flessibilita_uscita, arrot_giornaliero,
              arrot_tipo, oltre_action, usa_pausa_std, is_active, now, now, user))
    
    db.commit()
    app.logger.info("Admin %s ha salvato regole timbratura per gruppo %s: mode=%s", user, group_id, rounding_mode)
    return jsonify({"success": True})


@app.delete("/api/admin/group-timbratura-rules/<int:group_id>")
@login_required
def api_delete_group_timbratura_rules(group_id: int):
    """Elimina le regole timbrature per un gruppo (torna a usare quelle globali)."""
    if not session.get("is_admin"):
        return jsonify({"error": "Non autorizzato"}), 403
    
    db = get_db()
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    db.execute(
        f"DELETE FROM group_timbratura_rules WHERE group_id = {placeholder}",
        (group_id,)
    )
    db.commit()
    
    app.logger.info("Admin %s ha eliminato regole timbratura per gruppo %s", session.get('user'), group_id)
    return jsonify({"success": True})


# ═══════════════════════════════════════════════════════════════════════════════
#  RILEVAMENTO EXTRA TURNO
# ═══════════════════════════════════════════════════════════════════════════════

def _detect_extra_turno(
    ora_timbrata: str,
    ora_mod: str,
    tipo: str,
    turno_start: str = None,
    turno_end: str = None,
    rules: dict = None
) -> Optional[dict]:
    """
    Rileva se una timbratura genera Extra Turno.
    
    LOGICA:
    - INGRESSO (inizio_giornata): Se l'ora timbrata è PRIMA di (turno_start - anticipo_max)
      → Extra Turno anticipato. I minuti extra sono da ora_timbrata a (turno_start - anticipo_max)
    
    - USCITA (fine_giornata): Se l'ora arrotondata (ora_mod) è DOPO fine_turno
      → Extra Turno posticipato. I minuti extra sono da fine_turno a ora_mod
    
    Args:
        ora_timbrata: orario originale timbrato (HH:MM:SS o HH:MM)
        ora_mod: orario arrotondato (HH:MM:SS o HH:MM)
        tipo: tipo timbratura (inizio_giornata, fine_giornata)
        turno_start: orario inizio turno (HH:MM) - richiesto per ingresso
        turno_end: orario fine turno (HH:MM) - richiesto per uscita
        rules: dizionario con le regole (se None, usa default)
    
    Returns:
        dict con dettagli Extra Turno se rilevato, None altrimenti
        {
            "extra_type": "before_shift" | "after_shift",
            "extra_minutes": int,
            "turno_time": str (HH:MM),
            "ora_timbrata": str,
            "ora_mod": str
        }
    """
    if tipo not in ('inizio_giornata', 'fine_giornata'):
        return None
    
    if rules is None:
        rules = {
            'anticipo_max_minuti': 30,
            'tolleranza_ritardo_minuti': 5,
            'arrotondamento_ingresso_minuti': 15,
            'arrotondamento_uscita_minuti': 15
        }
    
    # Converte ora timbrata in minuti
    parts = ora_timbrata.split(':')
    timbrata_min = int(parts[0]) * 60 + int(parts[1])
    
    # Converte ora_mod in minuti
    mod_parts = ora_mod.split(':')
    mod_min = int(mod_parts[0]) * 60 + int(mod_parts[1])
    
    if tipo == 'inizio_giornata' and turno_start:
        # INGRESSO: Extra Turno se timbro PRIMA di (turno_start - anticipo_max)
        turno_parts = turno_start.split(':')
        turno_min = int(turno_parts[0]) * 60 + int(turno_parts[1])
        anticipo_max = rules.get('anticipo_max_minuti', 30)
        
        # Soglia: turno_start - anticipo_max
        soglia_min = turno_min - anticipo_max
        
        app.logger.info(
            "_detect_extra_turno INGRESSO: timbrata_min=%s, turno_min=%s, anticipo_max=%s, soglia_min=%s, check=%s",
            timbrata_min, turno_min, anticipo_max, soglia_min, timbrata_min < soglia_min
        )
        
        if timbrata_min < soglia_min:
            # Extra Turno anticipato!
            # I minuti extra sono calcolati dall'ora_mod (già arrotondata) all'inizio del turno
            # L'ora_mod usa arrotondamento in eccesso per l'ingresso (sfavorevole al dipendente)
            # quindi i minuti extra saranno coerenti con l'orario mostrato
            
            # Usa l'ora_mod per calcolare i minuti extra (coerenza con il valore mostrato)
            mod_parts_inner = ora_mod.split(':')
            mod_min_inner = int(mod_parts_inner[0]) * 60 + int(mod_parts_inner[1])
            
            # Extra minuti = dall'ora_mod (arrotondata in eccesso) all'inizio del turno
            extra_minutes = turno_min - mod_min_inner
            
            if extra_minutes > 0:
                h_soglia = soglia_min // 60
                m_soglia = soglia_min % 60
                return {
                    "extra_type": "before_shift",
                    "extra_minutes": extra_minutes,
                    "turno_time": turno_start,
                    "soglia_time": f"{h_soglia:02d}:{m_soglia:02d}",
                    "ora_timbrata": ora_timbrata,
                    "ora_mod": ora_mod
                }
    
    elif tipo == 'fine_giornata' and turno_end:
        # USCITA: Extra Turno se ora_mod > fine_turno (+ flessibilità se daily mode)
        turno_parts = turno_end.split(':')
        turno_end_min = int(turno_parts[0]) * 60 + int(turno_parts[1])
        
        # Per modalità daily, considera la flessibilità uscita
        rounding_mode = rules.get('rounding_mode', 'single')
        flessibilita_uscita = 0
        if rounding_mode == 'daily':
            flessibilita_uscita = rules.get('flessibilita_uscita_minuti', 30)
        
        # Soglia per Extra Turno = fine turno + flessibilità
        soglia_extra = turno_end_min + flessibilita_uscita
        
        app.logger.info(
            "_detect_extra_turno USCITA: timbrata_min=%s, mod_min=%s, turno_end_min=%s, flessibilita=%s, soglia_extra=%s, check=%s",
            timbrata_min, mod_min, turno_end_min, flessibilita_uscita, soglia_extra, mod_min > soglia_extra
        )
        
        if mod_min > soglia_extra:
            # Extra Turno posticipato!
            # I minuti extra partono dalla soglia (fine turno + flessibilità)
            extra_minutes = mod_min - soglia_extra
            
            if extra_minutes > 0:
                return {
                    "extra_type": "after_shift",
                    "extra_minutes": extra_minutes,
                    "turno_time": turno_end,
                    "ora_timbrata": ora_timbrata,
                    "ora_mod": ora_mod,
                    "flessibilita_usata": flessibilita_uscita
                }
    
    return None


def _create_auto_extra_turno_request(
    db: DatabaseLike,
    username: str,
    date_str: str,
    extra_data: dict,
    notes: str = ""
) -> Optional[int]:
    """
    Crea automaticamente una richiesta di Extra Turno quando rilevato.
    
    Args:
        db: connessione database
        username: username dell'utente
        date_str: data in formato YYYY-MM-DD
        extra_data: dati Extra Turno da _detect_extra_turno
        notes: note opzionali
    
    Returns:
        ID della richiesta creata, o None se errore
    """
    try:
        ensure_user_requests_table(db)
        overtime_type_id = get_overtime_request_type_id(db)
        
        now_ts = int(time.time() * 1000)
        
        # Prepara extra_data con tutti i dettagli
        # Mappa i campi per compatibilità con il frontend admin_user_requests
        extra_type = extra_data.get("extra_type")
        ora_timbrata = extra_data.get("ora_timbrata", "")[:5] if extra_data.get("ora_timbrata") else ""
        ora_mod = extra_data.get("ora_mod", "")[:5] if extra_data.get("ora_mod") else ""
        planned_start = extra_data.get("planned_start", "")
        planned_end = extra_data.get("planned_end", "")
        
        request_extra_data = {
            "extra_type": extra_type,
            "extra_minutes": extra_data.get("extra_minutes"),
            "turno_time": extra_data.get("turno_time"),
            "ora_timbrata": extra_data.get("ora_timbrata"),
            "ora_mod": extra_data.get("ora_mod"),
            "soglia_time": extra_data.get("soglia_time"),
            "auto_detected": True,
            "overtime_type": extra_type,
            # Campi per visualizzazione nel riepilogo admin
            "planned_start": planned_start,
            "planned_end": planned_end,
            # Per ingresso anticipato: actual_start = ora_timbrata, actual_end = TBD
            # Per uscita posticipata: actual_start = TBD, actual_end = ora_timbrata
            "actual_start": ora_timbrata if extra_type == "before_shift" else "",
            "actual_end": ora_timbrata if extra_type == "after_shift" else "",
            "rounded_start": ora_mod if extra_type == "before_shift" else "",
            "rounded_end": ora_mod if extra_type == "after_shift" else "",
            # Minuti extra per tipo
            "extra_minutes_before": extra_data.get("extra_minutes") if extra_type == "before_shift" else 0,
            "extra_minutes_after": extra_data.get("extra_minutes") if extra_type == "after_shift" else 0,
        }
        extra_data_json = json.dumps(request_extra_data)
        
        total_minutes = extra_data.get("extra_minutes", 0)
        
        placeholder = "%s" if DB_VENDOR == "mysql" else "?"
        
        db.execute(f"""
            INSERT INTO user_requests 
            (user_id, username, request_type_id, date_from, date_to, value_amount, 
             notes, cdc, attachment_path, tratte, extra_data, status, created_ts, updated_ts)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, 
                    {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, 'pending', {placeholder}, {placeholder})
        """, (0, username, overtime_type_id, date_str, date_str, total_minutes, 
              notes, None, None, None, extra_data_json, now_ts, now_ts))
        
        # Recupera l'ID appena inserito
        if DB_VENDOR == "mysql":
            row = db.execute("SELECT LAST_INSERT_ID() as id").fetchone()
        else:
            row = db.execute("SELECT last_insert_rowid() as id").fetchone()
        
        request_id = row['id'] if isinstance(row, dict) else row[0]
        
        db.commit()
        
        app.logger.info(
            "Auto-created Extra Turno request: id=%s, user=%s, date=%s, type=%s, minutes=%s",
            request_id, username, date_str, extra_data.get("extra_type"), total_minutes
        )
        
        # Notifica admin
        _send_overtime_notification_to_admins(db, username, date_str, total_minutes)
        
        return request_id
        
    except Exception as e:
        app.logger.error(f"Errore creazione richiesta Extra Turno automatica: {e}")
        return None


def _create_flex_request(
    db, 
    username: str, 
    date_str: str, 
    tipo: str, 
    ora_timbrata: str, 
    ora_mod: str,
    diff_minutes: int,
    message: str,
    placeholder: str,
    turno_start: str = None,
    turno_end: str = None,
    arrotondamento_minuti: int = 30
) -> Optional[int]:
    """
    Crea automaticamente una richiesta di tipo "Fuori Flessibilità" quando
    una timbratura viene registrata oltre la flessibilità consentita.
    
    Args:
        db: connessione database
        username: username dell'utente
        date_str: data della timbratura (YYYY-MM-DD)
        tipo: tipo timbratura (inizio_giornata, fine_giornata)
        ora_timbrata: ora effettiva della timbratura
        ora_mod: ora modificata/arrotondata
        diff_minutes: minuti di differenza dal turno
        message: messaggio descrittivo
        placeholder: placeholder SQL (%s o ?)
        turno_start: ora inizio turno (HH:MM)
        turno_end: ora fine turno (HH:MM)
        arrotondamento_minuti: minuti per arrotondamento (dalle regole gruppo)
    
    Returns:
        ID della richiesta creata, o None se errore
    """
    try:
        FLEX_TYPE_ID = 17  # ID del tipo "Fuori Flessibilità"
        
        now_ts = now_ms()
        
        # Note descrittive
        tipo_label = "Ingresso" if tipo == "inizio_giornata" else "Uscita"
        notes = f"{tipo_label} fuori flessibilità: {message}"
        
        # Extra data con dettagli
        extra_data = {
            "tipo_timbratura": tipo,
            "ora_timbrata": ora_timbrata,
            "ora_mod": ora_mod,
            "diff_minutes": diff_minutes,
            "turno_start": turno_start,
            "turno_end": turno_end,
            "arrotondamento_minuti": arrotondamento_minuti,
            "auto_created": True,
            "created_reason": "timbratura_fuori_flessibilita"
        }
        extra_data_json = json.dumps(extra_data)
        
        db.execute(f"""
            INSERT INTO user_requests 
            (user_id, username, request_type_id, date_from, date_to, value_amount, 
             notes, cdc, attachment_path, tratte, extra_data, status, created_ts, updated_ts)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, 
                    {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, 'pending', {placeholder}, {placeholder})
        """, (0, username, FLEX_TYPE_ID, date_str, date_str, abs(diff_minutes), 
              notes, None, None, None, extra_data_json, now_ts, now_ts))
        
        # Recupera l'ID appena inserito
        if DB_VENDOR == "mysql":
            row = db.execute("SELECT LAST_INSERT_ID() as id").fetchone()
        else:
            row = db.execute("SELECT last_insert_rowid() as id").fetchone()
        
        request_id = row['id'] if isinstance(row, dict) else row[0]
        
        db.commit()
        
        app.logger.info(
            "Auto-created Fuori Flessibilità request: id=%s, user=%s, date=%s, tipo=%s, diff=%smin",
            request_id, username, date_str, tipo, diff_minutes
        )
        
        # Notifica admin
        _send_flex_notification_to_admins(db, username, date_str, tipo, diff_minutes)
        
        return request_id
        
    except Exception as e:
        app.logger.error(f"Errore creazione richiesta Fuori Flessibilità: {e}")
        import traceback
        app.logger.error(traceback.format_exc())
        return None


def _send_flex_notification_to_admins(db, username: str, date_str: str, tipo: str, diff_minutes: int):
    """Invia notifica push agli admin per richiesta fuori flessibilità."""
    try:
        tipo_label = "Ingresso" if tipo == "inizio_giornata" else "Uscita"
        direction = "ritardo" if diff_minutes > 0 else "anticipo"
        
        title = f"⏰ Timbratura fuori flessibilità"
        body = f"{username}: {tipo_label} {abs(diff_minutes)} min di {direction} il {date_str}"
        
        # Recupera admin attivi
        placeholder = "%s" if DB_VENDOR == "mysql" else "?"
        admins = db.execute(
            f"SELECT username FROM app_users WHERE role IN ('admin', 'superadmin') AND is_active = 1"
        ).fetchall()
        
        for admin_row in admins:
            admin_username = admin_row['username'] if isinstance(admin_row, dict) else admin_row[0]
            try:
                send_push_notification(db, admin_username, title, body, data={"type": "flex_request"})
            except Exception as e:
                app.logger.debug(f"Push notification to {admin_username} failed: {e}")
    except Exception as e:
        app.logger.warning(f"Errore invio notifiche flex agli admin: {e}")


def _calcola_ora_fine_daily(db, username: str, today: str, ora: str, turno_start: str, turno_end: str, rules: dict, placeholder: str) -> str:
    """
    Calcola l'ora di fine giornata per il mode 'daily'.
    
    In modalità giornaliera, l'ora di uscita viene calcolata in modo che:
    - Le ore nette risultino esattamente quelle del turno
    - Considera l'ora di inizio effettiva e la pausa effettuata
    
    Formula: ora_mod_fine = ora_inizio + ore_turno + pausa
    
    Esempio:
        - Turno: 09:00-18:00 (8h nette con 1h pausa)
        - Inizio effettivo: 09:20
        - Pausa: 60 min
        - ora_mod_fine = 09:20 + 8h + 1h = 18:20
    """
    try:
        # 1. Recupera l'ora di inizio giornata di oggi (ora_mod se presente)
        inizio_row = db.execute(
            f"""SELECT ora, ora_mod FROM timbrature 
               WHERE username = {placeholder} AND data = {placeholder} AND tipo = 'inizio_giornata'
               ORDER BY created_ts ASC LIMIT 1""",
            (username, today)
        ).fetchone()
        
        if not inizio_row:
            app.logger.warning(f"Daily mode: nessun inizio giornata trovato per {username} il {today}")
            return ora if ':' in ora and ora.count(':') == 2 else f"{ora}:00"
        
        ora_inizio = inizio_row['ora_mod'] if isinstance(inizio_row, dict) else inizio_row[1]
        if not ora_inizio:
            ora_inizio = inizio_row['ora'] if isinstance(inizio_row, dict) else inizio_row[0]
        
        # Converti in stringa
        if hasattr(ora_inizio, 'strftime'):
            ora_inizio_str = ora_inizio.strftime("%H:%M")
        elif hasattr(ora_inizio, 'total_seconds'):
            total_sec = int(ora_inizio.total_seconds())
            ora_inizio_str = f"{total_sec // 3600:02d}:{(total_sec % 3600) // 60:02d}"
        else:
            ora_inizio_str = str(ora_inizio)[:5]
        
        # 2. Calcola la pausa effettuata oggi (somma di tutte le pause chiuse)
        pausa_minuti = 0
        pause_rows = db.execute(
            f"""SELECT tipo, ora, ora_mod FROM timbrature 
               WHERE username = {placeholder} AND data = {placeholder} AND tipo IN ('inizio_pausa', 'fine_pausa')
               ORDER BY created_ts ASC""",
            (username, today)
        ).fetchall()
        
        inizio_pausa_min = None
        for row in pause_rows:
            tipo_pausa = row['tipo'] if isinstance(row, dict) else row[0]
            ora_p = row['ora_mod'] if isinstance(row, dict) else row[2]
            if not ora_p:
                ora_p = row['ora'] if isinstance(row, dict) else row[1]
            
            if hasattr(ora_p, 'strftime'):
                ora_p_str = ora_p.strftime("%H:%M")
            elif hasattr(ora_p, 'total_seconds'):
                total_sec = int(ora_p.total_seconds())
                ora_p_str = f"{total_sec // 3600:02d}:{(total_sec % 3600) // 60:02d}"
            else:
                ora_p_str = str(ora_p)[:5]
            
            p_parts = ora_p_str.split(':')
            p_min = int(p_parts[0]) * 60 + int(p_parts[1])
            
            if tipo_pausa == 'inizio_pausa':
                inizio_pausa_min = p_min
            elif tipo_pausa == 'fine_pausa' and inizio_pausa_min is not None:
                pausa_minuti += p_min - inizio_pausa_min
                inizio_pausa_min = None
        
        # 3. Calcola le ore nette del turno
        ore_turno_minuti = 480  # Default 8 ore
        if turno_start and turno_end:
            ts_parts = turno_start.split(':')
            te_parts = turno_end.split(':')
            ts_min = int(ts_parts[0]) * 60 + int(ts_parts[1])
            te_min = int(te_parts[0]) * 60 + int(te_parts[1])
            
            # Ore lorde turno - pausa del turno (da employee_shifts)
            ore_lorde_turno = te_min - ts_min
            
            # Recupera la pausa prevista dal turno (break_start/break_end)
            pausa_turno_minuti = 0
            try:
                day_of_week = datetime.strptime(today, '%Y-%m-%d').weekday()
                shift_row = db.execute(
                    f"""SELECT break_start, break_end FROM employee_shifts 
                       WHERE username = {placeholder} AND day_of_week = {placeholder}""",
                    (username, day_of_week)
                ).fetchone()
                
                if shift_row:
                    break_start = shift_row['break_start'] if isinstance(shift_row, dict) else shift_row[0]
                    break_end = shift_row['break_end'] if isinstance(shift_row, dict) else shift_row[1]
                    
                    if break_start and break_end:
                        if hasattr(break_start, 'total_seconds'):
                            bs_min = int(break_start.total_seconds()) // 60
                        else:
                            bs_parts = str(break_start)[:5].split(':')
                            bs_min = int(bs_parts[0]) * 60 + int(bs_parts[1])
                        
                        if hasattr(break_end, 'total_seconds'):
                            be_min = int(break_end.total_seconds()) // 60
                        else:
                            be_parts = str(break_end)[:5].split(':')
                            be_min = int(be_parts[0]) * 60 + int(be_parts[1])
                        
                        pausa_turno_minuti = be_min - bs_min
            except Exception as e:
                app.logger.warning(f"Errore lettura pausa turno: {e}")
            
            ore_turno_minuti = ore_lorde_turno - pausa_turno_minuti
        
        # 4. Applica arrotondamento giornaliero alle ore nette (arrotonda per difetto)
        blocco = rules.get('arrotondamento_giornaliero_minuti', 15)
        tipo_arrot = rules.get('arrotondamento_giornaliero_tipo', 'floor')
        
        if tipo_arrot == 'floor':
            ore_arrotondate = (ore_turno_minuti // blocco) * blocco
        elif tipo_arrot == 'ceil':
            ore_arrotondate = ((ore_turno_minuti + blocco - 1) // blocco) * blocco
        else:  # round
            ore_arrotondate = round(ore_turno_minuti / blocco) * blocco
        
        # 5. Calcola ora_mod_fine = ora_inizio + ore_arrotondate + pausa
        inizio_parts = ora_inizio_str.split(':')
        inizio_min = int(inizio_parts[0]) * 60 + int(inizio_parts[1])
        
        ora_fine_mod_min = inizio_min + ore_arrotondate + pausa_minuti
        
        h = ora_fine_mod_min // 60
        m = ora_fine_mod_min % 60
        ora_mod = f"{h:02d}:{m:02d}:00"
        
        app.logger.info(
            f"Daily mode fine_giornata: inizio={ora_inizio_str}, pausa={pausa_minuti}min, "
            f"ore_turno={ore_turno_minuti}min, ore_arrot={ore_arrotondate}min, "
            f"ora_timbrata={ora}, ora_mod={ora_mod}"
        )
        
        return ora_mod
        
    except Exception as e:
        app.logger.error(f"Errore _calcola_ora_fine_daily: {e}")
        import traceback
        app.logger.error(traceback.format_exc())
        return ora if ':' in ora and ora.count(':') == 2 else f"{ora}:00"


# ═══════════════════════════════════════════════════════════════════════════════
#  CALCOLO ORA MODIFICATA (ora_mod)
# ═══════════════════════════════════════════════════════════════════════════════

def calcola_ora_mod(ora_originale: str, tipo: str, turno_start: str = None, rules: dict = None) -> str:
    """
    Calcola l'ora modificata in base alle regole.
    
    Args:
        ora_originale: orario originale (HH:MM:SS o HH:MM)
        tipo: tipo timbratura (inizio_giornata, fine_giornata, inizio_pausa, fine_pausa)
        turno_start: orario inizio turno se presente (HH:MM)
        rules: dizionario con le regole (se None, usa default)
    
    Returns:
        ora modificata (HH:MM:SS)
    
    Note:
        Per gruppi con rounding_mode='daily', questa funzione restituisce l'ora originale
        senza arrotondamento (l'arrotondamento viene fatto sul totale giornaliero).
    """
    if rules is None:
        rules = {
            'anticipo_max_minuti': 30,
            'tolleranza_ritardo_minuti': 5,
            'arrotondamento_ingresso_minuti': 15,
            'arrotondamento_uscita_minuti': 15,
            'rounding_mode': 'single'
        }
    
    # Converte ora originale in minuti
    parts = ora_originale.split(':')
    ora_min = int(parts[0]) * 60 + int(parts[1])
    
    # Se rounding_mode è 'daily', non arrotonda la singola timbrata
    # Verifica solo la flessibilità
    rounding_mode = rules.get('rounding_mode', 'single')
    
    if rounding_mode == 'daily':
        # Modalità giornaliera: non arrotonda, verifica solo flessibilità
        if turno_start and tipo == 'inizio_giornata':
            turno_parts = turno_start.split(':')
            turno_min = int(turno_parts[0]) * 60 + int(turno_parts[1])
            flessibilita = rules.get('flessibilita_ingresso_minuti', 30)
            
            # Finestra flessibilità: [turno - fless, turno + fless]
            oltre_action = rules.get('oltre_flessibilita_action', 'allow')
            
            diff = abs(ora_min - turno_min)
            if diff <= flessibilita:
                # Dentro la flessibilità: registra ora reale
                return ora_originale if ':' in ora_originale and ora_originale.count(':') == 2 else f"{ora_originale}:00"
            else:
                # Fuori flessibilità: dipende dall'azione configurata
                if oltre_action == 'block':
                    # In teoria dovrebbe bloccare, ma qui restituiamo comunque l'ora
                    # Il blocco va gestito a livello di API
                    pass
                # Per 'allow' e 'warn': registra ora reale
                return ora_originale if ':' in ora_originale and ora_originale.count(':') == 2 else f"{ora_originale}:00"
        
        # Per fine_giornata e altri tipi: registra ora reale
        return ora_originale if ':' in ora_originale and ora_originale.count(':') == 2 else f"{ora_originale}:00"
    
    # Modalità 'single': comportamento originale
    # Se c'è un turno e siamo in ingresso, prova normalizzazione
    if turno_start and tipo == 'inizio_giornata':
        turno_parts = turno_start.split(':')
        turno_min = int(turno_parts[0]) * 60 + int(turno_parts[1])
        
        anticipo = rules.get('anticipo_max_minuti', 30)
        tolleranza = rules.get('tolleranza_ritardo_minuti', 5)
        
        # Finestra di normalizzazione: [turno - anticipo, turno + tolleranza]
        min_window = turno_min - anticipo
        max_window = turno_min + tolleranza
        
        if min_window <= ora_min <= max_window:
            # Normalizza all'orario del turno
            h = turno_min // 60
            m = turno_min % 60
            return f"{h:02d}:{m:02d}:00"
    
    # Arrotondamento
    if tipo == 'inizio_giornata':
        # Arrotonda in eccesso (sfavorevole al dipendente per ingresso)
        blocco = rules.get('arrotondamento_ingresso_minuti', 15)
        ora_mod_min = ((ora_min + blocco - 1) // blocco) * blocco
    elif tipo == 'fine_giornata':
        # Arrotonda in difetto (sfavorevole al dipendente per uscita)
        blocco = rules.get('arrotondamento_uscita_minuti', 15)
        ora_mod_min = (ora_min // blocco) * blocco
    else:
        # Per inizio_pausa e fine_pausa, nessun arrotondamento sull'ora singola
        # Le pause vengono gestite sulla durata totale
        return ora_originale if ':' in ora_originale and ora_originale.count(':') == 2 else f"{ora_originale}:00"
    
    h = ora_mod_min // 60
    m = ora_mod_min % 60
    return f"{h:02d}:{m:02d}:00"


def calcola_ore_giornaliere_arrotondate(
    ora_inizio: str,
    ora_fine: str,
    pausa_minuti: int,
    rules: dict
) -> dict:
    """
    Calcola le ore giornaliere con arrotondamento a fine giornata (daily mode).
    
    Args:
        ora_inizio: orario inizio giornata (HH:MM o HH:MM:SS)
        ora_fine: orario fine giornata (HH:MM o HH:MM:SS)
        pausa_minuti: durata pausa in minuti
        rules: regole del gruppo con rounding_mode='daily'
    
    Returns:
        dict con:
        - ore_lorde: minuti lordi (fine - inizio)
        - ore_nette: minuti netti (lordi - pausa)
        - ore_arrotondate: minuti arrotondati secondo regola
        - ore_str: stringa formattata (es. "8:15")
        - turno_base_minuti: durata turno base in minuti
        - straordinario_lordo: minuti straordinario lordo (netto - turno)
        - straordinario_arrotondato: minuti straordinario arrotondato
        - blocco_minuti: blocco usato per arrotondamento
        - blocchi_straordinario: numero blocchi di straordinario
    """
    # Parse orari
    inizio_parts = ora_inizio.split(':')
    fine_parts = ora_fine.split(':')
    
    inizio_min = int(inizio_parts[0]) * 60 + int(inizio_parts[1])
    fine_min = int(fine_parts[0]) * 60 + int(fine_parts[1])
    
    # Calcola
    ore_lorde = fine_min - inizio_min
    ore_nette = ore_lorde - pausa_minuti
    
    # Arrotondamento
    blocco = rules.get('arrotondamento_giornaliero_minuti', 15)
    tipo_arrot = rules.get('arrotondamento_giornaliero_tipo', 'floor')
    
    # Turno base (default 8 ore = 480 minuti)
    turno_base = rules.get('turno_base_minuti', 480)
    
    # Calcola straordinario lordo
    straordinario_lordo = max(0, ore_nette - turno_base)
    
    # Arrotonda lo straordinario per blocchi
    if straordinario_lordo > 0:
        if tipo_arrot == 'floor':
            blocchi = straordinario_lordo // blocco
        elif tipo_arrot == 'ceil':
            blocchi = (straordinario_lordo + blocco - 1) // blocco
        else:  # nearest
            blocchi = round(straordinario_lordo / blocco)
        straordinario_arrotondato = blocchi * blocco
    else:
        blocchi = 0
        straordinario_arrotondato = 0
    
    # Ore totali arrotondate = turno base + straordinario arrotondato (ma non superiore alle ore nette)
    ore_arrotondate = min(turno_base + straordinario_arrotondato, ore_nette)
    
    # Se sotto il turno base, arrotonda anche quello
    if ore_nette < turno_base:
        if tipo_arrot == 'floor':
            ore_arrotondate = (ore_nette // blocco) * blocco
        elif tipo_arrot == 'ceil':
            ore_arrotondate = ((ore_nette + blocco - 1) // blocco) * blocco
        else:  # nearest
            ore_arrotondate = round(ore_nette / blocco) * blocco
    
    # Formatta stringa
    h = ore_arrotondate // 60
    m = ore_arrotondate % 60
    ore_str = f"{h}:{m:02d}"
    
    return {
        'ore_lorde': ore_lorde,
        'ore_nette': ore_nette,
        'ore_arrotondate': ore_arrotondate,
        'ore_str': ore_str,
        'turno_base_minuti': turno_base,
        'straordinario_lordo': straordinario_lordo,
        'straordinario_arrotondato': straordinario_arrotondato,
        'blocco_minuti': blocco,
        'blocchi_straordinario': blocchi,
        'tipo_arrotondamento': tipo_arrot
    }


def verifica_flessibilita_timbrata(
    ora_timbrata: str,
    tipo: str,
    turno_start: str,
    turno_end: str,
    rules: dict
) -> dict:
    """
    Verifica se una timbrata è dentro la flessibilità per gruppi con daily mode.
    
    Args:
        ora_timbrata: orario timbrato (HH:MM o HH:MM:SS)
        tipo: 'inizio_giornata' o 'fine_giornata'
        turno_start: orario inizio turno (HH:MM)
        turno_end: orario fine turno (HH:MM)
        rules: regole del gruppo
    
    Returns:
        dict con:
        - within_flex: True se dentro flessibilità
        - diff_minutes: differenza in minuti dal turno
        - action: 'allow', 'warn', 'block'
        - message: messaggio per l'utente
    """
    parts = ora_timbrata.split(':')
    ora_min = int(parts[0]) * 60 + int(parts[1])
    
    if tipo == 'inizio_giornata' and turno_start:
        turno_parts = turno_start.split(':')
        turno_min = int(turno_parts[0]) * 60 + int(turno_parts[1])
        flessibilita = rules.get('flessibilita_ingresso_minuti', 30)
        diff = ora_min - turno_min  # positivo = ritardo, negativo = anticipo
    elif tipo == 'fine_giornata' and turno_end:
        turno_parts = turno_end.split(':')
        turno_min = int(turno_parts[0]) * 60 + int(turno_parts[1])
        flessibilita = rules.get('flessibilita_uscita_minuti', 30)
        diff = ora_min - turno_min  # positivo = dopo turno, negativo = prima
    else:
        return {'within_flex': True, 'diff_minutes': 0, 'action': 'allow', 'message': ''}
    
    within_flex = abs(diff) <= flessibilita
    action = rules.get('oltre_flessibilita_action', 'allow')
    
    message = ''
    if not within_flex:
        if diff > 0:
            message = f"Timbrata {abs(diff)} minuti oltre la flessibilità"
        else:
            message = f"Timbrata {abs(diff)} minuti prima della flessibilità"
    
    return {
        'within_flex': within_flex,
        'diff_minutes': diff,
        'action': action if not within_flex else 'allow',
        'message': message
    }


def calcola_pausa_mod(inizio_pausa: str, fine_pausa: str, rules: dict = None) -> int:
    """
    Calcola la durata della pausa modificata in minuti.
    
    Args:
        inizio_pausa: orario inizio pausa (HH:MM:SS o HH:MM)
        fine_pausa: orario fine pausa (HH:MM:SS o HH:MM)
        rules: dizionario con le regole (se None, usa default)
    
    Returns:
        durata pausa in minuti (arrotondata secondo le regole)
    """
    if rules is None:
        rules = {
            'pausa_blocco_minimo_minuti': 30,
            'pausa_incremento_minuti': 15,
            'pausa_tolleranza_minuti': 5
        }
    
    # Converte in minuti
    ip = inizio_pausa.split(':')
    fp = fine_pausa.split(':')
    inizio_min = int(ip[0]) * 60 + int(ip[1])
    fine_min = int(fp[0]) * 60 + int(fp[1])
    
    durata_effettiva = fine_min - inizio_min
    
    blocco_min = rules.get('pausa_blocco_minimo_minuti', 30)
    incremento = rules.get('pausa_incremento_minuti', 15)
    tolleranza = rules.get('pausa_tolleranza_minuti', 5)
    
    # Se durata < blocco minimo, usa blocco minimo
    if durata_effettiva <= blocco_min:
        return blocco_min
    
    # Calcola eccesso rispetto al blocco minimo
    eccesso = durata_effettiva - blocco_min
    
    # Quanti blocchi di incremento sono necessari?
    blocchi_extra = eccesso // incremento
    resto = eccesso % incremento
    
    # Se il resto è > tolleranza, aggiungi un blocco
    if resto > tolleranza:
        blocchi_extra += 1
    
    return blocco_min + (blocchi_extra * incremento)


# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIGURAZIONE AZIENDA - COMPANY SETTINGS
# ═══════════════════════════════════════════════════════════════════════════════

COMPANY_SETTINGS_TABLE_MYSQL = """
CREATE TABLE IF NOT EXISTS company_settings (
    id INT PRIMARY KEY DEFAULT 1,
    company_name VARCHAR(200) NOT NULL DEFAULT 'La Mia Azienda',
    external_id VARCHAR(100),
    logo_path VARCHAR(500),
    address TEXT,
    phone VARCHAR(50),
    email VARCHAR(200),
    website VARCHAR(200),
    vat_number VARCHAR(50),
    fiscal_code VARCHAR(50),
    modules_enabled JSON,
    custom_settings JSON,
    created_ts BIGINT NOT NULL,
    updated_ts BIGINT NOT NULL,
    updated_by VARCHAR(100)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
"""

COMPANY_SETTINGS_TABLE_SQLITE = """
CREATE TABLE IF NOT EXISTS company_settings (
    id INTEGER PRIMARY KEY DEFAULT 1,
    company_name TEXT NOT NULL DEFAULT 'La Mia Azienda',
    external_id TEXT,
    logo_path TEXT,
    address TEXT,
    phone TEXT,
    email TEXT,
    website TEXT,
    vat_number TEXT,
    fiscal_code TEXT,
    modules_enabled TEXT,
    custom_settings TEXT,
    created_ts INTEGER NOT NULL,
    updated_ts INTEGER NOT NULL,
    updated_by TEXT
)
"""


def ensure_company_settings_table(db: DatabaseLike) -> None:
    """Crea la tabella company_settings se non esiste."""
    statement = (
        COMPANY_SETTINGS_TABLE_MYSQL if DB_VENDOR == "mysql" else COMPANY_SETTINGS_TABLE_SQLITE
    )
    for stmt in statement.strip().split(";"):
        sql = stmt.strip()
        if not sql:
            continue
        cursor = db.execute(sql)
        try:
            cursor.close()
        except AttributeError:
            pass
    db.commit()


def get_company_settings(db: DatabaseLike) -> dict:
    """Ottiene le impostazioni azienda dal database."""
    ensure_company_settings_table(db)
    
    cursor = db.execute("SELECT * FROM company_settings WHERE id = 1")
    row = cursor.fetchone()
    
    if not row:
        # Inserisci valori di default
        now_ts = int(time.time() * 1000)
        if DB_VENDOR == "mysql":
            db.execute(
                """INSERT INTO company_settings 
                   (id, company_name, modules_enabled, custom_settings, created_ts, updated_ts) 
                   VALUES (1, 'La Mia Azienda', '{}', '{}', %s, %s)""",
                (now_ts, now_ts)
            )
        else:
            db.execute(
                """INSERT INTO company_settings 
                   (id, company_name, modules_enabled, custom_settings, created_ts, updated_ts) 
                   VALUES (1, 'La Mia Azienda', '{}', '{}', ?, ?)""",
                (now_ts, now_ts)
            )
        db.commit()
        cursor = db.execute("SELECT * FROM company_settings WHERE id = 1")
        row = cursor.fetchone()
    
    # Converti in dizionario - gestisce sia tuple che dict
    if isinstance(row, dict):
        settings = dict(row)
    else:
        columns = [desc[0] for desc in cursor.description]
        settings = dict(zip(columns, row))
    
    # Parse JSON fields
    for json_field in ['modules_enabled', 'custom_settings']:
        if settings.get(json_field):
            try:
                if isinstance(settings[json_field], str):
                    settings[json_field] = json.loads(settings[json_field])
            except (json.JSONDecodeError, TypeError):
                settings[json_field] = {}
        else:
            settings[json_field] = {}
    
    return settings


def is_module_enabled(db: DatabaseLike, module_name: str) -> bool:
    """Verifica se un modulo è attivo nelle impostazioni azienda.
    
    Args:
        db: connessione database
        module_name: nome del modulo (es. 'straordinari', 'magazzino', 'requests')
    
    Returns:
        True se il modulo è attivo (default True per moduli non specificati)
    """
    settings = get_company_settings(db)
    modules = settings.get('modules_enabled', {})
    enabled = modules.get(module_name, True)
    app.logger.info(f"Modulo '{module_name}' attivo: {enabled} (modules_enabled: {modules})")
    return enabled


def save_company_settings(db: DatabaseLike, data: dict, updated_by: str) -> bool:
    """Salva le impostazioni azienda (INSERT o UPDATE)."""
    ensure_company_settings_table(db)
    now_ts = int(time.time() * 1000)
    
    # Prepara JSON fields
    modules_enabled = json.dumps(data.get('modules_enabled', {}))
    custom_settings = json.dumps(data.get('custom_settings', {}))
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Verifica se esiste già un record
    cursor = db.execute("SELECT id FROM company_settings WHERE id = 1")
    exists = cursor.fetchone() is not None
    
    if exists:
        # UPDATE
        db.execute(f"""
            UPDATE company_settings SET
                company_name = {placeholder},
                external_id = {placeholder},
                logo_path = {placeholder},
                address = {placeholder},
                phone = {placeholder},
                email = {placeholder},
                website = {placeholder},
                vat_number = {placeholder},
                fiscal_code = {placeholder},
                modules_enabled = {placeholder},
                custom_settings = {placeholder},
                updated_ts = {placeholder},
                updated_by = {placeholder}
            WHERE id = 1
        """, (
            data.get('company_name', 'La Mia Azienda'),
            data.get('external_id'),
            data.get('logo_path'),
            data.get('address'),
            data.get('phone'),
            data.get('email'),
            data.get('website'),
            data.get('vat_number'),
            data.get('fiscal_code'),
            modules_enabled,
            custom_settings,
            now_ts,
            updated_by
        ))
    else:
        # INSERT
        db.execute(f"""
            INSERT INTO company_settings (
                id, company_name, external_id, logo_path, address, phone, email, 
                website, vat_number, fiscal_code, modules_enabled, custom_settings, 
                created_ts, updated_ts, updated_by
            ) VALUES (
                1, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, 
                {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, 
                {placeholder}, {placeholder}, {placeholder}, {placeholder}
            )
        """, (
            data.get('company_name', 'La Mia Azienda'),
            data.get('external_id'),
            data.get('logo_path'),
            data.get('address'),
            data.get('phone'),
            data.get('email'),
            data.get('website'),
            data.get('vat_number'),
            data.get('fiscal_code'),
            modules_enabled,
            custom_settings,
            now_ts,
            now_ts,
            updated_by
        ))
    
    db.commit()
    return True


@app.get("/admin/company-settings")
@login_required
def admin_company_settings_page() -> ResponseReturnValue:
    """Pagina configurazione azienda (solo admin)."""
    if not session.get("is_admin"):
        abort(403)

    display_name = session.get("user_display") or session.get("user_name") or session.get("user")
    primary_name = session.get("user_name") or display_name or session.get("user")
    initials = session.get("user_initials") or compute_initials(primary_name or "")

    return render_template(
        "admin_company_settings.html",
        user_name=primary_name,
        user_display=display_name,
        user_initials=initials,
        is_admin=True,
    )


@app.get("/api/admin/company-settings")
@login_required
def api_get_company_settings() -> ResponseReturnValue:
    """API per ottenere le impostazioni azienda."""
    if not session.get("is_admin"):
        return jsonify({"error": "Non autorizzato"}), 403
    
    db = get_db()
    settings = get_company_settings(db)
    
    # Aggiungi flag per indicare se CedolinoWeb è configurato nel config.json
    config = load_config()
    cedolino_section = config.get("cedolino_web", {})
    settings["cedolino_configured"] = bool(
        cedolino_section.get("enabled") and 
        cedolino_section.get("username") and 
        cedolino_section.get("password")
    )
    
    return jsonify(settings)


@app.post("/api/admin/company-settings")
@login_required
def api_save_company_settings() -> ResponseReturnValue:
    """API per salvare le impostazioni azienda."""
    if not session.get("is_admin"):
        return jsonify({"error": "Non autorizzato"}), 403
    
    data = request.get_json()
    if not data:
        return jsonify({"error": "Dati mancanti"}), 400
    
    db = get_db()
    username = session.get("user") or session.get("username") or "admin"
    
    try:
        save_company_settings(db, data, username)
        return jsonify({"ok": True, "message": "Impostazioni salvate"})
    except Exception as e:
        app.logger.error(f"Errore salvataggio company settings: {e}")
        return jsonify({"error": str(e)}), 500


@app.post("/api/admin/company-settings/logo")
@login_required
def api_upload_company_logo() -> ResponseReturnValue:
    """API per caricare il logo aziendale."""
    if not session.get("is_admin"):
        return jsonify({"error": "Non autorizzato"}), 403
    
    if 'logo' not in request.files:
        return jsonify({"error": "Nessun file caricato"}), 400
    
    file = request.files['logo']
    if file.filename == '':
        return jsonify({"error": "Nessun file selezionato"}), 400
    
    # Verifica estensione
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'svg', 'webp'}
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in allowed_extensions:
        return jsonify({"error": f"Formato non supportato. Usa: {', '.join(allowed_extensions)}"}), 400
    
    # Salva il file
    logo_dir = os.path.join(app.root_path, 'static', 'uploads', 'logo')
    os.makedirs(logo_dir, exist_ok=True)
    
    # Nome file univoco
    filename = f"company_logo_{int(time.time())}.{ext}"
    filepath = os.path.join(logo_dir, filename)
    file.save(filepath)
    
    # Path relativo per il database
    logo_path = f"/static/uploads/logo/{filename}"
    
    # Aggiorna database
    db = get_db()
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    now_ts = int(time.time() * 1000)
    username = session.get("user") or session.get("username") or "admin"
    
    ensure_company_settings_table(db)
    db.execute(f"""
        UPDATE company_settings SET 
            logo_path = {placeholder}, 
            updated_ts = {placeholder},
            updated_by = {placeholder}
        WHERE id = 1
    """, (logo_path, now_ts, username))
    db.commit()
    
    return jsonify({"ok": True, "logo_path": logo_path})


@app.delete("/api/admin/company-settings/logo")
@login_required
def api_delete_company_logo() -> ResponseReturnValue:
    """API per eliminare il logo aziendale."""
    if not session.get("is_admin"):
        return jsonify({"error": "Non autorizzato"}), 403
    
    db = get_db()
    settings = get_company_settings(db)
    
    # Elimina file se esiste
    if settings.get('logo_path'):
        old_path = os.path.join(app.root_path, settings['logo_path'].lstrip('/'))
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except Exception as e:
                app.logger.warning(f"Errore eliminazione logo: {e}")
    
    # Aggiorna database
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    now_ts = int(time.time() * 1000)
    username = session.get("user") or session.get("username") or "admin"
    
    db.execute(f"""
        UPDATE company_settings SET 
            logo_path = NULL, 
            updated_ts = {placeholder},
            updated_by = {placeholder}
        WHERE id = 1
    """, (now_ts, username))
    db.commit()
    
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
#  GESTIONE TIPOLOGIE RICHIESTE - ADMIN UI
# ═══════════════════════════════════════════════════════════════════════════════

VALUE_TYPE_LABELS = {
    "hours": "Ore",
    "days": "Giorni",
    "amount": "Importo €",
    "km": "Chilometri",
    "minutes": "Minuti"
}


@app.get("/admin/request-types")
@login_required
def admin_request_types_page() -> ResponseReturnValue:
    """Pagina gestione tipologie richieste (solo admin)."""
    if not session.get("is_admin"):
        abort(403)

    display_name = session.get("user_display") or session.get("user_name") or session.get("user")
    primary_name = session.get("user_name") or display_name or session.get("user")
    initials = session.get("user_initials") or compute_initials(primary_name or "")

    return render_template(
        "admin_request_types.html",
        user_name=primary_name,
        user_display=display_name,
        user_initials=initials,
        is_admin=True,
    )


@app.get("/api/admin/request-types")
@login_required
def api_admin_request_types_list() -> ResponseReturnValue:
    """Lista tutte le tipologie di richiesta."""
    if not is_admin_or_supervisor():
        return jsonify({"error": "forbidden"}), 403

    try:
        db = get_db()
        ensure_request_types_table(db)
        
        rows = db.execute("""
            SELECT id, name, value_type, external_id, abbreviation, description, active, sort_order, created_ts, updated_ts, is_giustificativo
            FROM request_types
            ORDER BY sort_order ASC, name ASC
        """).fetchall()
    except Exception as e:
        app.logger.error(f"Errore in api_admin_request_types_list: {e}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

    types = []
    for row in rows:
        if isinstance(row, Mapping):
            types.append({
                "id": row["id"],
                "name": row["name"],
                "value_type": row["value_type"],
                "value_type_label": VALUE_TYPE_LABELS.get(row["value_type"], row["value_type"]),
                "external_id": row["external_id"],
                "abbreviation": row["abbreviation"],
                "description": row["description"],
                "active": bool(row["active"]),
                "sort_order": row["sort_order"],
                "created_ts": row["created_ts"],
                "updated_ts": row["updated_ts"],
                "is_giustificativo": bool(row.get("is_giustificativo", False)),
            })
        else:
            types.append({
                "id": row[0],
                "name": row[1],
                "value_type": row[2],
                "value_type_label": VALUE_TYPE_LABELS.get(row[2], row[2]),
                "external_id": row[3],
                "abbreviation": row[4],
                "description": row[5],
                "active": bool(row[6]),
                "sort_order": row[7],
                "created_ts": row[8],
                "updated_ts": row[9],
                "is_giustificativo": bool(row[10]) if len(row) > 10 else False,
            })

    return jsonify({"types": types, "value_types": VALUE_TYPE_LABELS})


@app.post("/api/admin/request-types")
@login_required
def api_admin_request_types_create() -> ResponseReturnValue:
    """Crea una nuova tipologia di richiesta."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403

    data = request.get_json()
    if not data:
        return jsonify({"error": "Dati non validi"}), 400

    name = (data.get("name") or "").strip()
    value_type = data.get("value_type", "hours")
    external_id = (data.get("external_id") or "").strip() or None
    abbreviation = (data.get("abbreviation") or "").strip() or None
    description = (data.get("description") or "").strip() or None
    active = data.get("active", True)
    sort_order = data.get("sort_order", 0)
    is_giustificativo = data.get("is_giustificativo", False)

    if not name:
        return jsonify({"error": "Il nome è obbligatorio"}), 400

    if value_type not in VALUE_TYPE_LABELS:
        return jsonify({"error": f"Tipo valore non valido. Valori ammessi: {list(VALUE_TYPE_LABELS.keys())}"}), 400

    db = get_db()
    ensure_request_types_table(db)
    now_ms = int(time.time() * 1000)

    if DB_VENDOR == "mysql":
        db.execute("""
            INSERT INTO request_types (name, value_type, external_id, abbreviation, description, active, sort_order, created_ts, updated_ts, is_giustificativo)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (name, value_type, external_id, abbreviation, description, 1 if active else 0, sort_order, now_ms, now_ms, 1 if is_giustificativo else 0))
    else:
        db.execute("""
            INSERT INTO request_types (name, value_type, external_id, abbreviation, description, active, sort_order, created_ts, updated_ts, is_giustificativo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (name, value_type, external_id, abbreviation, description, 1 if active else 0, sort_order, now_ms, now_ms, 1 if is_giustificativo else 0))
    
    db.commit()

    return jsonify({"ok": True, "message": f"Tipologia '{name}' creata con successo"})


@app.put("/api/admin/request-types/<int:type_id>")
@login_required
def api_admin_request_types_update(type_id: int) -> ResponseReturnValue:
    """Aggiorna una tipologia di richiesta."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403

    data = request.get_json()
    if not data:
        return jsonify({"error": "Dati non validi"}), 400

    name = (data.get("name") or "").strip()
    value_type = data.get("value_type", "hours")
    external_id = (data.get("external_id") or "").strip() or None
    abbreviation = (data.get("abbreviation") or "").strip() or None
    description = (data.get("description") or "").strip() or None
    active = data.get("active", True)
    sort_order = data.get("sort_order", 0)
    is_giustificativo = data.get("is_giustificativo", False)

    if not name:
        return jsonify({"error": "Il nome è obbligatorio"}), 400

    if value_type not in VALUE_TYPE_LABELS:
        return jsonify({"error": f"Tipo valore non valido"}), 400

    db = get_db()
    ensure_request_types_table(db)
    now_ms = int(time.time() * 1000)

    if DB_VENDOR == "mysql":
        db.execute("""
            UPDATE request_types
            SET name = %s, value_type = %s, external_id = %s, abbreviation = %s, description = %s, 
                active = %s, sort_order = %s, updated_ts = %s, is_giustificativo = %s
            WHERE id = %s
        """, (name, value_type, external_id, abbreviation, description, 1 if active else 0, sort_order, now_ms, 1 if is_giustificativo else 0, type_id))
    else:
        db.execute("""
            UPDATE request_types
            SET name = ?, value_type = ?, external_id = ?, abbreviation = ?, description = ?, 
                active = ?, sort_order = ?, updated_ts = ?, is_giustificativo = ?
            WHERE id = ?
        """, (name, value_type, external_id, abbreviation, description, 1 if active else 0, sort_order, now_ms, 1 if is_giustificativo else 0, type_id))
    
    db.commit()

    return jsonify({"ok": True, "message": f"Tipologia '{name}' aggiornata"})


@app.delete("/api/admin/request-types/<int:type_id>")
@login_required
def api_admin_request_types_delete(type_id: int) -> ResponseReturnValue:
    """Elimina una tipologia di richiesta."""
    if not session.get("is_admin"):
        return jsonify({"error": "forbidden"}), 403

    db = get_db()
    ensure_request_types_table(db)
    ensure_user_requests_table(db)

    # Verifica che non ci siano richieste collegate
    if DB_VENDOR == "mysql":
        count = db.execute("SELECT COUNT(*) as cnt FROM user_requests WHERE request_type_id = %s", (type_id,)).fetchone()
    else:
        count = db.execute("SELECT COUNT(*) as cnt FROM user_requests WHERE request_type_id = ?", (type_id,)).fetchone()
    
    cnt = count["cnt"] if isinstance(count, Mapping) else count[0]
    if cnt > 0:
        return jsonify({"error": f"Impossibile eliminare: ci sono {cnt} richieste collegate a questa tipologia"}), 400

    if DB_VENDOR == "mysql":
        db.execute("DELETE FROM request_types WHERE id = %s", (type_id,))
    else:
        db.execute("DELETE FROM request_types WHERE id = ?", (type_id,))
    
    db.commit()

    return jsonify({"ok": True, "message": "Tipologia eliminata"})


def _send_request_review_notification(
    db: DatabaseLike, 
    username: str, 
    type_name: str, 
    status: str, 
    review_notes: str,
    is_partial: bool = False,
    rounded_start: str = None,
    rounded_end: str = None
) -> None:
    """Invia notifica push all'utente quando la sua richiesta viene revisionata.
    
    Args:
        is_partial: True se l'admin ha modificato gli orari (approvazione parziale)
        rounded_start: Orario inizio arrotondato confermato
        rounded_end: Orario fine arrotondato confermato
    """
    settings = get_webpush_settings()
    if not settings:
        app.logger.info("Notifiche push non configurate, skip notifica revisione richiesta")
        return
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Recupera le subscription push dell'utente
    subscriptions = db.execute(
        f"SELECT endpoint, p256dh, auth FROM push_subscriptions WHERE username = {placeholder}",
        (username,)
    ).fetchall()
    
    if not subscriptions:
        app.logger.info("Nessuna subscription push per utente %s", username)
        return
    
    # Prepara il messaggio
    if status == "approved":
        if is_partial:
            title = "⚠️ Richiesta approvata parzialmente"
            body = f"La tua richiesta di {type_name} è stata approvata con modifiche"
            # Aggiungi gli orari arrotondati
            if rounded_start and rounded_end:
                body += f"\n🕐 Orari confermati: {rounded_start} - {rounded_end}"
            elif rounded_start:
                body += f"\n🕐 Inizio confermato: {rounded_start}"
            elif rounded_end:
                body += f"\n🕐 Fine confermata: {rounded_end}"
        else:
            title = "✅ Richiesta approvata"
            body = f"La tua richiesta di {type_name} è stata approvata"
    else:
        title = "❌ Richiesta respinta"
        body = f"La tua richiesta di {type_name} è stata respinta"
    
    if review_notes:
        body += f"\n📝 Note: {review_notes}"
    
    payload = {
        "title": title,
        "body": body,
        "icon": "/static/icons/icon-192x192.png",
        "badge": "/static/icons/icon-72x72.png",
        "tag": f"request-review-{status}",
        "data": {
            "url": "/user/requests",
            "type": "request_reviewed"
        }
    }
    
    # Invia a tutte le subscription dell'utente
    sent_ok = False
    for sub in subscriptions:
        endpoint = sub['endpoint'] if isinstance(sub, dict) else sub[0]
        p256dh = sub['p256dh'] if isinstance(sub, dict) else sub[1]
        auth = sub['auth'] if isinstance(sub, dict) else sub[2]
        
        subscription_info = {
            "endpoint": endpoint,
            "keys": {
                "p256dh": p256dh,
                "auth": auth
            }
        }
        
        try:
            webpush(
                subscription_info=subscription_info,
                data=json.dumps(payload),
                vapid_private_key=settings["vapid_private"],
                vapid_claims={"sub": settings["subject"]},
                ttl=86400,  # 24 ore
            )
            app.logger.info("Notifica revisione richiesta inviata a %s", username)
            sent_ok = True
                
        except WebPushException as e:
            app.logger.warning("Errore invio notifica revisione a %s: %s", username, e)
            if e.response and e.response.status_code in {404, 410}:
                remove_push_subscription(db, endpoint)
        except Exception as e:
            app.logger.error("Errore generico invio notifica revisione: %s", e)
    
    # Salva la notifica nel log (una volta per utente)
    if sent_ok:
        try:
            record_push_notification(
                db,
                kind="request_reviewed",
                title=title,
                body=body,
                payload=payload,
                username=username,
            )
            app.logger.info("Notifica revisione salvata nel log per %s", username)
        except Exception as e:
            app.logger.error("Errore salvataggio notifica revisione nel log: %s", e)


# =====================================================
# ADMIN DOCUMENTS - Gestione documenti aziendali
# =====================================================

@app.get("/admin/documents")
@login_required
def admin_documents_page() -> ResponseReturnValue:
    """Pagina admin per gestire i documenti aziendali."""
    if not session.get("is_admin"):
        return ("Forbidden", 403)
    
    username = session.get("username", "Admin")
    initials = "".join([p[0].upper() for p in username.split()[:2]]) if username else "?"
    
    return render_template(
        "admin_documents.html",
        is_admin=True,
        user_name=username,
        user_initials=initials
    )


@app.get("/api/admin/documents")
@login_required
def api_admin_documents_list() -> ResponseReturnValue:
    """Lista tutti i documenti caricati."""
    if not session.get("is_admin"):
        return jsonify({"error": "Accesso negato"}), 403
    
    db = get_db()
    ensure_user_documents_table(db)
    
    rows = db.execute("""
        SELECT id, category, title, description, file_path, file_name, 
               target_users, target_all, created_by, created_at, notified_at
        FROM user_documents
        ORDER BY created_at DESC
    """).fetchall()
    
    # Carica la lista utenti per mostrare i nomi completi
    users_dict = {}
    try:
        users_rows = db.execute("SELECT username, display_name FROM users").fetchall()
        for u in users_rows:
            if isinstance(u, Mapping):
                users_dict[u["username"]] = u["display_name"] or u["username"]
            else:
                users_dict[u[0]] = u[1] or u[0]
    except:
        pass
    
    documents = []
    for row in rows:
        if isinstance(row, Mapping):
            target_users = row["target_users"]
            file_name = row["file_name"]
        else:
            target_users = row[6]
            file_name = row[5]
            
        if target_users and isinstance(target_users, str):
            try:
                target_users = json.loads(target_users)
            except:
                target_users = []
        
        # Costruisci lista destinatari con nomi completi
        target_users_display = []
        if target_users:
            for username in target_users:
                display_name = users_dict.get(username, username)
                target_users_display.append({"username": username, "display_name": display_name})
        
        # Costruisci file_url usando file_name (come API user)
        file_url = None
        if file_name:
            file_url = f"/uploads/documents/{file_name}"
        
        if isinstance(row, Mapping):
            documents.append({
                "id": row["id"],
                "category": row["category"],
                "title": row["title"],
                "description": row["description"],
                "file_name": row["file_name"],
                "file_url": file_url,
                "target_users": target_users,
                "target_users_display": target_users_display,
                "target_all": bool(row["target_all"]),
                "created_by": row["created_by"],
                "created_at": row["created_at"],
                "notified_at": row["notified_at"]
            })
        else:
            documents.append({
                "id": row[0],
                "category": row[1],
                "title": row[2],
                "description": row[3],
                "file_name": row[5],
                "file_url": file_url,
                "target_users": target_users,
                "target_users_display": target_users_display,
                "target_all": bool(row[7]),
                "created_by": row[8],
                "created_at": row[9],
                "notified_at": row[10]
            })
    
    return jsonify({"documents": documents})


@app.post("/api/admin/documents")
@login_required
def api_admin_documents_create() -> ResponseReturnValue:
    """Carica un nuovo documento."""
    if not session.get("is_admin"):
        return jsonify({"error": "Accesso negato"}), 403
    
    category = request.form.get("category")
    title = request.form.get("title")
    description = request.form.get("description", "")
    target_all = request.form.get("target_all", "1") == "1"
    target_users_json = request.form.get("target_users", "[]")
    
    if not category or category not in ("circolare", "comunicazione", "busta_paga"):
        return jsonify({"error": "Categoria non valida"}), 400
    
    if not title:
        return jsonify({"error": "Titolo obbligatorio"}), 400
    
    # Gestione file allegato
    file_path = None
    file_name = None
    
    if "file" in request.files:
        file = request.files["file"]
        if file and file.filename:
            # Genera nome file univoco
            import uuid
            ext = os.path.splitext(file.filename)[1]
            file_name = f"{uuid.uuid4().hex}{ext}"
            file_path = os.path.join("uploads", "documents", file_name)
            
            # Salva il file
            full_path = os.path.join(os.path.dirname(__file__), file_path)
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
            file.save(full_path)
    
    db = get_db()
    ensure_user_documents_table(db)
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    created_by = session.get("user", "admin")
    
    db.execute(f"""
        INSERT INTO user_documents (category, title, description, file_path, file_name, 
                                    target_users, target_all, created_by)
        VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, 
                {placeholder}, {placeholder}, {placeholder})
    """, (category, title, description, file_path, file_name, 
          target_users_json if not target_all else None, 
          1 if target_all else 0, created_by))
    
    db.commit()
    
    # Le notifiche NON vengono inviate automaticamente
    # L'admin deve inviarle manualmente dal tab "Da Inviare"
    
    return jsonify({"success": True, "message": "Documento caricato. Vai su 'Da Inviare' per inviare le notifiche."})


@app.put("/api/admin/documents/<int:doc_id>")
@login_required
def api_admin_documents_update(doc_id: int) -> ResponseReturnValue:
    """Aggiorna titolo e descrizione di un documento."""
    if not session.get("is_admin"):
        return jsonify({"error": "Accesso negato"}), 403
    
    db = get_db()
    ensure_user_documents_table(db)
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Recupera il documento esistente
    row = db.execute(f"SELECT id FROM user_documents WHERE id = {placeholder}", (doc_id,)).fetchone()
    if not row:
        return jsonify({"error": "Documento non trovato"}), 404
    
    data = request.get_json() or {}
    title = data.get("title", "").strip()
    description = data.get("description", "").strip()
    
    if not title:
        return jsonify({"error": "Il titolo è obbligatorio"}), 400
    
    # Aggiorna il documento
    db.execute(
        f"UPDATE user_documents SET title = {placeholder}, description = {placeholder} WHERE id = {placeholder}",
        (title, description, doc_id)
    )
    db.commit()
    
    return jsonify({"success": True, "message": "Documento aggiornato"})


@app.delete("/api/admin/documents/<int:doc_id>")
@login_required
def api_admin_documents_delete(doc_id: int) -> ResponseReturnValue:
    """Elimina un documento."""
    if not session.get("is_admin"):
        return jsonify({"error": "Accesso negato"}), 403
    
    db = get_db()
    ensure_user_documents_table(db)
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Recupera il file path per eliminarlo
    row = db.execute(f"SELECT file_path FROM user_documents WHERE id = {placeholder}", (doc_id,)).fetchone()
    if row:
        file_path = row["file_path"] if isinstance(row, Mapping) else row[0]
        if file_path:
            full_path = os.path.join(os.path.dirname(__file__), file_path)
            try:
                os.remove(full_path)
            except:
                pass
    
    # Elimina dal database
    db.execute(f"DELETE FROM user_documents_read WHERE document_id = {placeholder}", (doc_id,))
    db.execute(f"DELETE FROM user_documents WHERE id = {placeholder}", (doc_id,))
    db.commit()
    
    return jsonify({"success": True, "message": "Documento eliminato"})


@app.post("/api/admin/documents/<int:doc_id>/notify")
@login_required
def api_admin_documents_notify(doc_id: int) -> ResponseReturnValue:
    """Reinvia la notifica per un documento."""
    if not session.get("is_admin"):
        return jsonify({"error": "Accesso negato"}), 403
    
    db = get_db()
    ensure_user_documents_table(db)
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Recupera il documento
    row = db.execute(
        f"SELECT category, title, target_all, target_users FROM user_documents WHERE id = {placeholder}",
        (doc_id,)
    ).fetchone()
    
    if not row:
        return jsonify({"error": "Documento non trovato"}), 404
    
    if isinstance(row, Mapping):
        category = row["category"]
        title = row["title"]
        target_all = row["target_all"]
        target_users_json = row["target_users"] or "[]"
    else:
        category, title, target_all, target_users_json = row[0], row[1], row[2], row[3] or "[]"
    
    # Invia le notifiche (passa doc_id per URL diretto)
    count = _send_document_notifications(db, category, title, bool(target_all), target_users_json, doc_id)
    
    # Aggiorna notified_at
    if count > 0:
        db.execute(
            f"UPDATE user_documents SET notified_at = {placeholder} WHERE id = {placeholder}",
            (now_ms(), doc_id)
        )
        db.commit()
    
    return jsonify({
        "success": True, 
        "message": f"Notifica inviata a {count} dispositivi",
        "count": count
    })


@app.get("/api/admin/documents/<int:doc_id>/recipients")
@login_required
def api_admin_documents_recipients(doc_id: int) -> ResponseReturnValue:
    """Restituisce la lista dei destinatari di un documento con stato di lettura."""
    if not session.get("is_admin"):
        return jsonify({"error": "Accesso negato"}), 403
    
    db = get_db()
    ensure_user_documents_table(db)
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Recupera il documento
    row = db.execute(
        f"SELECT target_all, target_users FROM user_documents WHERE id = {placeholder}",
        (doc_id,)
    ).fetchone()
    
    if not row:
        return jsonify({"error": "Documento non trovato"}), 404
    
    if isinstance(row, Mapping):
        target_all = row["target_all"]
        target_users_json = row["target_users"] or "[]"
    else:
        target_all, target_users_json = row[0], row[1] or "[]"
    
    # Determina i destinatari
    if target_all:
        # Tutti gli operatori (role = 'user')
        users_rows = db.execute(
            f"SELECT username, display_name FROM app_users WHERE role = 'user' AND is_active = 1"
        ).fetchall()
        target_usernames = []
        user_display_map = {}
        for u in users_rows:
            if isinstance(u, Mapping):
                target_usernames.append(u["username"])
                user_display_map[u["username"]] = u["display_name"]
            else:
                target_usernames.append(u[0])
                user_display_map[u[0]] = u[1]
    else:
        # Destinatari specifici
        try:
            target_usernames = json.loads(target_users_json)
        except json.JSONDecodeError:
            target_usernames = []
        
        # Recupera display_name per ciascuno
        user_display_map = {}
        if target_usernames:
            placeholders = ",".join([placeholder] * len(target_usernames))
            users_rows = db.execute(
                f"SELECT username, display_name FROM app_users WHERE username IN ({placeholders})",
                tuple(target_usernames)
            ).fetchall()
            for u in users_rows:
                if isinstance(u, Mapping):
                    user_display_map[u["username"]] = u["display_name"]
                else:
                    user_display_map[u[0]] = u[1]
    
    # Recupera chi ha letto il documento
    read_rows = db.execute(
        f"SELECT username, read_at FROM user_documents_read WHERE document_id = {placeholder}",
        (doc_id,)
    ).fetchall()
    
    read_map = {}
    for r in read_rows:
        if isinstance(r, Mapping):
            read_map[r["username"]] = r["read_at"]
        else:
            read_map[r[0]] = r[1]
    
    # Costruisci la lista destinatari con stato lettura
    recipients = []
    for username in target_usernames:
        display_name = user_display_map.get(username, username)
        read_at = read_map.get(username)
        recipients.append({
            "username": username,
            "display_name": display_name,
            "read": read_at is not None,
            "read_at": str(read_at) if read_at else None
        })
    
    # Ordina: prima non letti, poi letti
    recipients.sort(key=lambda x: (x["read"], x["display_name"].lower()))
    
    return jsonify({
        "target_all": bool(target_all),
        "recipients": recipients,
        "total": len(recipients),
        "read_count": len(read_map),
        "unread_count": len(recipients) - len(read_map)
    })


# =====================================================
# ADMIN EMPLOYEE SHIFTS - Turni settimanali impiegati
# =====================================================

@app.get("/admin/employee-shifts")
@login_required
def admin_employee_shifts_page() -> ResponseReturnValue:
    """Pagina admin per gestire i turni settimanali degli impiegati."""
    if not session.get("is_admin"):
        return ("Forbidden", 403)
    
    username = session.get("username", "Admin")
    initials = "".join([p[0].upper() for p in username.split()[:2]]) if username else "?"
    
    return render_template(
        "admin_employee_shifts.html",
        is_admin=True,
        user_name=username,
        user_initials=initials
    )


@app.get("/api/admin/employee-shifts/users")
@login_required
def api_admin_employee_shifts_users() -> ResponseReturnValue:
    """Restituisce gli utenti che NON hanno rentman_crew_id (impiegati non-Rentman)."""
    if not session.get("is_admin"):
        return jsonify({"error": "Forbidden"}), 403
    
    db = get_db()
    ensure_employee_shifts_table(db)
    
    # Utenti senza associazione Rentman
    rows = db.execute("""
        SELECT username, display_name, full_name, role
        FROM app_users
        WHERE is_active = 1 AND (rentman_crew_id IS NULL OR rentman_crew_id = 0)
        ORDER BY display_name ASC
    """).fetchall()
    
    users = []
    for row in rows:
        if isinstance(row, dict):
            users.append({
                "username": row["username"],
                "display_name": row["display_name"],
                "full_name": row.get("full_name") or row["display_name"],
                "role": row["role"]
            })
        else:
            users.append({
                "username": row[0],
                "display_name": row[1],
                "full_name": row[2] or row[1],
                "role": row[3]
            })
    
    return jsonify({"users": users})


@app.get("/api/admin/employee-shifts")
@login_required
def api_admin_employee_shifts_list() -> ResponseReturnValue:
    """Lista tutti i turni configurati, raggruppati per utente."""
    if not session.get("is_admin"):
        return jsonify({"error": "Forbidden"}), 403
    
    db = get_db()
    ensure_employee_shifts_table(db)
    
    rows = db.execute("""
        SELECT es.id, es.username, es.day_of_week, es.start_time, es.end_time,
               es.break_start, es.break_end, es.shift_name, es.location_name, es.is_active,
               au.display_name, au.full_name
        FROM employee_shifts es
        LEFT JOIN app_users au ON es.username = au.username
        ORDER BY au.display_name ASC, es.day_of_week ASC
    """).fetchall()
    
    shifts_by_user = {}
    for row in rows:
        if isinstance(row, dict):
            username = row["username"]
            shift = {
                "id": row["id"],
                "day_of_week": row["day_of_week"],
                "start_time": format_time_value(row["start_time"]),
                "end_time": format_time_value(row["end_time"]),
                "break_start": format_time_value(row["break_start"]),
                "break_end": format_time_value(row["break_end"]),
                "shift_name": row.get("shift_name"),
                "location_name": row.get("location_name"),
                "is_active": bool(row["is_active"])
            }
            display_name = row.get("display_name") or username
            full_name = row.get("full_name") or display_name
        else:
            username = row[1]
            shift = {
                "id": row[0],
                "day_of_week": row[2],
                "start_time": format_time_value(row[3]),
                "end_time": format_time_value(row[4]),
                "break_start": format_time_value(row[5]),
                "break_end": format_time_value(row[6]),
                "shift_name": row[7] if len(row) > 7 else None,
                "location_name": row[8] if len(row) > 8 else None,
                "is_active": bool(row[9]) if len(row) > 9 else True
            }
            display_name = row[10] if len(row) > 10 else username
            full_name = row[11] if len(row) > 11 else display_name
        
        if username not in shifts_by_user:
            shifts_by_user[username] = {
                "username": username,
                "display_name": display_name,
                "full_name": full_name,
                "shifts": []
            }
        shifts_by_user[username]["shifts"].append(shift)
    
    return jsonify({"users": list(shifts_by_user.values())})


@app.get("/api/admin/employee-shifts/<username>")
@login_required
def api_admin_employee_shifts_get(username: str) -> ResponseReturnValue:
    """Restituisce i turni di un utente specifico."""
    if not session.get("is_admin"):
        return jsonify({"error": "Forbidden"}), 403
    
    db = get_db()
    ensure_employee_shifts_table(db)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    rows = db.execute(f"""
        SELECT id, day_of_week, start_time, end_time, break_start, break_end, shift_name, location_name, is_active
        FROM employee_shifts
        WHERE username = {placeholder}
        ORDER BY day_of_week ASC
    """, (username,)).fetchall()
    
    shifts = []
    for row in rows:
        if isinstance(row, dict):
            shifts.append({
                "id": row["id"],
                "day_of_week": row["day_of_week"],
                "start_time": format_time_value(row["start_time"]),
                "end_time": format_time_value(row["end_time"]),
                "break_start": format_time_value(row["break_start"]),
                "break_end": format_time_value(row["break_end"]),
                "shift_name": row.get("shift_name"),
                "location_name": row.get("location_name"),
                "is_active": bool(row["is_active"])
            })
        else:
            shifts.append({
                "id": row[0],
                "day_of_week": row[1],
                "start_time": format_time_value(row[2]),
                "end_time": format_time_value(row[3]),
                "break_start": format_time_value(row[4]),
                "break_end": format_time_value(row[5]),
                "shift_name": row[6] if len(row) > 6 else None,
                "location_name": row[7] if len(row) > 7 else None,
                "is_active": bool(row[8]) if len(row) > 8 else True
            })
    
    return jsonify({"username": username, "shifts": shifts})


@app.post("/api/admin/employee-shifts/<username>")
@login_required
def api_admin_employee_shifts_save(username: str) -> ResponseReturnValue:
    """Salva i turni settimanali di un utente (sovrascrive tutti)."""
    if not session.get("is_admin"):
        return jsonify({"error": "Forbidden"}), 403
    
    data = request.get_json()
    if not data or "shifts" not in data:
        return jsonify({"error": "Dati mancanti"}), 400
    
    shifts = data["shifts"]  # Array di {day_of_week, start_time, end_time, break_start, break_end, shift_name, location_name, is_active}
    
    db = get_db()
    ensure_employee_shifts_table(db)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Elimina turni esistenti per l'utente
    db.execute(f"DELETE FROM employee_shifts WHERE username = {placeholder}", (username,))
    
    # Inserisce i nuovi turni
    for shift in shifts:
        day = shift.get("day_of_week")
        if day is None:
            continue
        
        start_time = shift.get("start_time") or None
        end_time = shift.get("end_time") or None
        break_start = shift.get("break_start") or None
        break_end = shift.get("break_end") or None
        shift_name = shift.get("shift_name") or None
        location_name = shift.get("location_name") or None
        is_active = 1 if shift.get("is_active", True) else 0
        
        # Salta se orari non validi
        if not start_time or not end_time:
            continue
        
        db.execute(f"""
            INSERT INTO employee_shifts (username, day_of_week, start_time, end_time, break_start, break_end, shift_name, location_name, is_active)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
        """, (username, day, start_time, end_time, break_start, break_end, shift_name, location_name, is_active))
    
    db.commit()
    
    return jsonify({"success": True, "message": "Turni salvati con successo"})


@app.delete("/api/admin/employee-shifts/<username>")
@login_required
def api_admin_employee_shifts_delete(username: str) -> ResponseReturnValue:
    """Elimina tutti i turni di un utente."""
    if not session.get("is_admin"):
        return jsonify({"error": "Forbidden"}), 403
    
    db = get_db()
    ensure_employee_shifts_table(db)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    db.execute(f"DELETE FROM employee_shifts WHERE username = {placeholder}", (username,))
    db.commit()
    
    return jsonify({"success": True, "message": "Turni eliminati"})


@app.post("/api/admin/employee-shifts/bulk")
@login_required
def api_admin_employee_shifts_bulk() -> ResponseReturnValue:
    """Salva gli stessi turni per più utenti contemporaneamente."""
    if not session.get("is_admin"):
        return jsonify({"error": "Forbidden"}), 403
    
    data = request.get_json()
    if not data:
        return jsonify({"error": "Dati mancanti"}), 400
    
    usernames = data.get("usernames", [])
    shifts = data.get("shifts", [])
    
    if not usernames:
        return jsonify({"error": "Nessun utente selezionato"}), 400
    
    if not shifts:
        return jsonify({"error": "Nessun turno configurato"}), 400
    
    db = get_db()
    ensure_employee_shifts_table(db)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    saved_count = 0
    
    for username in usernames:
        # Elimina turni esistenti per l'utente
        db.execute(f"DELETE FROM employee_shifts WHERE username = {placeholder}", (username,))
        
        # Inserisce i nuovi turni
        for shift in shifts:
            day = shift.get("day_of_week")
            if day is None:
                continue
            
            start_time = shift.get("start_time") or None
            end_time = shift.get("end_time") or None
            break_start = shift.get("break_start") or None
            break_end = shift.get("break_end") or None
            shift_name = shift.get("shift_name") or None
            location_name = shift.get("location_name") or None
            is_active = 1 if shift.get("is_active", True) else 0
            
            # Salta se orari non validi
            if not start_time or not end_time:
                continue
            
            db.execute(f"""
                INSERT INTO employee_shifts (username, day_of_week, start_time, end_time, break_start, break_end, shift_name, location_name, is_active)
                VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
            """, (username, day, start_time, end_time, break_start, break_end, shift_name, location_name, is_active))
        
        saved_count += 1
    
    db.commit()
    
    return jsonify({"success": True, "message": f"Turni salvati per {saved_count} utenti"})


# =====================================================
# ADMIN USER REQUESTS - Gestione richieste utenti
# =====================================================

@app.route("/admin/user-requests")
@login_required
def admin_user_requests_page() -> ResponseReturnValue:
    """Pagina admin per gestire le richieste degli utenti."""
    if not session.get("is_admin"):
        return ("Forbidden", 403)
    
    username = session.get("username", "Admin")
    initials = "".join([p[0].upper() for p in username.split()[:2]]) if username else "?"
    
    return render_template(
        "admin_user_requests.html",
        is_admin=True,
        user_name=username,
        user_initials=initials
    )


@app.get("/api/admin/user-requests/pending-count")
@login_required
def api_admin_pending_requests_count() -> ResponseReturnValue:
    """Restituisce il conteggio delle richieste in attesa."""
    if not session.get("is_admin"):
        return jsonify({"error": "Accesso negato"}), 403
    
    db = get_db()
    ensure_user_requests_table(db)
    
    row = db.execute("SELECT COUNT(*) as cnt FROM user_requests WHERE status = 'pending'").fetchone()
    count = row["cnt"] if isinstance(row, Mapping) else row[0]
    
    return jsonify({"count": count})


@app.get("/api/admin/user-requests")
@login_required
def api_admin_user_requests_list() -> ResponseReturnValue:
    """Restituisce tutte le richieste degli utenti per l'admin."""
    if not session.get("is_admin"):
        return jsonify({"error": "Accesso negato"}), 403
    
    db = get_db()
    ensure_user_requests_table(db)
    
    if DB_VENDOR == "mysql":
        rows = db.execute("""
            SELECT ur.id, ur.username, ur.request_type_id, rt.name as type_name, rt.value_type,
                   ur.date_from, ur.date_to, ur.value_amount, ur.notes, ur.status,
                   ur.reviewed_by, ur.reviewed_ts, ur.review_notes, ur.created_ts, ur.updated_ts,
                   ur.cdc, ur.attachment_path, ur.tratte, ur.extra_data
            FROM user_requests ur
            JOIN request_types rt ON ur.request_type_id = rt.id
            ORDER BY 
                CASE ur.status WHEN 'pending' THEN 0 ELSE 1 END,
                ur.created_ts DESC
        """).fetchall()
    else:
        rows = db.execute("""
            SELECT ur.id, ur.username, ur.request_type_id, rt.name as type_name, rt.value_type,
                   ur.date_from, ur.date_to, ur.value_amount, ur.notes, ur.status,
                   ur.reviewed_by, ur.reviewed_ts, ur.review_notes, ur.created_ts, ur.updated_ts,
                   ur.cdc, ur.attachment_path, ur.tratte, ur.extra_data
            FROM user_requests ur
            JOIN request_types rt ON ur.request_type_id = rt.id
            ORDER BY 
                CASE ur.status WHEN 'pending' THEN 0 ELSE 1 END,
                ur.created_ts DESC
        """).fetchall()

    requests = []
    for row in rows:
        # Parsing tratte JSON
        tratte_data = None
        if isinstance(row, Mapping):
            tratte_raw = row.get("tratte")
        else:
            tratte_raw = row[17] if len(row) > 17 else None
        
        if tratte_raw:
            try:
                if isinstance(tratte_raw, str):
                    tratte_data = json.loads(tratte_raw)
                else:
                    tratte_data = tratte_raw
            except:
                tratte_data = None
        
        # Parsing extra_data JSON (per straordinari e altri dati)
        extra_data = None
        if isinstance(row, Mapping):
            extra_raw = row.get("extra_data")
        else:
            extra_raw = row[18] if len(row) > 18 else None
        
        if extra_raw:
            try:
                if isinstance(extra_raw, str):
                    extra_data = json.loads(extra_raw)
                else:
                    extra_data = extra_raw
            except:
                extra_data = None
        
        if isinstance(row, Mapping):
            req_item = {
                "id": row["id"],
                "username": row["username"],
                "request_type_id": row["request_type_id"],
                "type_name": row["type_name"],
                "value_type": row["value_type"],
                "date_from": row["date_from"],
                "date_to": row["date_to"],
                "value": float(row["value_amount"]) if row["value_amount"] else None,
                "notes": row["notes"],
                "status": row["status"],
                "reviewed_by": row["reviewed_by"],
                "reviewed_ts": row["reviewed_ts"],
                "review_notes": row["review_notes"],
                "created_ts": row["created_ts"],
                "updated_ts": row["updated_ts"],
                "cdc": row["cdc"],
                "attachment_path": row["attachment_path"],
                "tratte": tratte_data,
                "extra_data": extra_data,
            }
        else:
            req_item = {
                "id": row[0],
                "username": row[1],
                "request_type_id": row[2],
                "type_name": row[3],
                "value_type": row[4],
                "date_from": row[5],
                "date_to": row[6],
                "value": float(row[7]) if row[7] else None,
                "notes": row[8],
                "status": row[9],
                "reviewed_by": row[10],
                "reviewed_ts": row[11],
                "review_notes": row[12],
                "created_ts": row[13],
                "updated_ts": row[14],
                "cdc": row[15] if len(row) > 15 else None,
                "attachment_path": row[16] if len(row) > 16 else None,
                "tratte": tratte_data,
                "extra_data": extra_data,
            }
        
        requests.append(req_item)
    
    # Per le richieste di tipo "timbratura", aggiungi i dati del turno previsto
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    for req in requests:
        if req.get("value_type") == "timbratura":
            req_username = req.get("username")
            date_from = req.get("date_from")
            app.logger.info(f"Cerco turno per {req_username} - date_from={date_from}")
            if req_username and date_from:
                # Converti date_from in data e giorno della settimana
                try:
                    if hasattr(date_from, 'weekday'):
                        check_date = date_from
                    else:
                        check_date = datetime.strptime(str(date_from)[:10], "%Y-%m-%d")
                    day_of_week = check_date.weekday()
                    app.logger.info(f"day_of_week={day_of_week} per data {date_from}")
                    
                    # Cerca il turno da employee_shifts
                    shift_row = db.execute(f"""
                        SELECT start_time, end_time, break_start, break_end, location_name
                        FROM employee_shifts
                        WHERE username = {placeholder} AND day_of_week = {placeholder} AND is_active = 1
                    """, (req_username, day_of_week)).fetchone()
                    
                    app.logger.info(f"shift_row trovato: {shift_row}")
                    if shift_row:
                        if isinstance(shift_row, Mapping):
                            turno_info = {
                                "start_time": str(shift_row['start_time'])[:5] if shift_row.get('start_time') else None,
                                "end_time": str(shift_row['end_time'])[:5] if shift_row.get('end_time') else None,
                                "break_start": str(shift_row['break_start'])[:5] if shift_row.get('break_start') else None,
                                "break_end": str(shift_row['break_end'])[:5] if shift_row.get('break_end') else None,
                                "location_name": shift_row.get('location_name')
                            }
                        else:
                            turno_info = {
                                "start_time": str(shift_row[0])[:5] if shift_row[0] else None,
                                "end_time": str(shift_row[1])[:5] if shift_row[1] else None,
                                "break_start": str(shift_row[2])[:5] if shift_row[2] else None,
                                "break_end": str(shift_row[3])[:5] if shift_row[3] else None,
                                "location_name": shift_row[4] if len(shift_row) > 4 else None
                            }
                        req["turno_previsto"] = turno_info
                except Exception as e:
                    app.logger.warning(f"Errore recupero turno per {req_username}: {e}")

    return jsonify({"requests": requests})


def _process_approved_mancata_timbratura(
    db: DatabaseLike,
    username: str,
    date_from: str,
    extra_data_str: str
) -> dict:
    """
    Processa una richiesta di 'Mancata Timbratura' approvata.
    
    1. Inserisce la timbratura nella tabella 'timbrature'
    2. Se CedolinoWeb è attivo, inserisce in 'cedolino_timbrature' e invia al webservice
    
    Args:
        db: connessione database
        username: username dell'utente
        date_from: data della timbratura (YYYY-MM-DD)
        extra_data_str: JSON con tipo_timbratura, ora_timbratura, motivazione
    
    Returns:
        dict con risultato operazione
    """
    result = {
        "inserted_timbratura": False,
        "cedolino_sent": False,
        "cedolino_error": None
    }
    
    # Parse extra_data
    try:
        extra_data = json.loads(extra_data_str) if extra_data_str else {}
    except:
        extra_data = {}
    
    tipo_timbratura = extra_data.get("tipo_timbratura")  # ingresso/uscita/pausa_in/pausa_out
    ora_timbratura = extra_data.get("ora_timbratura")    # HH:MM
    motivazione = extra_data.get("motivazione", "")
    
    if not tipo_timbratura or not ora_timbratura:
        result["error"] = "Dati timbratura mancanti"
        return result
    
    # Formatta data come stringa YYYY-MM-DD
    if hasattr(date_from, 'strftime'):
        date_str = date_from.strftime("%Y-%m-%d")
    else:
        date_str = str(date_from)[:10]
    
    # Mappa tipo timbratura utente -> tipo interno
    TIPO_MAP = {
        "ingresso": "inizio_giornata",
        "uscita": "fine_giornata",
        "pausa_in": "inizio_pausa",
        "pausa_out": "fine_pausa"
    }
    tipo_interno = TIPO_MAP.get(tipo_timbratura, tipo_timbratura)
    
    # Mappa tipo interno -> timeframe CedolinoWeb
    TIPO_TO_TIMEFRAME = {
        'inizio_giornata': TIMEFRAME_INIZIO_GIORNATA,  # 1
        'inizio_pausa': TIMEFRAME_INIZIO_PAUSA,        # 4
        'fine_pausa': TIMEFRAME_FINE_PAUSA,            # 5
        'fine_giornata': TIMEFRAME_FINE_GIORNATA,      # 8
    }
    timeframe_id = TIPO_TO_TIMEFRAME.get(tipo_interno)
    
    # Formatta ora con secondi
    ora_full = f"{ora_timbratura}:00" if len(ora_timbratura) == 5 else ora_timbratura
    
    now_ts = now_ms()
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # 1. Inserisci nella tabella 'timbrature'
    try:
        db.execute(f"""
            INSERT INTO timbrature (username, tipo, data, ora, ora_mod, created_ts, method, gps_lat, gps_lon, location_name)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
        """, (username, tipo_interno, date_str, ora_full, ora_full, now_ts, "manual_request", None, None, "Mancata Timbratura"))
        result["inserted_timbratura"] = True
        app.logger.info(f"Mancata Timbratura: inserita timbratura {tipo_interno} per {username} alle {ora_full} del {date_str}")
    except Exception as e:
        app.logger.error(f"Mancata Timbratura: errore inserimento timbrature - {e}")
        result["error"] = f"Errore inserimento timbratura: {e}"
        return result
    
    # 2. Se CedolinoWeb è attivo, invia la timbrata
    settings = get_cedolino_settings()
    if settings and timeframe_id:
        try:
            # Recupera display_name per il log
            user_row = db.execute(
                f"SELECT display_name FROM app_users WHERE username = {placeholder}",
                (username,)
            ).fetchone()
            display_name = username
            if user_row:
                display_name = (user_row['display_name'] if isinstance(user_row, dict) else user_row[0]) or username
            
            # Usa la funzione esistente per inviare a CedolinoWeb
            success, external_id, error, request_url = send_timbrata_utente(
                db=db,
                username=username,
                member_name=display_name,
                timeframe_id=timeframe_id,
                data_riferimento=date_str,
                ora_originale=ora_full,
                ora_modificata=ora_full,
                overtime_request_id=None  # Non bloccare, è già approvata
            )
            
            if success:
                result["cedolino_sent"] = True
                result["cedolino_external_id"] = external_id
                app.logger.info(f"Mancata Timbratura: inviata a CedolinoWeb per {username} (external_id={external_id})")
            else:
                result["cedolino_error"] = error
                app.logger.warning(f"Mancata Timbratura: errore CedolinoWeb per {username} - {error}")
        except Exception as e:
            result["cedolino_error"] = str(e)
            app.logger.error(f"Mancata Timbratura: eccezione CedolinoWeb - {e}")
    else:
        result["cedolino_sent"] = None  # CedolinoWeb non configurato
    
    db.commit()
    return result


def _update_timbrature_with_confirmed_times(
    db: DatabaseLike, 
    username: str, 
    date_from: str, 
    rounded_start: str, 
    rounded_end: str,
    extra_data_str: str
) -> None:
    """
    Aggiorna le timbrature con gli orari arrotondati confermati dall'admin.
    Chiamata quando uno straordinario viene approvato con orari modificati.
    Aggiorna sia la tabella 'timbrature' (per la visualizzazione utente)
    che 'cedolino_timbrature' (per l'export al gestionale).
    """
    if not date_from:
        return
    
    # Formatta data_from come stringa YYYY-MM-DD se necessario
    if hasattr(date_from, 'strftime'):
        date_str = date_from.strftime("%Y-%m-%d")
    else:
        date_str = str(date_from)[:10]
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Recupera session_id da extra_data se disponibile
    session_id = None
    if extra_data_str:
        try:
            extra_data = json.loads(extra_data_str) if isinstance(extra_data_str, str) else extra_data_str
            session_id = extra_data.get("session_id")
        except:
            pass
    
    app.logger.info(
        f"Aggiornamento timbrature per {username} del {date_str} con orari confermati: "
        f"inizio={rounded_start}, fine={rounded_end}, session_id={session_id}"
    )
    
    # ===== AGGIORNA TABELLA 'timbrature' (per visualizzazione utente) =====
    # Aggiorna la timbrata di INIZIO giornata
    if rounded_start:
        result = db.execute(f"""
            UPDATE timbrature 
            SET ora_mod = {placeholder}
            WHERE username = {placeholder} 
              AND data = {placeholder}
              AND tipo = 'inizio_giornata'
        """, (rounded_start, username, date_str))
        app.logger.info(f"Aggiornato timbrature.ora_mod inizio_giornata: {rounded_start}")
    
    # Aggiorna la timbrata di FINE giornata
    if rounded_end:
        result = db.execute(f"""
            UPDATE timbrature 
            SET ora_mod = {placeholder}
            WHERE username = {placeholder} 
              AND data = {placeholder}
              AND tipo = 'fine_giornata'
        """, (rounded_end, username, date_str))
        app.logger.info(f"Aggiornato timbrature.ora_mod fine_giornata: {rounded_end}")
    
    # ===== AGGIORNA TABELLA 'cedolino_timbrature' (per export gestionale) =====
    # Aggiorna la timbrata di INIZIO giornata (timeframe_id = 1)
    if rounded_start:
        if session_id:
            db.execute(f"""
                UPDATE cedolino_timbrature 
                SET ora_modificata = {placeholder}, synced_ts = NULL
                WHERE username = {placeholder} 
                  AND data_riferimento = {placeholder}
                  AND timeframe_id = 1
                  AND session_id = {placeholder}
            """, (rounded_start, username, date_str, session_id))
        else:
            db.execute(f"""
                UPDATE cedolino_timbrature 
                SET ora_modificata = {placeholder}, synced_ts = NULL
                WHERE username = {placeholder} 
                  AND data_riferimento = {placeholder}
                  AND timeframe_id = 1
            """, (rounded_start, username, date_str))
    
    # Aggiorna la timbrata di FINE giornata (timeframe_id = 8)
    if rounded_end:
        if session_id:
            db.execute(f"""
                UPDATE cedolino_timbrature 
                SET ora_modificata = {placeholder}, synced_ts = NULL
                WHERE username = {placeholder} 
                  AND data_riferimento = {placeholder}
                  AND timeframe_id = 8
                  AND session_id = {placeholder}
            """, (rounded_end, username, date_str, session_id))
        else:
            db.execute(f"""
                UPDATE cedolino_timbrature 
                SET ora_modificata = {placeholder}, synced_ts = NULL
                WHERE username = {placeholder} 
                  AND data_riferimento = {placeholder}
                  AND timeframe_id = 8
            """, (rounded_end, username, date_str))
    
    db.commit()
    
    app.logger.info(f"Timbrature aggiornate con successo per {username} del {date_str}")


@app.put("/api/admin/user-requests/<int:request_id>")
@login_required
def api_admin_user_request_review(request_id: int) -> ResponseReturnValue:
    """Approva o respinge una richiesta utente."""
    if not session.get("is_admin"):
        return jsonify({"error": "Accesso negato"}), 403
    
    data = request.get_json() or {}
    status = data.get("status")
    review_notes = data.get("review_notes", "").strip()
    confirmed_value = data.get("confirmed_value")  # Valore confermato (per straordinari)
    rounded_start = data.get("rounded_start")  # Orario arrotondato inizio confermato
    rounded_end = data.get("rounded_end")  # Orario arrotondato fine confermato
    flex_type = data.get("flex_type")  # Per distinguere "fuori_flessibilita"
    rounded_time = data.get("rounded_time")  # Orario singolo per fuori flessibilità
    
    app.logger.info(
        f"Review request: status={status}, confirmed_value={confirmed_value}, "
        f"rounded_start={rounded_start}, rounded_end={rounded_end}, flex_type={flex_type}, rounded_time={rounded_time}"
    )
    
    if status not in ("approved", "rejected"):
        return jsonify({"error": "Stato non valido. Usa 'approved' o 'rejected'"}), 400
    
    db = get_db()
    ensure_user_requests_table(db)
    
    # Verifica che la richiesta esista e sia pending
    if DB_VENDOR == "mysql":
        existing = db.execute("""
            SELECT ur.id, ur.status, ur.username, rt.name as type_name, rt.value_type,
                   ur.extra_data, ur.date_from
            FROM user_requests ur
            JOIN request_types rt ON ur.request_type_id = rt.id
            WHERE ur.id = %s
        """, (request_id,)).fetchone()
    else:
        existing = db.execute("""
            SELECT ur.id, ur.status, ur.username, rt.name as type_name, rt.value_type,
                   ur.extra_data, ur.date_from
            FROM user_requests ur
            JOIN request_types rt ON ur.request_type_id = rt.id
            WHERE ur.id = ?
        """, (request_id,)).fetchone()
    
    if not existing:
        return jsonify({"error": "Richiesta non trovata"}), 404
    
    if isinstance(existing, Mapping):
        current_status = existing["status"]
        target_username = existing["username"]
        type_name = existing["type_name"]
        value_type = existing.get("value_type")
        extra_data_str = existing.get("extra_data")
        date_from = existing.get("date_from")
    else:
        current_status = existing[1]
        target_username = existing[2]
        type_name = existing[3]
        value_type = existing[4] if len(existing) > 4 else None
        extra_data_str = existing[5] if len(existing) > 5 else None
        date_from = existing[6] if len(existing) > 6 else None
    
    if current_status != "pending":
        return jsonify({"error": "La richiesta è già stata revisionata"}), 400
    
    reviewed_by = session.get("user", "")
    now = int(datetime.now().timestamp() * 1000)
    
    # Estrai orari originali prima di modificarli (per determinare se approvazione parziale)
    original_rounded_start = None
    original_rounded_end = None
    if extra_data_str:
        try:
            orig_extra = json.loads(extra_data_str)
            original_rounded_start = orig_extra.get("rounded_start")
            original_rounded_end = orig_extra.get("rounded_end")
        except:
            pass
    
    # Se ci sono orari arrotondati confermati (modificati dall'admin), aggiorna extra_data
    if value_type == "minutes" and (rounded_start or rounded_end):
        try:
            extra_data = json.loads(extra_data_str) if extra_data_str else {}
            # Salva gli orari originali (calcolati dal sistema) se non già presenti
            if "original_rounded_start" not in extra_data and original_rounded_start:
                extra_data["original_rounded_start"] = original_rounded_start
            if "original_rounded_end" not in extra_data and original_rounded_end:
                extra_data["original_rounded_end"] = original_rounded_end
            # Ora aggiorna con i nuovi valori confermati dall'admin
            if rounded_start:
                extra_data["rounded_start"] = rounded_start
            if rounded_end:
                extra_data["rounded_end"] = rounded_end
            extra_data_str = json.dumps(extra_data)
            
            # Aggiorna extra_data nella richiesta
            placeholder = "%s" if DB_VENDOR == "mysql" else "?"
            db.execute(f"""
                UPDATE user_requests SET extra_data = {placeholder} WHERE id = {placeholder}
            """, (extra_data_str, request_id))
            
            app.logger.info(f"Extra data aggiornato con orari confermati: {extra_data_str}")
        except Exception as e:
            app.logger.warning(f"Errore aggiornamento extra_data: {e}")
    
    # Se è uno straordinario (minutes) e c'è un valore confermato, aggiorna anche value_amount
    if value_type == "minutes" and confirmed_value is not None:
        if DB_VENDOR == "mysql":
            db.execute("""
                UPDATE user_requests 
                SET status = %s, reviewed_by = %s, reviewed_ts = %s, review_notes = %s, 
                    value_amount = %s, updated_ts = %s
                WHERE id = %s
            """, (status, reviewed_by, now, review_notes, confirmed_value, now, request_id))
        else:
            db.execute("""
                UPDATE user_requests 
                SET status = ?, reviewed_by = ?, reviewed_ts = ?, review_notes = ?, 
                    value_amount = ?, updated_ts = ?
                WHERE id = ?
            """, (status, reviewed_by, now, review_notes, confirmed_value, now, request_id))
    else:
        if DB_VENDOR == "mysql":
            db.execute("""
                UPDATE user_requests 
                SET status = %s, reviewed_by = %s, reviewed_ts = %s, review_notes = %s, updated_ts = %s
                WHERE id = %s
            """, (status, reviewed_by, now, review_notes, now, request_id))
        else:
            db.execute("""
                UPDATE user_requests 
                SET status = ?, reviewed_by = ?, reviewed_ts = ?, review_notes = ?, updated_ts = ?
                WHERE id = ?
            """, (status, reviewed_by, now, review_notes, now, request_id))
    
    db.commit()
    
    # Se è uno straordinario approvato e ci sono orari arrotondati confermati, aggiorna le timbrature
    app.logger.info(
        f"Checking overtime update: value_type={value_type}, status={status}, "
        f"rounded_start={rounded_start}, rounded_end={rounded_end}"
    )
    # Determina se è un'approvazione parziale (orari modificati)
    is_partial_approval = False
    if value_type == "minutes" and status == "approved" and (rounded_start or rounded_end):
        _update_timbrature_with_confirmed_times(
            db, target_username, date_from, rounded_start, rounded_end, extra_data_str
        )
        # È parziale se l'admin ha cambiato almeno uno degli orari rispetto agli originali
        if (rounded_start and original_rounded_start and rounded_start != original_rounded_start) or \
           (rounded_end and original_rounded_end and rounded_end != original_rounded_end):
            is_partial_approval = True
            app.logger.info(
                f"Approvazione parziale: originale={original_rounded_start}-{original_rounded_end}, "
                f"confermato={rounded_start}-{rounded_end}"
            )
    
    # Se è Fuori Flessibilità, gestisci l'aggiornamento della timbratura
    flex_result = None
    if flex_type == "fuori_flessibilita" and rounded_time:
        try:
            app.logger.info(f"Processing Fuori Flessibilità: request_id={request_id}, status={status}, rounded_time={rounded_time}")
            flex_result = _process_fuori_flessibilita(
                db, request_id, target_username, date_from, status, rounded_time, extra_data_str
            )
            app.logger.info(f"Fuori Flessibilità result: {flex_result}")
        except Exception as e:
            import traceback
            app.logger.error(f"Errore processing Fuori Flessibilità: {e}\n{traceback.format_exc()}")
            flex_result = {"error": str(e)}
    
    # Se è uno straordinario (minutes) o Extra Turno, sincronizza le timbrature bloccate
    cedolino_debug = None
    is_overtime = (value_type == "minutes" or type_name == "Extra Turno") and flex_type != "fuori_flessibilita"
    app.logger.info(f"DEBUG: value_type={value_type}, type_name={type_name}, status={status}, is_overtime={is_overtime}")
    if is_overtime:
        try:
            app.logger.info(f"DEBUG: Calling _sync_overtime_blocked_timbrature for request {request_id}")
            cedolino_debug = _sync_overtime_blocked_timbrature(db, request_id, status, extra_data_str)
            app.logger.info(f"DEBUG: cedolino_debug result = {cedolino_debug}")
        except Exception as e:
            import traceback
            app.logger.error(f"Errore sync overtime: {e}\n{traceback.format_exc()}")
            cedolino_debug = {"error": str(e)}
    
    # Se è una Mancata Timbratura approvata, inserisci la timbratura e invia a CedolinoWeb
    timbratura_result = None
    if value_type == "timbratura" and status == "approved":
        try:
            app.logger.info(f"Processing approved Mancata Timbratura for request {request_id}")
            timbratura_result = _process_approved_mancata_timbratura(
                db, target_username, date_from, extra_data_str
            )
            app.logger.info(f"Mancata Timbratura result: {timbratura_result}")
        except Exception as e:
            import traceback
            app.logger.error(f"Errore processing Mancata Timbratura: {e}\n{traceback.format_exc()}")
            timbratura_result = {"error": str(e)}
    
    # Invia notifica push all'utente
    _send_request_review_notification(
        db, target_username, type_name, status, review_notes,
        is_partial=is_partial_approval,
        rounded_start=rounded_start if is_partial_approval else None,
        rounded_end=rounded_end if is_partial_approval else None
    )
    
    status_label = "approvata parzialmente" if is_partial_approval else ("approvata" if status == "approved" else "respinta")
    response = {
        "ok": True, 
        "message": f"Richiesta {status_label} con successo",
        "debug": {
            "value_type": value_type,
            "type_name": type_name,
            "is_overtime": is_overtime,
            "status": status,
            "flex_type": flex_type
        }
    }
    if cedolino_debug:
        response["cedolino_debug"] = cedolino_debug
    if timbratura_result:
        response["timbratura_result"] = timbratura_result
    if flex_result:
        response["flex_result"] = flex_result
    return jsonify(response)


@app.delete("/api/admin/user-requests/<int:request_id>")
@login_required
def api_admin_user_request_delete(request_id: int) -> ResponseReturnValue:
    """Elimina una richiesta utente (protetto da password)."""
    if not session.get("is_admin"):
        return jsonify({"error": "Accesso negato"}), 403
    
    data = request.get_json() or {}
    password = data.get("password", "")
    
    # Password di sicurezza per eliminazione
    DELETE_PASSWORD = "225524"
    
    if password != DELETE_PASSWORD:
        return jsonify({"error": "Password non corretta"}), 403
    
    db = get_db()
    ensure_user_requests_table(db)
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Verifica che la richiesta esista
    existing = db.execute(
        f"SELECT id, username, status FROM user_requests WHERE id = {placeholder}",
        (request_id,)
    ).fetchone()
    
    if not existing:
        return jsonify({"error": "Richiesta non trovata"}), 404
    
    # Elimina la richiesta
    db.execute(f"DELETE FROM user_requests WHERE id = {placeholder}", (request_id,))
    db.commit()
    
    username = existing["username"] if isinstance(existing, Mapping) else existing[1]
    app.logger.info(f"Richiesta {request_id} eliminata da {session.get('user')} (utente: {username})")
    
    return jsonify({"ok": True, "message": "Richiesta eliminata con successo"})


# =====================================================
# USER REQUESTS - Pagina e API per richieste utente
# =====================================================

@app.route("/user/requests")
@login_required
def user_requests_page() -> ResponseReturnValue:
    """Pagina utente per inviare richieste (ferie, permessi, rimborsi, ecc.)."""
    db = get_db()
    magazzino_enabled = is_module_enabled(db, "magazzino")
    return render_template(
        "user_requests.html",
        username=session.get("user"),
        user_display=session.get("user_display", session.get("user")),
        user_initials=session.get("user_initials", "U"),
        user_role=session.get("user_role", "Utente"),
        magazzino_enabled=magazzino_enabled,
    )


@app.route("/user/turni")
@login_required
def user_turni_page() -> ResponseReturnValue:
    """Pagina utente per visualizzare i propri turni."""
    db = get_db()
    magazzino_enabled = is_module_enabled(db, "magazzino")
    return render_template(
        "user_turni.html",
        username=session.get("user"),
        user_display=session.get("user_display", session.get("user")),
        user_initials=session.get("user_initials", "U"),
        user_role=session.get("user_role", "Utente"),
        magazzino_enabled=magazzino_enabled,
    )


@app.route("/user/notifications")
@login_required
def user_notifications_page() -> ResponseReturnValue:
    """Pagina utente per visualizzare lo storico delle notifiche push."""
    db = get_db()
    magazzino_enabled = is_module_enabled(db, "magazzino")
    return render_template(
        "user_notifications.html",
        username=session.get("user"),
        user_display=session.get("user_display", session.get("user")),
        user_initials=session.get("user_initials", "U"),
        user_role=session.get("user_role", "Utente"),
        magazzino_enabled=magazzino_enabled,
    )


@app.route("/user/storico-timbrature")
@login_required
def user_storico_timbrature_page() -> ResponseReturnValue:
    """Pagina utente per visualizzare lo storico delle timbrature per mese."""
    db = get_db()
    magazzino_enabled = is_module_enabled(db, "magazzino")
    return render_template(
        "user_storico_timbrature.html",
        username=session.get("user"),
        user_display=session.get("user_display", session.get("user")),
        user_initials=session.get("user_initials", "U"),
        user_role=session.get("user_role", "Utente"),
        magazzino_enabled=magazzino_enabled,
    )


@app.get("/api/user/storico-timbrature")
@login_required
def api_user_storico_timbrature() -> ResponseReturnValue:
    """Restituisce lo storico delle timbrature dell'utente per un mese specifico."""
    username = session.get("user")
    if not username:
        return jsonify({"error": "Non autenticato"}), 401
    
    # Parametri: year e month (default: mese corrente)
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    
    if not year or not month:
        today = datetime.now()
        year = year or today.year
        month = month or today.month
    
    print(f"[storico-timbrature] User: {username}, Year: {year}, Month: {month}")
    
    # Calcola primo e ultimo giorno del mese
    first_day = f"{year:04d}-{month:02d}-01"
    if month == 12:
        last_day = f"{year+1:04d}-01-01"
    else:
        last_day = f"{year:04d}-{month+1:02d}-01"
    
    db = get_db()
    ensure_timbrature_table(db)
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Recupera tutte le timbrature del mese
    rows = db.execute(f"""
        SELECT id, tipo, data, ora, ora_mod, created_ts, method, location_name
        FROM timbrature
        WHERE username = {placeholder}
          AND data >= {placeholder}
          AND data < {placeholder}
        ORDER BY data ASC, ora ASC
    """, (username, first_day, last_day)).fetchall()
    
    print(f"[storico-timbrature] Trovate {len(rows)} timbrature per {username}")
    
    # Helper per formattare ora in HH:MM
    def format_ora(ora_val):
        if not ora_val:
            return None
        ora_str = str(ora_val)
        # Gestisce formati come "9:30:00" o "09:30:00" o timedelta
        if ':' in ora_str:
            parts = ora_str.split(':')
            if len(parts) >= 2:
                return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
        return ora_str[:5]
    
    # Raggruppa per giorno
    days_data = {}
    for row in rows:
        if isinstance(row, Mapping):
            data = str(row["data"])
            timbratura = {
                "id": row["id"],
                "tipo": row["tipo"],
                "ora": format_ora(row["ora"]),
                "ora_mod": format_ora(row.get("ora_mod")),
                "method": row.get("method"),
                "location": row.get("location_name"),
            }
        else:
            data = str(row[2])
            timbratura = {
                "id": row[0],
                "tipo": row[1],
                "ora": format_ora(row[3]),
                "ora_mod": format_ora(row[4]) if len(row) > 4 else None,
                "method": row[5] if len(row) > 5 else None,
                "location": row[7] if len(row) > 7 else None,
            }
        
        if data not in days_data:
            days_data[data] = []
        days_data[data].append(timbratura)
    
    # Recupera le richieste dell'utente per il mese (solo giustificativi approvati)
    requests_by_date = {}
    
    # Recupera anche le richieste "Fuori Flessibilità" (qualunque stato) per mostrare i motivi delle normalizzazioni
    fuori_flex_by_date = {}
    try:
        # Query per richieste Fuori Flessibilità (tutte, non solo approvate)
        fuori_flex_rows = db.execute(f"""
            SELECT ur.id, ur.date_from, ur.status, ur.notes, ur.review_notes, ur.extra_data,
                   ur.reviewed_by, ur.reviewed_ts
            FROM user_requests ur
            LEFT JOIN request_types rt ON ur.request_type_id = rt.id
            WHERE ur.username = {placeholder}
              AND rt.name = 'Fuori Flessibilità'
              AND ur.date_from >= {placeholder}
              AND ur.date_from < {placeholder}
            ORDER BY ur.date_from ASC
        """, (username, first_day, last_day)).fetchall()
        
        for req in fuori_flex_rows:
            if isinstance(req, Mapping):
                date_str = str(req["date_from"])[:10]
                extra_data = {}
                if req.get("extra_data"):
                    try:
                        extra_data = json.loads(req["extra_data"]) if isinstance(req["extra_data"], str) else req["extra_data"]
                    except:
                        pass
                fuori_flex_data = {
                    "id": req["id"],
                    "status": req["status"],
                    "notes": req.get("notes"),  # Contiene il motivo originale (es. "Fine Giornata fuori flessibilità: +41 minuti oltre la flessibilità")
                    "review_notes": req.get("review_notes"),  # Note dell'admin quando approva/rifiuta
                    "reviewed_by": req.get("reviewed_by"),
                    "reviewed_ts": req.get("reviewed_ts"),
                    "tipo_timbratura": extra_data.get("tipo_timbratura"),
                    "ora_timbrata": extra_data.get("ora_timbrata"),
                    "ora_finale": extra_data.get("ora_finale"),  # Orario approvato dall'admin
                    "rounded_time": extra_data.get("rounded_time"),
                    "turno_end": extra_data.get("turno_end"),
                    "flessibilita": extra_data.get("flessibilita"),
                    "diff_minuti": extra_data.get("diff_minuti"),
                }
            else:
                date_str = str(req[1])[:10]
                extra_data = {}
                if req[5]:
                    try:
                        extra_data = json.loads(req[5]) if isinstance(req[5], str) else req[5]
                    except:
                        pass
                fuori_flex_data = {
                    "id": req[0],
                    "status": req[2],
                    "notes": req[3],
                    "review_notes": req[4],
                    "reviewed_by": req[6] if len(req) > 6 else None,
                    "reviewed_ts": req[7] if len(req) > 7 else None,
                    "tipo_timbratura": extra_data.get("tipo_timbratura"),
                    "ora_timbrata": extra_data.get("ora_timbrata"),
                    "ora_finale": extra_data.get("ora_finale"),
                    "rounded_time": extra_data.get("rounded_time"),
                    "turno_end": extra_data.get("turno_end"),
                    "flessibilita": extra_data.get("flessibilita"),
                    "diff_minuti": extra_data.get("diff_minuti"),
                }
            
            if date_str not in fuori_flex_by_date:
                fuori_flex_by_date[date_str] = []
            fuori_flex_by_date[date_str].append(fuori_flex_data)
    except Exception as e:
        print(f"[storico-timbrature] Errore recupero richieste fuori flessibilità: {e}")
    
    try:
        requests_rows = db.execute(f"""
            SELECT ur.id, ur.request_type_id, ur.date_from, ur.date_to, ur.value_amount, 
                   ur.notes, ur.status, ur.review_notes, rt.name as type_name, rt.abbreviation,
                   ur.created_ts, ur.reviewed_ts, ur.reviewed_by
            FROM user_requests ur
            LEFT JOIN request_types rt ON ur.request_type_id = rt.id
            WHERE ur.username = {placeholder}
              AND ur.status = 'approved'
              AND rt.is_giustificativo = 1
              AND (
                (ur.date_from >= {placeholder} AND ur.date_from < {placeholder})
                OR (ur.date_to IS NOT NULL AND ur.date_to >= {placeholder} AND ur.date_to < {placeholder})
                OR (ur.date_from < {placeholder} AND (ur.date_to IS NULL OR ur.date_to >= {placeholder}))
              )
            ORDER BY ur.date_from ASC
        """, (username, first_day, last_day, first_day, last_day, first_day, first_day)).fetchall()
        
        # Mappa richieste per data
        for req in requests_rows:
            if isinstance(req, Mapping):
                date_from = str(req["date_from"])
                date_to = str(req["date_to"]) if req.get("date_to") else date_from
                req_data = {
                    "id": req["id"],
                    "type_name": req.get("type_name") or "Richiesta",
                    "abbreviation": req.get("abbreviation"),
                    "value": float(req["value_amount"]) if req["value_amount"] else 0,
                    "status": req["status"],
                    "notes": req.get("notes"),
                    "review_notes": req.get("review_notes"),
                    "created_ts": req.get("created_ts"),
                    "reviewed_ts": req.get("reviewed_ts"),
                    "reviewed_by": req.get("reviewed_by"),
                }
            else:
                date_from = str(req[2])
                date_to = str(req[3]) if req[3] else date_from
                req_data = {
                    "id": req[0],
                    "type_name": req[8] if len(req) > 8 and req[8] else "Richiesta",
                    "abbreviation": req[9] if len(req) > 9 else None,
                    "value": float(req[4]) if req[4] else 0,
                    "status": req[6],
                    "notes": req[5],
                    "review_notes": req[7],
                    "created_ts": req[10] if len(req) > 10 else None,
                    "reviewed_ts": req[11] if len(req) > 11 else None,
                    "reviewed_by": req[12] if len(req) > 12 else None,
                }
            
            # Aggiungi la richiesta a ogni giorno nel range
            try:
                start = datetime.strptime(date_from[:10], "%Y-%m-%d")
                end = datetime.strptime(date_to[:10], "%Y-%m-%d")
                current = start
                while current <= end:
                    date_str = current.strftime("%Y-%m-%d")
                    if date_str not in requests_by_date:
                        requests_by_date[date_str] = []
                    requests_by_date[date_str].append(req_data)
                    current += timedelta(days=1)
            except:
                pass
    except Exception as e:
        # Se la tabella non esiste o c'è un errore, ignora le richieste
        print(f"[storico-timbrature] Errore recupero richieste: {e}")
    
    # Converti in formato per il calendario
    # Recupera le regole di timbratura specifiche dell'utente (considera il gruppo)
    user_rules = get_user_timbratura_rules(db, username)
    rounding_mode = user_rules.get('rounding_mode', 'single')
    
    # Recupera il turno dell'utente da employee_shifts (per mostrare nel riepilogo)
    user_shift = None
    try:
        ensure_employee_shifts_table(db)
        # Prendo il turno del lunedì come riferimento (day_of_week=0)
        shift_row = db.execute(f"""
            SELECT start_time, end_time, break_start, break_end
            FROM employee_shifts
            WHERE username = {placeholder} AND is_active = 1
            ORDER BY day_of_week ASC
            LIMIT 1
        """, (username,)).fetchone()
        
        if shift_row:
            if isinstance(shift_row, Mapping):
                user_shift = {
                    "start": str(shift_row['start_time'])[:5] if shift_row['start_time'] else None,
                    "end": str(shift_row['end_time'])[:5] if shift_row['end_time'] else None,
                    "break_start": str(shift_row['break_start'])[:5] if shift_row['break_start'] else None,
                    "break_end": str(shift_row['break_end'])[:5] if shift_row['break_end'] else None,
                }
            else:
                user_shift = {
                    "start": str(shift_row[0])[:5] if shift_row[0] else None,
                    "end": str(shift_row[1])[:5] if shift_row[1] else None,
                    "break_start": str(shift_row[2])[:5] if shift_row[2] else None,
                    "break_end": str(shift_row[3])[:5] if shift_row[3] else None,
                }
    except Exception as e:
        print(f"[storico-timbrature] Errore recupero turno utente: {e}")
    
    # Calcola ore lavorate per ogni giorno
    timbrature_by_day = {}
    debug_calcs = []  # Per debug
    for data in sorted(days_data.keys()):
        timbrature_list = days_data[data]
        
        # Calcola ore lavorate
        ora_inizio = None
        ora_fine = None
        pausa_minuti = 0
        inizio_pausa = None
        
        for t in timbrature_list:
            tipo = t.get("tipo", "")
            ora = t.get("ora_mod") or t.get("ora")
            
            if tipo == "inizio_giornata" and not ora_inizio:
                ora_inizio = ora
            elif tipo == "fine_giornata":
                ora_fine = ora
            elif tipo == "inizio_pausa":
                inizio_pausa = ora
            elif tipo == "fine_pausa" and inizio_pausa:
                try:
                    h1, m1 = map(int, inizio_pausa.split(':')[:2])
                    h2, m2 = map(int, ora.split(':')[:2])
                    pausa_minuti += (h2 * 60 + m2) - (h1 * 60 + m1)
                except Exception as e:
                    print(f"[storico] Errore calcolo pausa: {e} ({inizio_pausa} -> {ora})")
                inizio_pausa = None
        
        # Calcola ore nette
        ore_lavorate = None
        calcolo_dettagli = None  # Dettagli per UI
        if ora_inizio and ora_fine:
            try:
                h1, m1 = map(int, ora_inizio.split(':')[:2])
                h2, m2 = map(int, ora_fine.split(':')[:2])
                total_minutes = (h2 * 60 + m2) - (h1 * 60 + m1) - pausa_minuti
                
                # Se rounding_mode è 'daily', applica arrotondamento giornaliero
                if rounding_mode == 'daily' and total_minutes > 0:
                    result = calcola_ore_giornaliere_arrotondate(
                        ora_inizio, ora_fine, pausa_minuti, user_rules
                    )
                    ore_lavorate = result['ore_str']
                    
                    # Calcola pausa prevista dal turno (da break_start e break_end)
                    pausa_turno_minuti = 60  # Default 1 ora
                    if user_shift and user_shift.get('break_start') and user_shift.get('break_end'):
                        try:
                            bs_h, bs_m = map(int, user_shift['break_start'].split(':')[:2])
                            be_h, be_m = map(int, user_shift['break_end'].split(':')[:2])
                            pausa_turno_minuti = (be_h * 60 + be_m) - (bs_h * 60 + bs_m)
                        except:
                            pass
                    
                    # Salva dettagli per il frontend
                    calcolo_dettagli = {
                        "ore_lorde_minuti": result['ore_lorde'],
                        "ore_nette_minuti": result['ore_nette'],
                        "ore_arrotondate_minuti": result['ore_arrotondate'],
                        "turno_base_minuti": result['turno_base_minuti'],
                        "straordinario_lordo_minuti": result['straordinario_lordo'],
                        "straordinario_arrotondato_minuti": result['straordinario_arrotondato'],
                        "blocco_minuti": result['blocco_minuti'],
                        "blocchi_straordinario": result['blocchi_straordinario'],
                        "tipo_arrotondamento": result['tipo_arrotondamento'],
                        # Aggiungi dati turno per il riepilogo
                        "turno_inizio": user_shift.get('start') if user_shift else None,
                        "turno_fine": user_shift.get('end') if user_shift else None,
                        "pausa_turno_minuti": pausa_turno_minuti,  # Pausa prevista dal turno
                    }
                    debug_calcs.append(f"{data}: {ora_inizio}->{ora_fine} pausa={pausa_minuti} netto={total_minutes}m => arrotondato={ore_lavorate}")
                elif total_minutes > 0:
                    ore_lavorate = f"{total_minutes // 60}:{total_minutes % 60:02d}"
                    debug_calcs.append(f"{data}: {ora_inizio}->{ora_fine} pausa={pausa_minuti} => {ore_lavorate}")
            except Exception as e:
                print(f"[storico] Errore calcolo ore: {e} ({ora_inizio} -> {ora_fine})")
        
        timbrature_by_day[data] = {
            "timbrature": timbrature_list,
            "requests": requests_by_date.get(data, []),
            "fuori_flessibilita": fuori_flex_by_date.get(data, []),  # Richieste fuori flessibilità per mostrare motivi
            "ora_inizio": ora_inizio,
            "ora_fine": ora_fine,
            "pausa_minuti": pausa_minuti,
            "ore_lavorate": ore_lavorate,
            "calcolo_dettagli": calcolo_dettagli  # Dettagli arrotondamento per UI
        }
    
    # Aggiungi anche le richieste per giorni senza timbrature
    all_requests = []
    for date_str, reqs in requests_by_date.items():
        if date_str not in timbrature_by_day:
            timbrature_by_day[date_str] = {
                "timbrature": [],
                "requests": reqs,
                "ora_inizio": None,
                "ora_fine": None,
                "pausa_minuti": 0,
                "ore_lavorate": None
            }
        for req in reqs:
            req_copy = req.copy()
            req_copy["data"] = date_str
            all_requests.append(req_copy)
    
    # Debug: stampa le ore calcolate per i primi 5 giorni
    if debug_calcs:
        print(f"[storico-timbrature] Calcoli ore: {debug_calcs[:5]}...")
    
    arrotondamento_info = {
        "rounding_mode": rounding_mode,
        "source": user_rules.get('source', 'global'),
        "ingresso": {
            "minuti": user_rules.get("arrotondamento_ingresso_minuti", 15),
            "tipo": user_rules.get("arrotondamento_ingresso_tipo", "~")
        },
        "uscita": {
            "minuti": user_rules.get("arrotondamento_uscita_minuti", 15),
            "tipo": user_rules.get("arrotondamento_uscita_tipo", "~")
        }
    }
    
    # Per rounding_mode daily, aggiungi info sulla flessibilità
    if rounding_mode == 'daily':
        arrotondamento_info["flessibilita_ingresso"] = user_rules.get("flessibilita_ingresso_minuti", 30)
        arrotondamento_info["flessibilita_uscita"] = user_rules.get("flessibilita_uscita_minuti", 30)
        arrotondamento_info["arrotondamento_giornaliero"] = {
            "minuti": user_rules.get("arrotondamento_giornaliero_minuti", 15),
            "tipo": user_rules.get("arrotondamento_giornaliero_tipo", "floor")
        }
    
    return jsonify({
        "success": True,
        "year": year,
        "month": month,
        "timbrature_by_day": timbrature_by_day,
        "requests": all_requests,
        "arrotondamento": arrotondamento_info
    })


@app.route("/user/documents")
@login_required
def user_documents_page() -> ResponseReturnValue:
    """Pagina utente per visualizzare i documenti (circolari, comunicazioni, buste paga)."""
    db = get_db()
    magazzino_enabled = is_module_enabled(db, "magazzino")
    return render_template(
        "user_documents.html",
        username=session.get("user"),
        user_display=session.get("user_display", session.get("user")),
        user_initials=session.get("user_initials", "U"),
        user_role=session.get("user_role", "Utente"),
        magazzino_enabled=magazzino_enabled,
    )


# =====================================================
# USER DOCUMENTS - API per documenti utente
# =====================================================

@app.get("/api/user/documents")
@login_required
def api_user_documents_list() -> ResponseReturnValue:
    """Restituisce i documenti visibili all'utente (circolari, comunicazioni, buste paga)."""
    username = session.get("user")
    if not username:
        return jsonify({"error": "Non autenticato"}), 401
    
    db = get_db()
    ensure_user_documents_table(db)
    
    # Recupera documenti visibili all'utente (solo quelli già inviati: notified_at IS NOT NULL)
    # E che sono destinati all'utente (target_all=1 o username in target_users)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    rows = db.execute(f"""
        SELECT d.id, d.category, d.title, d.description, d.file_path, d.file_name, d.created_at,
               CASE WHEN r.id IS NOT NULL THEN 1 ELSE 0 END as is_read
        FROM user_documents d
        LEFT JOIN user_documents_read r ON d.id = r.document_id AND r.username = {placeholder}
        WHERE d.notified_at IS NOT NULL
          AND (d.target_all = 1 OR (d.target_users IS NOT NULL AND d.target_users LIKE {placeholder}))
        ORDER BY d.created_at DESC
    """, (username, f'%"{username}"%')).fetchall()
    
    documents = []
    for row in rows:
        if isinstance(row, Mapping):
            doc = {
                "id": row["id"],
                "category": row["category"],
                "title": row["title"],
                "description": row["description"],
                "file_url": f"/uploads/documents/{row['file_name']}" if row["file_name"] else None,
                "created_at": row["created_at"],
                "read": bool(row["is_read"])
            }
        else:
            doc = {
                "id": row[0],
                "category": row[1],
                "title": row[2],
                "description": row[3],
                "file_url": f"/uploads/documents/{row[5]}" if row[5] else None,
                "created_at": row[6],
                "read": bool(row[7])
            }
        documents.append(doc)
    
    return jsonify({"documents": documents})


@app.post("/api/user/documents/<int:doc_id>/read")
@login_required
def api_user_document_mark_read(doc_id: int) -> ResponseReturnValue:
    """Marca un documento come letto dall'utente."""
    username = session.get("user")
    if not username:
        return jsonify({"error": "Non autenticato"}), 401
    
    db = get_db()
    ensure_user_documents_table(db)
    
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    try:
        if DB_VENDOR == "mysql":
            db.execute(f"""
                INSERT IGNORE INTO user_documents_read (document_id, username)
                VALUES ({placeholder}, {placeholder})
            """, (doc_id, username))
        else:
            db.execute(f"""
                INSERT OR IGNORE INTO user_documents_read (document_id, username)
                VALUES ({placeholder}, {placeholder})
            """, (doc_id, username))
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        app.logger.error(f"Errore marcatura documento letto: {e}")
        return jsonify({"error": str(e)}), 500


@app.get("/api/user/request-types")
@login_required
def api_user_request_types_list() -> ResponseReturnValue:
    """Restituisce le tipologie di richiesta attive per l'utente (escluse quelle con ordine 99)."""
    db = get_db()
    ensure_request_types_table(db)
    
    rows = db.execute("""
        SELECT id, name, value_type, description
        FROM request_types
        WHERE active = 1 AND (sort_order IS NULL OR sort_order != 99)
        ORDER BY sort_order ASC, name ASC
    """).fetchall()

    types = []
    for row in rows:
        if isinstance(row, Mapping):
            types.append({
                "id": row["id"],
                "name": row["name"],
                "value_type": row["value_type"],
                "description": row["description"],
            })
        else:
            types.append({
                "id": row[0],
                "name": row[1],
                "value_type": row[2],
                "description": row[3],
            })

    return jsonify({"types": types})


@app.get("/api/user/residuals")
@login_required
def api_user_residuals() -> ResponseReturnValue:
    """Restituisce le ore residue di ferie e permessi per l'utente loggato."""
    # Per ora ritorna valori placeholder - in futuro si potrà integrare con sistema HR
    username = session.get("user", "")
    
    # TODO: Implementare calcolo reale basato su:
    # - Ore totali annuali assegnate
    # - Ore già godute/approvate
    
    return jsonify({
        "ferie_hours": 0,
        "permessi_hours": 0
    })


@app.get("/api/user/requests")
@login_required
def api_user_requests_list() -> ResponseReturnValue:
    """Restituisce lo storico delle richieste dell'utente loggato."""
    username = session.get("user", "")
    
    db = get_db()
    ensure_user_requests_table(db)
    
    if DB_VENDOR == "mysql":
        rows = db.execute("""
            SELECT ur.id, ur.request_type_id, rt.name as type_name, rt.value_type,
                   ur.date_from, ur.date_to, ur.value_amount, ur.notes, ur.status,
                   ur.review_notes, ur.created_ts, ur.updated_ts, ur.cdc, ur.attachment_path, ur.tratte, ur.extra_data
            FROM user_requests ur
            JOIN request_types rt ON ur.request_type_id = rt.id
            WHERE ur.username = %s
            ORDER BY ur.created_ts DESC
        """, (username,)).fetchall()
    else:
        rows = db.execute("""
            SELECT ur.id, ur.request_type_id, rt.name as type_name, rt.value_type,
                   ur.date_from, ur.date_to, ur.value_amount, ur.notes, ur.status,
                   ur.review_notes, ur.created_ts, ur.updated_ts, ur.cdc, ur.attachment_path, ur.tratte, ur.extra_data
            FROM user_requests ur
            JOIN request_types rt ON ur.request_type_id = rt.id
            WHERE ur.username = ?
            ORDER BY ur.created_ts DESC
        """, (username,)).fetchall()

    value_units = {"hours": "ore", "days": "giorni", "amount": "€", "km": "km", "minutes": "minuti"}
    
    requests = []
    for row in rows:
        if isinstance(row, Mapping):
            # Parse tratte JSON
            tratte_raw = row.get("tratte")
            tratte = None
            if tratte_raw:
                try:
                    tratte = json.loads(tratte_raw) if isinstance(tratte_raw, str) else tratte_raw
                except:
                    pass
            
            # Parse extra_data JSON (per straordinari)
            extra_data_raw = row.get("extra_data")
            extra_data = None
            if extra_data_raw:
                try:
                    extra_data = json.loads(extra_data_raw) if isinstance(extra_data_raw, str) else extra_data_raw
                except:
                    pass
            
            req_item = {
                "id": row["id"],
                "request_type_id": row["request_type_id"],
                "type_name": row["type_name"],
                "value_type": row["value_type"],
                "value_unit": value_units.get(row["value_type"], ""),
                "date_start": row["date_from"],
                "date_end": row["date_to"],
                "value": float(row["value_amount"]) if row["value_amount"] else None,
                "notes": row["notes"],
                "status": row["status"],
                "admin_notes": row["review_notes"],
                "created_ts": row["created_ts"],
                "updated_ts": row["updated_ts"],
                "cdc": row["cdc"],
                "attachment_path": row["attachment_path"],
                "tratte": tratte,
            }
            
            # Aggiungi dettagli straordinario se presenti
            if extra_data:
                req_item["planned_start"] = extra_data.get("planned_start")
                req_item["planned_end"] = extra_data.get("planned_end")
                req_item["actual_start"] = extra_data.get("actual_start")
                req_item["actual_end"] = extra_data.get("actual_end")
                req_item["rounded_start"] = extra_data.get("rounded_start")
                req_item["rounded_end"] = extra_data.get("rounded_end")
                req_item["extra_minutes_before"] = extra_data.get("extra_minutes_before", 0)
                req_item["extra_minutes_after"] = extra_data.get("extra_minutes_after", 0)
                req_item["shift_source"] = extra_data.get("shift_source")
            
            requests.append(req_item)
        else:
            # Parse tratte JSON
            tratte_raw = row[14] if len(row) > 14 else None
            tratte = None
            if tratte_raw:
                try:
                    tratte = json.loads(tratte_raw) if isinstance(tratte_raw, str) else tratte_raw
                except:
                    pass
            
            # Parse extra_data JSON (per straordinari)
            extra_data_raw = row[15] if len(row) > 15 else None
            extra_data = None
            if extra_data_raw:
                try:
                    extra_data = json.loads(extra_data_raw) if isinstance(extra_data_raw, str) else extra_data_raw
                except:
                    pass
            
            req_item = {
                "id": row[0],
                "request_type_id": row[1],
                "type_name": row[2],
                "value_type": row[3],
                "value_unit": value_units.get(row[3], ""),
                "date_start": row[4],
                "date_end": row[5],
                "value": float(row[6]) if row[6] else None,
                "notes": row[7],
                "status": row[8],
                "admin_notes": row[9],
                "created_ts": row[10],
                "updated_ts": row[11],
                "cdc": row[12] if len(row) > 12 else None,
                "attachment_path": row[13] if len(row) > 13 else None,
                "tratte": tratte,
            }
            
            # Aggiungi dettagli straordinario se presenti
            if extra_data:
                req_item["planned_start"] = extra_data.get("planned_start")
                req_item["planned_end"] = extra_data.get("planned_end")
                req_item["actual_start"] = extra_data.get("actual_start")
                req_item["actual_end"] = extra_data.get("actual_end")
                req_item["rounded_start"] = extra_data.get("rounded_start")
                req_item["rounded_end"] = extra_data.get("rounded_end")
                req_item["extra_minutes_before"] = extra_data.get("extra_minutes_before", 0)
                req_item["extra_minutes_after"] = extra_data.get("extra_minutes_after", 0)
                req_item["shift_source"] = extra_data.get("shift_source")
            
            requests.append(req_item)

    return jsonify({"requests": requests})


@app.put("/api/user/requests/<int:request_id>/notes")
@login_required
def api_user_request_update_notes(request_id: int) -> ResponseReturnValue:
    """Aggiorna le note di una richiesta utente (solo se pending e di proprietà dell'utente)."""
    username = session.get("user", "")
    if not username:
        return jsonify({"error": "Non autenticato"}), 401
    
    data = request.get_json() or {}
    notes = (data.get("notes") or "").strip()
    
    db = get_db()
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Verifica che la richiesta esista, sia dell'utente e sia pending
    row = db.execute(
        f"SELECT id, username, status FROM user_requests WHERE id = {placeholder}",
        (request_id,)
    ).fetchone()
    
    if not row:
        return jsonify({"error": "Richiesta non trovata"}), 404
    
    req_username = row['username'] if isinstance(row, dict) else row[1]
    req_status = row['status'] if isinstance(row, dict) else row[2]
    
    if req_username != username:
        return jsonify({"error": "Non autorizzato"}), 403
    
    if req_status != 'pending':
        return jsonify({"error": "Impossibile modificare una richiesta già processata"}), 400
    
    # Aggiorna le note
    db.execute(
        f"UPDATE user_requests SET notes = {placeholder}, updated_ts = {placeholder} WHERE id = {placeholder}",
        (notes, now_ms(), request_id)
    )
    db.commit()
    
    return jsonify({"ok": True, "message": "Note aggiornate"})


@app.post("/api/user/requests")
@login_required
def api_user_requests_create() -> ResponseReturnValue:
    """Crea una nuova richiesta utente. Supporta JSON o multipart/form-data per upload file."""
    username = session.get("user", "")
    
    # Determina se è multipart/form-data o JSON
    content_type = request.content_type or ""
    
    if 'multipart/form-data' in content_type:
        # Form con allegati (supporta multipli)
        request_type_id = request.form.get("request_type_id")
        date_start = request.form.get("date_start") or request.form.get("start_date")
        date_end = request.form.get("date_end") or request.form.get("end_date")
        value = request.form.get("value")
        notes = (request.form.get("notes") or "").strip()
        cdc = (request.form.get("cdc") or "").strip() or None
        tratte = None  # Non supportato via form per ora
        # Supporta sia 'attachment' singolo che 'attachments' multipli
        attachment_files = request.files.getlist("attachments") or []
        single_attachment = request.files.get("attachment")
        if single_attachment and single_attachment.filename:
            attachment_files.append(single_attachment)
    else:
        # JSON (senza allegato) - force=True per evitare errore 415
        data = request.get_json(force=True, silent=True) or {}
        request_type_id = data.get("request_type_id")
        date_start = data.get("date_start") or data.get("start_date")
        date_end = data.get("date_end") or data.get("end_date")
        value = data.get("value")
        notes = (data.get("notes") or "").strip()
        cdc = (data.get("cdc") or "").strip() or None
        tratte = data.get("tratte")  # Array di {da, a, km}
        extra_data_json = data.get("extra_data")  # JSON per dati aggiuntivi (es. orari permessi)
        attachment_files = []

    if not request_type_id or not date_start:
        return jsonify({"error": "Tipo richiesta e data inizio sono obbligatori"}), 400

    db = get_db()
    ensure_user_requests_table(db)
    
    # Verifica che il tipo richiesta esista e sia attivo
    if DB_VENDOR == "mysql":
        rt = db.execute("SELECT id, value_type FROM request_types WHERE id = %s AND active = 1", (request_type_id,)).fetchone()
    else:
        rt = db.execute("SELECT id, value_type FROM request_types WHERE id = ? AND active = 1", (request_type_id,)).fetchone()
    
    if not rt:
        return jsonify({"error": "Tipologia richiesta non valida o non attiva"}), 400

    value_type = rt["value_type"] if isinstance(rt, dict) else rt[1]
    
    # Validazioni specifiche per value_type
    if value_type == "amount":
        if not cdc:
            return jsonify({"error": "Centro di costo (CDC) è obbligatorio per i rimborsi"}), 400
        if not attachment_files or len(attachment_files) == 0:
            return jsonify({"error": "Almeno un allegato (foto ricevuta) è obbligatorio per i rimborsi"}), 400
    
    # Validazioni specifiche per km
    if value_type == "km":
        if not cdc:
            return jsonify({"error": "Centro di costo (CDC) è obbligatorio per i rimborsi km"}), 400
        if tratte and len(tratte) > 0:
            # Valida le tratte
            total_km = 0
            for t in tratte:
                if not t.get("da") or not t.get("a"):
                    return jsonify({"error": "Ogni tratta deve avere origine e destinazione"}), 400
                km = t.get("km", 0)
                if km <= 0:
                    return jsonify({"error": "Ogni tratta deve avere km > 0"}), 400
                total_km += km
            # Sovrascrivi value con il totale calcolato dalle tratte
            value = total_km

    # Gestione upload file multipli
    attachment_paths = []
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf'}
    
    for idx, attachment_file in enumerate(attachment_files):
        if attachment_file and attachment_file.filename:
            ext = attachment_file.filename.rsplit('.', 1)[-1].lower() if '.' in attachment_file.filename else ''
            if ext not in allowed_extensions:
                return jsonify({"error": f"Formato file non supportato: {attachment_file.filename}. Usa: {', '.join(allowed_extensions)}"}), 400
            
            # Crea directory se non esiste
            upload_dir = os.path.join(app.root_path, 'uploads', 'requests', username)
            os.makedirs(upload_dir, exist_ok=True)
            
            # Nome file univoco (con indice per multipli)
            timestamp = int(time.time() * 1000)  # millisecondi per unicità
            filename = f"request_{timestamp}_{idx}.{ext}"
            filepath = os.path.join(upload_dir, filename)
            attachment_file.save(filepath)
            attachment_paths.append(f"/uploads/requests/{username}/{filename}")
    
    # Salva come JSON array se ci sono più file, altrimenti stringa singola per retrocompatibilità
    if len(attachment_paths) == 0:
        attachment_path = None
    elif len(attachment_paths) == 1:
        attachment_path = attachment_paths[0]
    else:
        attachment_path = json.dumps(attachment_paths)

    now = int(datetime.now().timestamp() * 1000)  # timestamp in millisecondi
    value_amount = float(value) if value else 0.0
    tratte_json = json.dumps(tratte) if tratte else None
    
    # extra_data_json può essere valorizzato dal frontend (es. per permessi con orari)
    # Se non viene passato, rimane None
    
    if DB_VENDOR == "mysql":
        db.execute("""
            INSERT INTO user_requests (user_id, username, request_type_id, date_from, date_to, value_amount, notes, cdc, attachment_path, tratte, extra_data, status, created_ts, updated_ts)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'pending', %s, %s)
        """, (0, username, request_type_id, date_start, date_end, value_amount, notes, cdc, attachment_path, tratte_json, extra_data_json, now, now))
    else:
        db.execute("""
            INSERT INTO user_requests (user_id, username, request_type_id, date_from, date_to, value_amount, notes, cdc, attachment_path, tratte, extra_data, status, created_ts, updated_ts)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?, ?)
        """, (0, username, request_type_id, date_start, date_end, value_amount, notes, cdc, attachment_path, tratte_json, extra_data_json, now, now))
    
    db.commit()

    return jsonify({"ok": True, "message": "Richiesta inviata con successo"})


# Route per servire gli allegati delle richieste
@app.route('/uploads/requests/<path:filename>')
@login_required
def serve_request_attachment(filename):
    """Serve gli allegati delle richieste (solo per l'utente proprietario o admin)."""
    # Verifica accesso: l'utente può vedere solo i propri file, admin può vedere tutti
    username = session.get("user", "")
    is_admin = session.get("is_admin", False)
    
    # Estrai username dal path
    parts = filename.split('/')
    if len(parts) >= 1:
        file_owner = parts[0]
        if not is_admin and file_owner != username:
            return jsonify({"error": "Accesso negato"}), 403
    
    return send_from_directory(
        os.path.join(app.root_path, 'uploads', 'requests'),
        filename
    )


@app.route("/uploads/documents/<path:filename>")
@login_required
def serve_document_file(filename):
    """Serve i file dei documenti aziendali."""
    return send_from_directory(
        os.path.join(app.root_path, 'uploads', 'documents'),
        filename
    )


# ═══════════════════════════════════════════════════════════════════════════════
# OVERTIME (STRAORDINARI) - API
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/overtime")
@login_required
def admin_overtime_page() -> ResponseReturnValue:
    """Pagina admin per gestire le richieste di straordinario."""
    if not session.get("is_admin"):
        return redirect(url_for("home"))
    
    # Verifica se il modulo straordinari è attivo
    db = get_db()
    if not is_module_enabled(db, "straordinari"):
        flash("Modulo straordinari non attivo", "warning")
        return redirect(url_for("admin_home"))
    
    username = session.get("user", "")
    initials = "".join([p[0].upper() for p in username.split()[:2]]) if username else "A"
    
    return render_template(
        "admin_overtime.html",
        user_name=username,
        user_initials=initials
    )


@app.route("/user/overtime")
@login_required
def user_overtime_page() -> ResponseReturnValue:
    """Pagina utente per visualizzare i propri straordinari."""
    # Verifica se il modulo straordinari è attivo
    db = get_db()
    if not is_module_enabled(db, "straordinari"):
        flash("Modulo straordinari non attivo", "warning")
        return redirect(url_for("user_home"))
    
    return render_template(
        "user_overtime.html",
        username=session.get("user"),
        user_display=session.get("user_display", session.get("user")),
        user_initials=session.get("user_initials", "U"),
        user_role=session.get("user_role", "Utente"),
    )


@app.get("/api/admin/overtime")
@login_required
def api_admin_overtime_list() -> ResponseReturnValue:
    """
    Restituisce tutte le richieste di straordinario per l'admin.
    Legge dalla tabella user_requests filtrando per tipo 'Straordinario'.
    """
    if not session.get("is_admin"):
        return jsonify({"error": "Accesso negato"}), 403
    
    db = get_db()
    
    # Verifica se il modulo straordinari è attivo
    if not is_module_enabled(db, "straordinari"):
        return jsonify({"error": "Modulo straordinari non attivo", "overtime": []}), 200
    
    ensure_user_requests_table(db)
    
    # Ottieni l'ID del tipo "Straordinario"
    overtime_type_id = get_overtime_request_type_id(db)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    status_filter = request.args.get("status")
    username_filter = request.args.get("username")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    
    query = f"""
        SELECT ur.id, ur.username, ur.date_from as date, ur.value_amount as total_extra_minutes,
               ur.notes, ur.extra_data, ur.status, ur.reviewed_by, ur.reviewed_ts, ur.review_notes,
               ur.created_ts, ur.updated_ts,
               au.display_name, au.full_name
        FROM user_requests ur
        LEFT JOIN app_users au ON ur.username = au.username
        WHERE ur.request_type_id = {placeholder}
    """
    params = [overtime_type_id]
    
    if status_filter:
        query += f" AND ur.status = {placeholder}"
        params.append(status_filter)
    if username_filter:
        query += f" AND ur.username = {placeholder}"
        params.append(username_filter)
    if date_from:
        query += f" AND ur.date_from >= {placeholder}"
        params.append(date_from)
    if date_to:
        query += f" AND ur.date_from <= {placeholder}"
        params.append(date_to)
    
    query += " ORDER BY CASE ur.status WHEN 'pending' THEN 0 ELSE 1 END, ur.created_ts DESC"
    
    rows = db.execute(query, tuple(params)).fetchall()
    
    overtime_list = []
    for row in rows:
        if isinstance(row, Mapping):
            item = dict(row)
        else:
            columns = ['id', 'username', 'date', 'total_extra_minutes', 'notes', 'extra_data',
                      'status', 'reviewed_by', 'reviewed_ts', 'review_notes', 'created_ts', 'updated_ts',
                      'display_name', 'full_name']
            item = dict(zip(columns, row))
        
        # Parse extra_data per estrarre i dettagli dello straordinario
        extra_data = item.get('extra_data')
        if extra_data:
            try:
                if isinstance(extra_data, str):
                    extra_data = json.loads(extra_data)
                item.update({
                    'session_id': extra_data.get('session_id'),
                    'planning_id': extra_data.get('planning_id'),
                    'shift_source': extra_data.get('shift_source', 'none'),
                    'planned_start': extra_data.get('planned_start'),
                    'planned_end': extra_data.get('planned_end'),
                    'actual_start': extra_data.get('actual_start'),
                    'actual_end': extra_data.get('actual_end'),
                    'extra_minutes_before': extra_data.get('extra_minutes_before', 0),
                    'extra_minutes_after': extra_data.get('extra_minutes_after', 0),
                    'overtime_type': extra_data.get('overtime_type', 'after_shift'),
                    'auto_detected': extra_data.get('auto_detected', False)
                })
            except (json.JSONDecodeError, TypeError):
                pass
        
        # Rimuovi extra_data raw dalla risposta
        item.pop('extra_data', None)
        
        # Converti total_extra_minutes in intero
        item['total_extra_minutes'] = int(item.get('total_extra_minutes', 0))
        
        overtime_list.append(item)
    
    return jsonify({"overtime": overtime_list})


@app.get("/api/user/overtime")
@login_required
def api_user_overtime_list() -> ResponseReturnValue:
    """
    Restituisce gli straordinari dell'utente corrente.
    Legge dalla tabella user_requests filtrando per tipo 'Straordinario'.
    """
    username = session.get("user")
    if not username:
        return jsonify({"error": "Non autenticato"}), 401
    
    db = get_db()
    
    # Verifica se il modulo straordinari è attivo
    if not is_module_enabled(db, "straordinari"):
        return jsonify({"error": "Modulo straordinari non attivo", "overtime": []}), 200
    
    ensure_user_requests_table(db)
    
    # Ottieni l'ID del tipo "Straordinario"
    overtime_type_id = get_overtime_request_type_id(db)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    rows = db.execute(f"""
        SELECT ur.id, ur.username, ur.date_from as date, ur.value_amount as total_extra_minutes,
               ur.notes, ur.extra_data, ur.status, ur.reviewed_by, ur.reviewed_ts, ur.review_notes,
               ur.created_ts, ur.updated_ts
        FROM user_requests ur
        WHERE ur.username = {placeholder} AND ur.request_type_id = {placeholder}
        ORDER BY ur.date_from DESC, ur.created_ts DESC
    """, (username, overtime_type_id)).fetchall()
    
    overtime_list = []
    for row in rows:
        if isinstance(row, Mapping):
            item = dict(row)
        else:
            columns = ['id', 'username', 'date', 'total_extra_minutes', 'notes', 'extra_data',
                      'status', 'reviewed_by', 'reviewed_ts', 'review_notes', 'created_ts', 'updated_ts']
            item = dict(zip(columns, row))
        
        # Parse extra_data per estrarre i dettagli dello straordinario
        extra_data = item.get('extra_data')
        if extra_data:
            try:
                if isinstance(extra_data, str):
                    extra_data = json.loads(extra_data)
                item.update({
                    'session_id': extra_data.get('session_id'),
                    'planning_id': extra_data.get('planning_id'),
                    'shift_source': extra_data.get('shift_source', 'none'),
                    'planned_start': extra_data.get('planned_start'),
                    'planned_end': extra_data.get('planned_end'),
                    'actual_start': extra_data.get('actual_start'),
                    'actual_end': extra_data.get('actual_end'),
                    'extra_minutes_before': extra_data.get('extra_minutes_before', 0),
                    'extra_minutes_after': extra_data.get('extra_minutes_after', 0),
                    'overtime_type': extra_data.get('overtime_type', 'after_shift'),
                    'auto_detected': extra_data.get('auto_detected', False)
                })
            except (json.JSONDecodeError, TypeError):
                pass
        
        # Rimuovi extra_data raw dalla risposta
        item.pop('extra_data', None)
        
        # Converti total_extra_minutes in intero
        item['total_extra_minutes'] = int(item.get('total_extra_minutes', 0))
        
        overtime_list.append(item)
    
    return jsonify({"overtime": overtime_list})


@app.post("/api/user/overtime")
@login_required
def api_user_overtime_create() -> ResponseReturnValue:
    """
    Crea una nuova richiesta di straordinario come user_request.
    Usa la tabella user_requests con type='Straordinario' e extra_data per i dettagli.
    """
    username = session.get("user")
    if not username:
        return jsonify({"error": "Non autenticato"}), 401
    
    db = get_db()
    
    # Verifica se il modulo straordinari è attivo
    if not is_module_enabled(db, "straordinari"):
        return jsonify({"error": "Modulo straordinari non attivo"}), 400
    
    data = request.get_json() or {}
    
    required = ["date", "total_extra_minutes"]
    for field in required:
        if not data.get(field):
            return jsonify({"error": f"Campo obbligatorio mancante: {field}"}), 400
    
    ensure_user_requests_table(db)
    
    # Ottieni l'ID del tipo "Straordinario"
    overtime_type_id = get_overtime_request_type_id(db)
    
    now_ts = int(time.time() * 1000)
    
    # Prepara extra_data con tutti i dettagli dello straordinario
    extra_data = {
        "session_id": data.get("session_id"),
        "planning_id": data.get("planning_id"),
        "shift_source": data.get("shift_source", "none"),
        "planned_start": data.get("planned_start"),
        "planned_end": data.get("planned_end"),
        "actual_start": data.get("actual_start"),
        "actual_end": data.get("actual_end"),
        "rounded_start": data.get("rounded_start"),
        "rounded_end": data.get("rounded_end"),
        "extra_minutes_before": data.get("extra_minutes_before", 0),
        "extra_minutes_after": data.get("extra_minutes_after", 0),
        "overtime_type": data.get("overtime_type", "after_shift"),
        "auto_detected": data.get("auto_detected", False)
    }
    extra_data_json = json.dumps(extra_data)
    
    date_str = data.get("date")
    total_minutes = data.get("total_extra_minutes")
    notes = data.get("notes", "")
    
    if DB_VENDOR == "mysql":
        db.execute("""
            INSERT INTO user_requests 
            (user_id, username, request_type_id, date_from, date_to, value_amount, 
             notes, cdc, attachment_path, tratte, extra_data, status, created_ts, updated_ts)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'pending', %s, %s)
        """, (0, username, overtime_type_id, date_str, date_str, total_minutes, 
              notes, None, None, None, extra_data_json, now_ts, now_ts))
    else:
        db.execute("""
            INSERT INTO user_requests 
            (user_id, username, request_type_id, date_from, date_to, value_amount, 
             notes, cdc, attachment_path, tratte, extra_data, status, created_ts, updated_ts)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?, ?)
        """, (0, username, overtime_type_id, date_str, date_str, total_minutes, 
              notes, None, None, None, extra_data_json, now_ts, now_ts))
    
    db.commit()
    
    # Notifica admin
    _send_overtime_notification_to_admins(db, username, date_str, total_minutes)
    
    return jsonify({"ok": True, "message": "Richiesta Extra Turno inviata"})


def _process_fuori_flessibilita(
    db: DatabaseLike,
    request_id: int,
    username: str,
    date_from: str,
    status: str,
    rounded_time: str,
    extra_data_str: str = None
) -> dict:
    """
    Elabora una richiesta Fuori Flessibilità approvata o respinta.
    
    - Se APPROVATA: aggiorna la timbratura con l'orario arrotondato sull'ora reale
    - Se RESPINTA: mantiene l'orario calcolato sul turno (già salvato)
    
    Args:
        db: connessione database
        request_id: ID della richiesta
        username: username dell'utente
        date_from: data della timbratura
        status: 'approved' o 'rejected'
        rounded_time: orario confermato (per approvate) o orario turno (per respinte)
        extra_data_str: JSON con i dettagli della richiesta
    
    Returns:
        Dict con info di debug
    """
    result = {"request_id": request_id, "status": status, "updated": False}
    
    try:
        # Parsing extra_data
        extra_data = {}
        if extra_data_str:
            try:
                extra_data = json.loads(extra_data_str) if isinstance(extra_data_str, str) else extra_data_str
            except:
                pass
        
        tipo_timbratura = extra_data.get("tipo_timbratura", "fine_giornata")
        ora_timbrata = extra_data.get("ora_timbrata")
        ora_mod_originale = extra_data.get("ora_mod")
        
        app.logger.info(
            "Fuori Flessibilità: processing request %s, status=%s, tipo=%s, rounded_time=%s, ora_mod_originale=%s",
            request_id, status, tipo_timbratura, rounded_time, ora_mod_originale
        )
        
        # Determina quale ora usare per l'aggiornamento
        if status == 'approved':
            # APPROVATA: usa l'orario confermato dall'admin (basato sull'ora reale arrotondata)
            ora_da_usare = rounded_time
        else:
            # RESPINTA: usa l'orario calcolato sul turno (già presente in ora_mod)
            ora_da_usare = ora_mod_originale or rounded_time
        
        if not ora_da_usare:
            app.logger.warning("Fuori Flessibilità: nessun orario disponibile per request %s", request_id)
            return {"request_id": request_id, "error": "Nessun orario disponibile"}
        
        # Costruisci il datetime completo
        ora_da_usare_clean = ora_da_usare if len(ora_da_usare) == 5 else ora_da_usare[:5]
        datetime_str = f"{date_from} {ora_da_usare_clean}:00"
        
        placeholder = "%s" if DB_VENDOR == "mysql" else "?"
        
        # Aggiorna extra_data con il rounded_time finale per visualizzazione nella card
        extra_data["ora_finale"] = ora_da_usare_clean
        extra_data_updated = json.dumps(extra_data)
        db.execute(f"""
            UPDATE user_requests SET extra_data = {placeholder} WHERE id = {placeholder}
        """, (extra_data_updated, request_id))
        
        # Aggiorna la tabella timbrature (per il dettaglio giornata)
        timbratura_updated = db.execute(f"""
            UPDATE timbrature 
            SET ora_mod = {placeholder}
            WHERE username = {placeholder} AND data = {placeholder} AND tipo = {placeholder}
        """, (ora_da_usare_clean, username, date_from, tipo_timbratura)).rowcount
        
        if timbratura_updated:
            app.logger.info(
                "Fuori Flessibilità: aggiornata timbrature per %s, data=%s, tipo=%s, ora_mod=%s",
                username, date_from, tipo_timbratura, ora_da_usare_clean
            )
            result["timbrature_updated"] = True
        
        # Trova la timbratura da aggiornare in cedolino_timbrature
        timbratura_row = db.execute(f"""
            SELECT id, ora_modificata, synced_ts 
            FROM cedolino_timbrature 
            WHERE username = {placeholder} AND data_riferimento = {placeholder}
            ORDER BY id DESC LIMIT 1
        """, (username, date_from)).fetchone()
        
        if timbratura_row:
            timbr_id = timbratura_row['id'] if isinstance(timbratura_row, Mapping) else timbratura_row[0]
            synced_ts = timbratura_row['synced_ts'] if isinstance(timbratura_row, Mapping) else timbratura_row[2]
            
            # Aggiorna ora_modificata
            db.execute(f"""
                UPDATE cedolino_timbrature 
                SET ora_modificata = {placeholder}, synced_ts = NULL
                WHERE id = {placeholder}
            """, (datetime_str, timbr_id))
            
            result["cedolino_timbr_id"] = timbr_id
            result["updated"] = True
            result["ora_finale"] = ora_da_usare_clean
            
            app.logger.info(
                "Fuori Flessibilità: aggiornata cedolino_timbrature id=%s, ora_modificata=%s",
                timbr_id, datetime_str
            )
            
            # Se approvata, invia subito a CedolinoWeb
            if status == 'approved':
                try:
                    cedolino_result = _resync_flex_to_cedolino(db, timbr_id, datetime_str)
                    result["cedolino_sync"] = cedolino_result
                except Exception as e:
                    app.logger.error(f"Errore sync CedolinoWeb per flex: {e}")
                    result["cedolino_sync_error"] = str(e)
        
        db.commit()
        return result
        
    except Exception as e:
        import traceback
        app.logger.error(f"Errore in _process_fuori_flessibilita: {e}\n{traceback.format_exc()}")
        return {"request_id": request_id, "error": str(e)}


def _resync_flex_to_cedolino(db: DatabaseLike, timbr_id: int, new_ora_modificata: str) -> dict:
    """
    Invia a CedolinoWeb la timbratura aggiornata per Fuori Flessibilità.
    Usa call_cedolino_webservice con tutti i parametri corretti.
    
    Args:
        timbr_id: ID della riga in cedolino_timbrature
        new_ora_modificata: Nuovo orario modificato (HH:MM o HH:MM:SS)
    """
    settings = get_cedolino_settings()
    if not settings:
        return {"error": "CedolinoWeb non configurato"}
    
    endpoint = settings.get("endpoint") or CEDOLINO_WEB_ENDPOINT
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Recupera i dati della timbratura
    row = db.execute(f"""
        SELECT external_id, timeframe_id, data_riferimento, ora_originale, username
        FROM cedolino_timbrature 
        WHERE id = {placeholder}
    """, (timbr_id,)).fetchone()
    
    if not row:
        return {"error": "Timbratura non trovata"}
    
    if isinstance(row, Mapping):
        external_id = row.get("external_id")
        timeframe_id = row.get("timeframe_id")
        data_rif = row.get("data_riferimento")
        ora_originale = row.get("ora_originale")
        username = row.get("username")
    else:
        external_id = row[0]
        timeframe_id = row[1]
        data_rif = row[2]
        ora_originale = row[3]
        username = row[4]
    
    # Converti data_riferimento in stringa se è un oggetto date
    if hasattr(data_rif, 'strftime'):
        data_rif = data_rif.strftime("%Y-%m-%d")
    else:
        data_rif = str(data_rif)
    
    # Converti ora_originale se è timedelta
    if isinstance(ora_originale, timedelta):
        total_secs = int(ora_originale.total_seconds())
        h, remainder = divmod(total_secs, 3600)
        m, s = divmod(remainder, 60)
        ora_originale = f"{h:02d}:{m:02d}:{s:02d}"
    else:
        ora_originale = str(ora_originale)
    
    # Normalizza new_ora_modificata - potrebbe essere solo HH:MM o già con data
    new_ora_modificata = str(new_ora_modificata)
    
    # Se contiene già la data (es. "2026-01-17 19:00:00"), estrai solo l'ora
    if " " in new_ora_modificata and len(new_ora_modificata) > 10:
        # Formato "YYYY-MM-DD HH:MM:SS" - prendi solo l'ora
        new_ora_modificata = new_ora_modificata.split(" ")[1]
    
    # Aggiungi secondi se mancano (HH:MM -> HH:MM:00)
    if len(new_ora_modificata) == 5:  # HH:MM
        new_ora_modificata = f"{new_ora_modificata}:00"
    
    # Costruisci data_originale e data_modificata complete
    data_originale = f"{data_rif} {ora_originale}"
    data_modificata = f"{data_rif} {new_ora_modificata}"
    
    # Recupera gruppo_id esterno
    external_group_id = get_external_group_id_for_username(db, username)
    
    app.logger.info(
        "CedolinoWeb Flex resync: timbr_id=%s, external_id=%s, timeframe=%s, data_rif=%s, ora_orig=%s, ora_mod=%s, gruppo=%s",
        timbr_id, external_id, timeframe_id, data_rif, ora_originale, new_ora_modificata, external_group_id
    )
    
    # Chiama il webservice con tutti i parametri
    success, error, request_url = call_cedolino_webservice(
        external_id, timeframe_id, data_rif, data_originale, data_modificata, endpoint, external_group_id
    )
    
    now_ts = int(datetime.now().timestamp() * 1000)
    
    if success:
        # Aggiorna synced_ts e resetta sync_error e overtime_request_id
        db.execute(f"""
            UPDATE cedolino_timbrature 
            SET synced_ts = {placeholder},
                sync_error = NULL, overtime_request_id = NULL
            WHERE id = {placeholder}
        """, (now_ts, timbr_id))
        db.commit()
        
        app.logger.info(f"CedolinoWeb Flex sync OK: timbr_id={timbr_id}")
        
        return {
            "success": True,
            "url": request_url
        }
    else:
        app.logger.error(f"CedolinoWeb Flex sync FAILED: timbr_id={timbr_id}, error={error}")
        return {"error": error, "url": request_url}


def _sync_overtime_blocked_timbrature(
    db: DatabaseLike, 
    overtime_request_id: int,
    request_status: str = None,
    extra_data_passed: str = None
) -> dict:
    """
    Sincronizza le timbrature che erano bloccate in attesa della revisione di uno straordinario.
    Chiamata dopo che una richiesta di straordinario viene approvata o respinta.
    
    - Se APPROVATA: invia la timbrata con l'orario arrotondato (Extra Turno)
    - Se RESPINTA: invia la timbrata con l'orario del turno pianificato (no Extra Turno)
    
    Args:
        db: connessione database
        overtime_request_id: ID della richiesta straordinario revisionata
        request_status: status della richiesta (passed from caller to avoid re-query)
        extra_data_passed: extra_data della richiesta (passed from caller)
    
    Returns:
        Dict con info di debug sulla sincronizzazione
    """
    app.logger.warning(
        "DEBUG SYNC: INIZIO _sync_overtime_blocked_timbrature - overtime_request_id=%s, request_status=%r, extra_data_passed=%r",
        overtime_request_id, request_status, extra_data_passed[:200] if extra_data_passed else None
    )
    
    settings = get_cedolino_settings()
    if not settings:
        return {"synced_count": 0, "error": "CedolinoWeb non configurato"}
    
    endpoint = settings.get("endpoint") or CEDOLINO_WEB_ENDPOINT
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Usa i dati passati dal chiamante oppure fai una query
    if request_status is None or extra_data_passed is None:
        request_row = db.execute(f"""
            SELECT status, extra_data FROM user_requests WHERE id = {placeholder}
        """, (overtime_request_id,)).fetchone()
        
        if not request_row:
            return {"synced_count": 0, "error": "Richiesta non trovata"}
        
        if isinstance(request_row, Mapping):
            request_status = request_row.get("status")
            extra_data_str = request_row.get("extra_data")
        else:
            request_status = request_row[0]
            extra_data_str = request_row[1]
    else:
        extra_data_str = extra_data_passed
    
    app.logger.info(
        "CedolinoWeb _sync_overtime: overtime_request_id=%s, request_status=%s",
        overtime_request_id, request_status
    )
    
    # Estrai i dati del turno pianificato
    planned_start = None
    planned_end = None
    extra_type = None
    
    if extra_data_str:
        try:
            extra_data = json.loads(extra_data_str) if isinstance(extra_data_str, str) else extra_data_str
            planned_start = extra_data.get("planned_start")
            planned_end = extra_data.get("planned_end")
            extra_type = extra_data.get("extra_type")
            app.logger.info(
                "CedolinoWeb: overtime %s - status=%s, extra_type=%s, planned_start=%s, planned_end=%s",
                overtime_request_id, request_status, extra_type, planned_start, planned_end
            )
        except Exception as e:
            app.logger.error("CedolinoWeb: errore parsing extra_data per overtime %s: %s", overtime_request_id, e)
    
    # Recupera le timbrature bloccate per questo straordinario
    # Nota: includiamo anche quelle già sincronizzate se stiamo respingendo (per correggerle)
    if request_status == 'rejected':
        # Per le respinte, recupera TUTTE le timbrature associate, anche quelle già sincronizzate
        rows = db.execute(f"""
            SELECT id, external_id, timeframe_id, data_riferimento, ora_originale, ora_modificata, username, synced_ts
            FROM cedolino_timbrature
            WHERE overtime_request_id = {placeholder}
        """, (overtime_request_id,)).fetchall()
        app.logger.info("CedolinoWeb: REJECTED - cercando tutte le timbrature per overtime_request_id=%s, trovate %d", 
                       overtime_request_id, len(rows) if rows else 0)
    else:
        # Per le approvate, solo quelle non ancora sincronizzate
        rows = db.execute(f"""
            SELECT id, external_id, timeframe_id, data_riferimento, ora_originale, ora_modificata, username, synced_ts
            FROM cedolino_timbrature
            WHERE overtime_request_id = {placeholder} AND synced_ts IS NULL
        """, (overtime_request_id,)).fetchall()
        app.logger.info("CedolinoWeb: APPROVED - cercando timbrature non sincronizzate per overtime_request_id=%s, trovate %d", 
                       overtime_request_id, len(rows) if rows else 0)
    
    if not rows:
        app.logger.warning("CedolinoWeb: nessuna timbratura trovata per overtime_request_id=%s", overtime_request_id)
        return {"synced_count": 0, "error": "Nessuna timbratura trovata"}
    
    synced_count = 0
    for row in rows:
        if isinstance(row, Mapping):
            timbrata_id = row.get("id")
            external_id = row.get("external_id")
            timeframe_id = row.get("timeframe_id")
            data_rif = row.get("data_riferimento")
            ora_orig = row.get("ora_originale")
            ora_mod = row.get("ora_modificata")
            username = row.get("username")
        else:
            timbrata_id = row[0]
            external_id = row[1]
            timeframe_id = row[2]
            data_rif = row[3]
            ora_orig = row[4]
            ora_mod = row[5]
            username = row[6]
        
        # Formatta data_riferimento come stringa se necessario
        if hasattr(data_rif, 'strftime'):
            data_riferimento = data_rif.strftime("%Y-%m-%d")
        else:
            data_riferimento = str(data_rif)
        
        # Formatta ora come stringa se necessario
        if hasattr(ora_orig, 'strftime'):
            ora_originale = ora_orig.strftime("%H:%M:%S")
        else:
            ora_originale = str(ora_orig)
        
        if hasattr(ora_mod, 'strftime'):
            ora_modificata = ora_mod.strftime("%H:%M:%S")
        else:
            ora_modificata = str(ora_mod) if ora_mod else ora_originale
        
        # Se RESPINTA: usa l'orario del turno pianificato invece dell'extra
        ora_modificata_originale = ora_modificata  # Salva il valore originale per il log
        if request_status == 'rejected':
            app.logger.info(
                "CedolinoWeb: Richiesta RESPINTA - extra_type=%s, planned_start=%s, planned_end=%s, ora_mod_attuale=%s",
                extra_type, planned_start, planned_end, ora_modificata
            )
            # Determina quale orario usare in base al tipo di extra turno
            if extra_type == 'before_shift' and planned_start:
                # Ingresso anticipato respinto: usa l'orario di inizio turno
                ora_modificata = f"{planned_start}:00" if len(planned_start) == 5 else planned_start
                app.logger.info(
                    "CedolinoWeb: Extra Turno RESPINTO (before_shift) - uso orario turno %s invece di %s",
                    ora_modificata, ora_modificata_originale
                )
            elif extra_type == 'after_shift' and planned_end:
                # Uscita posticipata respinta: usa l'orario di fine turno
                ora_modificata = f"{planned_end}:00" if len(planned_end) == 5 else planned_end
                app.logger.info(
                    "CedolinoWeb: Extra Turno RESPINTO (after_shift) - uso orario turno %s invece di %s",
                    ora_modificata, ora_modificata_originale
                )
            else:
                app.logger.warning(
                    "CedolinoWeb: Extra Turno RESPINTO ma mancano dati - extra_type=%s, planned_start=%s, planned_end=%s",
                    extra_type, planned_start, planned_end
                )
            
            # Aggiorna anche il record cedolino_timbrature con l'ora corretta
            if ora_modificata != ora_modificata_originale:
                if DB_VENDOR == "mysql":
                    db.execute(
                        "UPDATE cedolino_timbrature SET ora_modificata = %s WHERE id = %s",
                        (ora_modificata, timbrata_id)
                    )
                else:
                    db.execute(
                        "UPDATE cedolino_timbrature SET ora_modificata = ? WHERE id = ?",
                        (ora_modificata, timbrata_id)
                    )
                
                # Aggiorna anche la tabella timbrature per mostrare l'ora corretta nello storico
                # Converte timeframe_id in tipo timbratura
                tipo_map = {1: 'inizio_giornata', 4: 'inizio_pausa', 5: 'fine_pausa', 8: 'fine_giornata'}
                tipo_timbratura = tipo_map.get(timeframe_id)
                
                if tipo_timbratura and username:
                    # Solo ore e minuti per ora_mod
                    ora_mod_short = ora_modificata[:5] if len(ora_modificata) >= 5 else ora_modificata
                    if DB_VENDOR == "mysql":
                        db.execute(
                            """UPDATE timbrature SET ora_mod = %s 
                               WHERE username = %s AND data = %s AND tipo = %s""",
                            (ora_mod_short, username, data_riferimento, tipo_timbratura)
                        )
                    else:
                        db.execute(
                            """UPDATE timbrature SET ora_mod = ? 
                               WHERE username = ? AND data = ? AND tipo = ?""",
                            (ora_mod_short, username, data_riferimento, tipo_timbratura)
                        )
                    app.logger.info(
                        "CedolinoWeb: aggiornato timbrature.ora_mod=%s per user=%s, data=%s, tipo=%s",
                        ora_mod_short, username, data_riferimento, tipo_timbratura
                    )
        
        # Componi data_originale e data_modificata
        data_originale = f"{data_riferimento} {ora_originale}"
        data_modificata = f"{data_riferimento} {ora_modificata}"
        
        app.logger.warning(
            "DEBUG SYNC: PRIMA DI CALL WEBSERVICE - data_originale=%s, data_modificata=%s, ora_originale=%s, ora_modificata=%s",
            data_originale, data_modificata, ora_originale, ora_modificata
        )
        
        # Recupera external_group_id se possibile
        external_group_id = get_external_group_id_for_username(db, username) if username else None
        
        app.logger.info(
            "CedolinoWeb: sincronizzazione timbrata %s bloccata per straordinario %s (status=%s)",
            timbrata_id, overtime_request_id, request_status
        )
        
        success, error, _url = call_cedolino_webservice(
            external_id, timeframe_id, data_riferimento, data_originale, data_modificata, endpoint, external_group_id
        )
        
        if success:
            if DB_VENDOR == "mysql":
                db.execute(
                    "UPDATE cedolino_timbrature SET synced_ts = %s, sync_error = NULL WHERE id = %s",
                    (now_ms(), timbrata_id)
                )
            else:
                db.execute(
                    "UPDATE cedolino_timbrature SET synced_ts = ?, sync_error = NULL WHERE id = ?",
                    (now_ms(), timbrata_id)
                )
            synced_count += 1
            app.logger.info("CedolinoWeb: timbrata %s sincronizzata con successo", timbrata_id)
        else:
            if DB_VENDOR == "mysql":
                db.execute(
                    "UPDATE cedolino_timbrature SET sync_error = %s, sync_attempts = sync_attempts + 1 WHERE id = %s",
                    (error, timbrata_id)
                )
            else:
                db.execute(
                    "UPDATE cedolino_timbrature SET sync_error = ?, sync_attempts = sync_attempts + 1 WHERE id = ?",
                    (error, timbrata_id)
                )
            app.logger.warning("CedolinoWeb: errore sincronizzazione timbrata %s: %s", timbrata_id, error)
    
    if rows:
        db.commit()
    
    # Ritorna dati di debug
    return {
        "synced_count": synced_count,
        "request_status": request_status,
        "extra_type": extra_type,
        "planned_start": planned_start,
        "planned_end": planned_end,
        "rows_found": len(rows) if rows else 0,
        "last_ora_modificata": ora_modificata if rows else None,
        "last_ora_originale": ora_originale if rows else None
    }


@app.put("/api/admin/overtime/<int:overtime_id>")
@login_required
def api_admin_overtime_review(overtime_id: int) -> ResponseReturnValue:
    """
    Approva o respinge una richiesta di straordinario.
    Aggiorna la tabella user_requests.
    """
    if not session.get("is_admin"):
        return jsonify({"error": "Accesso negato"}), 403
    
    data = request.get_json() or {}
    status = data.get("status")
    review_notes = data.get("review_notes", "").strip()
    
    if status not in ("approved", "rejected"):
        return jsonify({"error": "Stato non valido. Usa 'approved' o 'rejected'"}), 400
    
    db = get_db()
    ensure_user_requests_table(db)
    
    # Ottieni l'ID del tipo "Straordinario"
    overtime_type_id = get_overtime_request_type_id(db)
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    # Verifica che esista e sia pending
    existing = db.execute(f"""
        SELECT id, status, username, date_from, value_amount 
        FROM user_requests 
        WHERE id = {placeholder} AND request_type_id = {placeholder}
    """, (overtime_id, overtime_type_id)).fetchone()
    
    if not existing:
        return jsonify({"error": "Richiesta non trovata"}), 404
    
    if isinstance(existing, Mapping):
        current_status = existing.get("status")
        target_username = existing.get("username")
        ot_date = existing.get("date_from")
        ot_minutes = int(existing.get("value_amount", 0))
    else:
        current_status = existing[1]
        target_username = existing[2]
        ot_date = existing[3]
        ot_minutes = int(existing[4] or 0)
    
    if current_status != "pending":
        return jsonify({"error": "Questa richiesta è già stata gestita"}), 400
    
    now_ts = int(time.time() * 1000)
    admin_user = session.get("user", "admin")
    
    db.execute(f"""
        UPDATE user_requests 
        SET status = {placeholder}, reviewed_by = {placeholder}, reviewed_ts = {placeholder}, 
            review_notes = {placeholder}, updated_ts = {placeholder}
        WHERE id = {placeholder}
    """, (status, admin_user, now_ts, review_notes, now_ts, overtime_id))
    
    db.commit()
    
    # Sincronizza le timbrature bloccate per questo straordinario
    _sync_overtime_blocked_timbrature(db, overtime_id)
    
    # Notifica utente
    _send_overtime_review_notification(db, target_username, ot_date, ot_minutes, status, review_notes)
    
    status_label = "approvata" if status == "approved" else "respinta"
    return jsonify({"ok": True, "message": f"Richiesta Extra Turno {status_label}"})


@app.get("/api/user/check-overtime")
@login_required
def api_user_check_overtime() -> ResponseReturnValue:
    """
    Verifica se c'è uno straordinario da dichiarare basandosi sul turno pianificato.
    Chiamato al momento del checkout.
    """
    username = session.get("user")
    if not username:
        return jsonify({"error": "Non autenticato"}), 401
    
    # Parametri dalla richiesta
    session_id = request.args.get("session_id", type=int)
    actual_start = request.args.get("actual_start")  # HH:MM
    actual_end = request.args.get("actual_end")  # HH:MM
    date_str = request.args.get("date")  # YYYY-MM-DD
    
    if not all([actual_start, actual_end, date_str]):
        return jsonify({"overtime_detected": False, "reason": "Parametri mancanti"})
    
    db = get_db()
    
    # Verifica se il modulo straordinari è attivo
    if not is_module_enabled(db, "straordinari"):
        return jsonify({"overtime_detected": False, "reason": "Modulo straordinari non attivo"})
    
    # Carica impostazioni straordinari
    settings = get_company_settings(db)
    overtime_settings = settings.get("custom_settings", {}).get("overtime", {})
    
    if not overtime_settings.get("auto_detect", True):
        return jsonify({"overtime_detected": False, "reason": "Rilevamento automatico disabilitato"})
    
    threshold_minutes = overtime_settings.get("threshold_minutes", 15)
    rounding_minutes = overtime_settings.get("rounding_minutes", 15)
    
    # Cerca turno pianificato per questo utente in questa data
    planned_shift = _find_planned_shift(db, username, date_str)
    
    if not planned_shift:
        return jsonify({
            "overtime_detected": False, 
            "reason": "Nessun turno pianificato trovato",
            "can_request_manual": True  # L'utente può comunque richiedere manualmente
        })
    
    # Calcola differenza
    planned_start = planned_shift.get("start_time")  # HH:MM
    planned_end = planned_shift.get("end_time")  # HH:MM
    shift_source = planned_shift.get("source")  # 'rentman' o 'manual'
    planning_id = planned_shift.get("planning_id")
    
    # Converti in minuti dal mezzanotte per il calcolo
    def time_to_minutes(t):
        if not t:
            return 0
        parts = str(t).split(":")
        return int(parts[0]) * 60 + int(parts[1])
    
    planned_start_min = time_to_minutes(planned_start)
    planned_end_min = time_to_minutes(planned_end)
    actual_start_min = time_to_minutes(actual_start)
    actual_end_min = time_to_minutes(actual_end)
    
    # Calcola straordinari (valori effettivi non arrotondati)
    extra_before_raw = max(0, planned_start_min - actual_start_min)  # Arrivato prima
    extra_after_raw = max(0, actual_end_min - planned_end_min)  # Uscito dopo
    total_extra_raw = extra_before_raw + extra_after_raw
    
    # Applica arrotondamento
    extra_before = extra_before_raw
    extra_after = extra_after_raw
    total_extra = total_extra_raw
    
    if rounding_minutes > 1:
        total_extra = (total_extra_raw // rounding_minutes) * rounding_minutes
        extra_before = (extra_before_raw // rounding_minutes) * rounding_minutes
        extra_after = (extra_after_raw // rounding_minutes) * rounding_minutes
    
    if total_extra < threshold_minutes:
        return jsonify({
            "overtime_detected": False,
            "reason": f"Differenza ({total_extra} min) sotto la soglia ({threshold_minutes} min)",
            "extra_minutes": total_extra
        })
    
    # Determina tipo
    if extra_before > 0 and extra_after > 0:
        overtime_type = "both"
    elif extra_before > 0:
        overtime_type = "before_shift"
    else:
        overtime_type = "after_shift"
    
    # Calcola orari arrotondati da mostrare nel modal
    # Stessa logica usata per ora_mod nelle timbrature:
    # - INIZIO giornata: arrotondamento PER ECCESSO (es. 7:27 → 7:30)
    # - FINE giornata: arrotondamento PER DIFETTO (es. 17:12 → 17:00)
    def minutes_to_time(m):
        return f"{m // 60:02d}:{m % 60:02d}"
    
    # Calcola ora arrotondata inizio (sempre arrotondamento PER ECCESSO)
    if rounding_minutes > 1:
        # Arrotonda per eccesso: ceil(actual_start_min / rounding) * rounding
        rounded_start_min = ((actual_start_min + rounding_minutes - 1) // rounding_minutes) * rounding_minutes
        rounded_start = minutes_to_time(rounded_start_min)
    else:
        rounded_start = actual_start
    
    # Calcola ora arrotondata fine (sempre arrotondamento PER DIFETTO)
    if rounding_minutes > 1:
        # Arrotonda per difetto: floor(actual_end_min / rounding) * rounding
        rounded_end_min = (actual_end_min // rounding_minutes) * rounding_minutes
        rounded_end = minutes_to_time(rounded_end_min)
    else:
        rounded_end = actual_end
    
    return jsonify({
        "overtime_detected": True,
        "session_id": session_id,
        "planning_id": planning_id,
        "shift_source": shift_source,
        "date": date_str,
        "planned_start": planned_start,
        "planned_end": planned_end,
        "actual_start": actual_start,
        "actual_end": actual_end,
        "rounded_start": rounded_start,
        "rounded_end": rounded_end,
        "extra_minutes_before": extra_before,
        "extra_minutes_after": extra_after,
        "extra_minutes_before_raw": extra_before_raw,
        "extra_minutes_after_raw": extra_after_raw,
        "total_extra_minutes": total_extra,
        "total_extra_minutes_raw": total_extra_raw,
        "rounding_minutes": rounding_minutes,
        "overtime_type": overtime_type,
        "require_approval": overtime_settings.get("require_approval", True)
    })


def _find_planned_shift(db, username: str, date_str: str) -> dict | None:
    """
    Cerca il turno pianificato per un utente in una data specifica.
    Controlla sia Rentman Planning (via crew_id) che turni employee_shifts.
    
    IMPORTANTE: Se ci sono più turni nella stessa giornata (es. 07:30-15:30 e 15:30-16:30),
    restituisce l'intervallo complessivo (MIN start, MAX end) per calcolare
    correttamente gli straordinari.
    """
    placeholder = "%s" if DB_VENDOR == "mysql" else "?"
    
    def _extract_time_hhmm(dt_value) -> str | None:
        """Estrae l'ora in formato HH:MM da un valore datetime o stringa."""
        if not dt_value:
            return None
        if hasattr(dt_value, 'strftime'):
            return dt_value.strftime("%H:%M")
        dt_str = str(dt_value)
        if " " in dt_str:
            return dt_str.split(" ")[-1][:5]
        return dt_str[:5]
    
    # Recupera il crew_id dell'utente per cercare i turni Rentman
    user_row = db.execute(
        f"SELECT rentman_crew_id FROM app_users WHERE username = {placeholder}",
        (username,)
    ).fetchone()
    
    crew_id = None
    if user_row:
        crew_id = user_row['rentman_crew_id'] if isinstance(user_row, dict) else user_row[0]
    
    # 1. Se l'utente ha un crew_id, cerca nei turni Rentman Planning
    if crew_id:
        try:
            # Usa i campi corretti della tabella: plan_start, plan_end, planning_date
            rentman_agg = db.execute(f"""
                SELECT COUNT(*) as shift_count,
                       MIN(plan_start) as first_start,
                       MAX(plan_end) as last_end,
                       GROUP_CONCAT(id) as planning_ids
                FROM rentman_plannings 
                WHERE crew_id = {placeholder} AND planning_date = {placeholder} AND sent_to_webservice = 1
            """, (crew_id, date_str)).fetchone()
            
            if rentman_agg:
                if isinstance(rentman_agg, Mapping):
                    shift_count = rentman_agg.get("shift_count", 0)
                    first_start = rentman_agg.get("first_start")
                    last_end = rentman_agg.get("last_end")
                    planning_ids = rentman_agg.get("planning_ids")
                else:
                    shift_count = rentman_agg[0] or 0
                    first_start = rentman_agg[1]
                    last_end = rentman_agg[2]
                    planning_ids = rentman_agg[3]
                
                if shift_count > 0 and first_start and last_end:
                    start_time = _extract_time_hhmm(first_start)
                    end_time = _extract_time_hhmm(last_end)
                    
                    # Usa il primo ID come riferimento (o tutti se multipli)
                    planning_id = planning_ids.split(",")[0] if planning_ids else None
                    
                    app.logger.info(
                        f"Turni Rentman trovati per {username} (crew_id={crew_id}) il {date_str}: "
                        f"{shift_count} turni, intervallo {start_time}-{end_time}"
                    )
                    
                    return {
                        "source": "rentman",
                        "planning_id": planning_id,
                        "start_time": start_time,
                        "end_time": end_time,
                        "shift_count": shift_count
                    }
        except Exception as e:
            app.logger.warning(f"Errore ricerca turni Rentman: {e}")
    
    # 2. Cerca nei turni employee_shifts (per impiegati senza Rentman)
    try:
        # Trova il giorno della settimana dalla data
        from datetime import datetime as dt
        date_obj = dt.strptime(date_str, "%Y-%m-%d")
        day_of_week = date_obj.weekday()  # 0=Lunedì, 6=Domenica
        
        employee_agg = db.execute(f"""
            SELECT COUNT(*) as shift_count,
                   MIN(start_time) as first_start,
                   MAX(end_time) as last_end,
                   GROUP_CONCAT(id) as shift_ids
            FROM employee_shifts 
            WHERE username = {placeholder} AND day_of_week = {placeholder} AND is_active = 1
        """, (username, day_of_week)).fetchone()
        
        if employee_agg:
            if isinstance(employee_agg, Mapping):
                shift_count = employee_agg.get("shift_count", 0)
                first_start = employee_agg.get("first_start")
                last_end = employee_agg.get("last_end")
                shift_ids = employee_agg.get("shift_ids")
            else:
                shift_count = employee_agg[0] or 0
                first_start = employee_agg[1]
                last_end = employee_agg[2]
                shift_ids = employee_agg[3]
            
            if shift_count > 0 and first_start and last_end:
                start_time = _extract_time_hhmm(first_start)
                end_time = _extract_time_hhmm(last_end)
                
                planning_id = shift_ids.split(",")[0] if shift_ids else None
                
                app.logger.info(
                    f"Turni impiegato trovati per {username} il {date_str} (giorno {day_of_week}): "
                    f"{shift_count} turni, intervallo {start_time}-{end_time}"
                )
                
                return {
                    "source": "employee_shifts",
                    "planning_id": planning_id,
                    "start_time": start_time,
                    "end_time": end_time,
                    "shift_count": shift_count
                }
    except Exception as e:
        app.logger.debug(f"Tabella employee_shifts non trovata o errore: {e}")
    
    app.logger.info(f"Nessun turno pianificato trovato per {username} il {date_str}")
    return None


def _send_overtime_notification_to_admins(db, username: str, date: str, minutes: int):
    """Invia notifica push agli admin per nuova richiesta Extra Turno."""
    try:
        hours = minutes // 60
        mins = minutes % 60
        time_str = f"{hours}h {mins}m" if hours > 0 else f"{mins} minuti"
        
        # Trova tutti gli admin
        if DB_VENDOR == "mysql":
            admins = db.execute("SELECT username FROM users WHERE is_admin = 1").fetchall()
        else:
            admins = db.execute("SELECT username FROM users WHERE is_admin = 1").fetchall()
        
        for admin in admins:
            admin_username = admin[0] if not isinstance(admin, Mapping) else admin.get("username")
            try:
                send_push_notification(
                    db,
                    admin_username,
                    "⏰ Nuova Richiesta Extra Turno",
                    f"{username} ha richiesto {time_str} di Extra Turno per {date}",
                    url="/admin/overtime",
                    tag=f"overtime-request-{username}-{date}"
                )
            except Exception as e:
                app.logger.warning(f"Errore invio notifica admin {admin_username}: {e}")
    except Exception as e:
        app.logger.error(f"Errore invio notifiche Extra Turno agli admin: {e}")


def _send_overtime_review_notification(db, username: str, date: str, minutes: int, status: str, notes: str):
    """Invia notifica push all'utente per esito richiesta Extra Turno."""
    try:
        hours = minutes // 60
        mins = minutes % 60
        time_str = f"{hours}h {mins}m" if hours > 0 else f"{mins} minuti"
        
        if status == "approved":
            title = "✅ Extra Turno Approvato"
            body = f"Il tuo Extra Turno di {time_str} del {date} è stato approvato"
        else:
            title = "❌ Extra Turno Rifiutato"
            body = f"Il tuo Extra Turno di {time_str} del {date} è stato rifiutato"
        
        if notes:
            body += f". Note: {notes}"
        
        send_push_notification(
            db,
            username,
            title,
            body,
            url="/user/requests",
            tag=f"overtime-review-{date}"
        )
        
        # Salva nel log notifiche
        now_ts = int(time.time() * 1000)
        if DB_VENDOR == "mysql":
            db.execute("""
                INSERT INTO push_notifications_log 
                (username, title, body, url, tag, sent_ts, event_type)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (username, title, body, "/user/requests", f"overtime-review-{date}", now_ts, "overtime_review"))
        else:
            db.execute("""
                INSERT INTO push_notifications_log 
                (username, title, body, url, tag, sent_ts, event_type)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (username, title, body, "/user/requests", f"overtime-review-{date}", now_ts, "overtime_review"))
        db.commit()
        
        app.logger.info(f"Notifica revisione Extra Turno inviata a {username}")
    except Exception as e:
        app.logger.error(f"Errore invio notifica revisione Extra Turno: {e}")


if __name__ == "__main__":
    init_db()
    
    # HTTPS per permettere accesso camera da mobile in rete locale
    # Usa: python app.py --https
    import sys
    if "--https" in sys.argv:
        print("🔐 Avvio in modalità HTTPS (per accesso camera da mobile)")
        print("⚠️  Il browser mostrerà un avviso di sicurezza - clicca 'Avanzate' → 'Procedi'")
        app.run(host="0.0.0.0", port=5000, debug=True, use_reloader=False, ssl_context='adhoc')
    else:
        app.run(host="0.0.0.0", port=5000, debug=True, use_reloader=False)
