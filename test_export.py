import io
import requests
from openpyxl import load_workbook

url = "http://127.0.0.1:5000/api/export?format=excel"
response = requests.get(url)
print("status", response.status_code)
print("filename", response.headers.get("Content-Disposition"))
wb = load_workbook(io.BytesIO(response.content))
ws = wb.active
header = [ws.cell(row=5, column=i).value for i in range(1, 11)]
print("header_row", header)
print("max_row", ws.max_row)
